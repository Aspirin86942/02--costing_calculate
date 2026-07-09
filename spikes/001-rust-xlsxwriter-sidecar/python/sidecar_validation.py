from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EMPTY_VALUES = (None, '')
NUMERIC_TOLERANCE = Decimal('0.000001')
MAX_ERROR_DETAILS = 200


def validate_workbooks(
    python_workbook_path: Path,
    rust_workbook_path: Path,
    manifest: dict[str, Any],
) -> dict[str, Any]:
    python_workbook = load_workbook(python_workbook_path, data_only=False)
    rust_workbook = load_workbook(rust_workbook_path, data_only=False)
    errors: list[str] = []
    error_count = 0
    checked_cells = 0
    number_format_matched = True
    shape_matched = True
    auto_filter_matched = True

    def add_error(message: str) -> None:
        nonlocal error_count
        error_count += 1
        if len(errors) < MAX_ERROR_DETAILS:
            errors.append(message)

    manifest_sheets = manifest.get('sheets', [])
    expected_sheet_names = [sheet['sheet_name'] for sheet in manifest_sheets]
    if python_workbook.sheetnames != expected_sheet_names:
        add_error(f'python sheet order mismatch: {python_workbook.sheetnames} != {expected_sheet_names}')
    if rust_workbook.sheetnames != expected_sheet_names:
        add_error(f'rust sheet order mismatch: {rust_workbook.sheetnames} != {expected_sheet_names}')

    for sheet_manifest in manifest_sheets:
        sheet_name = sheet_manifest['sheet_name']
        if sheet_name not in python_workbook.sheetnames or sheet_name not in rust_workbook.sheetnames:
            add_error(f'missing sheet: {sheet_name}')
            continue

        python_sheet = python_workbook[sheet_name]
        rust_sheet = rust_workbook[sheet_name]
        columns = list(sheet_manifest.get('columns', []))
        column_types = dict(sheet_manifest.get('column_types', {}))
        number_formats = dict(sheet_manifest.get('number_formats', {}))
        write_types = dict(sheet_manifest.get('write_types', {}))
        expected_row_count = sheet_manifest.get('row_count')
        expected_column_count = sheet_manifest.get('column_count', len(columns))
        expected_max_row = expected_row_count + 1 if isinstance(expected_row_count, int) else None
        expected_max_column = expected_column_count if isinstance(expected_column_count, int) else len(columns)
        max_row = max(python_sheet.max_row, rust_sheet.max_row)

        if python_sheet.max_row != rust_sheet.max_row:
            shape_matched = False
            add_error(f'row count mismatch: sheet={sheet_name} python={python_sheet.max_row} rust={rust_sheet.max_row}')
        if python_sheet.max_column != rust_sheet.max_column:
            shape_matched = False
            add_error(
                f'column count mismatch: sheet={sheet_name} '
                f'python={python_sheet.max_column} rust={rust_sheet.max_column}'
            )
        if expected_max_row is not None:
            if python_sheet.max_row != expected_max_row:
                shape_matched = False
                add_error(
                    f'python row count does not match manifest: sheet={sheet_name} '
                    f'expected={expected_max_row} actual={python_sheet.max_row}'
                )
            if rust_sheet.max_row != expected_max_row:
                shape_matched = False
                add_error(
                    f'rust row count does not match manifest: sheet={sheet_name} '
                    f'expected={expected_max_row} actual={rust_sheet.max_row}'
                )
        if expected_max_column is not None:
            if python_sheet.max_column != expected_max_column:
                shape_matched = False
                add_error(
                    f'python column count does not match manifest: sheet={sheet_name} '
                    f'expected={expected_max_column} actual={python_sheet.max_column}'
                )
            if rust_sheet.max_column != expected_max_column:
                shape_matched = False
                add_error(
                    f'rust column count does not match manifest: sheet={sheet_name} '
                    f'expected={expected_max_column} actual={rust_sheet.max_column}'
                )

        for row_number in range(1, max_row + 1):
            for column_index, column_name in enumerate(columns, start=1):
                checked_cells += 1
                column_type = (
                    'text' if row_number == 1 else write_types.get(column_name, column_types.get(column_name, 'text'))
                )
                left = python_sheet.cell(row=row_number, column=column_index).value
                right = rust_sheet.cell(row=row_number, column=column_index).value
                if not cell_values_equivalent(left, right, column_type=column_type):
                    add_error(
                        f'value mismatch: sheet={sheet_name} row={row_number} col={column_name} '
                        f'python={left!r} rust={right!r}'
                    )

        for column_index, column_name in enumerate(columns, start=1):
            expected_format = number_formats.get(column_name)
            if expected_format is None:
                continue
            for row_number in range(2, max_row + 1):
                rust_cell = rust_sheet.cell(row=row_number, column=column_index)
                # 空单元格通常不会持久化格式，跳过避免把样式差异误判为数据错误。
                if rust_cell.value in EMPTY_VALUES:
                    continue
                if rust_cell.number_format != expected_format:
                    number_format_matched = False
                    add_error(
                        f'number format mismatch: sheet={sheet_name} row={row_number} col={column_name} '
                        f'expected={expected_format!r} actual={rust_cell.number_format!r}'
                    )

        expected_freeze_panes = sheet_manifest.get('freeze_panes')
        if expected_freeze_panes is not None and rust_sheet.freeze_panes != expected_freeze_panes:
            add_error(
                f'freeze panes mismatch: sheet={sheet_name} expected={expected_freeze_panes!r} '
                f'actual={rust_sheet.freeze_panes!r}'
            )

        if sheet_manifest.get('auto_filter'):
            expected_filter_ref = _expected_auto_filter_ref(
                column_count=len(columns),
                row_count=expected_row_count,
                fallback_max_row=max_row,
            )
            if rust_sheet.auto_filter.ref != expected_filter_ref:
                auto_filter_matched = False
                add_error(
                    f'auto filter mismatch: sheet={sheet_name} expected={expected_filter_ref!r} '
                    f'actual={rust_sheet.auto_filter.ref!r}'
                )

    return {
        'passed': error_count == 0,
        'sheet_count': len(manifest_sheets),
        'checked_cells': checked_cells,
        'number_format_matched': number_format_matched,
        'shape_matched': shape_matched,
        'auto_filter_matched': auto_filter_matched,
        'error_count': error_count,
        'errors': errors,
    }


def cell_values_equivalent(left: Any, right: Any, *, column_type: str) -> bool:
    if left in EMPTY_VALUES and right in EMPTY_VALUES:
        return True
    if left in EMPTY_VALUES or right in EMPTY_VALUES:
        return False
    if _is_numeric_column(column_type):
        left_decimal = _to_decimal(left)
        right_decimal = _to_decimal(right)
        if left_decimal is None or right_decimal is None:
            return False
        return abs(left_decimal - right_decimal) <= NUMERIC_TOLERANCE
    return str(left) == str(right)


def _is_numeric_column(column_type: str) -> bool:
    return column_type not in {'text', 'string', 'date', 'datetime'}


def _expected_auto_filter_ref(*, column_count: int, row_count: Any, fallback_max_row: int) -> str | None:
    if column_count <= 0:
        return None
    data_max_row = row_count + 1 if isinstance(row_count, int) else fallback_max_row
    last_row = max(data_max_row, 2)
    return f'A1:{get_column_letter(column_count)}{last_row}'


def _to_decimal(value: Any) -> Decimal | None:
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
