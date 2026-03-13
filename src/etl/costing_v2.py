"""ETL pipeline for costing workbooks."""

import logging
import sys
from decimal import Decimal
from numbers import Real
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

try:
    from src.analytics.pq_analysis import (
        QTY_DL_AMOUNT,
        QTY_DL_UNIT_COST,
        QTY_DM_AMOUNT,
        QTY_DM_UNIT_COST,
        QTY_MOH_AMOUNT,
        QTY_MOH_CONSUMABLES_AMOUNT,
        QTY_MOH_CONSUMABLES_UNIT_COST,
        QTY_MOH_DEPRECIATION_AMOUNT,
        QTY_MOH_DEPRECIATION_UNIT_COST,
        QTY_MOH_LABOR_AMOUNT,
        QTY_MOH_LABOR_UNIT_COST,
        QTY_MOH_OTHER_AMOUNT,
        QTY_MOH_OTHER_UNIT_COST,
        QTY_MOH_UNIT_COST,
        QTY_MOH_UTILITIES_AMOUNT,
        QTY_MOH_UTILITIES_UNIT_COST,
        FlatSheet,
        ProductAnomalySection,
        SectionBlock,
        build_report_artifacts,
        render_tables,
    )
    from src.config.settings import GB_PROCESSED_DIR, GB_RAW_DIR, ensure_directories
    from src.etl.utils import clean_column_name, format_period_col
except ModuleNotFoundError:
    # 直接执行 src/etl/costing_v2.py 时，解释器搜索路径不含项目根目录，补齐后重试导入。
    project_root = Path(__file__).resolve().parents[2]
    project_root_str = str(project_root)
    if project_root_str not in sys.path:
        sys.path.insert(0, project_root_str)
    from src.analytics.pq_analysis import (
        QTY_DL_AMOUNT,
        QTY_DL_UNIT_COST,
        QTY_DM_AMOUNT,
        QTY_DM_UNIT_COST,
        QTY_MOH_AMOUNT,
        QTY_MOH_CONSUMABLES_AMOUNT,
        QTY_MOH_CONSUMABLES_UNIT_COST,
        QTY_MOH_DEPRECIATION_AMOUNT,
        QTY_MOH_DEPRECIATION_UNIT_COST,
        QTY_MOH_LABOR_AMOUNT,
        QTY_MOH_LABOR_UNIT_COST,
        QTY_MOH_OTHER_AMOUNT,
        QTY_MOH_OTHER_UNIT_COST,
        QTY_MOH_UNIT_COST,
        QTY_MOH_UTILITIES_AMOUNT,
        QTY_MOH_UTILITIES_UNIT_COST,
        FlatSheet,
        ProductAnomalySection,
        SectionBlock,
        build_report_artifacts,
        render_tables,
    )
    from src.config.settings import GB_PROCESSED_DIR, GB_RAW_DIR, ensure_directories
    from src.etl.utils import clean_column_name, format_period_col

logger = logging.getLogger(__name__)


COL_PERIOD = '年期'
COL_MONTH = '月份'
COL_COST_CENTER = '成本中心名称'
COL_PRODUCT_CODE = '产品编码'
COL_PRODUCT_NAME = '产品名称'
COL_SPEC = '规格型号'
COL_ORDER_NO = '工单编号'
COL_ORDER_LINE = '工单行号'
COL_VENDOR_CODE = '供应商编码'
COL_VENDOR_NAME = '供应商名称'
COL_UNIT = '基本单位'
COL_PLAN_QTY = '计划产量'
COL_PRODUCTION_TYPE = '生产类型'
COL_DOC_TYPE = '单据类型'
COL_COST_ITEM = '成本项目名称'
COL_CHILD_MATERIAL = '子项物料编码'
COL_CHILD_MATERIAL_NAME = '子项物料名称'
COL_FILLED_COST_ITEM = 'Filled_成本项目'
COL_OPENING_WIP_QTY = '期初在产品数量'
COL_OPENING_WIP_AMOUNT = '期初在产品金额'
COL_OPENING_ADJUST_QTY = '期初调整数量'
COL_OPENING_ADJUST_AMOUNT = '期初调整金额'
COL_CURRENT_INPUT_QTY = '本期投入数量'
COL_CURRENT_INPUT_AMOUNT = '本期投入金额'
COL_CUMULATIVE_INPUT_QTY = '累计投入数量'
COL_CUMULATIVE_INPUT_AMOUNT = '累计投入金额'
COL_ENDING_WIP_QTY = '期末在产品数量'
COL_ENDING_WIP_AMOUNT = '期末在产品金额'
COL_CURRENT_COMPLETED_QTY = '本期完工数量'
COL_CURRENT_COMPLETED_CONSUMPTION = '本期完工单耗'
COL_CURRENT_COMPLETED_UNIT_COST = '本期完工单位成本'
COL_CURRENT_COMPLETED_AMOUNT = '本期完工金额'
COL_CUMULATIVE_COMPLETED_QTY = '累计完工数量'
COL_CUMULATIVE_COMPLETED_CONSUMPTION = '累计完工单耗'
COL_CUMULATIVE_COMPLETED_UNIT_COST = '累计完工单位成本'
COL_CUMULATIVE_COMPLETED_AMOUNT = '累计完工金额'
INTEGRATED_WORKSHOP_NAME = '集成车间'
ANALYSIS_PRODUCT_ORDER: tuple[tuple[str, str], ...] = (
    ('GB_C.D.B0048AA', 'BMS-400W驱动器'),
    ('GB_C.D.B0040AA', 'BMS-750W驱动器'),
    ('GB_C.D.B0041AA', 'BMS-1100W驱动器'),
    ('GB_C.D.B0042AA', 'BMS-1700W驱动器'),
    ('GB_C.D.B0043AA', 'BMS-2400W驱动器'),
    ('GB_C.D.B0044AA', 'BMS-3900W驱动器'),
    ('GB_C.D.B0045AA', 'BMS-5900W驱动器'),
    ('GB_C.D.B0046AA', 'BMS-7500W驱动器'),
)
ANALYSIS_PRODUCT_WHITELIST: set[tuple[str, str]] = set(ANALYSIS_PRODUCT_ORDER)
EXCEL_TWO_DECIMAL_FORMAT = '#,##0.00'
EXCEL_INTEGER_FORMAT = '#,##0'
EXCEL_SCORE_FORMAT = '0.0000'
EXCEL_PERCENT_FORMAT = '0.00%'
DETAIL_TWO_DECIMAL_COLUMNS = {
    COL_CURRENT_COMPLETED_UNIT_COST,
    COL_CURRENT_COMPLETED_AMOUNT,
}
QTY_TWO_DECIMAL_COLUMNS = {
    COL_CURRENT_COMPLETED_UNIT_COST,
    COL_CURRENT_COMPLETED_AMOUNT,
    QTY_DM_AMOUNT,
    QTY_DL_AMOUNT,
    QTY_MOH_AMOUNT,
    QTY_MOH_OTHER_AMOUNT,
    QTY_MOH_LABOR_AMOUNT,
    QTY_MOH_CONSUMABLES_AMOUNT,
    QTY_MOH_DEPRECIATION_AMOUNT,
    QTY_MOH_UTILITIES_AMOUNT,
    QTY_DM_UNIT_COST,
    QTY_DL_UNIT_COST,
    QTY_MOH_UNIT_COST,
    QTY_MOH_OTHER_UNIT_COST,
    QTY_MOH_LABOR_UNIT_COST,
    QTY_MOH_CONSUMABLES_UNIT_COST,
    QTY_MOH_DEPRECIATION_UNIT_COST,
    QTY_MOH_UTILITIES_UNIT_COST,
}
WORK_ORDER_HIGHLIGHT_COLUMNS: tuple[tuple[str, str], ...] = (
    ('直接材料单位完工成本', '直接材料异常标记'),
    ('直接人工单位完工成本', '直接人工异常标记'),
    ('制造费用单位完工成本', '制造费用异常标记'),
    ('制造费用_其他单位完工成本', '制造费用_其他异常标记'),
    ('制造费用_人工单位完工成本', '制造费用_人工异常标记'),
    ('制造费用_机物料及低耗单位完工成本', '制造费用_机物料及低耗异常标记'),
    ('制造费用_折旧单位完工成本', '制造费用_折旧异常标记'),
    ('制造费用_水电费单位完工成本', '制造费用_水电费异常标记'),
)


class CostingETL:
    """Process a costing workbook into detail/quantity sheets."""

    FILL_COLS = [
        COL_PERIOD,
        COL_COST_CENTER,
        COL_PRODUCT_CODE,
        COL_PRODUCT_NAME,
        COL_SPEC,
        COL_ORDER_NO,
        COL_ORDER_LINE,
        COL_VENDOR_CODE,
        COL_VENDOR_NAME,
        COL_UNIT,
        COL_PLAN_QTY,
        COL_PRODUCTION_TYPE,
        COL_DOC_TYPE,
    ]

    DETAIL_COLS = [
        COL_PERIOD,
        COL_MONTH,
        COL_COST_CENTER,
        COL_PRODUCT_CODE,
        COL_PRODUCT_NAME,
        COL_SPEC,
        COL_PRODUCTION_TYPE,
        COL_DOC_TYPE,
        COL_ORDER_NO,
        COL_ORDER_LINE,
        COL_VENDOR_CODE,
        COL_VENDOR_NAME,
        COL_UNIT,
        COL_PLAN_QTY,
        COL_COST_ITEM,
        COL_CHILD_MATERIAL,
        COL_CHILD_MATERIAL_NAME,
        COL_OPENING_WIP_QTY,
        COL_OPENING_WIP_AMOUNT,
        COL_OPENING_ADJUST_QTY,
        COL_OPENING_ADJUST_AMOUNT,
        COL_CURRENT_INPUT_QTY,
        COL_CURRENT_INPUT_AMOUNT,
        COL_CUMULATIVE_INPUT_QTY,
        COL_CUMULATIVE_INPUT_AMOUNT,
        COL_ENDING_WIP_QTY,
        COL_ENDING_WIP_AMOUNT,
        COL_CURRENT_COMPLETED_QTY,
        COL_CURRENT_COMPLETED_CONSUMPTION,
        COL_CURRENT_COMPLETED_UNIT_COST,
        COL_CURRENT_COMPLETED_AMOUNT,
        COL_CUMULATIVE_COMPLETED_QTY,
        COL_CUMULATIVE_COMPLETED_CONSUMPTION,
        COL_CUMULATIVE_COMPLETED_UNIT_COST,
        COL_CUMULATIVE_COMPLETED_AMOUNT,
    ]

    QTY_COLS = [
        COL_PERIOD,
        COL_MONTH,
        COL_COST_CENTER,
        COL_PRODUCT_CODE,
        COL_PRODUCT_NAME,
        COL_SPEC,
        COL_PRODUCTION_TYPE,
        COL_DOC_TYPE,
        COL_ORDER_NO,
        COL_ORDER_LINE,
        COL_UNIT,
        COL_PLAN_QTY,
        COL_OPENING_WIP_QTY,
        COL_OPENING_WIP_AMOUNT,
        COL_CURRENT_INPUT_QTY,
        COL_CURRENT_INPUT_AMOUNT,
        COL_CUMULATIVE_INPUT_QTY,
        COL_CUMULATIVE_INPUT_AMOUNT,
        COL_ENDING_WIP_QTY,
        COL_ENDING_WIP_AMOUNT,
        COL_CURRENT_COMPLETED_QTY,
        COL_CURRENT_COMPLETED_CONSUMPTION,
        COL_CURRENT_COMPLETED_UNIT_COST,
        COL_CURRENT_COMPLETED_AMOUNT,
        COL_CUMULATIVE_COMPLETED_QTY,
        COL_CUMULATIVE_COMPLETED_CONSUMPTION,
        COL_CUMULATIVE_COMPLETED_UNIT_COST,
        COL_CUMULATIVE_COMPLETED_AMOUNT,
    ]

    def __init__(self, skip_rows: int = 2):
        self.skip_rows = skip_rows
        ensure_directories()

    @staticmethod
    def _to_excel_number(value: object) -> object:
        """把 Decimal/数值对象转成 Excel 可稳定识别的 number 类型。"""
        if value is None or isinstance(value, bool):
            return value
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, Real):
            return value if pd.isna(value) else float(value)
        return value

    @staticmethod
    def _resolve_metric_number_format(metric_type: str, *, qty_format: str = EXCEL_INTEGER_FORMAT) -> str | None:
        if metric_type in {'amount', 'price'}:
            return EXCEL_TWO_DECIMAL_FORMAT
        if metric_type == 'qty':
            return qty_format
        if metric_type == 'score':
            return EXCEL_SCORE_FORMAT
        if metric_type == 'pct':
            return EXCEL_PERCENT_FORMAT
        return None

    def _coerce_excel_numeric_columns(self, df: pd.DataFrame, numeric_columns: set[str]) -> pd.DataFrame:
        """写出前统一转成数值单元格，避免 Decimal 被 Excel 当作文本。"""
        write_df = df.copy()
        for column_name in numeric_columns:
            if column_name in write_df.columns:
                write_df[column_name] = write_df[column_name].map(self._to_excel_number)
        return write_df

    def _apply_basic_sheet_styles(
        self,
        worksheet: Worksheet,
        columns: list[str],
        *,
        column_formats: dict[str, str],
        freeze_panes: str,
        fixed_width: int | None = None,
    ) -> None:
        header_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        for col_idx, column_name in enumerate(columns, start=1):
            header_cell = worksheet.cell(1, col_idx)
            header_cell.fill = header_fill
            header_cell.font = Font(bold=True)
            header_cell.alignment = align_center
            header_cell.border = border

            number_format = column_formats.get(column_name)
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell.border = border
                if number_format is None:
                    cell.alignment = align_left
                    continue

                cell.alignment = align_right
                cell.number_format = number_format

        worksheet.freeze_panes = freeze_panes
        if worksheet.max_column > 0:
            worksheet.auto_filter.ref = f'A1:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, 1)}'

        for col_idx in range(1, worksheet.max_column + 1):
            if fixed_width is not None:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = fixed_width
                continue

            max_length = 0
            for row_idx in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row_idx, col_idx).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_length + 2), 24)

    def _write_dataframe_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        *,
        numeric_columns: set[str],
        freeze_panes: str = 'A2',
        fixed_width: int | None = None,
    ) -> Worksheet:
        """写入普通 DataFrame sheet，并按列名套用数值格式。"""
        column_formats = {
            column_name: EXCEL_TWO_DECIMAL_FORMAT for column_name in numeric_columns if column_name in df.columns
        }
        write_df = self._coerce_excel_numeric_columns(df, set(column_formats))
        write_df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        self._apply_basic_sheet_styles(
            worksheet,
            write_df.columns.tolist(),
            column_formats=column_formats,
            freeze_panes=freeze_panes,
            fixed_width=fixed_width,
        )
        return worksheet

    def _apply_work_order_highlights(self, worksheet: Worksheet) -> None:
        """把异常颜色挂在“值列+标记列”上，降低业务人员定位异常字段的成本。"""
        header_map = {
            worksheet.cell(1, col_idx).value: col_idx
            for col_idx in range(1, worksheet.max_column + 1)
            if worksheet.cell(1, col_idx).value is not None
        }
        highlight_styles = {
            '关注': {
                'fill': PatternFill(fill_type='solid', fgColor='DDEBF7'),
                'font': None,
            },
            '高度可疑': {
                'fill': PatternFill(fill_type='solid', fgColor='4472C4'),
                'font': Font(color='FFFFFF'),
            },
        }

        for value_column, flag_column in WORK_ORDER_HIGHLIGHT_COLUMNS:
            value_idx = header_map.get(value_column)
            flag_idx = header_map.get(flag_column)
            if value_idx is None or flag_idx is None:
                logger.warning(
                    'Skip work-order highlight: missing columns sheet=%s value=%s flag=%s',
                    worksheet.title,
                    value_column,
                    flag_column,
                )
                continue

            for row_idx in range(2, worksheet.max_row + 1):
                flag_value = worksheet.cell(row_idx, flag_idx).value
                style = highlight_styles.get(str(flag_value).strip()) if flag_value is not None else None
                if style is None:
                    continue

                for col_idx in (value_idx, flag_idx):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.fill = style['fill']
                    if style['font'] is not None:
                        cell.font = style['font']

    def _auto_rename_columns(self, df: pd.DataFrame) -> dict[str, str]:
        """Infer key columns when source headers vary."""
        col_map: dict[str, str] = {}

        if COL_CHILD_MATERIAL not in df.columns:
            possible = [c for c in df.columns if '物料编码' in c or '子件' in c]
            if possible:
                col_map[possible[0]] = COL_CHILD_MATERIAL
                logger.info('Column rename: %s -> %s', possible[0], COL_CHILD_MATERIAL)

        if COL_COST_ITEM not in df.columns:
            possible = [c for c in df.columns if '成本项目' in c or '费用项目' in c]
            if possible:
                col_map[possible[0]] = COL_COST_ITEM
                logger.info('Column rename: %s -> %s', possible[0], COL_COST_ITEM)

        return col_map

    def _remove_total_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Drop summary rows containing '合计'."""
        initial_rows = len(df)
        cols_to_check = [c for c in df.columns[:3] if c in [COL_PERIOD, COL_COST_CENTER]]
        if not cols_to_check:
            return df

        mask_keep = pd.Series([True] * len(df), index=df.index)
        for col in cols_to_check:
            is_total = df[col].astype(str).str.contains('合计', na=False)
            mask_keep &= ~is_total

        result = df[mask_keep].copy()
        removed = initial_rows - len(result)
        if removed > 0:
            logger.info('Removed total rows: %s', removed)
        return result

    def _forward_fill_with_rules(self, df_raw: pd.DataFrame) -> pd.DataFrame:
        """按业务规则执行向下填充。"""
        df_filled = df_raw.copy()
        cols_to_fill = [c for c in df_filled.columns if c in self.FILL_COLS]
        if not cols_to_fill:
            return df_filled

        vendor_cols = [c for c in [COL_VENDOR_CODE, COL_VENDOR_NAME] if c in cols_to_fill]
        normal_fill_cols = [c for c in cols_to_fill if c not in vendor_cols]
        if normal_fill_cols:
            df_filled[normal_fill_cols] = df_filled[normal_fill_cols].ffill()

        if not vendor_cols:
            return df_filled
        if COL_COST_CENTER not in df_filled.columns:
            df_filled[vendor_cols] = df_filled[vendor_cols].ffill()
            return df_filled

        vendor_filled = df_filled[vendor_cols].ffill()
        # 集成车间供应商字段必须保留原值，避免把上方工单供应商错误继承到当前行。
        integrated_mask = df_filled[COL_COST_CENTER].astype(str).str.strip().eq(INTEGRATED_WORKSHOP_NAME)
        df_filled.loc[~integrated_mask, vendor_cols] = vendor_filled.loc[~integrated_mask, vendor_cols]
        return df_filled

    def _filter_fact_df_for_analysis(self, fact_df: pd.DataFrame) -> pd.DataFrame:
        """按白名单过滤分析数据，仅输出目标产品。"""
        return self._filter_dataframe_by_whitelist(
            fact_df,
            code_col='product_code',
            name_col='product_name',
            sort_cols=['period', 'cost_bucket'],
        )

    def _filter_dataframe_by_whitelist(
        self,
        df: pd.DataFrame,
        *,
        code_col: str,
        name_col: str,
        sort_cols: list[str] | None = None,
    ) -> pd.DataFrame:
        """按产品白名单过滤任意分析 DataFrame。"""
        actual_required_cols = {code_col, name_col}
        if not actual_required_cols.issubset(df.columns):
            missing_cols = sorted(actual_required_cols - set(df.columns))
            logger.warning('Skip analysis whitelist filter: missing columns=%s', missing_cols)
            return df

        if df.empty:
            return df

        product_pairs = pd.MultiIndex.from_frame(df[[code_col, name_col]].astype(str))
        matched_mask = product_pairs.isin(ANALYSIS_PRODUCT_WHITELIST)
        filtered_df = df.loc[matched_mask].copy()
        order_map = {pair: idx for idx, pair in enumerate(ANALYSIS_PRODUCT_ORDER)}
        filtered_pairs = pd.MultiIndex.from_frame(filtered_df[[code_col, name_col]].astype(str))
        filtered_df['_order_idx'] = filtered_pairs.map(order_map)
        order_cols = ['_order_idx']
        if sort_cols:
            order_cols.extend([col for col in sort_cols if col in filtered_df.columns])
        filtered_df = filtered_df.sort_values(order_cols).drop(columns=['_order_idx'])

        logger.info(
            'Analysis product whitelist filter applied: rows %s -> %s, products %s -> %s',
            len(df),
            len(filtered_df),
            df[[code_col, name_col]].drop_duplicates().shape[0],
            filtered_df[[code_col, name_col]].drop_duplicates().shape[0],
        )
        return filtered_df

    def _filter_product_anomaly_sections(
        self,
        sections: list[ProductAnomalySection],
    ) -> list[ProductAnomalySection]:
        """按白名单和既定顺序过滤兼容摘要分段。"""
        order_map = {pair: idx for idx, pair in enumerate(ANALYSIS_PRODUCT_ORDER)}
        filtered_sections = [
            section
            for section in sections
            if (str(section.product_code), str(section.product_name)) in ANALYSIS_PRODUCT_WHITELIST
        ]
        return sorted(
            filtered_sections,
            key=lambda section: order_map[(str(section.product_code), str(section.product_name))],
        )

    def _split_sheets(
        self,
        df_raw: pd.DataFrame,
        df_filled: pd.DataFrame,
        target_mat: str,
        target_item: str,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Split source rows into detail and quantity sheets."""
        cond_no_material = df_raw[target_mat].isna() | (df_raw[target_mat].astype(str).str.strip() == '')

        if target_item in df_raw.columns:
            cond_no_cost_item = df_raw[target_item].isna() | (df_raw[target_item].astype(str).str.strip() == '')
        else:
            cond_no_cost_item = True

        if COL_ORDER_NO in df_filled.columns:
            cond_has_order = df_filled[COL_ORDER_NO].notna()
        else:
            cond_has_order = True

        df_qty = df_filled[cond_no_material & cond_no_cost_item & cond_has_order].copy()
        df_qty = format_period_col(df_qty)
        qty_cols_final = [c for c in self.QTY_COLS if c in df_qty.columns]
        if qty_cols_final:
            df_qty = df_qty[qty_cols_final]

        cond_is_material = df_filled[target_mat].notna() & (df_filled[target_mat].astype(str).str.strip() != '')

        if target_item in df_raw.columns:
            cond_is_expense = (
                df_raw[target_mat].isna()
                & df_raw[target_item].notna()
                & (df_raw[target_item].astype(str).str.strip() != '直接材料')
            )
        else:
            cond_is_expense = False

        df_detail = df_filled[cond_is_material | cond_is_expense].copy()
        if COL_FILLED_COST_ITEM in df_detail.columns and target_item in df_detail.columns:
            df_detail[target_item] = df_detail[COL_FILLED_COST_ITEM]

        df_detail = format_period_col(df_detail)
        detail_cols_final = [c for c in self.DETAIL_COLS if c in df_detail.columns]
        if detail_cols_final:
            df_detail = df_detail[detail_cols_final]

        return df_detail, df_qty

    def _write_analysis_sheet(self, writer: pd.ExcelWriter, sheet_name: str, sections: list[SectionBlock]) -> None:
        """写入三段分析块并应用样式；禁止合并单元格。"""
        start_row = 0
        section_meta: list[dict[str, int | str | bool]] = []

        for section in sections:
            write_section_df = section.data.copy()
            if section.metric_type in {'amount', 'price', 'qty'}:
                numeric_columns = [
                    column_name
                    for column_name in write_section_df.columns
                    if column_name not in {'产品编码', '产品名称'}
                ]
                for column_name in numeric_columns:
                    write_section_df[column_name] = write_section_df[column_name].map(self._to_excel_number)

            title_frame = pd.DataFrame([[section.title]])
            title_frame.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=start_row,
            )
            write_section_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row + 1)

            title_row = start_row + 1
            header_row = start_row + 2
            data_start = start_row + 3
            data_end = data_start + len(write_section_df) - 1
            max_col = max(1, write_section_df.shape[1])

            section_meta.append(
                {
                    'title_row': title_row,
                    'header_row': header_row,
                    'data_start': data_start,
                    'data_end': data_end,
                    'max_col': max_col,
                    'metric_type': section.metric_type,
                    'has_total_row': section.has_total_row,
                }
            )
            start_row += len(write_section_df) + 3

        worksheet = writer.sheets[sheet_name]
        self._style_analysis_sheet(worksheet, section_meta)

    def _style_analysis_sheet(self, worksheet: Worksheet, section_meta: list[dict[str, int | str | bool]]) -> None:
        if not section_meta:
            return

        title_fill = PatternFill(fill_type='solid', fgColor='FFD966')
        header_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        total_fill = PatternFill(fill_type='solid', fgColor='BDD7EE')
        border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        max_col_overall = max(int(meta['max_col']) for meta in section_meta)

        for meta in section_meta:
            title_row = int(meta['title_row'])
            header_row = int(meta['header_row'])
            data_start = int(meta['data_start'])
            data_end = int(meta['data_end'])
            max_col = int(meta['max_col'])
            metric_type = str(meta['metric_type'])
            has_total_row = bool(meta['has_total_row'])

            title_cell = worksheet.cell(title_row, 1)
            title_cell.fill = title_fill
            title_cell.font = Font(bold=True)
            title_cell.alignment = align_left

            for col in range(1, max_col + 1):
                header_cell = worksheet.cell(header_row, col)
                header_cell.fill = header_fill
                header_cell.font = Font(bold=True)
                header_cell.alignment = align_center
                header_cell.border = border

            if data_end < data_start:
                continue

            for row in range(data_start, data_end + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row, col)
                    cell.border = border
                    if col <= 2:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_right
                        number_format = self._resolve_metric_number_format(
                            metric_type,
                            qty_format=EXCEL_TWO_DECIMAL_FORMAT,
                        )
                        if number_format is not None:
                            cell.number_format = number_format

            if has_total_row:
                total_row = data_end
                for col in range(1, max_col + 1):
                    total_cell = worksheet.cell(total_row, col)
                    total_cell.fill = total_fill
                    total_cell.font = Font(bold=True)

        first_meta = section_meta[0]
        worksheet.freeze_panes = 'C3'
        filter_end = int(first_meta['data_end'])
        if filter_end < int(first_meta['header_row']):
            filter_end = int(first_meta['header_row'])
        worksheet.auto_filter.ref = (
            f'A{int(first_meta["header_row"])}:{get_column_letter(int(first_meta["max_col"]))}{filter_end}'
        )

        for col in range(1, max_col_overall + 1):
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row, col).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            width = min(max(12, max_length + 2), 40)
            worksheet.column_dimensions[get_column_letter(col)].width = width

    def _write_flat_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        table: FlatSheet,
        *,
        freeze_panes: str = 'A2',
        fixed_width: int | None = None,
    ) -> Worksheet:
        """写入平铺数据 sheet 并应用基础样式。"""
        write_df = table.data.copy()
        column_formats: dict[str, str] = {}
        for column_name, metric_type in table.column_types.items():
            if column_name not in write_df.columns:
                continue
            if metric_type in {'amount', 'price', 'qty', 'score', 'pct'}:
                write_df[column_name] = write_df[column_name].map(self._to_excel_number)
            number_format = self._resolve_metric_number_format(metric_type)
            if number_format is not None:
                column_formats[column_name] = number_format

        write_df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        self._apply_basic_sheet_styles(
            worksheet,
            write_df.columns.tolist(),
            column_formats=column_formats,
            freeze_panes=freeze_panes,
            fixed_width=fixed_width,
        )
        return worksheet

    def _write_product_anomaly_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        sections: list[ProductAnomalySection],
    ) -> None:
        """写入按产品异常值分析页（不合并单元格）。"""
        worksheet = writer.book.create_sheet(title=sheet_name)
        writer.sheets[sheet_name] = worksheet

        title_fill = PatternFill(fill_type='solid', fgColor='FFD966')
        header_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        meta_fill = PatternFill(fill_type='solid', fgColor='B4C6E7')
        border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        worksheet.cell(1, 1, '四、按单个产品异常值分析')
        worksheet.cell(1, 1).fill = title_fill
        worksheet.cell(1, 1).font = Font(bold=True)
        worksheet.cell(1, 1).alignment = align_left

        current_row = 3
        max_col_overall = 1
        filter_set = False

        for section in sections:
            meta_header_row = current_row
            meta_value_row = current_row + 1
            table_header_row = current_row + 2
            data_start_row = current_row + 3
            data_end_row = data_start_row + len(section.data) - 1
            max_col = len(section.data.columns)
            max_col_overall = max(max_col_overall, max_col)

            worksheet.cell(meta_header_row, 1, '产品编码')
            worksheet.cell(meta_header_row, 2, '产品名称')
            worksheet.cell(meta_value_row, 1, section.product_code)
            worksheet.cell(meta_value_row, 2, section.product_name)

            for col in [1, 2]:
                header_cell = worksheet.cell(meta_header_row, col)
                header_cell.fill = header_fill
                header_cell.font = Font(bold=True)
                header_cell.alignment = align_center
                header_cell.border = border

                value_cell = worksheet.cell(meta_value_row, col)
                value_cell.fill = meta_fill
                value_cell.font = Font(bold=True)
                value_cell.alignment = align_left
                value_cell.border = border

            for col_idx, column_name in enumerate(section.data.columns, start=1):
                header_cell = worksheet.cell(table_header_row, col_idx, column_name)
                header_cell.fill = header_fill
                header_cell.font = Font(bold=True)
                header_cell.alignment = align_center
                header_cell.border = border

            for row_idx, row_data in section.data.iterrows():
                excel_row = data_start_row + row_idx
                for col_idx, column_name in enumerate(section.data.columns, start=1):
                    value = row_data[column_name]
                    cell = worksheet.cell(excel_row, col_idx, value)
                    cell.border = border
                    cell.alignment = align_left if col_idx == 1 else align_right

                    metric_type = section.column_types.get(column_name)
                    number_format = self._resolve_metric_number_format(metric_type or '')
                    if number_format is not None:
                        cell.number_format = number_format

            if not filter_set:
                worksheet.auto_filter.ref = (
                    f'A{table_header_row}:{get_column_letter(max_col)}{max(data_end_row, table_header_row)}'
                )
                filter_set = True

            current_row = data_end_row + 2

        worksheet.freeze_panes = 'A6'
        for col in range(1, max_col_overall + 1):
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row, col).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            width = 15
            worksheet.column_dimensions[get_column_letter(col)].width = width

    def process_file(self, input_path: Path, output_path: Path) -> bool:
        """Read one workbook and write split output workbook."""
        try:
            logger.info('Processing file: %s', input_path)
            df_raw = pd.read_excel(input_path, header=[0, 1], skiprows=self.skip_rows, engine='openpyxl')
            logger.info('Loaded rows=%s, cols=%s', len(df_raw), len(df_raw.columns))

            df_raw.columns = [clean_column_name(c) for c in df_raw.columns]
            col_map = self._auto_rename_columns(df_raw)
            if col_map:
                df_raw.rename(columns=col_map, inplace=True)

            target_mat = COL_CHILD_MATERIAL
            target_item = COL_COST_ITEM

            if target_mat not in df_raw.columns:
                logger.error("Missing required column '%s'; columns=%s", target_mat, df_raw.columns.tolist())
                return False

            df_raw = self._remove_total_rows(df_raw)
            df_filled = self._forward_fill_with_rules(df_raw)

            if target_item in df_filled.columns:
                df_filled[COL_FILLED_COST_ITEM] = df_filled[target_item].ffill()
            else:
                df_filled[COL_FILLED_COST_ITEM] = None

            df_detail, df_qty = self._split_sheets(df_raw, df_filled, target_mat, target_item)
            artifacts = build_report_artifacts(df_detail, df_qty)
            analysis_fact_df = self._filter_fact_df_for_analysis(artifacts.fact_df)
            analysis_tables = render_tables(analysis_fact_df)
            filtered_work_order_sheet = FlatSheet(
                data=self._filter_dataframe_by_whitelist(
                    artifacts.work_order_sheet.data,
                    code_col='产品编码',
                    name_col='产品名称',
                    sort_cols=['月份', '工单编号', '工单行'],
                ),
                column_types=artifacts.work_order_sheet.column_types,
            )
            product_anomaly_sections = self._filter_product_anomaly_sections(artifacts.product_anomaly_sections)
            error_log = artifacts.error_log.copy()

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self._write_dataframe_sheet(
                    writer,
                    '成本明细',
                    df_detail,
                    numeric_columns=DETAIL_TWO_DECIMAL_COLUMNS,
                    freeze_panes='A2',
                )
                self._write_dataframe_sheet(
                    writer,
                    '产品数量统计',
                    artifacts.qty_sheet_df,
                    numeric_columns=QTY_TWO_DECIMAL_COLUMNS,
                    freeze_panes='A2',
                )
                for sheet_name, sections in analysis_tables.items():
                    self._write_analysis_sheet(writer, sheet_name, sections)
                work_order_worksheet = self._write_flat_sheet(
                    writer,
                    '按工单按产品异常值分析',
                    filtered_work_order_sheet,
                    freeze_panes='A2',
                )
                self._apply_work_order_highlights(work_order_worksheet)
                self._write_product_anomaly_sheet(writer, '按产品异常值分析', product_anomaly_sections)
                self._write_flat_sheet(writer, '数据质量校验', artifacts.quality_sheet, freeze_panes='A2')
                error_log.to_excel(writer, sheet_name='error_log', index=False)

            logger.info(
                'Output saved: %s (detail=%s, qty=%s)', output_path, len(df_detail), len(artifacts.qty_sheet_df)
            )
            if not error_log.empty:
                logger.warning('Detected %s data quality issues, check sheet error_log', len(error_log))
            return True
        except Exception as exc:
            logger.error('Processing failed: %s', exc, exc_info=True)
            return False


def _find_input_files() -> list[Path]:
    """Match GB costing files; allow filename variants with/without a space."""
    patterns = [
        'GB-*成本计算单*.xlsx',
        'GB-* 成本计算单*.xlsx',
        'GB-*.xlsx',
    ]

    matched: list[Path] = []
    seen: set[Path] = set()
    for pattern in patterns:
        for path in sorted(GB_RAW_DIR.glob(pattern)):
            if path not in seen:
                seen.add(path)
                matched.append(path)
    return matched


def main() -> None:
    """Entry point for GB costing ETL."""
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    etl = CostingETL(skip_rows=2)
    input_files = _find_input_files()

    if not input_files:
        logger.warning('No GB costing file found under %s', GB_RAW_DIR)
        return

    input_file = input_files[0]
    logger.info('Matched input file: %s', input_file.name)
    output_file = GB_PROCESSED_DIR / f'{input_file.stem}_处理后.xlsx'

    if etl.process_file(input_file, output_file):
        print('处理成功')
    else:
        print('处理失败')


if __name__ == '__main__':
    main()
