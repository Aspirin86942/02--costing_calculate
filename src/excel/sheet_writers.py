"""单 sheet 写出与样式应用。"""

from __future__ import annotations

import logging

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.analytics.contracts import FlatSheet, ProductAnomalySection, SectionBlock
from src.excel.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BOLD_FONT,
    HEADER_FILL,
    HIGHLIGHT_ATTENTION_FILL,
    HIGHLIGHT_SUSPICIOUS_FILL,
    META_FILL,
    SECTION_TITLE_FILL,
    SECTION_TOTAL_FILL,
    THIN_BORDER,
    WHITE_FONT,
    estimate_analysis_column_widths,
    estimate_flat_column_widths,
    resolve_metric_number_format,
    to_excel_number,
)

logger = logging.getLogger(__name__)

WORK_ORDER_HIGHLIGHT_COLUMNS: tuple[tuple[str, str], ...] = (
    ('直接材料单位完工成本', '直接材料异常标记'),
    ('直接人工单位完工成本', '直接人工异常标记'),
    ('制造费用单位完工成本', '制造费用异常标记'),
    ('制造费用_其他单位完工成本', '制造费用_其他异常标记'),
    ('制造费用_人工单位完工成本', '制造费用_人工异常标记'),
    ('制造费用_机物料及低耗单位完工成本', '制造费用_机物料及低耗异常标记'),
    ('制造费用_折旧单位完工成本', '制造费用_折旧异常标记'),
    ('制造费用_水电费单位完工成本', '制造费用_水电费异常标记'),
)


class SheetWriter:
    """负责把 DataFrame/section 数据写成 workbook sheet。"""

    def write_dataframe_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        *,
        numeric_columns: set[str],
        freeze_panes: str = 'A2',
        fixed_width: int | None = None,
    ) -> Worksheet:
        """写入普通 DataFrame sheet，并按列名应用数值格式。"""
        column_formats = {column_name: '#,##0.00' for column_name in numeric_columns if column_name in df.columns}
        write_df = self._coerce_excel_numeric_columns(df, set(column_formats))
        write_df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        width_map = estimate_flat_column_widths(write_df, fixed_width=fixed_width)
        self._apply_basic_sheet_styles(
            worksheet,
            write_df.columns.tolist(),
            column_formats=column_formats,
            freeze_panes=freeze_panes,
            width_map=width_map,
        )
        return worksheet

    def write_analysis_sheet(self, writer: pd.ExcelWriter, sheet_name: str, sections: list[SectionBlock]) -> None:
        """写入三段分析块并应用样式；禁止合并单元格。"""
        start_row = 0
        section_meta: list[dict[str, int | str | bool]] = []

        for section in sections:
            write_section_df = section.data.copy()
            if section.metric_type in {'amount', 'price', 'qty'}:
                numeric_columns = [
                    column_name
                    for column_name in write_section_df.columns
                    if column_name not in {'产品编码', '产品名称'}
                ]
                for column_name in numeric_columns:
                    write_section_df[column_name] = write_section_df[column_name].map(to_excel_number)

            pd.DataFrame([[section.title]]).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=start_row,
            )
            write_section_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row + 1)

            title_row = start_row + 1
            header_row = start_row + 2
            data_start = start_row + 3
            data_end = data_start + len(write_section_df) - 1
            max_col = max(1, write_section_df.shape[1])

            section_meta.append(
                {
                    'title_row': title_row,
                    'header_row': header_row,
                    'data_start': data_start,
                    'data_end': data_end,
                    'max_col': max_col,
                    'metric_type': section.metric_type,
                    'has_total_row': section.has_total_row,
                }
            )
            start_row += len(write_section_df) + 3

        worksheet = writer.sheets[sheet_name]
        width_map = estimate_analysis_column_widths([(section.title, section.data) for section in sections])
        self._style_analysis_sheet(worksheet, section_meta, width_map)

    def write_flat_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        table: FlatSheet,
        *,
        freeze_panes: str = 'A2',
        fixed_width: int | None = None,
    ) -> Worksheet:
        """写入平铺数据 sheet 并应用基础样式。"""
        write_df = table.data.copy()
        column_formats: dict[str, str] = {}
        for column_name, metric_type in table.column_types.items():
            if column_name not in write_df.columns:
                continue
            if metric_type in {'amount', 'price', 'qty', 'score', 'pct'}:
                write_df[column_name] = write_df[column_name].map(to_excel_number)
            number_format = resolve_metric_number_format(metric_type)
            if number_format is not None:
                column_formats[column_name] = number_format

        write_df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        width_map = estimate_flat_column_widths(write_df, fixed_width=fixed_width)
        self._apply_basic_sheet_styles(
            worksheet,
            write_df.columns.tolist(),
            column_formats=column_formats,
            freeze_panes=freeze_panes,
            width_map=width_map,
        )
        return worksheet

    def write_product_anomaly_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        sections: list[ProductAnomalySection],
    ) -> None:
        """写入按产品异常值分析页（不合并单元格）。"""
        worksheet = writer.book.create_sheet(title=sheet_name)
        writer.sheets[sheet_name] = worksheet

        worksheet.cell(1, 1, '四、按单个产品异常值分析')
        worksheet.cell(1, 1).fill = SECTION_TITLE_FILL
        worksheet.cell(1, 1).font = BOLD_FONT
        worksheet.cell(1, 1).alignment = ALIGN_LEFT

        current_row = 3
        max_col_overall = 1
        filter_set = False

        for section in sections:
            meta_header_row = current_row
            meta_value_row = current_row + 1
            table_header_row = current_row + 2
            data_start_row = current_row + 3
            data_end_row = data_start_row + len(section.data) - 1
            max_col = len(section.data.columns)
            max_col_overall = max(max_col_overall, max_col)

            worksheet.cell(meta_header_row, 1, '产品编码')
            worksheet.cell(meta_header_row, 2, '产品名称')
            worksheet.cell(meta_value_row, 1, section.product_code)
            worksheet.cell(meta_value_row, 2, section.product_name)

            for col_idx in [1, 2]:
                header_cell = worksheet.cell(meta_header_row, col_idx)
                header_cell.fill = HEADER_FILL
                header_cell.font = BOLD_FONT
                header_cell.alignment = ALIGN_CENTER
                header_cell.border = THIN_BORDER

                value_cell = worksheet.cell(meta_value_row, col_idx)
                value_cell.fill = META_FILL
                value_cell.font = BOLD_FONT
                value_cell.alignment = ALIGN_LEFT
                value_cell.border = THIN_BORDER

            for col_idx, column_name in enumerate(section.data.columns, start=1):
                header_cell = worksheet.cell(table_header_row, col_idx, column_name)
                header_cell.fill = HEADER_FILL
                header_cell.font = BOLD_FONT
                header_cell.alignment = ALIGN_CENTER
                header_cell.border = THIN_BORDER

            for row_idx, row_data in section.data.iterrows():
                excel_row = data_start_row + row_idx
                for col_idx, column_name in enumerate(section.data.columns, start=1):
                    cell = worksheet.cell(excel_row, col_idx, row_data[column_name])
                    cell.border = THIN_BORDER
                    cell.alignment = ALIGN_LEFT if col_idx == 1 else ALIGN_RIGHT

                    metric_type = section.column_types.get(column_name)
                    number_format = resolve_metric_number_format(metric_type or '')
                    if number_format is not None:
                        cell.number_format = number_format

            if not filter_set:
                worksheet.auto_filter.ref = (
                    f'A{table_header_row}:{get_column_letter(max_col)}{max(data_end_row, table_header_row)}'
                )
                filter_set = True

            current_row = data_end_row + 2

        worksheet.freeze_panes = 'A6'
        for col_idx in range(1, max_col_overall + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 15

    def apply_work_order_highlights(self, worksheet: Worksheet) -> None:
        """把异常颜色挂在“值列+标记列”上。"""
        header_map = {
            worksheet.cell(1, col_idx).value: col_idx
            for col_idx in range(1, worksheet.max_column + 1)
            if worksheet.cell(1, col_idx).value is not None
        }
        highlight_styles = {
            '关注': {'fill': HIGHLIGHT_ATTENTION_FILL, 'font': None},
            '高度可疑': {'fill': HIGHLIGHT_SUSPICIOUS_FILL, 'font': WHITE_FONT},
        }

        for value_column, flag_column in WORK_ORDER_HIGHLIGHT_COLUMNS:
            value_idx = header_map.get(value_column)
            flag_idx = header_map.get(flag_column)
            if value_idx is None or flag_idx is None:
                logger.warning(
                    'Skip work-order highlight: missing columns sheet=%s value=%s flag=%s',
                    worksheet.title,
                    value_column,
                    flag_column,
                )
                continue

            for row_idx in range(2, worksheet.max_row + 1):
                flag_value = worksheet.cell(row_idx, flag_idx).value
                style = highlight_styles.get(str(flag_value).strip()) if flag_value is not None else None
                if style is None:
                    continue

                for col_idx in (value_idx, flag_idx):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.fill = style['fill']
                    if style['font'] is not None:
                        cell.font = style['font']

    def _apply_basic_sheet_styles(
        self,
        worksheet: Worksheet,
        columns: list[str],
        *,
        column_formats: dict[str, str],
        freeze_panes: str,
        width_map: dict[int, float],
    ) -> None:
        for col_idx, column_name in enumerate(columns, start=1):
            header_cell = worksheet.cell(1, col_idx)
            header_cell.fill = HEADER_FILL
            header_cell.font = BOLD_FONT
            header_cell.alignment = ALIGN_CENTER
            header_cell.border = THIN_BORDER

            number_format = column_formats.get(column_name)
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell.border = THIN_BORDER
                if number_format is None:
                    cell.alignment = ALIGN_LEFT
                    continue
                cell.alignment = ALIGN_RIGHT
                cell.number_format = number_format

        worksheet.freeze_panes = freeze_panes
        if worksheet.max_column > 0:
            worksheet.auto_filter.ref = f'A1:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, 1)}'

        for col_idx, width in width_map.items():
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _coerce_excel_numeric_columns(self, df: pd.DataFrame, numeric_columns: set[str]) -> pd.DataFrame:
        """写出前统一转成 Excel 数值单元格。"""
        write_df = df.copy()
        for column_name in numeric_columns:
            if column_name in write_df.columns:
                write_df[column_name] = write_df[column_name].map(to_excel_number)
        return write_df

    def _style_analysis_sheet(
        self,
        worksheet: Worksheet,
        section_meta: list[dict[str, int | str | bool]],
        width_map: dict[int, float],
    ) -> None:
        if not section_meta:
            return

        max_col_overall = max(int(meta['max_col']) for meta in section_meta)

        for meta in section_meta:
            title_row = int(meta['title_row'])
            header_row = int(meta['header_row'])
            data_start = int(meta['data_start'])
            data_end = int(meta['data_end'])
            max_col = int(meta['max_col'])
            metric_type = str(meta['metric_type'])
            has_total_row = bool(meta['has_total_row'])

            title_cell = worksheet.cell(title_row, 1)
            title_cell.fill = SECTION_TITLE_FILL
            title_cell.font = BOLD_FONT
            title_cell.alignment = ALIGN_LEFT

            for col_idx in range(1, max_col + 1):
                header_cell = worksheet.cell(header_row, col_idx)
                header_cell.fill = HEADER_FILL
                header_cell.font = BOLD_FONT
                header_cell.alignment = ALIGN_CENTER
                header_cell.border = THIN_BORDER

            if data_end < data_start:
                continue

            for row_idx in range(data_start, data_end + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.border = THIN_BORDER
                    if col_idx <= 2:
                        cell.alignment = ALIGN_LEFT
                    else:
                        cell.alignment = ALIGN_RIGHT
                        number_format = resolve_metric_number_format(metric_type, qty_format='#,##0.00')
                        if number_format is not None:
                            cell.number_format = number_format

            if has_total_row:
                total_row = data_end
                for col_idx in range(1, max_col + 1):
                    total_cell = worksheet.cell(total_row, col_idx)
                    total_cell.fill = SECTION_TOTAL_FILL
                    total_cell.font = BOLD_FONT

        first_meta = section_meta[0]
        worksheet.freeze_panes = 'C3'
        filter_end = int(first_meta['data_end'])
        if filter_end < int(first_meta['header_row']):
            filter_end = int(first_meta['header_row'])
        worksheet.auto_filter.ref = (
            f'A{int(first_meta["header_row"])}:{get_column_letter(int(first_meta["max_col"]))}{filter_end}'
        )

        for col_idx in range(1, max_col_overall + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width_map.get(col_idx, 12.0)
