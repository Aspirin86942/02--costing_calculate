"""Excel 样式与格式工具。"""

from __future__ import annotations

from decimal import Decimal
from numbers import Real

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

EXCEL_TWO_DECIMAL_FORMAT = '#,##0.00'
EXCEL_INTEGER_FORMAT = '#,##0'
EXCEL_SCORE_FORMAT = '0.0000'
EXCEL_PERCENT_FORMAT = '0.00%'

HEADER_FILL = PatternFill(fill_type='solid', fgColor='D9E1F2')
SECTION_TITLE_FILL = PatternFill(fill_type='solid', fgColor='FFD966')
SECTION_TOTAL_FILL = PatternFill(fill_type='solid', fgColor='BDD7EE')
META_FILL = PatternFill(fill_type='solid', fgColor='B4C6E7')
HIGHLIGHT_ATTENTION_FILL = PatternFill(fill_type='solid', fgColor='DDEBF7')
HIGHLIGHT_SUSPICIOUS_FILL = PatternFill(fill_type='solid', fgColor='4472C4')
WHITE_FONT = Font(color='FFFFFF')
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')


def to_excel_number(value: object) -> object:
    """把 Decimal/数值对象转成 Excel 可识别的数值类型。"""
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, Real):
        return value if pd.isna(value) else float(value)
    return value


def resolve_metric_number_format(metric_type: str, *, qty_format: str = EXCEL_INTEGER_FORMAT) -> str | None:
    """根据列的语义类型返回 Excel number format。"""
    if metric_type in {'amount', 'price'}:
        return EXCEL_TWO_DECIMAL_FORMAT
    if metric_type == 'qty':
        return qty_format
    if metric_type == 'score':
        return EXCEL_SCORE_FORMAT
    if metric_type == 'pct':
        return EXCEL_PERCENT_FORMAT
    return None


def estimate_flat_column_widths(
    df: pd.DataFrame,
    *,
    fixed_width: int | None = None,
    minimum: int = 12,
    maximum: int = 24,
) -> dict[int, float]:
    """基于 DataFrame 估算平铺 sheet 的列宽，避免回扫 worksheet。"""
    if fixed_width is not None:
        return {column_idx: float(fixed_width) for column_idx in range(1, len(df.columns) + 1)}

    widths: dict[int, float] = {}
    for column_idx, column_name in enumerate(df.columns, start=1):
        max_length = len(str(column_name))
        if column_name in df.columns:
            series = df[column_name]
            non_null_values = series[series.notna()]
            if not non_null_values.empty:
                max_length = max(max_length, max(len(str(value)) for value in non_null_values))
        widths[column_idx] = float(min(max(minimum, max_length + 2), maximum))
    return widths


def estimate_analysis_column_widths(
    sections: list[tuple[str, pd.DataFrame]],
    *,
    minimum: int = 12,
    maximum: int = 40,
) -> dict[int, float]:
    """按分析页所有分段内容预估列宽。"""
    column_lengths: dict[int, int] = {}
    for title, data in sections:
        column_lengths[1] = max(column_lengths.get(1, 0), len(str(title)))
        for column_idx, column_name in enumerate(data.columns, start=1):
            column_lengths[column_idx] = max(column_lengths.get(column_idx, 0), len(str(column_name)))
            series = data[column_name]
            non_null_values = series[series.notna()]
            if not non_null_values.empty:
                column_lengths[column_idx] = max(
                    column_lengths[column_idx],
                    max(len(str(value)) for value in non_null_values),
                )
    return {column_idx: float(min(max(minimum, length + 2), maximum)) for column_idx, length in column_lengths.items()}
