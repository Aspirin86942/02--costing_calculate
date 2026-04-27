from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import polars as pl
from openpyxl import load_workbook

from src.analytics.contracts import (
    AnalysisArtifacts,
    FlatSheet,
    NormalizedCostFrame,
    ProductAnomalySection,
    QualityMetric,
    RawWorkbookFrame,
    SplitResult,
)
from src.etl.costing_etl import CostingWorkbookETL

DEFAULT_WORKBOOK_BASENAME = 'workbook_contract_default.xlsx'
HIGHLIGHT_WORKBOOK_BASENAME = 'workbook_contract_highlight.xlsx'

DEFAULT_SHEETS = (
    '成本明细',
    '产品数量统计',
    '直接材料_价量比',
    '直接人工_价量比',
    '制造费用_价量比',
    '按工单按产品异常值分析',
    '按产品异常值分析',
)
ANALYSIS_SHEETS = {'直接材料_价量比', '直接人工_价量比', '制造费用_价量比'}


def load_contract_baseline(filename: str) -> dict[str, object]:
    """读取 contract baseline。"""
    baseline_path = Path(__file__).with_name('baselines') / filename
    return json.loads(baseline_path.read_text(encoding='utf-8'))


def build_default_contract_workbook(tmp_path: Path) -> Path:
    """用当前真实实现生成默认 workbook 语义样本。"""
    etl = CostingWorkbookETL(skip_rows=2)
    input_path = tmp_path / 'input.xlsx'
    output_path = tmp_path / DEFAULT_WORKBOOK_BASENAME

    df_detail = _build_default_detail_df()
    df_qty = _build_default_qty_df()
    with (
        patch.object(etl.pipeline, 'load_raw_workbook_frame', return_value=_build_stub_raw_workbook()),
        patch.object(etl.pipeline, 'build_normalized_cost_frame', return_value=_build_stub_normalized_frame()),
        patch.object(
            etl.pipeline,
            'split_normalized_frames',
            return_value=SplitResult(detail_df=df_detail, qty_df=df_qty),
        ),
    ):
        assert etl.process_file(input_path, output_path) is True

    return output_path


def build_highlight_contract_workbook(tmp_path: Path) -> Path:
    """生成带异常高亮的 workbook 语义样本。"""
    etl = CostingWorkbookETL(skip_rows=2)
    input_path = tmp_path / 'input.xlsx'
    output_path = tmp_path / HIGHLIGHT_WORKBOOK_BASENAME

    df_detail = pd.DataFrame(
        [
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接材料',
                '本期完工金额': 100,
            }
        ]
    )
    df_qty = pd.DataFrame(
        [
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '本期完工数量': 10,
                '本期完工金额': 100,
            }
        ]
    )
    work_order_df = pd.DataFrame(
        [
            {
                '月份': '2025年01月',
                '成本中心': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行': '1',
                '基本单位': 'PCS',
                '本期完工数量': 10,
                '直接材料单位完工成本': 18.0,
                '直接人工单位完工成本': 2.0,
                '制造费用单位完工成本': 3.0,
                '制造费用_其他单位完工成本': 0.0,
                '制造费用_人工单位完工成本': 30.0,
                '制造费用_机物料及低耗单位完工成本': 0.0,
                '制造费用_折旧单位完工成本': 0.0,
                '制造费用_水电费单位完工成本': 0.0,
                '直接材料异常标记': '关注',
                '直接人工异常标记': '正常',
                '制造费用异常标记': '正常',
                '制造费用_其他异常标记': '正常',
                '制造费用_人工异常标记': '高度可疑',
                '制造费用_机物料及低耗异常标记': '正常',
                '制造费用_折旧异常标记': '正常',
                '制造费用_水电费异常标记': '正常',
            }
        ]
    )
    work_order_column_types = {
        '月份': 'text',
        '成本中心': 'text',
        '产品编码': 'text',
        '产品名称': 'text',
        '规格型号': 'text',
        '工单编号': 'text',
        '工单行': 'text',
        '基本单位': 'text',
        '本期完工数量': 'qty',
        '直接材料单位完工成本': 'price',
        '直接人工单位完工成本': 'price',
        '制造费用单位完工成本': 'price',
        '制造费用_其他单位完工成本': 'price',
        '制造费用_人工单位完工成本': 'price',
        '制造费用_机物料及低耗单位完工成本': 'price',
        '制造费用_折旧单位完工成本': 'price',
        '制造费用_水电费单位完工成本': 'price',
        '直接材料异常标记': 'text',
        '直接人工异常标记': 'text',
        '制造费用异常标记': 'text',
        '制造费用_其他异常标记': 'text',
        '制造费用_人工异常标记': 'text',
        '制造费用_机物料及低耗异常标记': 'text',
        '制造费用_折旧异常标记': 'text',
        '制造费用_水电费异常标记': 'text',
    }
    artifacts = AnalysisArtifacts(
        fact_df=pd.DataFrame(
            [
                {
                    'period': '2025-01',
                    'product_code': 'GB_C.D.B0040AA',
                    'product_name': 'BMS-750W驱动器',
                    'cost_bucket': 'direct_material',
                    'amount': 100,
                    'qty': 10,
                    'price': 10,
                }
            ]
        ),
        qty_sheet_df=df_qty.copy(),
        work_order_sheet=FlatSheet(data=work_order_df, column_types=work_order_column_types),
        product_anomaly_sections=[
            ProductAnomalySection(
                product_code='GB_C.D.B0040AA',
                product_name='BMS-750W驱动器',
                data=pd.DataFrame([{'月份': '2025年01月', '总成本': 100.0, '完工数量': 10, '单位成本': 10.0}]),
                column_types={'月份': 'text', '总成本': 'amount', '完工数量': 'qty', '单位成本': 'price'},
                amount_columns=['总成本'],
                outlier_cells=set(),
            )
        ],
        quality_metrics=(
            QualityMetric(
                category='行数勾稽',
                metric='样例',
                value='1',
                description='测试',
            ),
        ),
        error_log=pd.DataFrame(),
    )

    with (
        patch.object(etl.pipeline, 'load_raw_workbook_frame', return_value=_build_stub_raw_workbook()),
        patch.object(etl.pipeline, 'build_normalized_cost_frame', return_value=_build_stub_normalized_frame()),
        patch.object(
            etl.pipeline,
            'split_normalized_frames',
            return_value=SplitResult(detail_df=df_detail, qty_df=df_qty),
        ),
        patch('src.etl.pipeline.build_report_artifacts', return_value=artifacts),
    ):
        assert etl.process_file(input_path, output_path) is True

    return output_path


def _build_stub_raw_workbook() -> RawWorkbookFrame:
    return RawWorkbookFrame(
        sheet_name='成本计算单',
        header_rows=(('年期', '产品编码'), ('', '')),
        frame=pl.DataFrame({'column_0': ['2025年1月'], 'column_1': ['GB_C.D.B0040AA']}),
    )


def _build_stub_normalized_frame() -> NormalizedCostFrame:
    return NormalizedCostFrame(
        frame=pl.DataFrame({'月份': ['2025年01月'], '产品编码': ['GB_C.D.B0040AA']}),
        key_columns=('月份', '产品编码'),
    )


def extract_workbook_semantics(workbook_path: Path) -> dict[str, object]:
    """提取 workbook 语义层快照，而不是二进制内容。"""
    workbook = load_workbook(workbook_path)
    semantics: dict[str, object] = {'sheet_order': workbook.sheetnames, 'sheets': {}}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if sheet_name in ANALYSIS_SHEETS:
            semantics['sheets'][sheet_name] = _extract_analysis_sheet(worksheet)
        elif sheet_name == '按产品异常值分析':
            semantics['sheets'][sheet_name] = _extract_product_anomaly_sheet(worksheet)
        else:
            semantics['sheets'][sheet_name] = _extract_flat_sheet(worksheet)

    return semantics


def extract_highlight_semantics(workbook_path: Path) -> dict[str, object]:
    """提取按工单异常页的条件格式规则语义。"""
    workbook = load_workbook(workbook_path)
    worksheet = workbook['按工单按产品异常值分析']
    rules: list[dict[str, object]] = []

    for conditional_range, rule_list in worksheet.conditional_formatting._cf_rules.items():
        sqref = _normalize_sqref(str(conditional_range.sqref))
        for rule in rule_list:
            rules.append(
                {
                    'sqref': sqref,
                    'formula': _normalize_rule_formulas(rule.formula),
                    'fill': _extract_rule_fill(rule),
                    'font': _extract_rule_font(rule),
                }
            )

    rules.sort(key=lambda item: (item['sqref'], tuple(item['formula'])))
    return {'sheet': worksheet.title, 'rules': rules}


def _build_default_detail_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接材料',
                '本期完工单位成本': 10,
                '本期完工金额': 100,
            },
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接人工',
                '本期完工单位成本': 2,
                '本期完工金额': 20,
            },
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '制造费用▪人工',
                '本期完工单位成本': 3,
                '本期完工金额': 30,
            },
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '委外加工费',
                '本期完工单位成本': 1.5,
                '本期完工金额': 15,
            },
        ]
    )


def _build_default_qty_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '单据类型': '汇报入库-普通生产',
                '基本单位': 'PCS',
                '本期完工数量': 10,
                '本期完工金额': 165,
            },
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-02',
                '工单编号': 'WO-002',
                '工单行号': 1,
                '单据类型': '汇报入库-普通生产',
                '基本单位': 'PCS',
                '本期完工数量': 0,
                '本期完工金额': 100,
            },
            {
                '月份': '2025年01月',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-03',
                '工单编号': 'WO-003',
                '工单行号': 1,
                '单据类型': '汇报入库-普通生产',
                '基本单位': 'PCS',
                '本期完工数量': 5,
                '本期完工金额': None,
            },
        ]
    )


def _extract_flat_sheet(worksheet) -> dict[str, object]:
    headers = [worksheet.cell(1, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
    first_data_formats = {
        header: worksheet.cell(2, col_idx).number_format
        for col_idx, header in enumerate(headers, start=1)
        if worksheet.max_row >= 2 and worksheet.cell(2, col_idx).number_format != 'General'
    }
    return {
        'kind': 'flat',
        'freeze_panes': worksheet.freeze_panes,
        'auto_filter': worksheet.auto_filter.ref,
        'columns': headers,
        'number_formats': first_data_formats,
        'column_widths': _extract_column_widths(worksheet),
    }


def _extract_analysis_sheet(worksheet) -> dict[str, object]:
    sections: list[dict[str, object]] = []
    row_idx = 1

    while row_idx <= worksheet.max_row:
        title = worksheet.cell(row_idx, 1).value
        next_row_has_headers = row_idx + 1 <= worksheet.max_row and worksheet.cell(row_idx + 1, 1).value is not None
        rest_empty = all(
            worksheet.cell(row_idx, col_idx).value is None for col_idx in range(2, worksheet.max_column + 1)
        )
        if title is None or not next_row_has_headers or not rest_empty:
            row_idx += 1
            continue

        header_row = row_idx + 1
        headers = []
        col_idx = 1
        while col_idx <= worksheet.max_column and worksheet.cell(header_row, col_idx).value is not None:
            headers.append(worksheet.cell(header_row, col_idx).value)
            col_idx += 1

        data_row = header_row + 1
        number_formats = {
            header: worksheet.cell(data_row, header_col).number_format
            for header_col, header in enumerate(headers, start=1)
            if data_row <= worksheet.max_row and worksheet.cell(data_row, header_col).number_format != 'General'
        }
        sections.append({'title': title, 'columns': headers, 'number_formats': number_formats})
        row_idx = header_row + 1

    return {
        'kind': 'analysis',
        'freeze_panes': worksheet.freeze_panes,
        'auto_filter': worksheet.auto_filter.ref,
        'sections': sections,
        'column_widths': _extract_column_widths(worksheet),
    }


def _extract_product_anomaly_sheet(worksheet) -> dict[str, object]:
    if worksheet['A5'].value == '分析口径':
        return _extract_scoped_product_anomaly_sheet(worksheet)
    return _extract_legacy_product_anomaly_sheet(worksheet)


def _extract_legacy_product_anomaly_sheet(worksheet) -> dict[str, object]:
    headers = [worksheet.cell(5, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
    while headers and headers[-1] is None:
        headers.pop()

    first_data_formats = {
        header: worksheet.cell(6, col_idx).number_format
        for col_idx, header in enumerate(headers, start=1)
        if worksheet.max_row >= 6 and worksheet.cell(6, col_idx).number_format != 'General'
    }
    return {
        'kind': 'product_anomaly',
        'layout': 'legacy',
        'freeze_panes': worksheet.freeze_panes,
        'auto_filter': worksheet.auto_filter.ref,
        'title': worksheet['A1'].value,
        'meta_labels': [worksheet['A3'].value, worksheet['B3'].value],
        'meta_values': [worksheet['A4'].value, worksheet['B4'].value],
        'columns': headers,
        'number_formats': first_data_formats,
        'column_widths': _extract_column_widths(worksheet),
    }


def _extract_scoped_product_anomaly_sheet(worksheet) -> dict[str, object]:
    sections: list[dict[str, object]] = []
    row_idx = 3
    while row_idx <= worksheet.max_row:
        meta_label_code = worksheet.cell(row_idx, 1).value
        meta_label_name = worksheet.cell(row_idx, 2).value
        if meta_label_code != '产品编码' or meta_label_name != '产品名称':
            row_idx += 1
            continue

        if worksheet.cell(row_idx + 2, 1).value != '分析口径':
            row_idx += 1
            continue

        header_row = row_idx + 3
        first_data_row = header_row + 1
        headers: list[object] = []
        col_idx = 1
        while col_idx <= worksheet.max_column and worksheet.cell(header_row, col_idx).value is not None:
            headers.append(worksheet.cell(header_row, col_idx).value)
            col_idx += 1

        number_formats = {
            header: worksheet.cell(first_data_row, header_col).number_format
            for header_col, header in enumerate(headers, start=1)
            if first_data_row <= worksheet.max_row
            and worksheet.cell(first_data_row, header_col).number_format != 'General'
        }
        sections.append(
            {
                'product_code': worksheet.cell(row_idx + 1, 1).value,
                'product_name': worksheet.cell(row_idx + 1, 2).value,
                'scope_label': worksheet.cell(row_idx + 2, 2).value,
                'columns': headers,
                'number_formats': number_formats,
            }
        )
        row_idx = first_data_row + 1

    first_section = sections[0] if sections else {'columns': [], 'number_formats': {}}
    return {
        'kind': 'product_anomaly',
        'layout': 'scoped',
        'freeze_panes': worksheet.freeze_panes,
        'auto_filter': worksheet.auto_filter.ref,
        'title': worksheet['A1'].value,
        'scope_labels': [section['scope_label'] for section in sections],
        'columns': first_section['columns'],
        'number_formats': first_section['number_formats'],
        'sections': sections,
        'column_widths': _extract_column_widths(worksheet),
    }


def _extract_column_widths(worksheet) -> dict[str, float]:
    return {
        dimension_letter: round(float(dimension.width), 2)
        for dimension_letter, dimension in worksheet.column_dimensions.items()
        if dimension.width is not None
    }


def _rgb_suffix(color) -> str | None:
    rgb = getattr(color, 'rgb', None)
    if rgb is None:
        return None
    if hasattr(rgb, 'value'):
        rgb = rgb.value
    return rgb[-6:]


def _normalize_sqref(sqref: str) -> str:
    return sqref if ':' in sqref else f'{sqref}:{sqref}'


def _normalize_rule_formulas(formulas) -> list[str]:
    normalized: list[str] = []
    for formula in formulas or []:
        formula_text = str(formula)
        if not formula_text.startswith('='):
            formula_text = f'={formula_text}'
        normalized.append(formula_text)
    return normalized


def _extract_rule_fill(rule) -> str | None:
    dxf = getattr(rule, 'dxf', None)
    fill = getattr(dxf, 'fill', None)
    if fill is None:
        return None
    return _rgb_suffix(fill.bgColor) or _rgb_suffix(fill.fgColor)


def _extract_rule_font(rule) -> str | None:
    dxf = getattr(rule, 'dxf', None)
    font = getattr(dxf, 'font', None)
    if font is None:
        return None
    return _rgb_suffix(font.color)
