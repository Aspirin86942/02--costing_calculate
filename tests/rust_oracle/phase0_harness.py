from __future__ import annotations

import argparse
import hashlib
import json
import os
import platform
import shutil
import subprocess
import sys
import uuid
import zipfile
from dataclasses import asdict, dataclass, field, replace
from decimal import ROUND_CEILING, Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from statistics import median
from typing import Any

from openpyxl import load_workbook

from tests.rust_oracle.benchmark_protocol import (
    COMPARISON_LIMITS,
    ENVIRONMENT_MEDIAN_DRIFT_LIMIT,
    PROFILE_RULES,
    AttemptState,
    BatchAttempt,
    BinaryRole,
    CalibrationGroup,
    CalibrationRound,
    ClosedBinaryLabel,
    ComparisonProfile,
    HarnessVerdict,
    MachineEvidence,
    MetricGroup,
    MetricName,
    MetricSample,
    NormalRunEvidence,
    PairedBenchmarkResult,
    PairedRound,
    Phase0AManifest,
    PipelineName,
    RoundPlan,
    RuntimeEvidence,
    RuntimeSchema,
    assert_same_benchmark_batch,
    build_round_plan,
    groups_have_conflicting_direction,
    merge_metric_groups,
    requires_mandatory_expansion,
    validate_calibration_group,
    validate_metric_group,
)
from tests.rust_oracle.evidence import (
    ApprovedSheet,
    BenchmarkManifestEvidence,
    BenchmarkMetric,
    BenchmarkMetricEvidence,
    BenchmarkRoundEvidence,
    EvidenceSanitizer,
    MachineArtifactEvidence,
    OutputBytesEvidence,
    PathAlias,
    RuntimeCount,
    RuntimeCountEvidence,
    SheetDimensionEvidence,
    SmokeSummaryEvidence,
    _batch_commit_marker,
)
from tests.rust_oracle.oracle_runner import (
    CapturedNormalRun,
    RustNormalProcessError,
    RustNormalValidationError,
    _io_path,
    _normal_path,
    _prepare_local_path,
    _reject_existing_reparse_components,
    parse_runtime_payload,
    run_rust_normal_captured,
)
from tests.rust_oracle.repo_paths import repo_root
from tests.rust_oracle.sanitized_fixture import build_raw_fixture

_PWS_DRIVER_TIMEOUT_SECONDS = 930.0
_PWS_DRIVER_TERMINATION_SECONDS = 30.0


class HarnessFailure(AssertionError):
    def __init__(
        self,
        verdict: HarnessVerdict,
        message: str,
        *,
        primary_verdict: HarnessVerdict | None = None,
        raw_log_sha256: str | None = None,
    ) -> None:
        super().__init__(message)
        self.verdict = verdict
        self.primary_verdict = primary_verdict
        self.raw_log_sha256 = raw_log_sha256


def _capture_boundary_failure(role: BinaryRole, context: str, error: Exception) -> HarnessFailure:
    if isinstance(error, RustNormalProcessError):
        verdict = HarnessVerdict.REFERENCE_FAILED if role == 'reference' else HarnessVerdict.CANDIDATE_FAILED
        return HarnessFailure(
            verdict,
            f'{context} {role} process failed with exit code {error.returncode}',
            raw_log_sha256=error.log_sha256,
        )
    if isinstance(error, RustNormalValidationError):
        verdict = HarnessVerdict.REFERENCE_FAILED if role == 'reference' else HarnessVerdict.CORRECTNESS_FAILED
        return HarnessFailure(
            verdict,
            f'{context} {role} runtime or workbook validation failed',
            raw_log_sha256=error.log_sha256,
        )
    verdict = HarnessVerdict.REFERENCE_FAILED if role == 'reference' else HarnessVerdict.CORRECTNESS_FAILED
    return HarnessFailure(verdict, f'{context} {role} capture boundary failed')


@dataclass(frozen=True)
class UnverifiedPriorEvidenceClaim:
    path_alias: str
    content_sha256: str

    def __post_init__(self) -> None:
        alias = PurePosixPath(self.path_alias)
        if alias.is_absolute() or '..' in alias.parts or '\\' in self.path_alias or not self.path_alias:
            raise ValueError('prior evidence path_alias must be repository-relative POSIX text')
        if not _is_sha256(self.content_sha256):
            raise ValueError('prior evidence content hash must be lowercase SHA-256')


@dataclass(frozen=True)
class BenchmarkIdentity:
    input_sha256: str
    reference_sha256: str
    candidate_sha256: str
    git_head: str
    repository_state_sha256: str
    machine_fingerprint_sha256: str


@dataclass(frozen=True)
class PairedBenchmarkRequest:
    pipeline: PipelineName
    input_path: Path
    reference_executable: Path
    candidate_executable: Path
    reference_label: ClosedBinaryLabel
    candidate_label: ClosedBinaryLabel
    comparison_profile: ComparisonProfile
    phase0a_manifest: Path
    local_root: Path
    evidence_path: Path
    attempt_ledger_root: Path
    prior_evidence_claims: tuple[UnverifiedPriorEvidenceClaim, ...] = ()


@dataclass(frozen=True)
class MetricGroupRequest:
    benchmark: PairedBenchmarkRequest
    batch_id: str
    metric: MetricName
    plans: tuple[RoundPlan, ...]
    attempt_directory: Path
    first_group_sha256: str | None = None

    def __post_init__(self) -> None:
        if self.metric not in ('wall', 'pws'):
            raise ValueError('metric must be wall or pws')
        rounds = tuple(plan.global_round for plan in self.plans)
        if rounds == (1, 2, 3, 4, 5):
            if self.first_group_sha256 is not None:
                raise ValueError('rounds one through five cannot be submitted as an expanded group')
        elif rounds == (6, 7, 8, 9, 10):
            if not _is_sha256(self.first_group_sha256):
                raise ValueError('expanded rounds require the original first-group SHA-256')
        else:
            raise ValueError('metric group plans must be global rounds one through five or six through ten')


@dataclass(frozen=True)
class Phase0HSmokeRequest:
    pipeline: PipelineName
    reference_executable: Path
    candidate_executable: Path
    local_root: Path


@dataclass(frozen=True)
class Phase0HSmokeResult:
    batch_id: str
    fixture_sha256: str
    verdict: HarnessVerdict


@dataclass(frozen=True)
class Phase0ARequest:
    gb_input_path: Path
    sk_input_path: Path
    reference_executable: Path
    fork_revision: str
    local_root: Path
    output_path: Path


@dataclass(frozen=True)
class _ApprovedPhase0ABaseline:
    machine_fingerprint_sha256: str
    output_size_bytes: int
    wall_values: tuple[Decimal, ...]
    pws_values: tuple[Decimal, ...]


_PROFILE_LABEL_PAIRS: dict[ComparisonProfile, frozenset[tuple[ClosedBinaryLabel, ClosedBinaryLabel]]] = {
    ComparisonProfile.PHASE0B_VS_PHASE0A: frozenset({(ClosedBinaryLabel.PHASE0A, ClosedBinaryLabel.PHASE0B)}),
    ComparisonProfile.PHASE1_VS_PHASE0B: frozenset({(ClosedBinaryLabel.PHASE0B, ClosedBinaryLabel.PHASE1)}),
    ComparisonProfile.PHASE1_VS_PHASE0A: frozenset({(ClosedBinaryLabel.PHASE0A, ClosedBinaryLabel.PHASE1)}),
    ComparisonProfile.PHASE2_B_VS_A: frozenset({(ClosedBinaryLabel.PHASE2_A, ClosedBinaryLabel.PHASE2_B)}),
    ComparisonProfile.PHASE2_C_VS_A: frozenset({(ClosedBinaryLabel.PHASE2_A, ClosedBinaryLabel.PHASE2_C)}),
    ComparisonProfile.PHASE2_D_VS_C: frozenset({(ClosedBinaryLabel.PHASE2_C, ClosedBinaryLabel.PHASE2_D)}),
    ComparisonProfile.PHASE2_D_VS_B: frozenset({(ClosedBinaryLabel.PHASE2_B, ClosedBinaryLabel.PHASE2_D)}),
    ComparisonProfile.PHASE2_B_VS_C: frozenset({(ClosedBinaryLabel.PHASE2_C, ClosedBinaryLabel.PHASE2_B)}),
    ComparisonProfile.PHASE2_SELECTED_VS_PHASE0A: frozenset(
        (ClosedBinaryLabel.PHASE0A, candidate)
        for candidate in (ClosedBinaryLabel.PHASE2_B, ClosedBinaryLabel.PHASE2_C, ClosedBinaryLabel.PHASE2_D)
    ),
    ComparisonProfile.PHASE3_VS_PHASE0A: frozenset({(ClosedBinaryLabel.PHASE0A, ClosedBinaryLabel.PHASE3)}),
    ComparisonProfile.PHASE3_ZLIB_ON_VS_OFF: frozenset(
        {
            (ClosedBinaryLabel.LOW_MEMORY_DEFAULT, ClosedBinaryLabel.LOW_MEMORY_ZLIB),
            (ClosedBinaryLabel.LOW_MEMORY_ZMIJ, ClosedBinaryLabel.LOW_MEMORY_ZLIB_ZMIJ),
        }
    ),
    ComparisonProfile.PHASE3_ZMIJ_ON_VS_OFF: frozenset(
        {
            (ClosedBinaryLabel.LOW_MEMORY_DEFAULT, ClosedBinaryLabel.LOW_MEMORY_ZMIJ),
            (ClosedBinaryLabel.LOW_MEMORY_ZLIB, ClosedBinaryLabel.LOW_MEMORY_ZLIB_ZMIJ),
        }
    ),
    ComparisonProfile.PHASE4_VS_PHASE3: frozenset({(ClosedBinaryLabel.PHASE3, ClosedBinaryLabel.PHASE4)}),
    ComparisonProfile.PHASE4_VS_PHASE0A: frozenset({(ClosedBinaryLabel.PHASE0A, ClosedBinaryLabel.PHASE4)}),
    ComparisonProfile.PHASE5_VS_PHASE0A: frozenset({(ClosedBinaryLabel.PHASE0A, ClosedBinaryLabel.PHASE5)}),
}


@dataclass(frozen=True)
class PwsLocalArtifacts:
    result_path: Path
    stdout_path: Path
    stderr_path: Path
    driver_log_path: Path
    log_root: Path


@dataclass
class AppendOnlyAttemptLedger:
    attempt_directory: Path
    local_root: Path
    identity: BenchmarkIdentity
    comparison_key: str
    attempt_number: int
    previous_attempt_head_sha256: str | None
    head_sha256: str
    state: AttemptState = AttemptState.CREATED
    terminal_verdict: HarnessVerdict | None = None
    terminal_raw_log_sha256: str | None = None
    terminal_primary_verdict: HarnessVerdict | None = None
    cleanup_only: bool = False
    recovery_primary_verdict: HarnessVerdict | None = None
    journal_head_sha256: str = ''
    first_group_sha256: str | None = None
    expanded_group_sha256: str | None = None
    _record_head_sha256: str = ''
    _checkpoint_head_sha256: str = ''
    _record_count: int = 0
    _sample_payloads: dict[tuple[str, int, str], dict[str, Any]] = field(default_factory=dict)
    _plan_payloads: dict[tuple[str, int, str], dict[str, Any]] = field(default_factory=dict)
    _inherited_plan_payloads: tuple[dict[str, Any], ...] = ()
    _prepared_evidence: dict[str, str] | None = None

    @classmethod
    def create(
        cls,
        root: Path,
        identity: BenchmarkIdentity,
        *,
        comparison_key: str,
    ) -> AppendOnlyAttemptLedger:
        if not _is_sha256(comparison_key):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'comparison_key must be 64 lowercase hex')
        trusted_local_root = _trusted_local_root()
        safe_local_root = _safe_harness_path(
            trusted_local_root,
            allowed_roots=(trusted_local_root,),
            purpose='attempt local root is invalid',
            create_parent=True,
        )
        if _normal_path(root).absolute() != (trusted_local_root / 'batches').absolute():
            raise HarnessFailure(
                HarnessVerdict.INCOMPLETE_EVIDENCE, 'attempt root must equal trusted local root/batches'
            )
        safe_root = _safe_harness_path(
            root,
            allowed_roots=(safe_local_root,),
            purpose='attempt root must stay below the ignored local root',
            create_parent=True,
        )
        comparison_directory = _safe_harness_path(
            safe_root / comparison_key,
            allowed_roots=(safe_root,),
            purpose='comparison_key escaped the attempt root',
            create_parent=True,
        )
        attempts = sorted(path for path in _io_path(comparison_directory).glob('attempt-*') if path.is_dir())
        previous_head: str | None = None
        inherited: tuple[dict[str, Any], ...] = ()
        cleanup_only = False
        recovery_primary: HarnessVerdict | None = None
        if attempts:
            previous = cls.load(attempts[-1], identity, strict_identity=False)
            if previous.terminal_verdict is None:
                return cls.load(attempts[-1], identity)
            if previous.terminal_verdict not in (
                HarnessVerdict.ENVIRONMENT_DRIFT,
                HarnessVerdict.REFERENCE_FAILED,
                HarnessVerdict.CLEANUP_FAILED,
            ):
                raise HarnessFailure(
                    HarnessVerdict.INCOMPLETE_EVIDENCE,
                    'failed candidate SHA cannot be retried after candidate, correctness, '
                    'gate, or inconclusive failure',
                )
            previous_head = previous.head_sha256
            if previous.terminal_verdict is HarnessVerdict.CLEANUP_FAILED:
                inherited = previous.all_planned_output_payloads()
                cleanup_only = True
                recovery_primary = previous.terminal_primary_verdict

        number = len(attempts) + 1
        attempt_directory = _io_path(comparison_directory / f'attempt-{number:04d}')
        attempt_directory.mkdir()
        (attempt_directory / 'records').mkdir()
        (attempt_directory / 'checkpoints').mkdir()
        _io_path(comparison_directory / 'journal').mkdir(exist_ok=True)
        _safe_harness_path(
            attempt_directory,
            allowed_roots=(safe_root,),
            purpose='attempt directory escaped local root after creation',
            create_parent=False,
        )
        metadata = {
            'comparison_key': comparison_key,
            'attempt_number': number,
            'identity': asdict(identity),
            'previous_attempt_head_sha256': previous_head,
            'reason': 'ENVIRONMENT_RECOVERED' if previous_head else 'FORMAL_START',
            'inherited_planned_outputs': inherited,
            'cleanup_only': cleanup_only,
            'recovery_primary_verdict': recovery_primary.value if recovery_primary else None,
        }
        metadata_bytes = _canonical_json(metadata)
        _write_create_new(attempt_directory / 'metadata.json', metadata_bytes, allowed_root=attempt_directory)
        metadata_sha = hashlib.sha256(metadata_bytes).hexdigest()
        ledger = cls(
            attempt_directory=attempt_directory,
            local_root=safe_local_root,
            identity=identity,
            comparison_key=comparison_key,
            attempt_number=number,
            previous_attempt_head_sha256=previous_head,
            head_sha256=metadata_sha,
            _record_head_sha256=metadata_sha,
            _checkpoint_head_sha256=metadata_sha,
            _inherited_plan_payloads=inherited,
            cleanup_only=cleanup_only,
            recovery_primary_verdict=recovery_primary,
        )
        ledger._append_journal_anchor()
        return ledger

    @classmethod
    def load(
        cls,
        directory: Path,
        identity: BenchmarkIdentity,
        *,
        strict_identity: bool = True,
    ) -> AppendOnlyAttemptLedger:
        normal_directory = _normal_path(directory).absolute()
        try:
            local_root = normal_directory.parents[2]
        except IndexError as exc:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'attempt directory is outside local root') from exc
        if normal_directory.parent.parent.name != 'batches':
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'attempt directory layout is invalid')
        if local_root.absolute() != _trusted_local_root():
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'attempt directory is outside trusted local root')
        normal_directory = _safe_harness_path(
            normal_directory,
            allowed_roots=(local_root,),
            purpose='attempt directory escaped local root',
            create_parent=False,
        )
        directory = _io_path(normal_directory)
        metadata_path = directory / 'metadata.json'
        metadata_raw = metadata_path.read_bytes()
        metadata = json.loads(metadata_raw)
        comparison_key = metadata.get('comparison_key')
        if not _is_sha256(comparison_key):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'attempt metadata comparison_key is invalid')
        stored_identity = metadata.get('identity')
        if (
            not isinstance(stored_identity, dict)
            or stored_identity.get('candidate_sha256') != identity.candidate_sha256
        ):
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'candidate SHA differs from attempt metadata')
        if strict_identity and stored_identity != asdict(identity):
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'attempt identity changed during resume')

        metadata_sha = hashlib.sha256(metadata_raw).hexdigest()
        record_head = metadata_sha
        checkpoint_head = metadata_sha
        sample_payloads: dict[tuple[str, int, str], dict[str, Any]] = {}
        plan_payloads: dict[tuple[str, int, str], dict[str, Any]] = {}
        first_group_sha256: str | None = None
        expanded_group_sha256: str | None = None
        cleanup_complete = False
        prepared_evidence: dict[str, str] | None = None
        evidence_committed = False
        committed_prior_record_head: str | None = None
        committed_prior_checkpoint_head: str | None = None
        records = sorted((directory / 'records').glob('*.json'))
        checkpoints = sorted((directory / 'checkpoints').glob('*.json'))
        missing_committed_checkpoint = len(records) == len(checkpoints) + 1
        if len(records) != len(checkpoints) and not missing_committed_checkpoint:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'record/checkpoint count mismatch')
        paired_records = records[:-1] if missing_committed_checkpoint else records
        for index, (record_path, checkpoint_path) in enumerate(zip(paired_records, checkpoints, strict=True), start=1):
            expected_prefix = f'{index:04d}-'
            if not record_path.name.startswith(expected_prefix) or checkpoint_path.name != f'{index:04d}.json':
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'ledger record sequence is not contiguous')
            raw = record_path.read_bytes()
            record = json.loads(raw)
            previous_record_head = record_head
            if record.get('previous_record_sha256') != previous_record_head:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'attempt ledger hash chain is broken')
            record_head = hashlib.sha256(raw).hexdigest()
            checkpoint_raw = checkpoint_path.read_bytes()
            checkpoint = json.loads(checkpoint_raw)
            previous_checkpoint_head = checkpoint_head
            expected_checkpoint = {
                'record_count': index,
                'record_sha256': record_head,
                'previous_checkpoint_sha256': previous_checkpoint_head,
            }
            if checkpoint != expected_checkpoint:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'durable checkpoint chain is broken')
            checkpoint_head = hashlib.sha256(checkpoint_raw).hexdigest()
            kind = record.get('kind')
            if evidence_committed or (cleanup_complete and kind not in ('evidence-prepared', 'evidence-committed')):
                raise HarnessFailure(
                    HarnessVerdict.INCOMPLETE_EVIDENCE,
                    'success record sequence has a trailing record',
                )
            if kind == 'sample':
                key = (record['metric'], record['global_round'], record['role'])
                if key in sample_payloads:
                    raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'sample record is duplicated')
                sample_payloads[key] = record['payload']
            elif kind == 'planned-output':
                key = (record['metric'], record['global_round'], record['role'])
                if key in plan_payloads:
                    raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output record is duplicated')
                plan_payloads[key] = _validate_planned_output_payload(record['payload'])
            elif kind == 'first-group':
                if first_group_sha256 is not None:
                    raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'first group record is duplicated')
                first_group_sha256 = record_head
            elif kind == 'expanded-group':
                if expanded_group_sha256 is not None:
                    raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'expanded group record is duplicated')
                expanded_group_sha256 = record_head
            elif kind == 'cleanup-complete':
                if cleanup_complete or evidence_committed or first_group_sha256 is None:
                    raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'cleanup success record order is invalid')
                cleanup_complete = True
            elif kind == 'evidence-prepared':
                if not cleanup_complete or prepared_evidence is not None or evidence_committed:
                    raise HarnessFailure(
                        HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence record order is invalid'
                    )
                prepared_evidence = _validate_prepared_evidence(record)
            elif kind == 'evidence-committed':
                if evidence_committed or not cleanup_complete or prepared_evidence is None:
                    raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'evidence success record order is invalid')
                _validate_committed_tail_record(
                    record_path,
                    raw,
                    record,
                    index=index,
                    previous_record_head=previous_record_head,
                    prepared_evidence=prepared_evidence,
                )
                evidence_committed = True
                committed_prior_record_head = previous_record_head
                committed_prior_checkpoint_head = previous_checkpoint_head

        if missing_committed_checkpoint:
            if (directory / 'terminal.json').exists() or not cleanup_complete or prepared_evidence is None:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'record/checkpoint count mismatch')
            index = len(records)
            record_path = records[-1]
            raw = record_path.read_bytes()
            record = json.loads(raw)
            _validate_committed_tail_record(
                record_path,
                raw,
                record,
                index=index,
                previous_record_head=record_head,
                prepared_evidence=prepared_evidence,
            )
            committed_prior_record_head = record_head
            committed_prior_checkpoint_head = checkpoint_head
            record_head = hashlib.sha256(raw).hexdigest()
            checkpoint = {
                'record_count': index,
                'record_sha256': record_head,
                'previous_checkpoint_sha256': checkpoint_head,
            }
            checkpoint_raw = _canonical_json(checkpoint)
            _write_create_new(
                directory / 'checkpoints' / f'{index:04d}.json',
                checkpoint_raw,
                allowed_root=directory,
            )
            checkpoint_head = hashlib.sha256(checkpoint_raw).hexdigest()
            evidence_committed = True

        state = AttemptState.CREATED
        terminal_verdict: HarnessVerdict | None = None
        terminal_raw: str | None = None
        terminal_primary: HarnessVerdict | None = None
        head = record_head
        terminal_path = directory / 'terminal.json'
        if terminal_path.exists():
            terminal_raw_bytes = terminal_path.read_bytes()
            terminal = json.loads(terminal_raw_bytes)
            if terminal.get('record_count') != len(records) or terminal.get('record_head_sha256') != record_head:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'sealed attempt record count/head mismatch')
            if terminal.get('checkpoint_head_sha256') != checkpoint_head:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'sealed attempt checkpoint head mismatch')
            try:
                terminal_verdict = HarnessVerdict(terminal['verdict'])
                terminal_primary = (
                    HarnessVerdict(terminal['primary_verdict']) if terminal.get('primary_verdict') else None
                )
            except (KeyError, ValueError) as exc:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal verdict is invalid') from exc
            terminal_raw = terminal.get('raw_log_sha256')
            if terminal_raw is not None and not _is_sha256(terminal_raw):
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal raw log SHA is invalid')
            state = AttemptState.FAILED
            head = hashlib.sha256(terminal_raw_bytes).hexdigest()

        inherited_raw = metadata.get('inherited_planned_outputs', [])
        if not isinstance(inherited_raw, list):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'inherited planned-output list is invalid')
        inherited = tuple(_validate_planned_output_payload(item) for item in inherited_raw)
        cleanup_only = metadata.get('cleanup_only', False)
        if not isinstance(cleanup_only, bool):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'cleanup_only metadata must be boolean')
        try:
            recovery_primary = (
                HarnessVerdict(metadata['recovery_primary_verdict'])
                if metadata.get('recovery_primary_verdict')
                else None
            )
        except ValueError as exc:
            raise HarnessFailure(
                HarnessVerdict.INCOMPLETE_EVIDENCE, 'cleanup recovery primary verdict is invalid'
            ) from exc
        if inherited and not cleanup_only:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'inherited outputs require cleanup-only metadata')
        if terminal_verdict is None:
            if evidence_committed:
                state = AttemptState.EVIDENCE_COMMITTED
            elif cleanup_complete:
                state = AttemptState.CLEANUP_COMPLETE
            elif expanded_group_sha256:
                state = AttemptState.EXPANDED_GROUP_COMPLETE
            elif first_group_sha256:
                state = AttemptState.FIRST_GROUP_COMPLETE
        attempt_number = int(metadata['attempt_number'])
        previous_head = metadata.get('previous_attempt_head_sha256')
        if attempt_number > 1:
            previous_directory = directory.parent / f'attempt-{attempt_number - 1:04d}'
            previous = cls.load(previous_directory, identity, strict_identity=False)
            if previous_head != previous.head_sha256:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'previous attempt head link is broken')
        elif previous_head is not None:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'first attempt cannot link a previous head')
        journal_head, latest_anchors = _load_comparison_journal(directory.parent)
        expected_anchor = _journal_state_payload(
            attempt_number=attempt_number,
            record_count=len(records),
            record_head_sha256=record_head,
            checkpoint_head_sha256=checkpoint_head,
            terminal_present=terminal_verdict is not None,
            terminal_head_sha256=head if terminal_verdict is not None else None,
            verdict=terminal_verdict,
        )
        if latest_anchors.get(attempt_number) != expected_anchor:
            if (
                terminal_verdict is not None
                or not evidence_committed
                or committed_prior_record_head is None
                or committed_prior_checkpoint_head is None
            ):
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'comparison journal anchor mismatch')
            prior_anchor = _journal_state_payload(
                attempt_number=attempt_number,
                record_count=len(records) - 1,
                record_head_sha256=committed_prior_record_head,
                checkpoint_head_sha256=committed_prior_checkpoint_head,
                terminal_present=False,
                terminal_head_sha256=None,
                verdict=None,
            )
            if latest_anchors.get(attempt_number) != prior_anchor:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'comparison journal anchor mismatch')
            journal_head = _append_recovered_journal_anchor(directory.parent, journal_head, expected_anchor)
        return cls(
            attempt_directory=directory,
            local_root=local_root.resolve(),
            identity=identity,
            comparison_key=comparison_key,
            attempt_number=attempt_number,
            previous_attempt_head_sha256=previous_head,
            head_sha256=head,
            state=state,
            terminal_verdict=terminal_verdict,
            terminal_raw_log_sha256=terminal_raw,
            terminal_primary_verdict=terminal_primary,
            cleanup_only=cleanup_only,
            recovery_primary_verdict=recovery_primary,
            journal_head_sha256=journal_head,
            first_group_sha256=first_group_sha256,
            expanded_group_sha256=expanded_group_sha256,
            _record_head_sha256=record_head,
            _checkpoint_head_sha256=checkpoint_head,
            _record_count=len(records),
            _sample_payloads=sample_payloads,
            _plan_payloads=plan_payloads,
            _inherited_plan_payloads=inherited,
            _prepared_evidence=prepared_evidence,
        )

    def _append(self, kind: str, payload: dict[str, Any]) -> str:
        if self.terminal_verdict is not None:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal attempt cannot accept more records')
        if self.state is AttemptState.CLEANUP_COMPLETE and kind not in ('evidence-prepared', 'evidence-committed'):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'cleanup success cannot accept benchmark records')
        if self.state is AttemptState.EVIDENCE_COMMITTED:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'committed success cannot accept more records')
        if self.cleanup_only:
            raise HarnessFailure(
                HarnessVerdict.INCOMPLETE_EVIDENCE, 'cleanup-only attempt cannot accept benchmark records'
            )
        sequence = self._record_count + 1
        record = {'kind': kind, 'previous_record_sha256': self._record_head_sha256, **payload}
        raw = _canonical_json(record)
        record_sha = hashlib.sha256(raw).hexdigest()
        _write_create_new(
            self.attempt_directory / 'records' / f'{sequence:04d}-{kind}.json',
            raw,
            allowed_root=self.attempt_directory,
        )
        checkpoint = {
            'record_count': sequence,
            'record_sha256': record_sha,
            'previous_checkpoint_sha256': self._checkpoint_head_sha256,
        }
        checkpoint_raw = _canonical_json(checkpoint)
        _write_create_new(
            self.attempt_directory / 'checkpoints' / f'{sequence:04d}.json',
            checkpoint_raw,
            allowed_root=self.attempt_directory,
        )
        self._record_count = sequence
        self._record_head_sha256 = record_sha
        self._checkpoint_head_sha256 = hashlib.sha256(checkpoint_raw).hexdigest()
        self.head_sha256 = record_sha
        self._append_journal_anchor()
        return record_sha

    def _append_journal_anchor(self) -> str:
        comparison_directory = self.attempt_directory.parent
        journal_head, _ = _load_comparison_journal(comparison_directory)
        state = _journal_state_payload(
            attempt_number=self.attempt_number,
            record_count=self._record_count,
            record_head_sha256=self._record_head_sha256,
            checkpoint_head_sha256=self._checkpoint_head_sha256,
            terminal_present=self.terminal_verdict is not None,
            terminal_head_sha256=self.head_sha256 if self.terminal_verdict is not None else None,
            verdict=self.terminal_verdict,
        )
        entry = {'previous_journal_sha256': journal_head or None, **state}
        raw = _canonical_json(entry)
        journal_directory = comparison_directory / 'journal'
        sequence = len(tuple(journal_directory.glob('*.json'))) + 1
        _write_create_new(
            journal_directory / f'{sequence:06d}.json',
            raw,
            allowed_root=comparison_directory,
        )
        self.journal_head_sha256 = hashlib.sha256(raw).hexdigest()
        return self.journal_head_sha256

    def record_sample(
        self,
        metric: MetricName,
        global_round: int,
        role: BinaryRole,
        payload: dict[str, Any],
    ) -> str:
        key = (metric, global_round, role)
        if key in self._sample_payloads:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'existing round record overwrite is forbidden')
        digest = self._append(
            'sample',
            {'metric': metric, 'global_round': global_round, 'role': role, 'payload': payload},
        )
        self._sample_payloads[key] = payload
        return digest

    def record_planned_output(
        self,
        metric: MetricName,
        global_round: int,
        role: BinaryRole,
        payload: dict[str, Any],
    ) -> str:
        validated = _validate_planned_output_payload(payload)
        key = (metric, global_round, role)
        if (validated['metric'], validated['global_round'], validated['role']) != key:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output key/payload mismatch')
        existing = self._plan_payloads.get(key)
        if existing is not None:
            if existing != validated:
                raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned output changed during resume')
            return self.head_sha256
        digest = self._append(
            'planned-output',
            {'metric': metric, 'global_round': global_round, 'role': role, 'payload': validated},
        )
        self._plan_payloads[key] = validated
        return digest

    def sample_payload(self, metric: MetricName, global_round: int, role: BinaryRole) -> dict[str, Any] | None:
        return self._sample_payloads.get((metric, global_round, role))

    def missing_samples(
        self,
        metrics: tuple[MetricName, ...],
        rounds: tuple[int, ...],
        roles: tuple[BinaryRole, ...],
    ) -> tuple[tuple[MetricName, int, BinaryRole], ...]:
        return tuple(
            (metric, round_number, role)
            for metric in metrics
            for round_number in rounds
            for role in roles
            if (metric, round_number, role) not in self._sample_payloads
        )

    def all_planned_output_payloads(self) -> tuple[dict[str, Any], ...]:
        combined: dict[str, dict[str, Any]] = {}
        for payload in (*self._inherited_plan_payloads, *self._plan_payloads.values()):
            combined[payload['relative_path']] = payload
        return tuple(combined[key] for key in sorted(combined))

    def commit_first_group(self, groups: dict[str, Any]) -> str:
        if self.state is not AttemptState.CREATED or self.first_group_sha256 is not None:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'first group record overwrite is forbidden')
        if set(groups) != {'wall', 'pws'}:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'first group must commit wall and pws together')
        digest = self._append('first-group', {'groups': groups})
        self.first_group_sha256 = digest
        self.state = AttemptState.FIRST_GROUP_COMPLETE
        return digest

    def commit_expanded_group(self, groups: dict[str, Any], *, first_group_sha256: str) -> str:
        if self.state is not AttemptState.FIRST_GROUP_COMPLETE or self.expanded_group_sha256 is not None:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'expanded group record overwrite is forbidden')
        if first_group_sha256 != self.first_group_sha256:
            raise HarnessFailure(
                HarnessVerdict.INCOMPLETE_EVIDENCE, 'expanded group does not link original first group SHA'
            )
        if set(groups) != {'wall', 'pws'}:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'expanded group must commit wall and pws together')
        digest = self._append('expanded-group', {'first_group_sha256': first_group_sha256, 'groups': groups})
        self.expanded_group_sha256 = digest
        self.state = AttemptState.EXPANDED_GROUP_COMPLETE
        return digest

    def mark_cleanup_complete(self) -> str:
        if self.state not in (AttemptState.FIRST_GROUP_COMPLETE, AttemptState.EXPANDED_GROUP_COMPLETE):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'cleanup success transition is invalid')
        digest = self._append('cleanup-complete', {'planned_output_count': len(self.all_planned_output_payloads())})
        self.state = AttemptState.CLEANUP_COMPLETE
        return digest

    def prepare_evidence(
        self,
        *,
        artifact_basename: str,
        artifact_sha256: str,
        artifact_content: str,
    ) -> str:
        if self.state is not AttemptState.CLEANUP_COMPLETE or self._prepared_evidence is not None:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence transition is invalid')
        payload = _validate_prepared_evidence(
            {
                'kind': 'evidence-prepared',
                'previous_record_sha256': self._record_head_sha256,
                'artifact_basename': artifact_basename,
                'artifact_sha256': artifact_sha256,
                'artifact_content': artifact_content,
            }
        )
        _rebuild_prepared_artifact(payload)
        digest = self._append('evidence-prepared', payload)
        self._prepared_evidence = payload
        return digest

    def prepared_evidence(self) -> dict[str, str] | None:
        return dict(self._prepared_evidence) if self._prepared_evidence is not None else None

    def mark_evidence_committed(self, *, artifact_sha256: str) -> str:
        if (
            self.state is not AttemptState.CLEANUP_COMPLETE
            or self._prepared_evidence is None
            or artifact_sha256 != self._prepared_evidence['artifact_sha256']
        ):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'evidence success transition is invalid')
        digest = self._append('evidence-committed', {'artifact_sha256': artifact_sha256})
        self.state = AttemptState.EVIDENCE_COMMITTED
        return digest

    def finish(
        self,
        verdict: HarnessVerdict,
        *,
        raw_log_sha256: str | None = None,
        primary_verdict: HarnessVerdict | None = None,
    ) -> str:
        if self.terminal_verdict is not None or (self.attempt_directory / 'terminal.json').exists():
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal attempt is already sealed')
        if not isinstance(verdict, HarnessVerdict) or verdict is HarnessVerdict.VALIDATED:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal failure verdict must be closed')
        if primary_verdict is not None and not isinstance(primary_verdict, HarnessVerdict):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal primary verdict must be closed')
        if raw_log_sha256 is not None and not _is_sha256(raw_log_sha256):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal raw log SHA is invalid')
        terminal = {
            'verdict': verdict.value,
            'primary_verdict': primary_verdict.value if primary_verdict else None,
            'raw_log_sha256': raw_log_sha256,
            'record_count': self._record_count,
            'record_head_sha256': self._record_head_sha256,
            'checkpoint_head_sha256': self._checkpoint_head_sha256,
        }
        raw = _canonical_json(terminal)
        _write_create_new(self.attempt_directory / 'terminal.json', raw, allowed_root=self.attempt_directory)
        self.terminal_verdict = verdict
        self.terminal_raw_log_sha256 = raw_log_sha256
        self.terminal_primary_verdict = primary_verdict
        self.state = AttemptState.FAILED
        self.head_sha256 = hashlib.sha256(raw).hexdigest()
        self._append_journal_anchor()
        return self.head_sha256


def _validate_prepared_evidence(record: dict[str, Any]) -> dict[str, str]:
    if set(record) != {
        'kind',
        'previous_record_sha256',
        'artifact_basename',
        'artifact_sha256',
        'artifact_content',
    }:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence record schema is invalid')
    basename = record['artifact_basename']
    sha256 = record['artifact_sha256']
    content = record['artifact_content']
    if (
        not isinstance(basename, str)
        or Path(basename).name != basename
        or not basename.startswith('benchmark-')
        or not basename.endswith('.json')
        or not _is_sha256(sha256)
        or not isinstance(content, str)
        or hashlib.sha256(content.encode('utf-8')).hexdigest() != sha256
    ):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence values are invalid')
    return {'artifact_basename': basename, 'artifact_sha256': sha256, 'artifact_content': content}


def _rebuild_prepared_artifact(payload: dict[str, str]) -> Any:
    policy = EvidenceSanitizer.closed_policy()
    try:
        source = policy.read_benchmark_manifest(
            payload['artifact_basename'],
            payload['artifact_content'].encode('utf-8'),
        )
        artifact = policy.build_benchmark_manifest(source)
    except ValueError as exc:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence typed rebuild failed') from exc
    if (
        artifact.file_name != payload['artifact_basename']
        or artifact.content != payload['artifact_content']
        or hashlib.sha256(artifact.content.encode('utf-8')).hexdigest() != payload['artifact_sha256']
    ):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence changed during typed rebuild')
    return artifact


def _validate_committed_tail_record(
    record_path: Path,
    raw: bytes,
    record: object,
    *,
    index: int,
    previous_record_head: str,
    prepared_evidence: dict[str, str],
) -> None:
    if (
        record_path.name != f'{index:04d}-evidence-committed.json'
        or not isinstance(record, dict)
        or set(record) != {'kind', 'previous_record_sha256', 'artifact_sha256'}
        or record.get('kind') != 'evidence-committed'
        or record.get('previous_record_sha256') != previous_record_head
        or record.get('artifact_sha256') != prepared_evidence['artifact_sha256']
        or raw != _canonical_json(record)
    ):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'committed evidence tail is invalid')


def _append_recovered_journal_anchor(
    comparison_directory: Path,
    journal_head: str,
    state: dict[str, Any],
) -> str:
    journal_directory = comparison_directory / 'journal'
    sequence = len(tuple(journal_directory.glob('*.json'))) + 1
    raw = _canonical_json({'previous_journal_sha256': journal_head or None, **state})
    _write_create_new(
        journal_directory / f'{sequence:06d}.json',
        raw,
        allowed_root=comparison_directory,
    )
    return hashlib.sha256(raw).hexdigest()


def _journal_state_payload(
    *,
    attempt_number: int,
    record_count: int,
    record_head_sha256: str,
    checkpoint_head_sha256: str,
    terminal_present: bool,
    terminal_head_sha256: str | None,
    verdict: HarnessVerdict | None,
) -> dict[str, Any]:
    return {
        'attempt_number': attempt_number,
        'record_count': record_count,
        'record_head_sha256': record_head_sha256,
        'checkpoint_head_sha256': checkpoint_head_sha256,
        'terminal_present': terminal_present,
        'terminal_head_sha256': terminal_head_sha256,
        'verdict': verdict.value if verdict else None,
    }


def _load_comparison_journal(comparison_directory: Path) -> tuple[str, dict[int, dict[str, Any]]]:
    # 外部 journal 用于识别 attempt 目录内的简单回退；Task 6 会把最终 journal head 绑定到版本化 manifest。
    # 恶意同时删除 journal 与 attempt 尾部不属于本地 Task 2 威胁模型。
    journal_directory = comparison_directory / 'journal'
    if not journal_directory.is_dir():
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'comparison journal is missing')
    head = ''
    latest: dict[int, dict[str, Any]] = {}
    for index, path in enumerate(sorted(journal_directory.glob('*.json')), start=1):
        if path.name != f'{index:06d}.json':
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'comparison journal sequence is not contiguous')
        raw = path.read_bytes()
        entry = json.loads(raw)
        if entry.get('previous_journal_sha256') != (head or None):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'comparison journal hash chain is broken')
        state = {key: value for key, value in entry.items() if key != 'previous_journal_sha256'}
        attempt_number = state.get('attempt_number')
        if not isinstance(attempt_number, int) or isinstance(attempt_number, bool) or attempt_number < 1:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'comparison journal attempt number is invalid')
        latest[attempt_number] = state
        head = hashlib.sha256(raw).hexdigest()
    return head, latest


def derive_batch_id(request: PairedBenchmarkRequest, identity: BenchmarkIdentity) -> str:
    payload = {'profile': request.comparison_profile.value, 'pipeline': request.pipeline, **asdict(identity)}
    return hashlib.sha256(_canonical_json(payload)).hexdigest()


def _mandatory_paired_expansion_plans(
    *,
    wall_requires_expansion: bool,
    pws_requires_expansion: bool,
) -> tuple[tuple[RoundPlan, ...], tuple[RoundPlan, ...]]:
    if not (wall_requires_expansion or pws_requires_expansion):
        return (), ()
    plans = build_round_plan(global_round_start=6, round_count=5)
    return plans, plans


def run_normal_wall_group(request: MetricGroupRequest) -> MetricGroup:
    if request.metric != 'wall':
        raise ValueError('normal wall runner accepts wall metric only')
    _validate_trusted_request_paths(request.benchmark)
    identity = _capture_identity(request.benchmark)
    ledger = AppendOnlyAttemptLedger.load(request.attempt_directory, identity)
    if ledger.terminal_verdict is not None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal attempt cannot be resumed')
    if request.first_group_sha256 is not None and request.first_group_sha256 != ledger.first_group_sha256:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'expanded group first-group SHA does not match ledger')

    role_executables = {
        'reference': request.benchmark.reference_executable,
        'candidate': request.benchmark.candidate_executable,
    }
    rule = PROFILE_RULES[request.benchmark.comparison_profile][request.benchmark.pipeline]
    schemas: dict[BinaryRole, RuntimeSchema] = {
        'reference': rule.reference_schema,
        'candidate': rule.candidate_schema,
    }
    cleanup_paths = list(_planned_paths(ledger.all_planned_output_payloads()))
    if ledger.cleanup_only:
        cleanup_errors = _cleanup_all(cleanup_paths)
        if cleanup_errors:
            final_verdict = HarnessVerdict.CLEANUP_FAILED
            primary_verdict = ledger.recovery_primary_verdict
        else:
            final_verdict = ledger.recovery_primary_verdict or HarnessVerdict.ENVIRONMENT_DRIFT
            primary_verdict = None
        ledger.finish(final_verdict, primary_verdict=primary_verdict)
        raise HarnessFailure(
            final_verdict,
            'cleanup-only recovery completed' if not cleanup_errors else 'cleanup-only recovery failed',
            primary_verdict=primary_verdict,
        )
    pairs: list[PairedRound] = []
    result_group: MetricGroup | None = None
    primary_error: HarnessFailure | None = None

    initial_cleanup_errors = _cleanup_all(cleanup_paths)
    if initial_cleanup_errors:
        primary_error = HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'historical workbook cleanup was transient')
    else:
        try:
            for plan in request.plans:
                captured: dict[BinaryRole, MetricSample] = {}
                for role in plan.order:
                    _assert_identity_unchanged(identity, _capture_identity(request.benchmark))
                    existing = ledger.sample_payload(request.metric, plan.global_round, role)
                    if existing is not None:
                        captured[role] = _sample_from_payload(existing)
                        continue
                    payload = _planned_output_payload(request, identity, plan.global_round, role)
                    ledger.record_planned_output(request.metric, plan.global_round, role, payload)
                    output = next(iter(_planned_paths((payload,))))
                    cleanup_paths.append(output)
                    try:
                        capture = run_rust_normal_captured(
                            role_executables[role],
                            request.benchmark.pipeline,
                            request.benchmark.input_path,
                            output,
                            schema=schemas[role],
                            local_log_root=request.benchmark.local_root / 'raw-logs',
                            workbook_oracle_fn=_stable_workbook_oracle,
                        )
                    except RustNormalProcessError as exc:
                        cleanup_paths.append(request.benchmark.local_root / 'raw-logs' / f'{exc.log_sha256}.json')
                        verdict = (
                            HarnessVerdict.REFERENCE_FAILED if role == 'reference' else HarnessVerdict.CANDIDATE_FAILED
                        )
                        raise HarnessFailure(
                            verdict,
                            f'{role} process failed with exit code {exc.returncode}',
                            raw_log_sha256=exc.log_sha256,
                        ) from exc
                    except RustNormalValidationError as exc:
                        cleanup_paths.append(request.benchmark.local_root / 'raw-logs' / f'{exc.log_sha256}.json')
                        verdict = (
                            HarnessVerdict.REFERENCE_FAILED
                            if role == 'reference'
                            else HarnessVerdict.CORRECTNESS_FAILED
                        )
                        raise HarnessFailure(
                            verdict,
                            f'{role} runtime or workbook validation failed',
                            raw_log_sha256=exc.log_sha256,
                        ) from exc
                    except Exception as exc:
                        verdict = (
                            HarnessVerdict.REFERENCE_FAILED
                            if role == 'reference'
                            else HarnessVerdict.CORRECTNESS_FAILED
                        )
                        raise HarnessFailure(verdict, f'{role} capture boundary failed') from exc
                    _assert_identity_unchanged(identity, _capture_identity(request.benchmark))
                    cleanup_paths.append(
                        request.benchmark.local_root / 'raw-logs' / f'{capture.local_unversioned_log_sha256}.json'
                    )
                    sample = _metric_sample(
                        role,
                        plan,
                        identity,
                        (capture.normal_run, capture.local_unversioned_log_sha256),
                    )
                    ledger.record_sample(request.metric, plan.global_round, role, _sample_to_payload(sample))
                    captured[role] = sample
                    try:
                        _remove_workbook(output)
                    except OSError as exc:
                        raise HarnessFailure(
                            HarnessVerdict.ENVIRONMENT_DRIFT,
                            'immediate workbook cleanup failed; outer cleanup will retry',
                        ) from exc
                if (
                    captured['reference'].normal_run.workbook_oracle_sha256
                    != captured['candidate'].normal_run.workbook_oracle_sha256
                ):
                    raise HarnessFailure(
                        HarnessVerdict.CORRECTNESS_FAILED,
                        'reference/candidate workbook oracle mismatch',
                        raw_log_sha256=captured['candidate'].local_unversioned_log_sha256,
                    )
                pairs.append(PairedRound(plan, captured['reference'], captured['candidate']))
            result_group = MetricGroup(
                request.batch_id,
                request.benchmark.pipeline,
                'wall',
                request.plans[0].global_round,  # type: ignore[arg-type]
                tuple(pairs),
            )
            validate_metric_group(result_group)
        except (KeyboardInterrupt, SystemExit) as interruption:
            cleanup_errors = _cleanup_all(cleanup_paths)
            if cleanup_errors:
                ledger.finish(HarnessVerdict.CLEANUP_FAILED)
                raise HarnessFailure(
                    HarnessVerdict.CLEANUP_FAILED,
                    f'workbook cleanup failed during interruption: {cleanup_errors!r}',
                ) from interruption
            raise
        except HarnessFailure as exc:
            primary_error = exc
        except Exception as exc:
            primary_error = HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'wall runner failed closed')
            primary_error.__cause__ = exc

    cleanup_errors = _cleanup_all(cleanup_paths)
    final_error = primary_error
    if cleanup_errors:
        final_error = HarnessFailure(
            HarnessVerdict.CLEANUP_FAILED,
            f'workbook cleanup failed: {cleanup_errors!r}',
            primary_verdict=primary_error.verdict if primary_error else None,
            raw_log_sha256=primary_error.raw_log_sha256 if primary_error else None,
        )
    if final_error is not None:
        ledger.finish(
            final_error.verdict,
            raw_log_sha256=final_error.raw_log_sha256,
            primary_verdict=final_error.primary_verdict,
        )
        raise final_error
    if result_group is None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'wall group produced no result')
    return result_group


def _metric_sample(
    role: BinaryRole,
    plan: RoundPlan,
    identity: BenchmarkIdentity,
    captured: tuple[NormalRunEvidence, str],
    *,
    metric_value: Decimal | None = None,
) -> MetricSample:
    normal_run, log_sha = captured
    return MetricSample(
        role=role,
        global_round=plan.global_round,
        metric_value=normal_run.external_wall_seconds if metric_value is None else metric_value,
        exit_code=0,
        input_sha256=identity.input_sha256,
        binary_sha256=identity.reference_sha256 if role == 'reference' else identity.candidate_sha256,
        git_head=identity.git_head,
        repository_state_sha256=identity.repository_state_sha256,
        machine_fingerprint_sha256=identity.machine_fingerprint_sha256,
        local_unversioned_log_sha256=log_sha,
        normal_run=normal_run,
    )


def _planned_output_payload(
    request: MetricGroupRequest,
    identity: BenchmarkIdentity,
    global_round: int,
    role: BinaryRole,
) -> dict[str, Any]:
    binary_sha = identity.reference_sha256 if role == 'reference' else identity.candidate_sha256
    relative = PurePosixPath(
        request.benchmark.pipeline,
        '.perf-runs',
        request.batch_id,
        request.metric,
        binary_sha,
        str(global_round),
        f'{role}.xlsx',
    ).as_posix()
    payload = {
        'pipeline': request.benchmark.pipeline,
        'batch_id': request.batch_id,
        'metric': request.metric,
        'binary_sha256': binary_sha,
        'global_round': global_round,
        'role': role,
        'relative_path': relative,
    }
    return _validate_planned_output_payload(payload)


def _validate_planned_output_payload(payload: object) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output payload must be an object')
    required = {
        'pipeline',
        'batch_id',
        'metric',
        'binary_sha256',
        'global_round',
        'role',
        'relative_path',
    }
    if set(payload) != required:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output payload fields are not closed')
    pipeline = payload['pipeline']
    metric = payload['metric']
    role = payload['role']
    batch_id = payload['batch_id']
    binary_sha = payload['binary_sha256']
    global_round = payload['global_round']
    if pipeline not in ('gb', 'sk') or metric not in ('wall', 'pws') or role not in ('reference', 'candidate'):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output identity is invalid')
    if not _is_lower_hex(batch_id, minimum=8, maximum=64) or not _is_lower_hex(binary_sha, minimum=8, maximum=64):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output hashes are invalid')
    if not isinstance(global_round, int) or isinstance(global_round, bool) or global_round not in range(1, 11):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output global round is invalid')
    expected = PurePosixPath(
        pipeline,
        '.perf-runs',
        batch_id,
        metric,
        binary_sha,
        str(global_round),
        f'{role}.xlsx',
    ).as_posix()
    if payload['relative_path'] != expected:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'planned-output relative path is not derivable')
    return dict(payload)


def _planned_paths(payloads: tuple[dict[str, Any], ...]) -> tuple[Path, ...]:
    root = repo_root()
    paths: list[Path] = []
    for payload in payloads:
        validated = _validate_planned_output_payload(payload)
        raw = root / 'data' / 'processed' / Path(validated['relative_path'])
        path = _safe_harness_path(
            raw,
            allowed_roots=(root / 'data' / 'processed',),
            purpose='planned output escaped data/processed',
            create_parent=False,
        )
        _io_path(path.parent).mkdir(parents=True, exist_ok=True)
        path = _safe_harness_path(
            path,
            allowed_roots=(root / 'data' / 'processed',),
            purpose='planned output escaped data/processed',
            create_parent=False,
        )
        paths.append(_io_path(path))
    return tuple(paths)


def _sample_to_payload(sample: MetricSample) -> dict[str, Any]:
    runtime = sample.normal_run.runtime
    return {
        'role': sample.role,
        'global_round': sample.global_round,
        'metric_value': str(sample.metric_value),
        'exit_code': sample.exit_code,
        'input_sha256': sample.input_sha256,
        'binary_sha256': sample.binary_sha256,
        'git_head': sample.git_head,
        'repository_state_sha256': sample.repository_state_sha256,
        'machine_fingerprint_sha256': sample.machine_fingerprint_sha256,
        'local_unversioned_log_sha256': sample.local_unversioned_log_sha256,
        'normal_run': {
            'external_wall_seconds': str(sample.normal_run.external_wall_seconds),
            'peak_working_set_bytes': sample.normal_run.peak_working_set_bytes,
            'workbook_oracle_sha256': sample.normal_run.workbook_oracle_sha256,
            'runtime': {
                **asdict(runtime),
                'stage_timings': [[name, str(value)] for name, value in runtime.stage_timings],
            },
        },
    }


def _sample_from_payload(payload: dict[str, Any]) -> MetricSample:
    normal_payload = payload['normal_run']
    runtime_payload = normal_payload['runtime']
    runtime = RuntimeEvidence(
        pipeline=runtime_payload['pipeline'],
        output_written=runtime_payload['output_written'],
        request_id_present=runtime_payload['request_id_present'],
        sheet_count=runtime_payload['sheet_count'],
        error_log_count=runtime_payload['error_log_count'],
        issue_type_counts=tuple(tuple(item) for item in runtime_payload['issue_type_counts']),
        quality_metrics=tuple(tuple(item) for item in runtime_payload['quality_metrics']),
        run_counts=tuple(tuple(item) for item in runtime_payload['run_counts']),
        stage_timings=tuple((item[0], Decimal(item[1])) for item in runtime_payload['stage_timings']),
        output_size_bytes=runtime_payload['output_size_bytes'],
        sheet_dimensions=tuple(runtime_payload['sheet_dimensions']),
        reader_snapshot_sha256=runtime_payload['reader_snapshot_sha256'],
    )
    normal_run = NormalRunEvidence(
        external_wall_seconds=Decimal(normal_payload['external_wall_seconds']),
        peak_working_set_bytes=normal_payload['peak_working_set_bytes'],
        runtime=runtime,
        workbook_oracle_sha256=normal_payload['workbook_oracle_sha256'],
    )
    return MetricSample(
        role=payload['role'],
        global_round=payload['global_round'],
        metric_value=Decimal(payload['metric_value']),
        exit_code=payload['exit_code'],
        input_sha256=payload['input_sha256'],
        binary_sha256=payload['binary_sha256'],
        git_head=payload['git_head'],
        repository_state_sha256=payload['repository_state_sha256'],
        machine_fingerprint_sha256=payload['machine_fingerprint_sha256'],
        local_unversioned_log_sha256=payload['local_unversioned_log_sha256'],
        normal_run=normal_run,
    )


def _capture_identity(request: PairedBenchmarkRequest) -> BenchmarkIdentity:
    root = repo_root()
    git_head = _run_git(root, 'rev-parse', 'HEAD').strip()
    status = _run_git(root, 'status', '--porcelain=v1', '--untracked-files=all')
    diff = _run_git(root, 'diff', '--binary', 'HEAD', '--')
    prior_claims = _prior_claim_fingerprint(root, request.prior_evidence_claims)
    repository_state = hashlib.sha256(f'{status}\n{diff}\n{prior_claims}'.encode()).hexdigest()
    return BenchmarkIdentity(
        _sha256(request.input_path),
        _sha256(request.reference_executable),
        _sha256(request.candidate_executable),
        git_head,
        repository_state,
        _capture_machine_evidence().fingerprint_sha256,
    )


def _prior_claim_fingerprint(root: Path, claims: tuple[UnverifiedPriorEvidenceClaim, ...]) -> str:
    rows: list[dict[str, str]] = []
    for claim in sorted(claims, key=lambda item: item.path_alias):
        path = _safe_harness_path(
            root / Path(claim.path_alias),
            allowed_roots=(root,),
            purpose='prior evidence escaped repository root',
            create_parent=False,
        )
        try:
            actual = _sha256(path)
        except OSError as exc:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'prior evidence is missing or unreadable') from exc
        if actual != claim.content_sha256:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'prior evidence content changed')
        rows.append({'path_alias': claim.path_alias, 'content_sha256': actual})
    return hashlib.sha256(_canonical_json(rows)).hexdigest()


def _assert_identity_unchanged(expected: BenchmarkIdentity, actual: BenchmarkIdentity) -> None:
    if actual != expected:
        raise HarnessFailure(
            HarnessVerdict.ENVIRONMENT_DRIFT,
            'input, executable, Git, repository, prior evidence, or machine drift detected',
        )


def validate_formal_repository_state(
    status_entries: tuple[str, ...],
    *,
    evidence_root: Path,
    prior_evidence_claims: tuple[UnverifiedPriorEvidenceClaim, ...],
    root: Path | None = None,
) -> None:
    root = (root or repo_root()).resolve()
    evidence_root = evidence_root.resolve()
    if not prior_evidence_claims:
        if status_entries:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'first formal batch requires a clean worktree')
        return
    approved = {claim.path_alias: claim for claim in prior_evidence_claims}
    seen: set[str] = set()
    for entry in status_entries:
        if not entry.startswith('?? '):
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'formal batch contains non-evidence worktree change')
        alias = entry[3:].replace('\\', '/')
        claim = approved.get(alias)
        if claim is None:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'formal batch contains non-evidence worktree change')
        path = _safe_harness_path(
            root / Path(alias),
            allowed_roots=(evidence_root,),
            purpose='prior evidence escaped evidence root',
            create_parent=False,
        )
        try:
            actual = _sha256(path)
        except OSError as exc:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'prior evidence is missing or unreadable') from exc
        if actual != claim.content_sha256:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'prior evidence content changed')
        seen.add(alias)
    if seen != set(approved):
        raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'approved prior evidence repository state is incomplete')


def _cleanup_all(paths: list[Path]) -> tuple[str, ...]:
    failures: list[str] = []
    for path in dict.fromkeys(paths):
        try:
            _remove_workbook(path)
        except Exception as exc:
            failures.append(f'{type(exc).__name__}:{getattr(exc, "errno", None)}')
    return tuple(failures)


def _remove_workbook(path: Path) -> None:
    _io_path(path).unlink(missing_ok=True)


def _remove_local_tree(path: Path, *, allowed_root: Path) -> None:
    safe = _safe_harness_path(
        path,
        allowed_roots=(allowed_root,),
        purpose='cleanup tree escaped its local root',
        create_parent=False,
    )
    if _io_path(safe).exists():
        shutil.rmtree(_io_path(safe))


def _remove_new_batch_evidence(path: Path) -> None:
    root = repo_root() / 'docs' / 'performance'
    safe_path = _safe_harness_path(
        path,
        allowed_roots=(root,),
        purpose='batch evidence escaped docs/performance',
        create_parent=False,
    )
    safe_path.unlink(missing_ok=True)


def _safe_harness_path(
    path: Path,
    *,
    allowed_roots: tuple[Path, ...],
    purpose: str,
    create_parent: bool,
) -> Path:
    try:
        return _prepare_local_path(
            _normal_path(path),
            allowed_roots=allowed_roots,
            purpose=purpose,
            create_parent=create_parent,
        )
    except AssertionError as exc:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, str(exc)) from exc


def _trusted_local_root() -> Path:
    return (repo_root() / 'rust' / 'target' / 'perf-local').absolute()


def _validate_trusted_request_paths(request: PairedBenchmarkRequest) -> None:
    trusted = _trusted_local_root()
    if _normal_path(request.local_root).absolute() != trusted:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'request.local_root must equal trusted local root')
    if _normal_path(request.attempt_ledger_root).absolute() != trusted / 'batches':
        raise HarnessFailure(
            HarnessVerdict.INCOMPLETE_EVIDENCE,
            'request.attempt_ledger_root must equal trusted local root/batches',
        )


def _run_git(root: Path, *args: str) -> str:
    git = shutil.which('git')
    if git is None:
        raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'git executable not found')
    completed = subprocess.run(  # noqa: S603 - resolved local Git executable with closed harness arguments.
        [git, '-C', str(root), *args],
        check=False,
        capture_output=True,
        encoding='utf-8',
        errors='replace',
    )
    if completed.returncode != 0:
        raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, f'git command failed: {args!r}')
    return completed.stdout


def _sha256(path: Path) -> str:
    with path.open('rb') as stream:
        return hashlib.file_digest(stream, 'sha256').hexdigest()


def _stable_workbook_oracle(path: Path) -> str:
    """Hash workbook package content while excluding volatile document properties."""
    try:
        with zipfile.ZipFile(_io_path(path)) as archive:
            names = tuple(sorted(name for name in archive.namelist() if not name.startswith('docProps/')))
            if '[Content_Types].xml' not in names or 'xl/workbook.xml' not in names:
                raise AssertionError('workbook package is incomplete')
            digest = hashlib.sha256()
            for name in names:
                encoded = name.encode('utf-8')
                content = archive.read(name)
                digest.update(len(encoded).to_bytes(8, 'big'))
                digest.update(encoded)
                digest.update(len(content).to_bytes(8, 'big'))
                digest.update(content)
            return digest.hexdigest()
    except (OSError, zipfile.BadZipFile, KeyError) as exc:
        raise AssertionError('workbook is not a readable XLSX package') from exc


def workbook_oracle(path: Path) -> str:
    """Compatibility seam for tests and resumed PWS validation."""
    return _stable_workbook_oracle(path)


def _workbook_sheet_names(path: Path) -> tuple[str, ...]:
    try:
        workbook = load_workbook(_io_path(path), read_only=True, data_only=True)
        try:
            return tuple(workbook.sheetnames)
        finally:
            workbook.close()
    except (OSError, KeyError, ValueError) as exc:
        raise HarnessFailure(HarnessVerdict.CORRECTNESS_FAILED, 'smoke workbook package is invalid') from exc


def _with_actual_workbook_dimensions(captured: CapturedNormalRun, output: Path) -> CapturedNormalRun:
    try:
        workbook = load_workbook(_io_path(output), read_only=True, data_only=True)
        try:
            expected_names = tuple(item.value for item in ApprovedSheet)
            if tuple(workbook.sheetnames) != expected_names:
                raise HarnessFailure(HarnessVerdict.CORRECTNESS_FAILED, 'Phase 0A Sheet contract mismatch')
            actual = tuple(workbook[name].calculate_dimension() for name in expected_names)
        finally:
            workbook.close()
    except HarnessFailure:
        raise
    except Exception as exc:
        raise HarnessFailure(HarnessVerdict.CORRECTNESS_FAILED, 'Phase 0A workbook dimensions are unreadable') from exc
    reported = captured.normal_run.runtime.sheet_dimensions
    if reported and reported != actual:
        raise HarnessFailure(HarnessVerdict.CORRECTNESS_FAILED, 'Phase 0A runtime dimensions mismatch workbook')
    if reported:
        return captured
    runtime = replace(captured.normal_run.runtime, sheet_dimensions=actual)
    normal_run = replace(captured.normal_run, runtime=runtime)
    return replace(captured, normal_run=normal_run)


def _canonical_json(payload: object) -> bytes:
    return json.dumps(payload, sort_keys=True, separators=(',', ':'), ensure_ascii=True).encode('utf-8')


def _write_create_new(path: Path, payload: bytes, *, allowed_root: Path) -> None:
    raw = _safe_harness_path(
        _normal_path(path),
        allowed_roots=(_normal_path(allowed_root),),
        purpose='append-only record escaped attempt root',
        create_parent=False,
    )
    path = _io_path(raw)
    path.parent.mkdir(parents=True, exist_ok=True)
    _reject_existing_reparse_components(raw.parent)
    try:
        with path.open('xb') as stream:
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
    except FileExistsError as exc:
        raise HarnessFailure(
            HarnessVerdict.INCOMPLETE_EVIDENCE,
            f'append-only record overwrite refused: {path}',
        ) from exc


def _is_sha256(value: object) -> bool:
    return isinstance(value, str) and len(value) == 64 and all(char in '0123456789abcdef' for char in value)


def _is_lower_hex(value: object, *, minimum: int, maximum: int) -> bool:
    return (
        isinstance(value, str)
        and minimum <= len(value) <= maximum
        and all(char in '0123456789abcdef' for char in value)
    )


def build_pws_cli_arguments(
    mode: str,
    pipeline: PipelineName,
    input_path: Path,
    output_path: Path | None,
) -> tuple[str, ...]:
    if mode == 'Normal':
        if output_path is None:
            raise ValueError('Normal PWS mode requires one output path')
        return (pipeline, '--input', str(input_path), '--output', str(output_path), '--benchmark')
    if mode == 'CheckOnly':
        if output_path is not None:
            raise ValueError('CheckOnly PWS mode forbids an output path')
        return (pipeline, '--input', str(input_path), '--check-only', '--benchmark')
    raise ValueError('PWS mode must be Normal or CheckOnly')


def _build_pws_script_command(
    *,
    mode: str,
    pipeline: PipelineName,
    input_path: Path,
    executable: Path,
    role: BinaryRole,
    batch_id: str,
    global_round: int,
    output_path: Path | None,
    local_log_root: Path,
    local_result_path: Path,
) -> tuple[str, ...]:
    if not _is_sha256(batch_id):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS batch ID must be lowercase SHA-256')
    if role not in ('reference', 'candidate') or global_round not in range(1, 11):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS round identity is invalid')
    powershell = shutil.which('powershell')
    if powershell is None:
        raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'Windows PowerShell executable not found')
    script = Path(__file__).with_name('measure_peak_working_set.ps1')
    command = [
        powershell,
        '-NoProfile',
        '-File',
        str(script),
        '-Mode',
        mode,
        '-Pipeline',
        pipeline,
        '-InputPath',
        str(input_path),
        '-Executable',
        str(executable),
        '-Role',
        role,
        '-BatchId',
        batch_id,
        '-GlobalRound',
        str(global_round),
        '-LocalLogRoot',
        str(local_log_root),
        '-LocalResultPath',
        str(local_result_path),
    ]
    if mode == 'Normal':
        if output_path is None:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'Normal PWS sample requires output path')
        command.extend(('-OutputPath', str(output_path)))
    elif mode != 'CheckOnly' or output_path is not None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'invalid PWS mode/output combination')
    return tuple(command)


def _pws_local_result_path(local_root: Path, batch_id: str, global_round: int, role: BinaryRole) -> Path:
    return _pws_local_artifact_paths(local_root, batch_id, global_round, role).result_path


def _pws_local_artifact_paths(
    local_root: Path,
    batch_id: str,
    global_round: int,
    role: BinaryRole,
) -> PwsLocalArtifacts:
    trusted = _trusted_local_root()
    if _normal_path(local_root).absolute() != trusted:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS raw root must equal trusted local root')
    if not _is_sha256(batch_id) or global_round not in range(1, 11) or role not in ('reference', 'candidate'):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS local artifact identity is invalid')
    result = trusted / 'pws-results' / batch_id / str(global_round) / f'{role}.json'
    log_root = trusted / 'pws-logs'
    log_directory = log_root / batch_id / str(global_round)
    raw_paths = (
        result,
        log_directory / f'{role}.stdout.log',
        log_directory / f'{role}.stderr.log',
        result.with_suffix('.powershell.json'),
    )
    validated = tuple(
        _safe_harness_path(
            path,
            allowed_roots=(trusted,),
            purpose='PWS raw artifact escaped trusted local root',
            create_parent=False,
        )
        for path in raw_paths
    )
    for parent in {path.parent for path in validated}:
        _io_path(parent).mkdir(parents=True, exist_ok=True)
        _reject_existing_reparse_components(parent)
    return PwsLocalArtifacts(*validated, log_root)


def _reject_duplicate_json_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, f'duplicate JSON key: {key}')
        result[key] = value
    return result


def _parse_pws_local_result(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding='utf-8'), object_pairs_hook=_reject_duplicate_json_object)
    except HarnessFailure:
        raise
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS local result is unreadable') from exc
    if not isinstance(payload, dict):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS local result must be an object')
    required = {
        'mode',
        'pipeline',
        'role',
        'batch_id',
        'global_round',
        'exit_code',
        'timed_out',
        'external_wall_seconds',
        'peak_working_set_bytes',
        'input_sha256',
        'binary_sha256',
        'command_arguments',
        'stdout_log_sha256',
        'stderr_log_sha256',
        'local_unversioned_log_sha256',
    }
    if set(payload) != required:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS local result fields are not closed')
    try:
        wall = Decimal(payload['external_wall_seconds'])
    except Exception as exc:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS wall time is invalid') from exc
    peak = payload['peak_working_set_bytes']
    exit_code = payload['exit_code']
    timed_out = payload['timed_out']
    if not wall.is_finite() or wall <= 0:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS wall time must be finite and positive')
    if not isinstance(peak, int) or isinstance(peak, bool) or peak < 0:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PeakWorkingSet64 must be a non-negative integer')
    if not isinstance(exit_code, int) or isinstance(exit_code, bool):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS exit code must be an integer')
    if not isinstance(timed_out, bool):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS timed_out must be boolean')
    if exit_code == 0 and not timed_out and peak <= 0:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'successful PWS sample must have a positive peak')
    if not isinstance(payload['command_arguments'], list) or not all(
        isinstance(item, str) for item in payload['command_arguments']
    ):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS command arguments must be strings')
    for field_name in (
        'input_sha256',
        'binary_sha256',
        'stdout_log_sha256',
        'stderr_log_sha256',
        'local_unversioned_log_sha256',
    ):
        if not _is_sha256(payload[field_name]):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, f'PWS {field_name} is invalid')
    return payload


def _terminate_windows_process_tree(pid: int) -> None:
    system_root = os.environ.get('SystemRoot')
    if not system_root:
        raise OSError('SystemRoot is unavailable')
    taskkill = Path(system_root) / 'System32' / 'taskkill.exe'
    if not taskkill.is_file():
        raise OSError('taskkill.exe is unavailable')
    completed = subprocess.run(  # noqa: S603 - absolute Windows system utility and numeric PID only.
        [str(taskkill), '/PID', str(pid), '/T', '/F'],
        check=False,
        capture_output=True,
        text=True,
        timeout=_PWS_DRIVER_TERMINATION_SECONDS,
    )
    if completed.returncode != 0:
        raise OSError(f'taskkill.exe returned {completed.returncode}')


def _launch_pws_driver(
    command: tuple[str, ...],
    *,
    driver_log_path: Path,
    local_root: Path,
) -> subprocess.CompletedProcess[str]:
    try:
        process = subprocess.Popen(  # noqa: S603 - closed local PowerShell/script/typed arguments.
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            cwd=repo_root(),
            encoding='utf-8',
            errors='replace',
            text=True,
        )
    except OSError as exc:
        raw = _canonical_json(
            {
                'returncode': None,
                'timed_out': False,
                'launch_failed': True,
                'tree_termination_failed': False,
                'driver_reaped': False,
                'stdout': '',
                'stderr': f'{type(exc).__name__}:{getattr(exc, "errno", None)}',
            }
        )
        _write_create_new(driver_log_path, raw, allowed_root=local_root)
        raise RustNormalProcessError(-1, hashlib.sha256(raw).hexdigest()) from exc
    try:
        stdout, stderr = process.communicate(timeout=_PWS_DRIVER_TIMEOUT_SECONDS)
    except subprocess.TimeoutExpired as exc:
        termination_errors: list[str] = []
        tree_termination_failed = False
        try:
            _terminate_windows_process_tree(process.pid)
        except Exception as termination_exc:
            tree_termination_failed = True
            termination_errors.append(
                f'tree={type(termination_exc).__name__}:{getattr(termination_exc, "errno", None)}'
            )
            try:
                process.kill()
            except Exception as kill_exc:
                termination_errors.append(f'driver_kill={type(kill_exc).__name__}:{getattr(kill_exc, "errno", None)}')
        try:
            stdout, stderr = process.communicate(timeout=_PWS_DRIVER_TERMINATION_SECONDS)
        except subprocess.TimeoutExpired:
            stdout, stderr = '', ''
            try:
                process.kill()
            except Exception as kill_exc:
                termination_errors.append(f'driver_kill={type(kill_exc).__name__}:{getattr(kill_exc, "errno", None)}')
            try:
                process.wait(timeout=_PWS_DRIVER_TERMINATION_SECONDS)
            except Exception as wait_exc:
                termination_errors.append(f'driver_wait={type(wait_exc).__name__}:{getattr(wait_exc, "errno", None)}')
        driver_reaped = process.poll() is not None
        if not driver_reaped:
            termination_errors.append('driver_reap=unconfirmed')
        if termination_errors:
            stderr = f'{stderr}\nprocess_termination={";".join(termination_errors)}'.strip()
        raw = _canonical_json(
            {
                'returncode': process.returncode,
                'timed_out': True,
                'launch_failed': False,
                'tree_termination_failed': tree_termination_failed,
                'driver_reaped': driver_reaped,
                'stdout': stdout,
                'stderr': stderr,
            }
        )
        _write_create_new(driver_log_path, raw, allowed_root=local_root)
        raise RustNormalProcessError(-1, hashlib.sha256(raw).hexdigest()) from exc
    raw = _canonical_json(
        {
            'returncode': process.returncode,
            'timed_out': False,
            'launch_failed': False,
            'tree_termination_failed': False,
            'driver_reaped': process.poll() is not None,
            'stdout': stdout,
            'stderr': stderr,
        }
    )
    _write_create_new(driver_log_path, raw, allowed_root=local_root)
    return subprocess.CompletedProcess(command, process.returncode, stdout, stderr)


def _read_closed_driver_log(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(
            _io_path(path).read_text(encoding='utf-8'),
            object_pairs_hook=_reject_duplicate_json_object,
        )
    except Exception as exc:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS driver log is unreadable') from exc
    if not isinstance(payload, dict) or set(payload) != {
        'returncode',
        'timed_out',
        'launch_failed',
        'tree_termination_failed',
        'driver_reaped',
        'stdout',
        'stderr',
    }:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS driver log fields are not closed')
    state_flags = ('timed_out', 'launch_failed', 'tree_termination_failed', 'driver_reaped')
    if not all(isinstance(payload[field], bool) for field in state_flags):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS driver state flags must be boolean')
    if not all(isinstance(payload[field], str) for field in ('stdout', 'stderr')):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS driver streams must be strings')
    if payload['returncode'] is not None and (
        not isinstance(payload['returncode'], int) or isinstance(payload['returncode'], bool)
    ):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS driver returncode must be integer or null')
    return payload


def _artifact_audit_sha(artifacts: PwsLocalArtifacts) -> str:
    hashes: list[str] = []
    for path in (artifacts.result_path, artifacts.stdout_path, artifacts.stderr_path, artifacts.driver_log_path):
        io_path = _io_path(path)
        hashes.append(_sha256(io_path) if io_path.is_file() else 'missing')
    return hashlib.sha256('\n'.join(hashes).encode()).hexdigest()


def _validate_complete_pws_artifacts(
    *,
    artifacts: PwsLocalArtifacts,
    executable: Path,
    pipeline: PipelineName,
    input_path: Path,
    output_path: Path,
    role: BinaryRole,
    batch_id: str,
    global_round: int,
    schema: RuntimeSchema,
) -> CapturedNormalRun:
    payload = _parse_pws_local_result(_io_path(artifacts.result_path))
    log_sha = payload['local_unversioned_log_sha256']
    expected = {
        'mode': 'Normal',
        'pipeline': pipeline,
        'role': role,
        'batch_id': batch_id,
        'global_round': global_round,
        'input_sha256': _sha256(input_path),
        'binary_sha256': _sha256(executable),
        'command_arguments': list(
            build_pws_cli_arguments('Normal', pipeline, input_path.resolve(), output_path.resolve())
        ),
    }
    for field_name, expected_value in expected.items():
        if payload[field_name] != expected_value:
            raise RustNormalValidationError(f'PWS result {field_name} mismatch', log_sha)

    stdout_io = _io_path(artifacts.stdout_path)
    stderr_io = _io_path(artifacts.stderr_path)
    stdout_sha = _sha256(stdout_io)
    stderr_sha = _sha256(stderr_io)
    combined_sha = hashlib.sha256(f'{stdout_sha}\n{stderr_sha}'.encode()).hexdigest()
    if payload['stdout_log_sha256'] != stdout_sha or payload['stderr_log_sha256'] != stderr_sha:
        raise RustNormalValidationError('PWS raw log SHA mismatch', combined_sha)
    if payload['local_unversioned_log_sha256'] != combined_sha:
        raise RustNormalValidationError('PWS combined raw log SHA mismatch', combined_sha)
    driver = _read_closed_driver_log(artifacts.driver_log_path)
    if (
        driver['timed_out']
        or driver['launch_failed']
        or driver['tree_termination_failed']
        or not driver['driver_reaped']
        or driver['returncode'] != payload['exit_code']
    ):
        raise RustNormalValidationError('PWS driver/result state mismatch', combined_sha)
    if payload['timed_out'] or payload['exit_code'] != 0:
        raise RustNormalProcessError(payload['exit_code'] or -1, combined_sha)

    try:
        stdout = stdout_io.read_text(encoding='utf-8')
        runtime_payload = json.loads(stdout, object_pairs_hook=_reject_duplicate_json_object)
        if not isinstance(runtime_payload, dict):
            raise AssertionError('Rust runtime stdout must be one JSON object')
        reported_path = _normal_path(Path(str(runtime_payload.get('workbook_path')))).resolve()
        expected_path = _normal_path(output_path).resolve()
        if reported_path != expected_path:
            raise AssertionError('Rust PWS runtime reported an unexpected workbook path')
        if not output_path.is_file():
            raise AssertionError('Rust PWS runtime did not create its workbook')
        output_size = output_path.stat().st_size
        if output_size <= 0:
            raise AssertionError('Rust PWS workbook must contain positive bytes')
        runtime_payload['output_size_bytes'] = output_size
        runtime = parse_runtime_payload(runtime_payload, schema=schema)
        oracle_sha256 = workbook_oracle(output_path)
    except Exception as exc:
        raise RustNormalValidationError('Rust PWS runtime or workbook validation failed', combined_sha) from exc
    return CapturedNormalRun(
        NormalRunEvidence(
            external_wall_seconds=Decimal(payload['external_wall_seconds']),
            peak_working_set_bytes=payload['peak_working_set_bytes'],
            runtime=runtime,
            workbook_oracle_sha256=oracle_sha256,
        ),
        0,
        combined_sha,
    )


def _invoke_pws_single_sample(
    *,
    executable: Path,
    pipeline: PipelineName,
    input_path: Path,
    output_path: Path,
    role: BinaryRole,
    batch_id: str,
    global_round: int,
    schema: RuntimeSchema,
    local_root: Path,
) -> CapturedNormalRun:
    artifacts = _pws_local_artifact_paths(local_root, batch_id, global_round, role)
    command = _build_pws_script_command(
        mode='Normal',
        pipeline=pipeline,
        input_path=input_path,
        executable=executable,
        role=role,
        batch_id=batch_id,
        global_round=global_round,
        output_path=output_path,
        local_log_root=artifacts.log_root,
        local_result_path=artifacts.result_path,
    )
    core_paths = (
        artifacts.result_path,
        artifacts.stdout_path,
        artifacts.stderr_path,
        artifacts.driver_log_path,
    )
    present = tuple(_io_path(path).is_file() for path in core_paths)
    if present[0]:
        if not all(present):
            raise RustNormalValidationError('PWS resume artifacts are incomplete', _artifact_audit_sha(artifacts))
    elif any(present) or output_path.is_file():
        raise RustNormalValidationError(
            'PWS raw collision is incomplete and cannot be rerun', _artifact_audit_sha(artifacts)
        )
    else:
        _launch_pws_driver(
            command,
            driver_log_path=artifacts.driver_log_path,
            local_root=local_root,
        )
        if not _io_path(artifacts.result_path).is_file():
            raise RustNormalProcessError(-1, _artifact_audit_sha(artifacts))
    artifacts = _pws_local_artifact_paths(local_root, batch_id, global_round, role)
    try:
        return _validate_complete_pws_artifacts(
            artifacts=artifacts,
            executable=executable,
            pipeline=pipeline,
            input_path=input_path,
            output_path=output_path,
            role=role,
            batch_id=batch_id,
            global_round=global_round,
            schema=schema,
        )
    except (RustNormalProcessError, RustNormalValidationError):
        raise
    except Exception as exc:
        raise RustNormalValidationError('PWS raw artifacts failed closed', _artifact_audit_sha(artifacts)) from exc


def run_pws_group(request: MetricGroupRequest) -> MetricGroup:
    if request.metric != 'pws':
        raise ValueError('PWS runner accepts pws metric only')
    _validate_trusted_request_paths(request.benchmark)
    _safe_harness_path(
        request.benchmark.evidence_path,
        allowed_roots=(repo_root() / 'docs' / 'performance',),
        purpose='PWS batch evidence escaped docs/performance',
        create_parent=False,
    )
    identity = _capture_identity(request.benchmark)
    ledger = AppendOnlyAttemptLedger.load(request.attempt_directory, identity)
    if ledger.terminal_verdict is not None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'terminal attempt cannot be resumed')
    if request.first_group_sha256 is not None and request.first_group_sha256 != ledger.first_group_sha256:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'expanded group first-group SHA does not match ledger')

    role_executables = {
        'reference': request.benchmark.reference_executable,
        'candidate': request.benchmark.candidate_executable,
    }
    rule = PROFILE_RULES[request.benchmark.comparison_profile][request.benchmark.pipeline]
    schemas: dict[BinaryRole, RuntimeSchema] = {
        'reference': rule.reference_schema,
        'candidate': rule.candidate_schema,
    }
    evidence_existed_before = request.benchmark.evidence_path.exists()
    planned_payloads = ledger.all_planned_output_payloads()
    cleanup_paths = list(_planned_paths(planned_payloads))
    if ledger.cleanup_only:
        cleanup_errors = _cleanup_all(cleanup_paths)
        if cleanup_errors:
            final_verdict = HarnessVerdict.CLEANUP_FAILED
            primary_verdict = ledger.recovery_primary_verdict
        else:
            final_verdict = ledger.recovery_primary_verdict or HarnessVerdict.ENVIRONMENT_DRIFT
            primary_verdict = None
        ledger.finish(final_verdict, primary_verdict=primary_verdict)
        raise HarnessFailure(
            final_verdict,
            'cleanup-only recovery completed' if not cleanup_errors else 'cleanup-only recovery failed',
            primary_verdict=primary_verdict,
        )

    pairs: list[PairedRound] = []
    result_group: MetricGroup | None = None
    primary_error: HarnessFailure | None = None
    recorded_cleanup_paths = [
        _planned_paths((payload,))[0]
        for payload in planned_payloads
        if ledger.sample_payload(payload['metric'], payload['global_round'], payload['role']) is not None
    ]
    initial_cleanup_errors = _cleanup_all(recorded_cleanup_paths)
    if initial_cleanup_errors:
        primary_error = HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'historical workbook cleanup was transient')
    else:
        try:
            for plan in request.plans:
                captured: dict[BinaryRole, MetricSample] = {}
                for role in plan.order:
                    _assert_identity_unchanged(identity, _capture_identity(request.benchmark))
                    payload = _planned_output_payload(request, identity, plan.global_round, role)
                    ledger.record_planned_output(request.metric, plan.global_round, role, payload)
                    output = next(iter(_planned_paths((payload,))))
                    if output not in cleanup_paths:
                        cleanup_paths.append(output)
                    raw_artifacts = _pws_local_artifact_paths(
                        request.benchmark.local_root,
                        request.batch_id,
                        plan.global_round,
                        role,
                    )
                    cleanup_paths.extend(
                        (
                            raw_artifacts.result_path,
                            raw_artifacts.stdout_path,
                            raw_artifacts.stderr_path,
                            raw_artifacts.driver_log_path,
                        )
                    )
                    existing = ledger.sample_payload(request.metric, plan.global_round, role)
                    if existing is not None:
                        captured[role] = _sample_from_payload(existing)
                        try:
                            _remove_workbook(output)
                        except OSError as exc:
                            raise HarnessFailure(
                                HarnessVerdict.ENVIRONMENT_DRIFT,
                                'recorded PWS sample residual workbook cleanup failed',
                            ) from exc
                        continue
                    try:
                        capture = _invoke_pws_single_sample(
                            executable=role_executables[role],
                            pipeline=request.benchmark.pipeline,
                            input_path=request.benchmark.input_path,
                            output_path=output,
                            role=role,
                            batch_id=request.batch_id,
                            global_round=plan.global_round,
                            schema=schemas[role],
                            local_root=request.benchmark.local_root,
                        )
                    except RustNormalProcessError as exc:
                        verdict = (
                            HarnessVerdict.REFERENCE_FAILED if role == 'reference' else HarnessVerdict.CANDIDATE_FAILED
                        )
                        raise HarnessFailure(
                            verdict,
                            f'{role} PWS process failed with exit code {exc.returncode}',
                            raw_log_sha256=exc.log_sha256,
                        ) from exc
                    except RustNormalValidationError as exc:
                        verdict = (
                            HarnessVerdict.REFERENCE_FAILED
                            if role == 'reference'
                            else HarnessVerdict.CORRECTNESS_FAILED
                        )
                        raise HarnessFailure(
                            verdict,
                            f'{role} PWS runtime or workbook validation failed',
                            raw_log_sha256=exc.log_sha256,
                        ) from exc
                    except Exception as exc:
                        verdict = (
                            HarnessVerdict.REFERENCE_FAILED
                            if role == 'reference'
                            else HarnessVerdict.CORRECTNESS_FAILED
                        )
                        raise HarnessFailure(verdict, f'{role} PWS capture boundary failed') from exc
                    _assert_identity_unchanged(identity, _capture_identity(request.benchmark))
                    if capture.normal_run.peak_working_set_bytes is None:
                        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS capture omitted PeakWorkingSet64')
                    sample = _metric_sample(
                        role,
                        plan,
                        identity,
                        (capture.normal_run, capture.local_unversioned_log_sha256),
                        metric_value=Decimal(capture.normal_run.peak_working_set_bytes),
                    )
                    ledger.record_sample(request.metric, plan.global_round, role, _sample_to_payload(sample))
                    captured[role] = sample
                    try:
                        _remove_workbook(output)
                    except OSError as exc:
                        raise HarnessFailure(
                            HarnessVerdict.ENVIRONMENT_DRIFT,
                            'immediate workbook cleanup failed; outer cleanup will retry',
                        ) from exc
                if (
                    captured['reference'].normal_run.workbook_oracle_sha256
                    != captured['candidate'].normal_run.workbook_oracle_sha256
                ):
                    raise HarnessFailure(
                        HarnessVerdict.CORRECTNESS_FAILED,
                        'reference/candidate workbook oracle mismatch',
                        raw_log_sha256=captured['candidate'].local_unversioned_log_sha256,
                    )
                pairs.append(PairedRound(plan, captured['reference'], captured['candidate']))
            result_group = MetricGroup(
                request.batch_id,
                request.benchmark.pipeline,
                'pws',
                request.plans[0].global_round,  # type: ignore[arg-type]
                tuple(pairs),
            )
            validate_metric_group(result_group)
        except (KeyboardInterrupt, SystemExit) as interruption:
            cleanup_errors = _cleanup_all(cleanup_paths)
            if not evidence_existed_before:
                try:
                    _remove_new_batch_evidence(request.benchmark.evidence_path)
                except OSError as exc:
                    cleanup_errors = (*cleanup_errors, f'{type(exc).__name__}:{getattr(exc, "errno", None)}')
            if cleanup_errors:
                ledger.finish(HarnessVerdict.CLEANUP_FAILED)
                raise HarnessFailure(
                    HarnessVerdict.CLEANUP_FAILED,
                    f'workbook cleanup failed during interruption: {cleanup_errors!r}',
                ) from interruption
            raise
        except HarnessFailure as exc:
            primary_error = exc
        except Exception as exc:
            primary_error = HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS runner failed closed')
            primary_error.__cause__ = exc

    cleanup_errors = _cleanup_all(cleanup_paths)
    if not evidence_existed_before:
        try:
            _remove_new_batch_evidence(request.benchmark.evidence_path)
        except OSError as exc:
            cleanup_errors = (*cleanup_errors, f'{type(exc).__name__}:{getattr(exc, "errno", None)}')
    final_error = primary_error
    if cleanup_errors:
        final_error = HarnessFailure(
            HarnessVerdict.CLEANUP_FAILED,
            f'workbook cleanup failed: {cleanup_errors!r}',
            primary_verdict=primary_error.verdict if primary_error else None,
            raw_log_sha256=primary_error.raw_log_sha256 if primary_error else None,
        )
    if final_error is not None:
        ledger.finish(
            final_error.verdict,
            raw_log_sha256=final_error.raw_log_sha256,
            primary_verdict=final_error.primary_verdict,
        )
        raise final_error
    if result_group is None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'PWS group produced no result')
    return result_group


def run_paired_normal_batch(request: PairedBenchmarkRequest) -> PairedBenchmarkResult:
    _validate_closed_pair_request(request)
    _validate_trusted_request_paths(request)
    evidence_root = repo_root() / 'docs' / 'performance'
    _safe_harness_path(
        request.evidence_path,
        allowed_roots=(evidence_root,),
        purpose='paired evidence escaped docs/performance',
        create_parent=False,
    )
    status = tuple(
        line for line in _run_git(repo_root(), 'status', '--porcelain=v1', '--untracked-files=all').splitlines()
    )
    validate_formal_repository_state(
        status,
        evidence_root=evidence_root,
        prior_evidence_claims=request.prior_evidence_claims,
    )
    identity = _capture_identity(request)
    baseline = _load_approved_phase0a_baseline(request, identity)
    batch_id = derive_batch_id(request, identity)
    comparison_key = hashlib.sha256(
        _canonical_json(
            {
                'pipeline': request.pipeline,
                'profile': request.comparison_profile.value,
                'reference_label': request.reference_label.value,
                'candidate_label': request.candidate_label.value,
                'input_sha256': identity.input_sha256,
                'reference_sha256': identity.reference_sha256,
                'candidate_sha256': identity.candidate_sha256,
            }
        )
    ).hexdigest()
    ledger = AppendOnlyAttemptLedger.create(request.attempt_ledger_root, identity, comparison_key=comparison_key)
    if ledger.prepared_evidence() is not None:
        return _recover_prepared_evidence(request, ledger, batch_id)
    plans = build_round_plan(global_round_start=1, round_count=5)
    try:
        wall_first = run_normal_wall_group(
            MetricGroupRequest(request, batch_id, 'wall', plans, ledger.attempt_directory)
        )
        pws_first = run_pws_group(MetricGroupRequest(request, batch_id, 'pws', plans, ledger.attempt_directory))
        assert_same_benchmark_batch(wall_first, pws_first)
        _validate_phase0a_drift(wall_first, pws_first, baseline, identity)
        ledger = AppendOnlyAttemptLedger.load(ledger.attempt_directory, identity)
        first_sha = ledger.commit_first_group(_group_commit_payload(wall_first, pws_first))
        wall, pws = wall_first, pws_first
        if _paired_groups_require_expansion(request, wall_first, pws_first):
            expanded_plans, _ = _mandatory_paired_expansion_plans(
                wall_requires_expansion=True,
                pws_requires_expansion=False,
            )
            wall_second = run_normal_wall_group(
                MetricGroupRequest(
                    request,
                    batch_id,
                    'wall',
                    expanded_plans,
                    ledger.attempt_directory,
                    first_group_sha256=first_sha,
                )
            )
            pws_second = run_pws_group(
                MetricGroupRequest(
                    request,
                    batch_id,
                    'pws',
                    expanded_plans,
                    ledger.attempt_directory,
                    first_group_sha256=first_sha,
                )
            )
            ledger = AppendOnlyAttemptLedger.load(ledger.attempt_directory, identity)
            ledger.commit_expanded_group(_group_commit_payload(wall_second, pws_second), first_group_sha256=first_sha)
            if groups_have_conflicting_direction(wall_first, wall_second) or groups_have_conflicting_direction(
                pws_first, pws_second
            ):
                raise HarnessFailure(HarnessVerdict.INCONCLUSIVE, 'expanded group changed metric direction')
            wall = merge_metric_groups(wall_first, wall_second)
            pws = merge_metric_groups(pws_first, pws_second)
        assert_same_benchmark_batch(wall, pws)
        _assert_identity_unchanged(identity, _capture_identity(request))
        _evaluate_closed_profile(request, wall, pws, baseline)
        raw_cleanup_errors = _cleanup_all(list(_formal_raw_evidence_paths(request, wall, pws)))
        if raw_cleanup_errors:
            raise HarnessFailure(
                HarnessVerdict.CLEANUP_FAILED,
                f'formal raw evidence cleanup failed: {raw_cleanup_errors!r}',
            )
        ledger.mark_cleanup_complete()
        cleaned_attempt = _batch_attempt(ledger, batch_id, AttemptState.CLEANUP_COMPLETE)
        evidence = _build_paired_evidence(request, wall, pws, cleaned_attempt, identity)
        artifact = EvidenceSanitizer.closed_policy().build_benchmark_manifest(evidence)
        if request.evidence_path.name != artifact.file_name:
            raise HarnessFailure(
                HarnessVerdict.INCOMPLETE_EVIDENCE, 'evidence path basename does not match typed artifact'
            )
        artifact_sha256 = hashlib.sha256(artifact.content.encode('utf-8')).hexdigest()
        ledger.prepare_evidence(
            artifact_basename=artifact.file_name,
            artifact_sha256=artifact_sha256,
            artifact_content=artifact.content,
        )
        try:
            EvidenceSanitizer.closed_policy().write_batch(
                destination_root=request.evidence_path.parent,
                artifacts=(artifact,),
                cleanup_state=AttemptState.CLEANUP_COMPLETE,
                scan_staged=True,
                sensitive_names=(request.input_path.name,),
            )
        except (OSError, ValueError) as exc:
            raise HarnessFailure(HarnessVerdict.SENSITIVE_EVIDENCE, 'paired evidence publication failed') from exc
        ledger.mark_evidence_committed(artifact_sha256=artifact_sha256)
        return PairedBenchmarkResult(
            wall,
            pws,
            _batch_attempt(ledger, batch_id, AttemptState.EVIDENCE_COMMITTED),
            HarnessVerdict.VALIDATED,
        )
    except HarnessFailure as exc:
        current = AppendOnlyAttemptLedger.load(ledger.attempt_directory, identity)
        if current.prepared_evidence() is not None:
            raise
        if current.terminal_verdict is None:
            current.finish(exc.verdict, raw_log_sha256=exc.raw_log_sha256, primary_verdict=exc.primary_verdict)
        raise


def _recover_prepared_evidence(
    request: PairedBenchmarkRequest,
    ledger: AppendOnlyAttemptLedger,
    batch_id: str,
) -> PairedBenchmarkResult:
    payload = ledger.prepared_evidence()
    if payload is None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence recovery was not requested')
    artifact = _rebuild_prepared_artifact(payload)
    if request.evidence_path.name != artifact.file_name:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence basename changed on resume')
    marker_name, marker_content = _batch_commit_marker((artifact,))
    artifact_path = request.evidence_path
    marker_path = request.evidence_path.parent / marker_name
    artifact_exists = _io_path(artifact_path).exists()
    marker_exists = _io_path(marker_path).exists()
    if artifact_exists != marker_exists:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence publication is incomplete')
    if not artifact_exists:
        try:
            EvidenceSanitizer.closed_policy().write_batch(
                destination_root=request.evidence_path.parent,
                artifacts=(artifact,),
                cleanup_state=AttemptState.CLEANUP_COMPLETE,
                scan_staged=True,
                sensitive_names=(request.input_path.name,),
            )
        except (OSError, ValueError) as exc:
            raise HarnessFailure(HarnessVerdict.SENSITIVE_EVIDENCE, 'prepared evidence republication failed') from exc
    try:
        artifact_raw = _io_path(artifact_path).read_bytes()
        marker_raw = _io_path(marker_path).read_bytes()
        published_source = EvidenceSanitizer.closed_policy().read_benchmark_manifest(artifact.file_name, artifact_raw)
        published_artifact = EvidenceSanitizer.closed_policy().build_benchmark_manifest(published_source)
    except (OSError, ValueError) as exc:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence publication is invalid') from exc
    if (
        published_artifact.file_name != artifact.file_name
        or published_artifact.content != artifact.content
        or hashlib.sha256(artifact_raw).hexdigest() != payload['artifact_sha256']
        or marker_raw != marker_content.encode('utf-8')
    ):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'prepared evidence publication does not match ledger')
    ledger.mark_evidence_committed(artifact_sha256=payload['artifact_sha256'])
    return PairedBenchmarkResult(
        None,
        None,
        _batch_attempt(ledger, batch_id, AttemptState.EVIDENCE_COMMITTED),
        HarnessVerdict.VALIDATED,
    )


def run_phase0h_smoke(request: Phase0HSmokeRequest) -> Phase0HSmokeResult:
    if request.pipeline not in ('gb', 'sk'):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'smoke pipeline must be gb or sk')
    root = _safe_harness_path(
        request.local_root,
        allowed_roots=(_trusted_local_root(),),
        purpose='smoke root must stay below rust/target/perf-local',
        create_parent=True,
    )
    if any(root.rglob('*.xlsx')):
        raise HarnessFailure(HarnessVerdict.CLEANUP_FAILED, 'smoke root contains a pre-existing workbook')
    for stale_directory in (root / 'raw-logs', root / 'outputs'):
        try:
            _remove_local_tree(stale_directory, allowed_root=root)
        except OSError as exc:
            raise HarnessFailure(HarnessVerdict.CLEANUP_FAILED, 'historical smoke cleanup failed') from exc
    if _sha256(request.reference_executable) != _sha256(request.candidate_executable):
        raise HarnessFailure(HarnessVerdict.CORRECTNESS_FAILED, 'Phase 0H smoke requires one binary in both roles')

    fixture = root / f'synthetic-{request.pipeline}.xlsx'
    cleanup: list[Path] = [fixture]
    temp_root = root / 'temp-canary'
    temp_root.mkdir(exist_ok=False)
    temp_marker = temp_root / 'sentinel.txt'
    temp_marker.write_text('synthetic sentinel', encoding='utf-8')
    previous_temp = {name: os.environ.get(name) for name in ('TEMP', 'TMP')}
    os.environ['TEMP'] = str(temp_root)
    os.environ['TMP'] = str(temp_root)
    batch_id = hashlib.sha256(
        f'phase0h|{request.pipeline}|{_sha256(request.reference_executable)}'.encode()
    ).hexdigest()
    all_log_sha: list[str] = []
    try:
        build_raw_fixture(fixture, request.pipeline, 'small')
        fixture_sha = _sha256(fixture)
        plans = build_round_plan(global_round_start=1, round_count=5)
        for metric in ('wall', 'pws'):
            for plan in plans:
                paired_hashes: dict[BinaryRole, str] = {}
                for role in plan.order:
                    executable = request.reference_executable if role == 'reference' else request.candidate_executable
                    output = root / 'outputs' / metric / f'{plan.global_round:02d}-{role}.xlsx'
                    cleanup.append(output)
                    try:
                        if metric == 'wall':
                            captured = run_rust_normal_captured(
                                executable,
                                request.pipeline,
                                fixture,
                                output,
                                schema=RuntimeSchema.BASE,
                                local_log_root=root / 'raw-logs',
                                workbook_oracle_fn=_stable_workbook_oracle,
                            )
                        else:
                            artifacts = _pws_local_artifact_paths(
                                _trusted_local_root(), batch_id, plan.global_round, role
                            )
                            cleanup.extend(
                                (
                                    artifacts.result_path,
                                    artifacts.stdout_path,
                                    artifacts.stderr_path,
                                    artifacts.driver_log_path,
                                )
                            )
                            captured = _invoke_pws_single_sample(
                                executable=executable,
                                pipeline=request.pipeline,
                                input_path=fixture,
                                output_path=output,
                                role=role,
                                batch_id=batch_id,
                                global_round=plan.global_round,
                                schema=RuntimeSchema.BASE,
                                local_root=_trusted_local_root(),
                            )
                    except HarnessFailure:
                        raise
                    except Exception as exc:
                        raise _capture_boundary_failure(role, 'smoke', exc) from exc
                    if _workbook_sheet_names(output) != tuple(item.value for item in ApprovedSheet):
                        raise HarnessFailure(HarnessVerdict.CORRECTNESS_FAILED, 'smoke Sheet contract mismatch')
                    paired_hashes[role] = captured.normal_run.workbook_oracle_sha256
                    all_log_sha.append(captured.local_unversioned_log_sha256)
                if paired_hashes['reference'] != paired_hashes['candidate']:
                    raise HarnessFailure(HarnessVerdict.CORRECTNESS_FAILED, 'smoke workbook oracle mismatch')

        temp_canary_created = temp_marker.is_file()
        temp_residue_count = (
            0 if not temp_canary_created else sum(1 for path in temp_root.iterdir() if path != temp_marker)
        )
        if not temp_canary_created or temp_residue_count:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'smoke temp canary changed during capture')
        smoke = SmokeSummaryEvidence(
            candidate_exe_sha256=_sha256(request.candidate_executable),
            fixture_sha256=fixture_sha,
            pipeline=request.pipeline,
            exit_code=0,
            approved_sheets=tuple(ApprovedSheet),
            temp_canary_created=temp_canary_created,
            temp_residue_count=temp_residue_count,
            missing_dll=False,
            local_log_sha256=hashlib.sha256('|'.join(all_log_sha).encode()).hexdigest(),
            verdict=HarnessVerdict.VALIDATED,
        )
        artifact = EvidenceSanitizer.closed_policy().build_smoke(smoke)
        sanitized_root = root / 'sanitized-check'
        sanitized_root.mkdir(parents=True, exist_ok=False)
        sanitized_path = sanitized_root / artifact.file_name
        sanitized_path.write_text(artifact.content, encoding='utf-8', newline='\n')
        # Typed builder is the sanitizer boundary; its legacy field name contains "canary",
        # which the generic tree scanner intentionally rejects as a versioned-data alarm.
        return Phase0HSmokeResult(batch_id, fixture_sha, HarnessVerdict.VALIDATED)
    finally:
        for name, value in previous_temp.items():
            if value is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = value
        failures = _cleanup_all(cleanup)
        for directory in (root / 'sanitized-check', temp_root, root / 'raw-logs', root / 'outputs'):
            try:
                _remove_local_tree(directory, allowed_root=root)
            except OSError as exc:
                failures = (*failures, f'{type(exc).__name__}:{getattr(exc, "errno", None)}')
        if failures:
            raise HarnessFailure(HarnessVerdict.CLEANUP_FAILED, f'smoke workbook cleanup failed: {failures!r}')


def capture_phase0a(request: Phase0ARequest) -> Phase0AManifest:
    if request.output_path.exists():
        raise FileExistsError(request.output_path)
    _validate_phase0a_inputs(request)
    local_root = _safe_harness_path(
        request.local_root,
        allowed_roots=(_trusted_local_root(),),
        purpose='Phase 0A local root must stay below rust/target/perf-local',
        create_parent=True,
    )
    evidence_root = repo_root() / 'docs' / 'performance'
    _safe_harness_path(
        request.output_path.parent,
        allowed_roots=(evidence_root,),
        purpose='Phase 0A manifest parent escaped docs/performance',
        create_parent=True,
    )
    output = _safe_harness_path(
        request.output_path,
        allowed_roots=(evidence_root,),
        purpose='Phase 0A manifest escaped docs/performance',
        create_parent=False,
    )
    machine = _capture_machine_evidence()
    groups: dict[str, CalibrationGroup] = {}
    manifest: Phase0AManifest | None = None
    staging_root: Path | None = None
    output_created = False
    primary: HarnessFailure | None = None
    phase = 'capture'
    try:
        environment_before = _phase0a_environment_state(request.reference_executable, machine)
        for pipeline, input_path in (('gb', request.gb_input_path), ('sk', request.sk_input_path)):
            for metric in ('wall', 'pws'):
                phase = 'capture'
                try:
                    group = _capture_phase0a_group(
                        pipeline=pipeline,
                        metric=metric,
                        input_path=input_path,
                        executable=request.reference_executable,
                        local_root=local_root,
                        machine=machine,
                    )
                except HarnessFailure:
                    raise
                except Exception as exc:
                    raise _capture_boundary_failure('reference', 'Phase 0A', exc) from exc
                phase = 'validation'
                validate_calibration_group(group)
                groups[f'{pipeline}_{metric}'] = group
        current_machine = _capture_machine_evidence()
        if current_machine != machine:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'Phase 0A machine changed before publication')
        if _phase0a_environment_state(request.reference_executable, current_machine) != environment_before:
            raise HarnessFailure(
                HarnessVerdict.ENVIRONMENT_DRIFT, 'Phase 0A environment changed across calibration groups'
            )
        manifest = Phase0AManifest(
            reference_exe_sha256=_sha256(request.reference_executable),
            fork_revision=request.fork_revision,
            git_head=_run_git(repo_root(), 'rev-parse', 'HEAD').strip(),
            machine=machine,
            gb_wall=groups['gb_wall'],
            gb_pws=groups['gb_pws'],
            sk_wall=groups['sk_wall'],
            sk_pws=groups['sk_pws'],
        )
        payload = _phase0a_payload(manifest, request)
        staging_root = local_root / f'phase0a-sanitized-{uuid.uuid4().hex}'
        staging_root.mkdir()
        staging = staging_root / 'phase0a.json'
        staging.write_bytes(_canonical_json(payload))
        phase = 'scan'
        EvidenceSanitizer.closed_policy().scan_tree(
            staging_root,
            sensitive_names=(
                request.gb_input_path.name,
                request.sk_input_path.name,
                platform.node(),
            ),
        )
        phase = 'publish'
        if output.exists():
            raise FileExistsError(output)
        os.link(staging, output)
        output_created = True
    except HarnessFailure as exc:
        primary = exc
    except Exception:
        verdict = (
            HarnessVerdict.SENSITIVE_EVIDENCE if phase in ('scan', 'publish') else HarnessVerdict.INCOMPLETE_EVIDENCE
        )
        primary = HarnessFailure(verdict, f'Phase 0A {phase} failed')
    finally:
        cleanup_failures: list[str] = []
        cleanup_trees = [local_root / 'raw-logs', local_root / 'outputs']
        if staging_root is not None:
            cleanup_trees.append(staging_root)
        for cleanup_tree in cleanup_trees:
            try:
                _remove_local_tree(cleanup_tree, allowed_root=local_root)
            except OSError as exc:
                cleanup_failures.append(f'{type(exc).__name__}:{getattr(exc, "errno", None)}')
        if (primary is not None or cleanup_failures) and output_created:
            try:
                _io_path(output).unlink(missing_ok=True)
            except OSError as exc:
                cleanup_failures.append(f'{type(exc).__name__}:{getattr(exc, "errno", None)}')
        if cleanup_failures:
            raise HarnessFailure(
                HarnessVerdict.CLEANUP_FAILED,
                f'Phase 0A cleanup failed: {cleanup_failures!r}',
                primary_verdict=(primary.primary_verdict or primary.verdict) if primary else None,
                raw_log_sha256=primary.raw_log_sha256 if primary else None,
            ) from primary
    if primary is not None:
        raise primary
    if manifest is None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'Phase 0A produced no manifest')
    return manifest


def main(argv: list[str] | None = None) -> int:
    try:
        arguments = _argument_parser().parse_args(argv)
        if arguments.command == 'smoke':
            result = run_phase0h_smoke(
                Phase0HSmokeRequest(
                    arguments.pipeline,
                    Path(arguments.reference_executable),
                    Path(arguments.candidate_executable),
                    Path(arguments.local_root),
                )
            )
            print(result.verdict.value)
            return _exit_code(result.verdict)
        if arguments.command == 'phase0a':
            capture_phase0a(
                Phase0ARequest(
                    Path(arguments.gb_input),
                    Path(arguments.sk_input),
                    Path(arguments.reference_executable),
                    arguments.fork_revision,
                    Path(arguments.local_root),
                    Path(arguments.output),
                )
            )
            print(HarnessVerdict.VALIDATED.value)
            return 0
        request = PairedBenchmarkRequest(
            pipeline=arguments.pipeline,
            input_path=Path(arguments.input),
            reference_executable=Path(arguments.reference_executable),
            candidate_executable=Path(arguments.candidate_executable),
            reference_label=ClosedBinaryLabel(arguments.reference_label),
            candidate_label=ClosedBinaryLabel(arguments.candidate_label),
            comparison_profile=ComparisonProfile(arguments.comparison_profile),
            phase0a_manifest=Path(arguments.phase0a_manifest),
            local_root=Path(arguments.local_root),
            evidence_path=Path(arguments.evidence_path),
            attempt_ledger_root=repo_root() / 'rust' / 'target' / 'perf-local' / 'batches',
        )
        result = run_paired_normal_batch(request)
        print(result.verdict.value)
        return _exit_code(result.verdict)
    except HarnessFailure as exc:
        print(exc.verdict.value, file=sys.stderr)
        return _exit_code(exc.verdict)
    except _CliUsageError as exc:
        print(f'USAGE_ERROR: {exc}', file=sys.stderr)
        return 5
    except (ValueError, FileExistsError, OSError) as exc:
        print(f'{HarnessVerdict.INCOMPLETE_EVIDENCE.value}: {type(exc).__name__}', file=sys.stderr)
        return 3


def _group_commit_payload(wall: MetricGroup, pws: MetricGroup) -> dict[str, str]:
    def digest(group: MetricGroup) -> str:
        payload = [
            {
                'global_round': item.plan.global_round,
                'reference': str(item.reference.metric_value),
                'candidate': str(item.candidate.metric_value),
            }
            for item in group.rounds
        ]
        return hashlib.sha256(_canonical_json(payload)).hexdigest()

    return {'wall': digest(wall), 'pws': digest(pws)}


def _validate_closed_pair_request(request: PairedBenchmarkRequest) -> None:
    if request.pipeline not in PROFILE_RULES.get(request.comparison_profile, {}):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'profile does not support the selected pipeline')
    pair = (request.reference_label, request.candidate_label)
    if pair not in _PROFILE_LABEL_PAIRS.get(request.comparison_profile, frozenset()):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'binary labels do not match the closed profile')


def _paired_groups_require_expansion(
    request: PairedBenchmarkRequest,
    wall: MetricGroup,
    pws: MetricGroup,
) -> bool:
    limits = COMPARISON_LIMITS[request.comparison_profile][request.pipeline]
    measured = {
        'wall_ratio': median(item.candidate.metric_value for item in wall.rounds)
        / median(item.reference.metric_value for item in wall.rounds),
        'pws_ratio': median(item.candidate.metric_value for item in pws.rounds)
        / median(item.reference.metric_value for item in pws.rounds),
        'wall_seconds': median(item.candidate.metric_value for item in wall.rounds),
        'pws_bytes': median(item.candidate.metric_value for item in pws.rounds),
    }
    if any(
        key in limits and requires_mandatory_expansion(measured=value, limit=Decimal(limits[key]))
        for key, value in measured.items()
    ):
        return True
    for stage in ('ingest', 'writer_populate', 'xlsx_save'):
        key = f'{stage}_ratio'
        if key in limits and requires_mandatory_expansion(
            measured=_stage_ratio(wall, stage),
            limit=Decimal(limits[key]),
        ):
            return True
    if 'ingest_or_pws_ratio' in limits:
        combined = min(_stage_ratio(wall, 'ingest'), measured['pws_ratio'])
        if requires_mandatory_expansion(measured=combined, limit=Decimal(limits['ingest_or_pws_ratio'])):
            return True
    if 'writer_populate_or_export_ratio' in limits:
        limit = Decimal(limits['writer_populate_or_export_ratio'])
        paired_ratios = tuple(
            min(_paired_stage_ratio(item, 'writer_populate'), _paired_stage_ratio(item, 'export'))
            for item in wall.rounds
        )
        if any(requires_mandatory_expansion(measured=value, limit=limit) for value in paired_ratios):
            return True
        wins = sum(value <= limit for value in paired_ratios)
        required_wins = int(limits.get('minimum_wins', 0))
        if (
            wins > 0
            and required_wins > 0
            and requires_mandatory_expansion(
                measured=Decimal(wins),
                limit=Decimal(required_wins),
            )
        ):
            return True
    if request.comparison_profile is ComparisonProfile.PHASE2_B_VS_C:
        return any(
            requires_mandatory_expansion(measured=measured[key], limit=Decimal(1))
            for key in ('wall_ratio', 'pws_ratio')
        )
    return False


def _batch_attempt(ledger: AppendOnlyAttemptLedger, batch_id: str, state: AttemptState) -> BatchAttempt:
    return BatchAttempt(
        comparison_key=ledger.comparison_key,
        batch_id=batch_id,
        attempt_number=ledger.attempt_number,
        state=state,
        previous_attempt_head_sha256=ledger.previous_attempt_head_sha256,
        first_group_sha256=ledger.first_group_sha256,
        expanded_group_sha256=ledger.expanded_group_sha256,
        ledger_head_sha256=ledger.head_sha256,
        attempt_directory=ledger.attempt_directory,
    )


def _load_approved_phase0a_baseline(
    request: PairedBenchmarkRequest,
    identity: BenchmarkIdentity,
) -> _ApprovedPhase0ABaseline:
    path = _safe_harness_path(
        request.phase0a_manifest,
        allowed_roots=(repo_root() / 'docs' / 'performance',),
        purpose='Phase 0A manifest escaped docs/performance',
        create_parent=False,
    )
    try:
        payload = json.loads(
            path.read_text(encoding='utf-8'),
            object_pairs_hook=_reject_duplicate_json_object,
        )
        if not isinstance(payload, dict) or type(payload.get('schema_version')) is not int:
            raise ValueError('Phase 0A manifest schema is invalid')
        _require_exact_object_keys(
            payload,
            {
                'schema_version',
                'approval_state',
                'reference_exe_sha256',
                'fork_revision',
                'git_head',
                'machine',
                'pipelines',
            },
            'Phase 0A manifest',
        )
        if payload['schema_version'] != 1 or payload.get('approval_state') != 'APPROVED':
            raise ValueError('Phase 0A manifest is not approved')
        if (
            payload.get('reference_exe_sha256') != identity.reference_sha256
            or not _is_lower_hex(payload.get('fork_revision'), minimum=40, maximum=40)
            or not _is_lower_hex(payload.get('git_head'), minimum=40, maximum=40)
        ):
            raise ValueError('Phase 0A reference executable SHA changed')
        machine = payload['machine']
        _require_exact_object_keys(
            machine,
            {
                'windows_build_sha256',
                'architecture',
                'cpu_model_sha256',
                'logical_cpu_count',
                'physical_memory_bytes',
                'system_drive_media_type',
                'system_drive_size_bytes',
                'fingerprint_sha256',
            },
            'Phase 0A machine',
        )
        if (
            not _is_sha256(machine['windows_build_sha256'])
            or machine['architecture'] != 'x86_64'
            or not _is_sha256(machine['cpu_model_sha256'])
            or type(machine['logical_cpu_count']) is not int
            or machine['logical_cpu_count'] <= 0
            or type(machine['physical_memory_bytes']) is not int
            or machine['physical_memory_bytes'] <= 0
            or machine['system_drive_media_type'] not in ('SSD', 'HDD', 'UNKNOWN')
            or type(machine['system_drive_size_bytes']) is not int
            or machine['system_drive_size_bytes'] <= 0
            or not _is_sha256(machine['fingerprint_sha256'])
        ):
            raise ValueError('Phase 0A machine values are invalid')
        pipelines = payload['pipelines']
        _require_exact_object_keys(pipelines, {'gb', 'sk'}, 'Phase 0A pipelines')
        validated_pipelines = {
            name: _validate_phase0a_pipeline_payload(value, name) for name, value in pipelines.items()
        }
        pipeline, wall_values, pws_values = validated_pipelines[request.pipeline]
        if pipeline['input_sha256'] != identity.input_sha256:
            raise ValueError('Phase 0A input SHA changed')
        machine_sha = machine['fingerprint_sha256']
        output_size = pipeline['output_size_bytes']
    except (
        InvalidOperation,
        KeyError,
        TypeError,
        json.JSONDecodeError,
        UnicodeDecodeError,
        OSError,
        ValueError,
    ) as exc:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'approved Phase 0A manifest is invalid') from exc
    return _ApprovedPhase0ABaseline(machine_sha, output_size, wall_values, pws_values)


def _require_exact_object_keys(value: object, keys: set[str], name: str) -> None:
    if not isinstance(value, dict) or set(value) != keys:
        raise ValueError(f'{name} keys are invalid')


def _validate_phase0a_pipeline_payload(
    value: object,
    pipeline: str,
) -> tuple[dict[str, object], tuple[Decimal, ...], tuple[Decimal, ...]]:
    _require_exact_object_keys(
        value,
        {
            'input_alias',
            'input_sha256',
            'output_size_bytes',
            'sheet_dimensions',
            'runtime',
            'calibration',
        },
        f'Phase 0A {pipeline} pipeline',
    )
    assert isinstance(value, dict)
    expected_alias = '$GB_INPUT' if pipeline == 'gb' else '$SK_INPUT'
    dimensions = value['sheet_dimensions']
    if (
        value['input_alias'] != expected_alias
        or not _is_sha256(value['input_sha256'])
        or type(value['output_size_bytes']) is not int
        or value['output_size_bytes'] <= 0
        or not isinstance(dimensions, list)
        or len(dimensions) != 3
        or any(not isinstance(item, str) or not item for item in dimensions)
    ):
        raise ValueError(f'Phase 0A {pipeline} pipeline values are invalid')
    runtime = value['runtime']
    _require_exact_object_keys(
        runtime,
        {'sheet_count', 'error_log_count', 'run_counts', 'stage_timings'},
        f'Phase 0A {pipeline} runtime',
    )
    assert isinstance(runtime, dict)
    run_counts = runtime['run_counts']
    timings = runtime['stage_timings']
    if (
        type(runtime['sheet_count']) is not int
        or runtime['sheet_count'] != 3
        or type(runtime['error_log_count']) is not int
        or runtime['error_log_count'] < 0
        or not isinstance(run_counts, dict)
        or any(not isinstance(key, str) or type(item) is not int or item < 0 for key, item in run_counts.items())
        or not isinstance(timings, dict)
        or any(
            not isinstance(key, str) or not isinstance(item, str) or Decimal(item) < 0 for key, item in timings.items()
        )
    ):
        raise ValueError(f'Phase 0A {pipeline} runtime values are invalid')
    calibration = value['calibration']
    _require_exact_object_keys(calibration, {'wall', 'pws'}, f'Phase 0A {pipeline} calibration')
    assert isinstance(calibration, dict)
    return value, _phase0a_metric_values(calibration['wall'], 'wall'), _phase0a_metric_values(calibration['pws'], 'pws')


def _phase0a_metric_values(value: object, metric: str) -> tuple[Decimal, ...]:
    if not isinstance(value, dict) or set(value) != {'batch_id_sha256', 'values', 'local_log_sha256'}:
        raise ValueError(f'Phase 0A {metric} calibration schema is invalid')
    values = value['values']
    logs = value['local_log_sha256']
    if not isinstance(values, list) or len(values) != 5 or not isinstance(logs, list) or len(logs) != 5:
        raise ValueError(f'Phase 0A {metric} calibration requires five rounds')
    if not _is_sha256(value['batch_id_sha256']) or any(not _is_sha256(item) for item in logs):
        raise ValueError(f'Phase 0A {metric} calibration hashes are invalid')
    if any(not isinstance(item, str) for item in values):
        raise ValueError(f'Phase 0A {metric} values must be decimal strings')
    decimals = tuple(Decimal(item) for item in values)
    if any(not item.is_finite() or item <= 0 for item in decimals):
        raise ValueError(f'Phase 0A {metric} values must be positive finite decimals')
    return decimals


def _validate_phase0a_drift(
    wall: MetricGroup,
    pws: MetricGroup,
    baseline: _ApprovedPhase0ABaseline,
    identity: BenchmarkIdentity,
) -> None:
    if identity.machine_fingerprint_sha256 != baseline.machine_fingerprint_sha256:
        raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'machine fingerprint changed from Phase 0A')
    for group, approved in ((wall, baseline.wall_values), (pws, baseline.pws_values)):
        current = median(item.reference.metric_value for item in group.rounds)
        expected = median(approved)
        if abs(current / expected - Decimal(1)) > ENVIRONMENT_MEDIAN_DRIFT_LIMIT:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'reference median drift exceeds ten percent')
    reference_sizes = _output_sizes(wall, pws, 'reference')
    if reference_sizes != {baseline.output_size_bytes}:
        raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'reference output bytes changed from Phase 0A')


def _evaluate_closed_profile(
    request: PairedBenchmarkRequest,
    wall: MetricGroup,
    pws: MetricGroup,
    baseline: _ApprovedPhase0ABaseline,
) -> None:
    limits = COMPARISON_LIMITS[request.comparison_profile][request.pipeline]
    ratios = {
        'wall_ratio': _metric_ratio(wall),
        'pws_ratio': _metric_ratio(pws),
    }
    for key in ('wall_ratio', 'pws_ratio'):
        if key in limits and ratios[key] > Decimal(limits[key]):
            raise HarnessFailure(HarnessVerdict.CANDIDATE_FAILED, f'{key} exceeded the closed profile limit')
    absolute = {
        'wall_seconds': median(item.candidate.metric_value for item in wall.rounds),
        'pws_bytes': median(item.candidate.metric_value for item in pws.rounds),
    }
    for key, value in absolute.items():
        if key in limits and value > Decimal(limits[key]):
            raise HarnessFailure(HarnessVerdict.CANDIDATE_FAILED, f'{key} exceeded the closed profile limit')
    for stage in ('ingest', 'writer_populate', 'xlsx_save'):
        key = f'{stage}_ratio'
        if key in limits and _stage_ratio(wall, stage) > Decimal(limits[key]):
            raise HarnessFailure(HarnessVerdict.CANDIDATE_FAILED, f'{key} exceeded the closed profile limit')
    if 'ingest_or_pws_ratio' in limits:
        measured = min(_stage_ratio(wall, 'ingest'), ratios['pws_ratio'])
        if measured > Decimal(limits['ingest_or_pws_ratio']):
            raise HarnessFailure(HarnessVerdict.CANDIDATE_FAILED, 'ingest/PWS improvement gate failed')
    if 'writer_populate_or_export_ratio' in limits:
        limit = Decimal(limits['writer_populate_or_export_ratio'])
        required = int(limits.get('minimum_wins', 0))
        if len(wall.rounds) not in (5, 10):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'writer/export rounds must use five-round windows')
        for offset in range(0, len(wall.rounds), 5):
            window = wall.rounds[offset : offset + 5]
            wins = sum(
                min(_paired_stage_ratio(item, 'writer_populate'), _paired_stage_ratio(item, 'export')) <= limit
                for item in window
            )
            if wins < required:
                raise HarnessFailure(HarnessVerdict.CANDIDATE_FAILED, 'writer/export minimum-wins gate failed')
    candidate_sizes = _output_sizes(wall, pws, 'candidate')
    output_limit = Decimal(limits['output_bytes_ratio']) * Decimal(baseline.output_size_bytes)
    if any(Decimal(value) > output_limit for value in candidate_sizes):
        raise HarnessFailure(HarnessVerdict.CANDIDATE_FAILED, 'candidate output bytes exceeded the closed limit')


def _metric_ratio(group: MetricGroup) -> Decimal:
    return median(item.candidate.metric_value for item in group.rounds) / median(
        item.reference.metric_value for item in group.rounds
    )


def _paired_stage_ratio(paired: PairedRound, stage: str) -> Decimal:
    reference = dict(paired.reference.normal_run.runtime.stage_timings).get(stage)
    candidate = dict(paired.candidate.normal_run.runtime.stage_timings).get(stage)
    if reference is None or candidate is None or reference <= 0 or candidate < 0:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, f'{stage} runtime evidence is incomplete')
    return candidate / reference


def _stage_ratio(group: MetricGroup, stage: str) -> Decimal:
    return median(_paired_stage_ratio(item, stage) for item in group.rounds)


def _output_sizes(wall: MetricGroup, pws: MetricGroup, role: BinaryRole) -> set[int]:
    values = {
        getattr(item, role).normal_run.runtime.output_size_bytes for group in (wall, pws) for item in group.rounds
    }
    if None in values or any(type(value) is not int or value <= 0 for value in values if value is not None):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, f'{role} output bytes are incomplete')
    return {int(value) for value in values if value is not None}


def _formal_raw_evidence_paths(
    request: PairedBenchmarkRequest,
    wall: MetricGroup,
    pws: MetricGroup,
) -> tuple[Path, ...]:
    paths: list[Path] = [
        request.local_root / 'raw-logs' / f'{sample.local_unversioned_log_sha256}.json'
        for paired in wall.rounds
        for sample in (paired.reference, paired.candidate)
    ]
    for paired in pws.rounds:
        for role in ('reference', 'candidate'):
            artifacts = _pws_local_artifact_paths(
                request.local_root,
                pws.batch_id,
                paired.plan.global_round,
                role,
            )
            paths.extend(
                (
                    artifacts.result_path,
                    artifacts.stdout_path,
                    artifacts.stderr_path,
                    artifacts.driver_log_path,
                )
            )
    return tuple(paths)


def _build_paired_evidence(
    request: PairedBenchmarkRequest,
    wall: MetricGroup,
    pws: MetricGroup,
    attempt: BatchAttempt,
    identity: BenchmarkIdentity,
) -> BenchmarkManifestEvidence:
    machine = _capture_machine_evidence()
    rounds = tuple(
        BenchmarkRoundEvidence(
            BenchmarkMetric.WALL if group.metric == 'wall' else BenchmarkMetric.PWS,
            item.plan.global_round,
            item.plan.order,
            item.reference.metric_value,
            item.candidate.metric_value,
        )
        for group in (wall, pws)
        for item in group.rounds
    )
    metrics = (
        BenchmarkMetricEvidence(
            BenchmarkMetric.WALL_MEDIAN, median(item.candidate.metric_value for item in wall.rounds)
        ),
        BenchmarkMetricEvidence(BenchmarkMetric.PWS_MEDIAN, median(item.candidate.metric_value for item in pws.rounds)),
        BenchmarkMetricEvidence(BenchmarkMetric.WALL_RATIO, _metric_ratio(wall)),
        BenchmarkMetricEvidence(BenchmarkMetric.PWS_RATIO, _metric_ratio(pws)),
    )
    runtime = wall.rounds[0].candidate.normal_run.runtime
    runtime_counts = tuple(
        RuntimeCountEvidence(RuntimeCount(name), value)
        for name, value in runtime.run_counts
        if name in {item.value for item in RuntimeCount}
    ) + (RuntimeCountEvidence(RuntimeCount.ERROR_LOG_COUNT, runtime.error_log_count),)
    if len(runtime.sheet_dimensions) != len(ApprovedSheet):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'paired Sheet dimensions are incomplete')
    dimensions = tuple(
        SheetDimensionEvidence(sheet, dimension)
        for sheet, dimension in zip(ApprovedSheet, runtime.sheet_dimensions, strict=True)
    )
    reference_sizes = _output_sizes(wall, pws, 'reference')
    candidate_sizes = _output_sizes(wall, pws, 'candidate')
    if len(reference_sizes) != 1 or len(candidate_sizes) != 1 or attempt.first_group_sha256 is None:
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'paired output or group evidence is inconsistent')
    logs = tuple(
        sample.local_unversioned_log_sha256
        for group in (wall, pws)
        for paired in group.rounds
        for sample in (paired.reference, paired.candidate)
    )
    return BenchmarkManifestEvidence(
        schema_version=1,
        profile=request.comparison_profile,
        pipeline=request.pipeline,
        input_alias=PathAlias.GB_INPUT if request.pipeline == 'gb' else PathAlias.SK_INPUT,
        input_sha256=identity.input_sha256,
        reference_label=request.reference_label,
        reference_exe_sha256=identity.reference_sha256,
        candidate_label=request.candidate_label,
        candidate_exe_sha256=identity.candidate_sha256,
        machine=MachineArtifactEvidence(
            hashlib.sha256(machine.windows_build.encode()).hexdigest(),
            machine.architecture,
            hashlib.sha256(machine.cpu_model.encode()).hexdigest(),
            machine.logical_cpu_count,
            machine.physical_memory_bytes,
            machine.system_drive_media_type,
            machine.system_drive_size_bytes,
            identity.machine_fingerprint_sha256,
        ),
        attempt_count=attempt.attempt_number,
        prior_safe_verdicts=(),
        ledger_head_sha256=attempt.ledger_head_sha256,
        first_group_sha256=attempt.first_group_sha256,
        expanded_group_sha256=attempt.expanded_group_sha256,
        rounds=rounds,
        metrics=metrics,
        runtime_counts=runtime_counts,
        sheet_dimensions=dimensions,
        output_bytes=(
            OutputBytesEvidence('reference', next(iter(reference_sizes))),
            OutputBytesEvidence('candidate', next(iter(candidate_sizes))),
        ),
        mismatches=(),
        local_log_sha256=logs,
        verdict=HarnessVerdict.VALIDATED,
    )


def _validate_phase0a_inputs(request: Phase0ARequest) -> None:
    for environment_name, actual in (
        ('COSTING_GB_SAMPLE', request.gb_input_path),
        ('COSTING_SK_SAMPLE', request.sk_input_path),
    ):
        configured = os.environ.get(environment_name)
        if not configured or Path(configured).resolve(strict=False) != actual.resolve(strict=False):
            raise HarnessFailure(
                HarnessVerdict.ENVIRONMENT_DRIFT,
                f'{environment_name} must equal the explicit Phase 0A input path',
            )
        if not actual.is_file():
            raise HarnessFailure(HarnessVerdict.REFERENCE_FAILED, f'{environment_name} input is missing')
    if not request.reference_executable.is_file():
        raise HarnessFailure(HarnessVerdict.REFERENCE_FAILED, 'Phase 0A reference executable is missing')
    if not _is_lower_hex(request.fork_revision, minimum=40, maximum=40):
        raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'fork revision must be a full lowercase Git SHA')


def _capture_machine_evidence() -> MachineEvidence:
    windows_build = platform.version() or platform.release() or 'unknown'
    cpu_model = platform.processor() or platform.machine() or 'unknown'
    logical_cpu_count = os.cpu_count() or 1
    physical_memory = _physical_memory_bytes()
    drive_root = Path(repo_root().anchor or Path.cwd().anchor or '/')
    drive_size = shutil.disk_usage(drive_root).total
    fingerprint_payload = '|'.join(
        (
            windows_build,
            platform.machine(),
            cpu_model,
            str(logical_cpu_count),
            str(physical_memory),
            str(drive_size),
        )
    )
    return MachineEvidence(
        windows_build=windows_build,
        architecture='x86_64',
        cpu_model=cpu_model,
        logical_cpu_count=logical_cpu_count,
        physical_memory_bytes=physical_memory,
        system_drive_media_type='UNKNOWN',
        system_drive_size_bytes=drive_size,
        fingerprint_sha256=hashlib.sha256(fingerprint_payload.encode()).hexdigest(),
    )


def _physical_memory_bytes() -> int:
    if os.name != 'nt':
        page_size = int(os.sysconf('SC_PAGE_SIZE'))
        pages = int(os.sysconf('SC_PHYS_PAGES'))
        return page_size * pages
    import ctypes

    class _MemoryStatus(ctypes.Structure):
        _fields_ = [
            ('length', ctypes.c_ulong),
            ('memory_load', ctypes.c_ulong),
            ('total_physical', ctypes.c_ulonglong),
            ('available_physical', ctypes.c_ulonglong),
            ('total_page_file', ctypes.c_ulonglong),
            ('available_page_file', ctypes.c_ulonglong),
            ('total_virtual', ctypes.c_ulonglong),
            ('available_virtual', ctypes.c_ulonglong),
            ('available_extended_virtual', ctypes.c_ulonglong),
        ]

    status = _MemoryStatus()
    status.length = ctypes.sizeof(status)
    if not ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(status)):
        raise OSError('GlobalMemoryStatusEx failed')
    return int(status.total_physical)


def _phase0a_environment_state(executable: Path, machine: MachineEvidence) -> tuple[str, str, str, str]:
    head = _run_git(repo_root(), 'rev-parse', 'HEAD').strip()
    status = _run_git(repo_root(), 'status', '--porcelain=v1', '--untracked-files=all')
    diff = _run_git(repo_root(), 'diff', '--binary', 'HEAD', '--')
    return (
        head,
        hashlib.sha256(f'{status}\n{diff}'.encode()).hexdigest(),
        _sha256(executable),
        machine.fingerprint_sha256,
    )


def _capture_phase0a_identity(
    input_path: Path,
    executable: Path,
    machine: MachineEvidence,
) -> BenchmarkIdentity:
    head, repository_state, executable_sha, machine_sha = _phase0a_environment_state(executable, machine)
    return BenchmarkIdentity(
        _sha256(input_path),
        executable_sha,
        executable_sha,
        head,
        repository_state,
        machine_sha,
    )


def _capture_phase0a_group(
    *,
    pipeline: str,
    metric: str,
    input_path: Path,
    executable: Path,
    local_root: Path,
    machine: MachineEvidence,
) -> CalibrationGroup:
    if pipeline not in ('gb', 'sk') or metric not in ('wall', 'pws'):
        raise ValueError('Phase 0A calibration labels are closed')
    identity = _capture_phase0a_identity(input_path, executable, machine)
    input_sha = identity.input_sha256
    executable_sha = identity.reference_sha256
    git_head = identity.git_head
    repository_state = identity.repository_state_sha256
    batch_id = hashlib.sha256(f'phase0a|{pipeline}|{metric}|{input_sha}|{executable_sha}'.encode()).hexdigest()
    cleanup: list[Path] = []

    def capture(round_number: int, *, warmup: bool) -> CapturedNormalRun:
        suffix = 'warmup' if warmup else f'{round_number:02d}'
        output = local_root / 'outputs' / pipeline / metric / f'{suffix}.xlsx'
        cleanup.append(output)
        if metric == 'wall':
            captured = run_rust_normal_captured(
                executable,
                pipeline,
                input_path,
                output,
                schema=RuntimeSchema.BASE,
                local_log_root=local_root / 'raw-logs',
            )
        else:
            sample_batch_id = hashlib.sha256(f'{batch_id}|warmup'.encode()).hexdigest() if warmup else batch_id
            artifacts = _pws_local_artifact_paths(
                _trusted_local_root(), sample_batch_id, 1 if warmup else round_number, 'reference'
            )
            cleanup.extend(
                (
                    artifacts.result_path,
                    artifacts.stdout_path,
                    artifacts.stderr_path,
                    artifacts.driver_log_path,
                )
            )
            captured = _invoke_pws_single_sample(
                executable=executable,
                pipeline=pipeline,
                input_path=input_path,
                output_path=output,
                role='reference',
                batch_id=sample_batch_id,
                global_round=1 if warmup else round_number,
                schema=RuntimeSchema.BASE,
                local_root=_trusted_local_root(),
            )
        return _with_actual_workbook_dimensions(captured, output)

    primary: BaseException | None = None
    rounds: list[CalibrationRound] = []
    try:
        try:
            capture(1, warmup=True)
        except HarnessFailure:
            raise
        except Exception as exc:
            raise _capture_boundary_failure('reference', 'Phase 0A', exc) from exc
        if _capture_machine_evidence() != machine:
            raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'Phase 0A machine changed during calibration')
        for round_number in range(1, 6):
            try:
                captured = capture(round_number, warmup=False)
            except HarnessFailure:
                raise
            except Exception as exc:
                raise _capture_boundary_failure('reference', 'Phase 0A', exc) from exc
            if _capture_machine_evidence() != machine:
                raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'Phase 0A machine changed during calibration')
            if _capture_phase0a_identity(input_path, executable, machine) != identity:
                raise HarnessFailure(HarnessVerdict.ENVIRONMENT_DRIFT, 'Phase 0A identity changed during calibration')
            metric_value = (
                captured.normal_run.external_wall_seconds
                if metric == 'wall'
                else Decimal(captured.normal_run.peak_working_set_bytes or 0)
            )
            sample = MetricSample(
                role='reference',
                global_round=round_number,
                metric_value=metric_value,
                exit_code=captured.exit_code,
                input_sha256=input_sha,
                binary_sha256=executable_sha,
                git_head=git_head,
                repository_state_sha256=repository_state,
                machine_fingerprint_sha256=machine.fingerprint_sha256,
                local_unversioned_log_sha256=captured.local_unversioned_log_sha256,
                normal_run=captured.normal_run,
            )
            rounds.append(CalibrationRound(round_number, sample))
    except BaseException as exc:
        primary = exc
    failures = _cleanup_all(cleanup)
    if failures:
        primary_verdict = primary.verdict if isinstance(primary, HarnessFailure) else None
        raw_log_sha256 = primary.raw_log_sha256 if isinstance(primary, HarnessFailure) else None
        raise HarnessFailure(
            HarnessVerdict.CLEANUP_FAILED,
            f'Phase 0A cleanup failed: {failures!r}',
            primary_verdict=primary_verdict,
            raw_log_sha256=raw_log_sha256,
        ) from primary
    if primary is not None:
        raise primary
    return CalibrationGroup(batch_id, pipeline, metric, True, tuple(rounds))  # type: ignore[arg-type]


def _phase0a_payload(manifest: Phase0AManifest, request: Phase0ARequest) -> dict[str, object]:
    pipelines: dict[str, object] = {}
    for pipeline, input_path in (('gb', request.gb_input_path), ('sk', request.sk_input_path)):
        wall = getattr(manifest, f'{pipeline}_wall')
        pws = getattr(manifest, f'{pipeline}_pws')
        samples = tuple(item.reference for item in (*wall.rounds, *pws.rounds))
        output_sizes = tuple(item.normal_run.runtime.output_size_bytes for item in samples)
        if any(type(value) is not int or value <= 0 for value in output_sizes):
            raise HarnessFailure(HarnessVerdict.INCOMPLETE_EVIDENCE, 'Phase 0A output bytes are inconsistent')
        decimal_median = median(Decimal(value) for value in output_sizes)
        # 偶数样本的中位数可能落在半字节；向上取整可避免低估后续 output-bytes 容量门禁。
        output_size = int(decimal_median.to_integral_value(rounding=ROUND_CEILING))
        runtime = wall.rounds[0].reference.normal_run.runtime
        pipelines[pipeline] = {
            'input_alias': f'${pipeline.upper()}_INPUT',
            'input_sha256': _sha256(input_path),
            'output_size_bytes': output_size,
            'sheet_dimensions': list(runtime.sheet_dimensions),
            'runtime': {
                'sheet_count': runtime.sheet_count,
                'error_log_count': runtime.error_log_count,
                'run_counts': dict(runtime.run_counts),
                'stage_timings': {name: str(value) for name, value in runtime.stage_timings},
            },
            'calibration': {
                metric: {
                    'batch_id_sha256': hashlib.sha256(group.batch_id.encode()).hexdigest(),
                    'values': [str(item.reference.metric_value) for item in group.rounds],
                    'local_log_sha256': [item.reference.local_unversioned_log_sha256 for item in group.rounds],
                }
                for metric, group in (('wall', wall), ('pws', pws))
            },
        }
    return {
        'schema_version': 1,
        'approval_state': 'CAPTURED_NOT_APPROVED',
        'reference_exe_sha256': manifest.reference_exe_sha256,
        'fork_revision': manifest.fork_revision,
        'git_head': manifest.git_head,
        'machine': {
            'windows_build_sha256': hashlib.sha256(manifest.machine.windows_build.encode()).hexdigest(),
            'architecture': manifest.machine.architecture,
            'cpu_model_sha256': hashlib.sha256(manifest.machine.cpu_model.encode()).hexdigest(),
            'logical_cpu_count': manifest.machine.logical_cpu_count,
            'physical_memory_bytes': manifest.machine.physical_memory_bytes,
            'system_drive_media_type': manifest.machine.system_drive_media_type,
            'system_drive_size_bytes': manifest.machine.system_drive_size_bytes,
            'fingerprint_sha256': manifest.machine.fingerprint_sha256,
        },
        'pipelines': pipelines,
    }


class _CliUsageError(ValueError):
    pass


class _ClosedArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        raise _CliUsageError(message)


def _argument_parser() -> argparse.ArgumentParser:
    parser = _ClosedArgumentParser(prog='python -m tests.rust_oracle.phase0_harness')
    subparsers = parser.add_subparsers(dest='command', required=True)

    paired = subparsers.add_parser('paired')
    paired.add_argument('--pipeline', required=True, choices=('gb', 'sk'))
    paired.add_argument('--input', required=True)
    paired.add_argument('--reference-executable', required=True)
    paired.add_argument('--candidate-executable', required=True)
    paired.add_argument('--reference-label', required=True, choices=tuple(item.value for item in ClosedBinaryLabel))
    paired.add_argument('--candidate-label', required=True, choices=tuple(item.value for item in ClosedBinaryLabel))
    paired.add_argument('--comparison-profile', required=True, choices=tuple(item.value for item in ComparisonProfile))
    paired.add_argument('--phase0a-manifest', required=True)
    paired.add_argument('--local-root', required=True)
    paired.add_argument('--evidence-path', required=True)

    smoke = subparsers.add_parser('smoke')
    smoke.add_argument('--pipeline', required=True, choices=('gb', 'sk'))
    smoke.add_argument('--reference-executable', required=True)
    smoke.add_argument('--candidate-executable', required=True)
    smoke.add_argument('--local-root', required=True)

    phase0a = subparsers.add_parser('phase0a')
    phase0a.add_argument('--gb-input', required=True)
    phase0a.add_argument('--sk-input', required=True)
    phase0a.add_argument('--reference-executable', required=True)
    phase0a.add_argument('--fork-revision', required=True)
    phase0a.add_argument('--local-root', required=True)
    phase0a.add_argument('--output', required=True)
    return parser


def _exit_code(verdict: HarnessVerdict) -> int:
    if verdict is HarnessVerdict.VALIDATED:
        return 0
    if verdict in (
        HarnessVerdict.CANDIDATE_FAILED,
        HarnessVerdict.CORRECTNESS_FAILED,
        HarnessVerdict.INCONCLUSIVE,
    ):
        return 2
    if verdict in (
        HarnessVerdict.REFERENCE_FAILED,
        HarnessVerdict.ENVIRONMENT_DRIFT,
        HarnessVerdict.INCOMPLETE_EVIDENCE,
    ):
        return 3
    if verdict in (HarnessVerdict.CLEANUP_FAILED, HarnessVerdict.SENSITIVE_EVIDENCE):
        return 4
    raise ValueError('unmapped harness verdict')


if __name__ == '__main__':
    try:
        raise SystemExit(main())
    except _CliUsageError as exc:
        print(f'USAGE_ERROR: {exc}', file=sys.stderr)
        raise SystemExit(5) from None
