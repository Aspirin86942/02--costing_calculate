from __future__ import annotations

import getpass
import hashlib
import importlib.util
import json
import socket
import zipfile
from collections.abc import Mapping
from dataclasses import asdict, dataclass, fields, replace
from decimal import Decimal
from pathlib import Path

import pytest

from tests.rust_oracle import phase0_harness
from tests.rust_oracle.benchmark_protocol import (
    AttemptState,
    BatchAttempt,
    CalibrationGroup,
    CalibrationRound,
    ClosedBinaryLabel,
    ComparisonProfile,
    HarnessVerdict,
    MachineEvidence,
    MetricGroup,
    MetricSample,
    NormalRunEvidence,
    PairedRound,
    RecoveryProvenance,
    RecoveryReason,
    RuntimeEvidence,
    UpstreamGateProvenance,
    assert_same_benchmark_batch,
    build_round_plan,
)
from tests.rust_oracle.evidence import (
    ApprovedSheet,
    BenchmarkManifestEvidence,
    EvidenceSanitizer,
    SmokeSummaryEvidence,
    expected_benchmark_artifact_name,
)
from tests.rust_oracle.oracle_runner import (
    CapturedNormalRun,
    RustNormalProcessError,
    RustNormalValidationError,
    build_rust_cli_release,
    run_python_oracle,
    run_rust_cli_release,
)
from tests.rust_oracle.phase0_harness import (
    AppendOnlyAttemptLedger,
    ApprovedRecoveryParent,
    BenchmarkIdentity,
    HarnessFailure,
    MetricGroupRequest,
    PairedBenchmarkRequest,
    PairedBenchmarkResult,
    Phase0ARequest,
    Phase0HSmokeRequest,
    StaticComparisonInputs,
    UnverifiedPriorEvidenceClaim,
    derive_batch_id,
    parse_and_validate_ledger_snapshot,
    run_normal_wall_group,
    run_phase0h_smoke,
    validate_formal_repository_state,
)


@pytest.fixture(autouse=True)
def _trusted_repo_root(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setattr(phase0_harness, 'repo_root', lambda: tmp_path)


def _runtime(pipeline: str = 'gb') -> RuntimeEvidence:
    return RuntimeEvidence(
        pipeline=pipeline,  # type: ignore[arg-type]
        output_written=True,
        request_id_present=True,
        sheet_count=3,
        error_log_count=0,
        issue_type_counts=(),
        quality_metrics=(),
        run_counts=(('reader_rows', 1),),
        stage_timings=tuple(
            (name, Decimal('0.1'))
            for name in ('ingest', 'normalize', 'split', 'fact', 'presentation', 'total', 'export')
        ),
        output_size_bytes=8,
        sheet_dimensions=_APPROVED_TEST_DIMENSIONS,
        reader_snapshot_sha256='',
    )


_APPROVED_TEST_DIMENSIONS = ('A1:B2', 'A1:C3', 'A1:D4')


def _write_approved_test_workbook(path: Path) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        for index, (sheet, dimension) in enumerate(zip(ApprovedSheet, _APPROVED_TEST_DIMENSIONS, strict=True)):
            worksheet = workbook.active if index == 0 else workbook.create_sheet()
            worksheet.title = sheet.value
            worksheet['A1'] = 'start'
            worksheet[dimension.split(':')[1]] = 'end'
        workbook.save(path)
    finally:
        workbook.close()


def _request(tmp_path: Path) -> PairedBenchmarkRequest:
    input_path = tmp_path / 'input.xlsx'
    reference = tmp_path / 'reference.exe'
    candidate = tmp_path / 'candidate.exe'
    manifest = tmp_path / 'docs' / 'performance' / 'phase0a.json'
    manifest.parent.mkdir(parents=True, exist_ok=True)
    for path, content in ((input_path, b'input'), (reference, b'ref'), (candidate, b'candidate'), (manifest, b'{}')):
        path.write_bytes(content)
    return PairedBenchmarkRequest(
        pipeline='gb',
        input_path=input_path,
        reference_executable=reference,
        candidate_executable=candidate,
        reference_label=ClosedBinaryLabel.PHASE0A,
        candidate_label=ClosedBinaryLabel.PHASE0B,
        comparison_profile=ComparisonProfile.PHASE0B_VS_PHASE0A,
        phase0a_manifest=manifest,
        local_root=tmp_path / 'rust' / 'target' / 'perf-local',
        evidence_path=tmp_path / 'docs' / 'performance' / 'batch.json',
        attempt_ledger_root=tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
    )


def _identity() -> BenchmarkIdentity:
    return BenchmarkIdentity('3' * 8, '1' * 8, '2' * 8, 'head', '4' * 8, '5' * 8)


def _full_identity() -> BenchmarkIdentity:
    return BenchmarkIdentity('3' * 64, '1' * 64, '2' * 64, '4' * 40, '5' * 64, '6' * 64)


def _test_recovery_provenance() -> RecoveryProvenance:
    return RecoveryProvenance(
        parent_protocol_version=2,
        parent_comparison_key='0' * 64,
        parent_attempt=1,
        parent_terminal_sha256='1' * 64,
        parent_comparison_tree_sha256='2' * 64,
        parent_journal_head_sha256='3' * 64,
        parent_inventory_entry_count=134,
        reason=RecoveryReason.MISSING_FORMAL_SHEET_DIMENSIONS,
    )


def _test_upstream_gate_provenance() -> UpstreamGateProvenance:
    return UpstreamGateProvenance(
        pipeline='gb',
        protocol_version=3,
        schema_version=3,
        comparison_key='4' * 64,
        artifact_basename=f'benchmark-v3-{"4" * 16}.json',
        artifact_sha256='5' * 64,
        marker_basename=f'batch-{"4" * 16}.commit.json',
        marker_sha256='6' * 64,
        validated_commit_sha='7' * 40,
    )


def _recovery_payload(value: RecoveryProvenance) -> dict[str, object]:
    return {
        'parent_protocol_version': value.parent_protocol_version,
        'parent_comparison_key': value.parent_comparison_key,
        'parent_attempt': value.parent_attempt,
        'parent_terminal_sha256': value.parent_terminal_sha256,
        'parent_comparison_tree_sha256': value.parent_comparison_tree_sha256,
        'parent_journal_head_sha256': value.parent_journal_head_sha256,
        'parent_inventory_entry_count': value.parent_inventory_entry_count,
        'reason': value.reason.value,
    }


def _comparison_key(request: PairedBenchmarkRequest, identity: BenchmarkIdentity) -> str:
    return phase0_harness.derive_v2_comparison_key(
        pipeline=request.pipeline,
        comparison_profile=request.comparison_profile,
        reference_label=request.reference_label,
        candidate_label=request.candidate_label,
        input_sha256=identity.input_sha256,
        reference_sha256=identity.reference_sha256,
        candidate_sha256=identity.candidate_sha256,
    )


def _ledger(tmp_path: Path) -> AppendOnlyAttemptLedger:
    local_root = tmp_path / 'rust' / 'target' / 'perf-local'
    return AppendOnlyAttemptLedger.create(
        local_root / 'batches',
        _identity(),
        comparison_key='a' * 64,
    )


def _rewrite_metadata_and_matching_empty_journal(attempt: Path, **changes: object) -> None:
    metadata_path = attempt / 'metadata.json'
    metadata = json.loads(metadata_path.read_text(encoding='utf-8'))
    metadata.update(changes)
    metadata_raw = phase0_harness._canonical_json(metadata)
    metadata_path.write_bytes(metadata_raw)
    metadata_sha = hashlib.sha256(metadata_raw).hexdigest()
    state = phase0_harness._journal_state_payload(
        attempt_number=1,
        record_count=0,
        record_head_sha256=metadata_sha,
        checkpoint_head_sha256=metadata_sha,
        terminal_present=False,
        terminal_head_sha256=None,
        verdict=None,
    )
    journal_raw = phase0_harness._canonical_json({'previous_journal_sha256': None, **state})
    (attempt.parent / 'journal' / '000001.json').write_bytes(journal_raw)


def _write_synthetic_v1_terminal(tmp_path: Path, *, comparison_key: str) -> Path:
    comparison = tmp_path / 'rust' / 'target' / 'perf-local' / 'batches' / comparison_key
    attempt = comparison / 'attempt-0001'
    (attempt / 'records').mkdir(parents=True)
    (attempt / 'checkpoints').mkdir()
    (comparison / 'journal').mkdir()
    metadata = {
        'comparison_key': comparison_key,
        'attempt_number': 1,
        'identity': asdict(_identity()),
        'previous_attempt_head_sha256': None,
        'reason': 'FORMAL_START',
        'inherited_planned_outputs': [],
        'cleanup_only': False,
        'recovery_primary_verdict': None,
    }
    metadata_raw = phase0_harness._canonical_json(metadata)
    (attempt / 'metadata.json').write_bytes(metadata_raw)
    metadata_sha = hashlib.sha256(metadata_raw).hexdigest()
    terminal = {
        'checkpoint_head_sha256': metadata_sha,
        'primary_verdict': None,
        'raw_log_sha256': None,
        'record_count': 0,
        'record_head_sha256': metadata_sha,
        'verdict': HarnessVerdict.INCONCLUSIVE.value,
    }
    terminal_raw = phase0_harness._canonical_json(terminal)
    (attempt / 'terminal.json').write_bytes(terminal_raw)
    terminal_sha = hashlib.sha256(terminal_raw).hexdigest()
    state = phase0_harness._journal_state_payload(
        attempt_number=1,
        record_count=0,
        record_head_sha256=metadata_sha,
        checkpoint_head_sha256=metadata_sha,
        terminal_present=True,
        terminal_head_sha256=terminal_sha,
        verdict=HarnessVerdict.INCONCLUSIVE,
    )
    journal_raw = phase0_harness._canonical_json({'previous_journal_sha256': None, **state})
    (comparison / 'journal' / '000001.json').write_bytes(journal_raw)
    return attempt


def _write_synthetic_record(
    attempt: Path,
    *,
    sequence: int,
    kind: str,
    previous_record_sha256: str,
    previous_checkpoint_sha256: str,
    write_checkpoint: bool = True,
    **payload: object,
) -> tuple[str, str]:
    record = {'kind': kind, 'previous_record_sha256': previous_record_sha256, **payload}
    record_raw = phase0_harness._canonical_json(record)
    (attempt / 'records' / f'{sequence:04d}-{kind}.json').write_bytes(record_raw)
    record_sha = hashlib.sha256(record_raw).hexdigest()
    checkpoint = {
        'record_count': sequence,
        'record_sha256': record_sha,
        'previous_checkpoint_sha256': previous_checkpoint_sha256,
    }
    checkpoint_raw = phase0_harness._canonical_json(checkpoint)
    if write_checkpoint:
        (attempt / 'checkpoints' / f'{sequence:04d}.json').write_bytes(checkpoint_raw)
    return record_sha, hashlib.sha256(checkpoint_raw).hexdigest()


def _write_recoverable_v1_success(
    tmp_path: Path,
    *,
    comparison_key: str,
    missing_checkpoint: bool,
    stale_journal: bool,
) -> Path:
    if missing_checkpoint == stale_journal:
        raise ValueError('select exactly one v1 recovery edge')
    comparison = tmp_path / 'rust' / 'target' / 'perf-local' / 'batches' / comparison_key
    attempt = comparison / 'attempt-0001'
    (attempt / 'records').mkdir(parents=True)
    (attempt / 'checkpoints').mkdir()
    (comparison / 'journal').mkdir()
    metadata = {
        'comparison_key': comparison_key,
        'attempt_number': 1,
        'identity': asdict(_identity()),
        'previous_attempt_head_sha256': None,
        'reason': 'FORMAL_START',
        'inherited_planned_outputs': [],
        'cleanup_only': False,
        'recovery_primary_verdict': None,
    }
    metadata_raw = phase0_harness._canonical_json(metadata)
    (attempt / 'metadata.json').write_bytes(metadata_raw)
    metadata_sha = hashlib.sha256(metadata_raw).hexdigest()
    record1, checkpoint1 = _write_synthetic_record(
        attempt,
        sequence=1,
        kind='first-group',
        previous_record_sha256=metadata_sha,
        previous_checkpoint_sha256=metadata_sha,
        groups={'wall': {}, 'pws': {}},
    )
    record2, checkpoint2 = _write_synthetic_record(
        attempt,
        sequence=2,
        kind='cleanup-complete',
        previous_record_sha256=record1,
        previous_checkpoint_sha256=checkpoint1,
        planned_output_count=0,
    )
    artifact_content = '{}'
    artifact_sha256 = hashlib.sha256(artifact_content.encode('utf-8')).hexdigest()
    record3, checkpoint3 = _write_synthetic_record(
        attempt,
        sequence=3,
        kind='evidence-prepared',
        previous_record_sha256=record2,
        previous_checkpoint_sha256=checkpoint2,
        artifact_basename='benchmark-v1-audit.json',
        artifact_sha256=artifact_sha256,
        artifact_content=artifact_content,
    )
    record4, checkpoint4 = _write_synthetic_record(
        attempt,
        sequence=4,
        kind='evidence-committed',
        previous_record_sha256=record3,
        previous_checkpoint_sha256=checkpoint3,
        write_checkpoint=not missing_checkpoint,
        artifact_sha256=artifact_sha256,
    )
    state = phase0_harness._journal_state_payload(
        attempt_number=1,
        record_count=3 if stale_journal else 4,
        record_head_sha256=record3 if stale_journal else record4,
        checkpoint_head_sha256=checkpoint3 if stale_journal else checkpoint4,
        terminal_present=False,
        terminal_head_sha256=None,
        verdict=None,
    )
    journal_raw = phase0_harness._canonical_json({'previous_journal_sha256': None, **state})
    (comparison / 'journal' / '000001.json').write_bytes(journal_raw)
    return attempt


def _file_snapshot(root: Path) -> dict[str, bytes]:
    return {path.relative_to(root).as_posix(): path.read_bytes() for path in root.rglob('*') if path.is_file()}


@dataclass(frozen=True)
class SyntheticV2Parent:
    comparison: Path
    approved: ApprovedRecoveryParent
    static: StaticComparisonInputs


def _synthetic_tree_digest(comparison: Path) -> tuple[str, str, int]:
    entries: list[dict[str, object]] = []
    for item in sorted(comparison.rglob('*'), key=lambda path: path.relative_to(comparison).as_posix()):
        relative = item.relative_to(comparison).as_posix()
        if item.is_dir():
            entries.append({'path': relative, 'kind': 'directory'})
        elif item.is_file():
            entries.append(
                {
                    'path': relative,
                    'kind': 'file',
                    'size': item.stat().st_size,
                    'sha256': hashlib.sha256(item.read_bytes()).hexdigest(),
                }
            )
    raw = json.dumps(entries, ensure_ascii=True, sort_keys=True, separators=(',', ':')).encode('utf-8')
    journal = sorted((comparison / 'journal').glob('*.json'))
    return hashlib.sha256(raw).hexdigest(), hashlib.sha256(journal[-1].read_bytes()).hexdigest(), len(entries)


def _write_synthetic_journal_entry(
    comparison: Path,
    *,
    previous_journal_sha256: str | None,
    record_count: int,
    record_head_sha256: str,
    checkpoint_head_sha256: str,
    terminal_head_sha256: str | None = None,
    verdict: HarnessVerdict | None = None,
) -> str:
    state = phase0_harness._journal_state_payload(
        attempt_number=1,
        record_count=record_count,
        record_head_sha256=record_head_sha256,
        checkpoint_head_sha256=checkpoint_head_sha256,
        terminal_present=terminal_head_sha256 is not None,
        terminal_head_sha256=terminal_head_sha256,
        verdict=verdict,
    )
    raw = phase0_harness._canonical_json({'previous_journal_sha256': previous_journal_sha256, **state})
    sequence = len(tuple((comparison / 'journal').glob('*.json'))) + 1
    (comparison / 'journal' / f'{sequence:06d}.json').write_bytes(raw)
    return hashlib.sha256(raw).hexdigest()


def _synthetic_v2_sample_payload(
    identity: BenchmarkIdentity,
    *,
    metric: str,
    global_round: int,
    role: str,
) -> dict[str, object]:
    oracle = hashlib.sha256(f'{metric}:{global_round}'.encode()).hexdigest()
    metric_value = str(global_round) if metric == 'wall' else str(1000 + global_round)
    return {
        'role': role,
        'global_round': global_round,
        'metric_value': metric_value,
        'exit_code': 0,
        'input_sha256': identity.input_sha256,
        'binary_sha256': identity.reference_sha256 if role == 'reference' else identity.candidate_sha256,
        'git_head': identity.git_head,
        'repository_state_sha256': identity.repository_state_sha256,
        'machine_fingerprint_sha256': identity.machine_fingerprint_sha256,
        'local_unversioned_log_sha256': hashlib.sha256(f'{metric}:{global_round}:{role}:log'.encode()).hexdigest(),
        'normal_run': {
            'external_wall_seconds': str(global_round),
            'peak_working_set_bytes': 1000 + global_round if metric == 'pws' else None,
            'workbook_oracle_sha256': oracle,
            'runtime': {
                'pipeline': 'gb',
                'output_written': True,
                'request_id_present': True,
                'sheet_count': 3,
                'error_log_count': 0,
                'issue_type_counts': [],
                'quality_metrics': [],
                'run_counts': [['reader_rows', 1]],
                'stage_timings': [['total', '0.1']],
                'output_size_bytes': 8,
                'sheet_dimensions': [],
                'reader_snapshot_sha256': '',
            },
        },
    }


def _write_synthetic_v2_recovery_parent(
    root: Path,
    *,
    semantic_override: Mapping[str, object] | None = None,
) -> SyntheticV2Parent:
    local_root = phase0_harness._trusted_local_root()
    identity = _full_identity()
    static = StaticComparisonInputs(
        pipeline='gb',
        comparison_profile=ComparisonProfile.PHASE0B_VS_PHASE0A,
        reference_label=ClosedBinaryLabel.PHASE0A,
        candidate_label=ClosedBinaryLabel.PHASE0B,
        phase0a_manifest_sha256='7' * 64,
        input_sha256=identity.input_sha256,
        reference_sha256=identity.reference_sha256,
        candidate_sha256=identity.candidate_sha256,
    )
    comparison_key = phase0_harness.derive_v2_comparison_key(
        pipeline=static.pipeline,
        comparison_profile=static.comparison_profile,
        reference_label=static.reference_label,
        candidate_label=static.candidate_label,
        input_sha256=static.input_sha256,
        reference_sha256=static.reference_sha256,
        candidate_sha256=static.candidate_sha256,
    )
    comparison = local_root / 'batches' / comparison_key
    attempt = comparison / 'attempt-0001'
    (attempt / 'records').mkdir(parents=True)
    (attempt / 'checkpoints').mkdir()
    (comparison / 'journal').mkdir()
    metadata = {
        'protocol_version': 2,
        'comparison_key': comparison_key,
        'attempt_number': 1,
        'identity': asdict(identity),
        'previous_attempt_head_sha256': None,
        'reason': 'FORMAL_START',
        'inherited_planned_outputs': [],
        'cleanup_only': False,
        'recovery_primary_verdict': None,
    }
    metadata_raw = phase0_harness._canonical_json(metadata)
    (attempt / 'metadata.json').write_bytes(metadata_raw)
    record_head = hashlib.sha256(metadata_raw).hexdigest()
    checkpoint_head = record_head
    journal_head = _write_synthetic_journal_entry(
        comparison,
        previous_journal_sha256=None,
        record_count=0,
        record_head_sha256=record_head,
        checkpoint_head_sha256=checkpoint_head,
    )
    sequence = 0
    groups: dict[str, str] = {}
    override_key = ('wall', 1, 'reference')
    for metric in ('wall', 'pws'):
        for plan in build_round_plan(global_round_start=1, round_count=5):
            for role in plan.order:
                sequence += 1
                planned_payload = {
                    'pipeline': 'gb',
                    'batch_id': '8' * 64,
                    'metric': metric,
                    'binary_sha256': identity.reference_sha256 if role == 'reference' else identity.candidate_sha256,
                    'global_round': plan.global_round,
                    'role': role,
                    'relative_path': (
                        f'gb/.perf-runs/{"8" * 64}/{metric}/'
                        f'{identity.reference_sha256 if role == "reference" else identity.candidate_sha256}/'
                        f'{plan.global_round}/{role}.xlsx'
                    ),
                }
                record_head, checkpoint_head = _write_synthetic_record(
                    attempt,
                    sequence=sequence,
                    kind='planned-output',
                    previous_record_sha256=record_head,
                    previous_checkpoint_sha256=checkpoint_head,
                    metric=metric,
                    global_round=plan.global_round,
                    role=role,
                    payload=planned_payload,
                )
                journal_head = _write_synthetic_journal_entry(
                    comparison,
                    previous_journal_sha256=journal_head,
                    record_count=sequence,
                    record_head_sha256=record_head,
                    checkpoint_head_sha256=checkpoint_head,
                )
                sequence += 1
                sample_payload = _synthetic_v2_sample_payload(
                    identity,
                    metric=metric,
                    global_round=plan.global_round,
                    role=role,
                )
                if semantic_override is not None and (metric, plan.global_round, role) == override_key:
                    normal_run = sample_payload['normal_run']
                    assert isinstance(normal_run, dict)
                    runtime = normal_run['runtime']
                    assert isinstance(runtime, dict)
                    override = dict(semantic_override)
                    if 'remove_runtime_key' in override:
                        runtime.pop(override.pop('remove_runtime_key'))
                    runtime_update = override.pop('runtime', {})
                    normal_run_update = override.pop('normal_run', {})
                    assert isinstance(runtime_update, Mapping)
                    assert isinstance(normal_run_update, Mapping)
                    runtime.update(runtime_update)
                    normal_run.update(normal_run_update)
                    sample_payload.update(override)
                record_head, checkpoint_head = _write_synthetic_record(
                    attempt,
                    sequence=sequence,
                    kind='sample',
                    previous_record_sha256=record_head,
                    previous_checkpoint_sha256=checkpoint_head,
                    metric=metric,
                    global_round=plan.global_round,
                    role=role,
                    payload=sample_payload,
                )
                journal_head = _write_synthetic_journal_entry(
                    comparison,
                    previous_journal_sha256=journal_head,
                    record_count=sequence,
                    record_head_sha256=record_head,
                    checkpoint_head_sha256=checkpoint_head,
                )
        groups[metric] = record_head
    sequence += 1
    record_head, checkpoint_head = _write_synthetic_record(
        attempt,
        sequence=sequence,
        kind='first-group',
        previous_record_sha256=record_head,
        previous_checkpoint_sha256=checkpoint_head,
        groups=groups,
    )
    journal_head = _write_synthetic_journal_entry(
        comparison,
        previous_journal_sha256=journal_head,
        record_count=sequence,
        record_head_sha256=record_head,
        checkpoint_head_sha256=checkpoint_head,
    )
    sequence += 1
    record_head, checkpoint_head = _write_synthetic_record(
        attempt,
        sequence=sequence,
        kind='cleanup-complete',
        previous_record_sha256=record_head,
        previous_checkpoint_sha256=checkpoint_head,
        planned_output_count=20,
    )
    journal_head = _write_synthetic_journal_entry(
        comparison,
        previous_journal_sha256=journal_head,
        record_count=sequence,
        record_head_sha256=record_head,
        checkpoint_head_sha256=checkpoint_head,
    )
    terminal = {
        'verdict': HarnessVerdict.INCOMPLETE_EVIDENCE.value,
        'primary_verdict': None,
        'raw_log_sha256': None,
        'record_count': sequence,
        'record_head_sha256': record_head,
        'checkpoint_head_sha256': checkpoint_head,
    }
    terminal_raw = phase0_harness._canonical_json(terminal)
    (attempt / 'terminal.json').write_bytes(terminal_raw)
    terminal_sha = hashlib.sha256(terminal_raw).hexdigest()
    _write_synthetic_journal_entry(
        comparison,
        previous_journal_sha256=journal_head,
        record_count=sequence,
        record_head_sha256=record_head,
        checkpoint_head_sha256=checkpoint_head,
        terminal_head_sha256=terminal_sha,
        verdict=HarnessVerdict.INCOMPLETE_EVIDENCE,
    )
    tree_sha, journal_sha, entry_count = _synthetic_tree_digest(comparison)
    approved = ApprovedRecoveryParent(
        pipeline='gb',
        comparison_profile=ComparisonProfile.PHASE0B_VS_PHASE0A,
        reference_label=ClosedBinaryLabel.PHASE0A,
        candidate_label=ClosedBinaryLabel.PHASE0B,
        phase0a_manifest_sha256=static.phase0a_manifest_sha256,
        input_sha256=identity.input_sha256,
        reference_sha256=identity.reference_sha256,
        candidate_sha256=identity.candidate_sha256,
        parent_protocol_version=2,
        parent_comparison_key=comparison_key,
        parent_attempt=1,
        parent_terminal_sha256=terminal_sha,
        parent_comparison_tree_sha256=tree_sha,
        parent_journal_head_sha256=journal_sha,
        parent_inventory_entry_count=entry_count,
        reason=phase0_harness.RecoveryReason.MISSING_FORMAL_SHEET_DIMENSIONS,
    )
    return SyntheticV2Parent(comparison, approved, static)


def _set_synthetic_journal_shape(parent: SyntheticV2Parent, journal_shape: str) -> None:
    journal = parent.comparison / 'journal'
    if journal_shape == 'missing':
        for item in journal.iterdir():
            item.unlink()
        journal.rmdir()
    elif journal_shape == 'empty':
        for item in journal.iterdir():
            item.unlink()
    elif journal_shape == 'invalid-name':
        latest = sorted(journal.glob('*.json'))[-1]
        latest.rename(journal / 'invalid.json')
    elif journal_shape == 'junk-file':
        (journal / 'junk.tmp').write_bytes(b'junk')
    elif journal_shape == 'directory':
        (journal / 'extra').mkdir()
    else:
        raise ValueError(f'unknown journal shape: {journal_shape}')


def _mutate_recovery_parent(parent: SyntheticV2Parent, mutation: str) -> None:
    if mutation == 'journal':
        latest = sorted((parent.comparison / 'journal').glob('*.json'))[-1]
        latest.write_bytes(latest.read_bytes() + b' ')
    elif mutation == 'attempt-0002':
        (parent.comparison / 'attempt-0002').mkdir()
    elif mutation == 'unknown-file':
        (parent.comparison / 'unknown.bin').write_bytes(b'unknown')
    else:
        raise ValueError(f'unknown parent mutation: {mutation}')


def _tree_bytes(path: Path) -> dict[str, bytes]:
    return {
        item.relative_to(path).as_posix(): item.read_bytes()
        for item in sorted(path.rglob('*'), key=lambda entry: entry.relative_to(path).as_posix())
        if item.is_file()
    }


def _mutate_v2_sample_semantics(root: Path, mutation: str) -> SyntheticV2Parent:
    if mutation == 'missing-dimensions':
        override: Mapping[str, object] = {'remove_runtime_key': 'sheet_dimensions'}
    elif mutation == 'none-dimensions':
        override = {'runtime': {'sheet_dimensions': None}}
    elif mutation == 'list-dimensions':
        override = {'runtime': {'sheet_dimensions': ['A1:A1', 'A1:A1', 'A1:A1']}}
    elif mutation == 'partial-dimensions':
        override = {'runtime': {'sheet_dimensions': ['A1:A1']}}
    elif mutation == 'identity-drift':
        override = {'input_sha256': '9' * 64}
    elif mutation == 'oracle-mismatch':
        override = {'normal_run': {'workbook_oracle_sha256': '9' * 64}}
    else:
        raise ValueError(f'unknown semantic mutation: {mutation}')
    return _write_synthetic_v2_recovery_parent(root, semantic_override=override)


def test_comparison_tree_digest_is_path_order_stable_and_rejects_reparse(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setattr(phase0_harness, '_trusted_local_root', lambda: tmp_path)
    parent = _write_synthetic_v2_recovery_parent(tmp_path)
    first = phase0_harness.comparison_tree_digest(parent.comparison)
    second = phase0_harness.comparison_tree_digest(parent.comparison)
    assert first == second
    assert first.entry_count > 0
    monkeypatch.setattr(phase0_harness, '_is_reparse_point', lambda path: path.name == 'metadata.json')
    with pytest.raises(HarnessFailure, match='reparse'):
        phase0_harness.comparison_tree_digest(parent.comparison)


@pytest.mark.parametrize('journal_shape', ('missing', 'empty', 'invalid-name', 'junk-file', 'directory'))
def test_comparison_tree_digest_rejects_invalid_journal(
    journal_shape: str,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setattr(phase0_harness, '_trusted_local_root', lambda: tmp_path)
    parent = _write_synthetic_v2_recovery_parent(tmp_path)
    _set_synthetic_journal_shape(parent, journal_shape)
    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.comparison_tree_digest(parent.comparison)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_comparison_tree_digest_maps_root_validation_oserror_to_incomplete_evidence(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setattr(
        phase0_harness,
        '_safe_harness_path',
        lambda *args, **kwargs: (_ for _ in ()).throw(OSError('root validation failed')),
    )
    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.comparison_tree_digest(tmp_path / 'batches' / ('a' * 64))
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_recovery_parent_manifest_drift_fails_before_v3_create_or_write(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    parent = _write_synthetic_v2_recovery_parent(tmp_path)
    drifted = replace(parent.static, phase0a_manifest_sha256='6' + parent.static.phase0a_manifest_sha256[1:])
    monkeypatch.setattr(
        AppendOnlyAttemptLedger,
        'create_v3_once',
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError('v3 create must not run')),
        raising=False,
    )
    monkeypatch.setattr(
        phase0_harness,
        '_write_create_new',
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError('authorization must not write')),
    )
    before = _tree_bytes(parent.comparison)
    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._authorize_v3_recovery(drifted, approved=parent.approved)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert _tree_bytes(parent.comparison) == before


@pytest.mark.parametrize('mutation', ('journal', 'attempt-0002', 'unknown-file'))
def test_recovery_parent_mutation_fails_before_v3_create(
    mutation: str,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    parent = _write_synthetic_v2_recovery_parent(tmp_path)
    _mutate_recovery_parent(parent, mutation)
    monkeypatch.setattr(
        AppendOnlyAttemptLedger,
        'create_v3_once',
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError('v3 create must not run')),
        raising=False,
    )
    before = _tree_bytes(parent.comparison)
    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._authorize_v3_recovery(parent.static, approved=parent.approved)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert _tree_bytes(parent.comparison) == before


def test_authorized_parent_uses_exact_v2_sample_parser_and_never_writes(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    parent = _write_synthetic_v2_recovery_parent(tmp_path)
    monkeypatch.setattr(
        phase0_harness,
        '_write_create_new',
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError('legacy audit attempted a write')),
    )
    before = phase0_harness.comparison_tree_digest(parent.comparison)
    result = phase0_harness._authorize_v3_recovery(parent.static, approved=parent.approved)
    assert result.parent_comparison_tree_sha256 == before.sha256
    assert phase0_harness.comparison_tree_digest(parent.comparison) == before


@pytest.mark.parametrize(
    'mutation',
    (
        'missing-dimensions',
        'none-dimensions',
        'list-dimensions',
        'partial-dimensions',
        'identity-drift',
        'oracle-mismatch',
    ),
)
def test_resealed_parent_rejects_non_exact_v2_sample_semantics(mutation: str, tmp_path: Path) -> None:
    parent = _mutate_v2_sample_semantics(tmp_path, mutation)
    snapshot = parse_and_validate_ledger_snapshot(
        parent.comparison / 'attempt-0001',
        expected_protocol_version=2,
    )
    assert snapshot.terminal_verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._authorize_v3_recovery(parent.static, approved=parent.approved)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def _claim(tmp_path: Path, path: Path, sha256: str) -> UnverifiedPriorEvidenceClaim:
    return UnverifiedPriorEvidenceClaim(
        path_alias=path.relative_to(tmp_path).as_posix(),
        content_sha256=sha256,
    )


def _install_runner(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    *,
    fail_role: str | None = None,
    validation_fail_role: str | None = None,
    exception_role: str | None = None,
    interrupt_role: str | None = None,
    oracle_sha: str = 'oracle',
) -> list[tuple[str, ...]]:
    commands: list[tuple[str, ...]] = []

    def fake_capture(
        executable: Path, pipeline: str, input_path: Path, output_path: Path, **kwargs: object
    ) -> CapturedNormalRun:
        role = 'reference' if executable.name.startswith('reference') else 'candidate'
        commands.append((role, pipeline, str(input_path), str(output_path), *kwargs.keys()))
        _write_approved_test_workbook(output_path)
        if role == interrupt_role:
            raise KeyboardInterrupt('simulated process interruption')
        if role == validation_fail_role:
            raise RustNormalValidationError('invalid workbook', 'e' * 64)
        if role == exception_role:
            raise RuntimeError('unexpected adapter failure')
        if role == fail_role:
            raise RustNormalProcessError(9, 'f' * 64)
        return CapturedNormalRun(
            normal_run=NormalRunEvidence(Decimal('1.0'), None, _runtime(pipeline), oracle_sha),
            exit_code=0,
            local_unversioned_log_sha256='l' * 64,
        )

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', fake_capture)
    monkeypatch.setattr(phase0_harness, '_capture_identity', lambda request: _identity())
    monkeypatch.setattr(phase0_harness, 'repo_root', lambda: tmp_path)
    return commands


def _group_request(tmp_path: Path) -> MetricGroupRequest:
    ledger = _ledger(tmp_path)
    return MetricGroupRequest(
        _request(tmp_path),
        'b' * 8,
        'wall',
        build_round_plan(global_round_start=1, round_count=5),
        ledger.attempt_directory,
    )


@pytest.mark.parametrize('metric', ('wall', 'pws'))
def test_formal_group_capture_records_actual_dimensions_when_runtime_omits_them(
    metric: str,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _install_runner(monkeypatch, tmp_path)
    request = _group_request(tmp_path)
    request = replace(request, batch_id='b' * 64, metric=metric)

    def captured(output: Path, *, peak: int | None) -> CapturedNormalRun:
        _write_approved_test_workbook(output)
        runtime = replace(_runtime(), sheet_dimensions=())
        return CapturedNormalRun(
            NormalRunEvidence(Decimal('1'), peak, runtime, '8' * 64),
            0,
            '7' * 64,
        )

    def wall_capture(*args: object, **_kwargs: object) -> CapturedNormalRun:
        return captured(Path(args[3]), peak=None)

    def pws_capture(**kwargs: object) -> CapturedNormalRun:
        return captured(Path(kwargs['output_path']), peak=100)

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', wall_capture)
    monkeypatch.setattr(phase0_harness, '_invoke_pws_single_sample', pws_capture)

    group = run_normal_wall_group(request) if metric == 'wall' else phase0_harness.run_pws_group(request)

    assert {paired.reference.normal_run.runtime.sheet_dimensions for paired in group.rounds} == {
        _APPROVED_TEST_DIMENSIONS
    }
    assert {paired.candidate.normal_run.runtime.sheet_dimensions for paired in group.rounds} == {
        _APPROVED_TEST_DIMENSIONS
    }
    ledger = AppendOnlyAttemptLedger.load(request.attempt_directory, _identity())
    recorded = {
        phase0_harness._sample_from_payload(
            ledger.sample_payload(metric, round_number, role)
        ).normal_run.runtime.sheet_dimensions
        for round_number in range(1, 6)
        for role in ('reference', 'candidate')
    }
    assert recorded == {_APPROVED_TEST_DIMENSIONS}


def _metric_group(metric: str, *, start: int = 1, reference: str = '1.0', candidate: str = '1.0') -> MetricGroup:
    plans = build_round_plan(global_round_start=start, round_count=5)  # type: ignore[arg-type]
    rounds: list[PairedRound] = []
    for plan in plans:
        common = {
            'global_round': plan.global_round,
            'exit_code': 0,
            'input_sha256': '3' * 64,
            'git_head': '4' * 40,
            'repository_state_sha256': '5' * 64,
            'machine_fingerprint_sha256': '6' * 64,
            'local_unversioned_log_sha256': '7' * 64,
        }
        runtime = replace(_runtime(), sheet_dimensions=('A1:A1', 'A1:A1', 'A1:A1'))
        reference_run = NormalRunEvidence(Decimal('1'), 100, runtime, '8' * 64)
        candidate_run = NormalRunEvidence(Decimal('1'), 100, runtime, '8' * 64)
        rounds.append(
            PairedRound(
                plan,
                MetricSample(
                    role='reference',
                    metric_value=Decimal(reference),
                    binary_sha256='1' * 64,
                    normal_run=reference_run,
                    **common,
                ),
                MetricSample(
                    role='candidate',
                    metric_value=Decimal(candidate),
                    binary_sha256='2' * 64,
                    normal_run=candidate_run,
                    **common,
                ),
            )
        )
    return MetricGroup('9' * 64, 'gb', metric, start, tuple(rounds))  # type: ignore[arg-type]


def _with_candidate_stage_ratios(group: MetricGroup, ratios: tuple[dict[str, Decimal], ...]) -> MetricGroup:
    rounds: list[PairedRound] = []
    for paired, round_ratios in zip(group.rounds, ratios, strict=True):
        reference_timings = dict(paired.reference.normal_run.runtime.stage_timings)
        candidate_timings = dict(paired.candidate.normal_run.runtime.stage_timings)
        for stage, ratio in round_ratios.items():
            reference_timings.setdefault(stage, Decimal('0.1'))
            candidate_timings[stage] = reference_timings[stage] * ratio
        reference = replace(
            paired.reference,
            normal_run=replace(
                paired.reference.normal_run,
                runtime=replace(
                    paired.reference.normal_run.runtime,
                    stage_timings=tuple(reference_timings.items()),
                ),
            ),
        )
        candidate = replace(
            paired.candidate,
            normal_run=replace(
                paired.candidate.normal_run,
                runtime=replace(
                    paired.candidate.normal_run.runtime,
                    stage_timings=tuple(candidate_timings.items()),
                ),
            ),
        )
        rounds.append(replace(paired, reference=reference, candidate=candidate))
    return replace(group, rounds=tuple(rounds))


def _with_output_sizes(
    group: MetricGroup,
    *,
    reference_sizes: tuple[int, ...],
    candidate_sizes: tuple[int, ...],
) -> MetricGroup:
    assert len(group.rounds) == len(reference_sizes) == len(candidate_sizes)

    def with_size(sample: MetricSample, size: int) -> MetricSample:
        runtime = replace(sample.normal_run.runtime, output_size_bytes=size)
        return replace(sample, normal_run=replace(sample.normal_run, runtime=runtime))

    return replace(
        group,
        rounds=tuple(
            replace(
                paired,
                reference=with_size(paired.reference, reference_size),
                candidate=with_size(paired.candidate, candidate_size),
            )
            for paired, reference_size, candidate_size in zip(
                group.rounds, reference_sizes, candidate_sizes, strict=True
            )
        ),
    )


def _phase0a_baseline(output_size: int = 100) -> phase0_harness._ApprovedPhase0ABaseline:
    return phase0_harness._ApprovedPhase0ABaseline(
        '5' * 8,
        output_size,
        (Decimal('1'),) * 5,
        (Decimal('100'),) * 5,
    )


def _calibration_group(pipeline: str, metric: str, output_size: int) -> CalibrationGroup:
    paired = _metric_group(metric)
    rounds = tuple(
        CalibrationRound(
            item.plan.global_round,
            replace(
                item.reference,
                normal_run=replace(
                    item.reference.normal_run,
                    runtime=replace(
                        item.reference.normal_run.runtime,
                        pipeline=pipeline,
                        output_size_bytes=output_size,
                        sheet_dimensions=('A1:L8', 'A1:AZ3', 'A1:AZ1'),
                    ),
                ),
            ),
        )
        for item in paired.rounds
    )
    return CalibrationGroup(f'{pipeline}-{metric}', pipeline, metric, True, rounds)  # type: ignore[arg-type]


def _with_calibration_output_sizes(
    group: CalibrationGroup,
    output_sizes: tuple[int | None, ...],
) -> CalibrationGroup:
    return replace(
        group,
        rounds=tuple(
            replace(
                item,
                reference=replace(
                    item.reference,
                    normal_run=replace(
                        item.reference.normal_run,
                        runtime=replace(item.reference.normal_run.runtime, output_size_bytes=output_size),
                    ),
                ),
            )
            for item, output_size in zip(group.rounds, output_sizes, strict=True)
        ),
    )


def test_sanitized_fixture_contains_no_erp_or_host_canary(tmp_path: Path) -> None:
    spec = importlib.util.find_spec('tests.rust_oracle.sanitized_fixture')
    assert spec is not None, 'sanitized fixture builder is required'
    from tests.rust_oracle.sanitized_fixture import build_raw_fixture

    fixture = tmp_path / 'synthetic-gb.xlsx'
    build_raw_fixture(fixture, 'gb', 'small')

    with zipfile.ZipFile(fixture) as archive:
        package_text = b'\n'.join(archive.read(name) for name in archive.namelist()).decode('utf-8', errors='ignore')
    folded = package_text.casefold()
    assert 'erp' not in folded
    assert getpass.getuser().casefold() not in folded
    assert socket.gethostname().casefold() not in folded


def test_stable_workbook_oracle_ignores_volatile_core_properties(tmp_path: Path) -> None:
    import xlsxwriter

    paths = (tmp_path / 'first.xlsx', tmp_path / 'second.xlsx')
    for index, path in enumerate(paths):
        workbook = xlsxwriter.Workbook(path)
        workbook.set_properties({'comments': f'volatile-{index}'})
        workbook.add_worksheet('Sheet1').write(0, 0, 'same')
        workbook.close()

    assert phase0_harness._stable_workbook_oracle(paths[0]) == phase0_harness._stable_workbook_oracle(paths[1])


def test_low_memory_fixture_produces_at_least_five_million_output_slots(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from tests.rust_oracle.sanitized_fixture import (
        LOW_MEMORY_OUTPUT_COLUMNS_LOWER_BOUND,
        LOW_MEMORY_QUANTITY_ROWS,
        build_raw_fixture,
    )

    fixture = tmp_path / 'synthetic-low-memory-sk.xlsx'
    output = tmp_path / 'synthetic-low-memory-sk-output.xlsx'
    build_raw_fixture(fixture, 'sk', 'low-memory')

    workbook = load_workbook(fixture, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        assert sheet.max_row >= LOW_MEMORY_QUANTITY_ROWS + 4
        assert LOW_MEMORY_QUANTITY_ROWS >= 100_000
        assert LOW_MEMORY_QUANTITY_ROWS * LOW_MEMORY_OUTPUT_COLUMNS_LOWER_BOUND >= 5_000_000
    finally:
        workbook.close()

    run_rust_cli_release(build_rust_cli_release(), 'sk', fixture, output)
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        sheet = workbook[ApprovedSheet.QUANTITY.value]
        assert sheet.max_row >= LOW_MEMORY_QUANTITY_ROWS + 1
        assert sheet.max_row * sheet.max_column >= 5_000_000
    finally:
        workbook.close()


def _install_smoke_capture(monkeypatch: pytest.MonkeyPatch) -> list[str]:
    calls: list[str] = []

    def fake_wall(
        executable: Path, pipeline: str, input_path: Path, output_path: Path, **kwargs: object
    ) -> CapturedNormalRun:
        del executable, input_path, kwargs
        calls.append('wall')
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b'synthetic workbook')
        return CapturedNormalRun(
            NormalRunEvidence(Decimal('1'), None, _runtime(pipeline), '8' * 64),
            0,
            '7' * 64,
        )

    def fake_pws(**kwargs: object) -> CapturedNormalRun:
        calls.append('pws')
        assert kwargs['local_root'] == phase0_harness._trusted_local_root()
        artifacts = phase0_harness._pws_local_artifact_paths(
            kwargs['local_root'], kwargs['batch_id'], kwargs['global_round'], kwargs['role']
        )
        for path in (artifacts.result_path, artifacts.stdout_path, artifacts.stderr_path, artifacts.driver_log_path):
            phase0_harness._io_path(path).write_text('{}', encoding='utf-8')
        output_path = kwargs['output_path']
        assert isinstance(output_path, Path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b'synthetic workbook')
        pipeline = kwargs['pipeline']
        assert isinstance(pipeline, str)
        return CapturedNormalRun(
            NormalRunEvidence(Decimal('1'), 123, _runtime(pipeline), '8' * 64),
            0,
            '7' * 64,
        )

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', fake_wall)
    monkeypatch.setattr(phase0_harness, '_invoke_pws_single_sample', fake_pws)
    monkeypatch.setattr(
        phase0_harness, '_workbook_sheet_names', lambda path: tuple(item.value for item in ApprovedSheet)
    )
    return calls


def test_phase0h_smoke_runs_normal_wall_and_normal_pws(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    calls = _install_smoke_capture(monkeypatch)
    executable = tmp_path / 'reference.exe'
    executable.write_bytes(b'synthetic executable')

    smoke_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0h-smoke'
    result = run_phase0h_smoke(Phase0HSmokeRequest('gb', executable, executable, smoke_root))

    assert result.verdict is HarnessVerdict.VALIDATED
    assert calls.count('wall') == 10
    assert calls.count('pws') == 10


def test_phase0h_smoke_cleans_every_workbook(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _install_smoke_capture(monkeypatch)
    executable = tmp_path / 'reference.exe'
    executable.write_bytes(b'synthetic executable')
    smoke_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0h-smoke'

    run_phase0h_smoke(Phase0HSmokeRequest('sk', executable, executable, smoke_root))

    assert not list(smoke_root.rglob('*.xlsx'))
    assert not list(phase0_harness._io_path(phase0_harness._trusted_local_root() / 'pws-results').rglob('*.json'))
    assert not list(phase0_harness._io_path(phase0_harness._trusted_local_root() / 'pws-logs').rglob('*.log'))


def test_phase0h_smoke_records_observed_canary_residue_and_sheet_contract(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_smoke_capture(monkeypatch)
    executable = tmp_path / 'reference.exe'
    executable.write_bytes(b'synthetic executable')
    observed: list[SmokeSummaryEvidence] = []
    original = EvidenceSanitizer.build_smoke

    def capture(self: EvidenceSanitizer, value: SmokeSummaryEvidence) -> object:
        observed.append(value)
        return original(self, value)

    monkeypatch.setattr(EvidenceSanitizer, 'build_smoke', capture)
    run_phase0h_smoke(
        Phase0HSmokeRequest(
            'gb',
            executable,
            executable,
            tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0h-smoke',
        )
    )

    smoke = observed[0]
    assert smoke.temp_canary_created is True
    assert smoke.temp_residue_count == 0
    assert smoke.approved_sheets == tuple(ApprovedSheet)


@pytest.mark.parametrize('mutation', ('missing-sentinel', 'extra-artifact'))
def test_phase0h_smoke_rejects_temp_canary_mutation_and_still_cleans_environment(
    mutation: str, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_smoke_capture(monkeypatch)
    original_capture = phase0_harness.run_rust_normal_captured
    mutated = False

    def mutate_canary(*args: object, **kwargs: object) -> CapturedNormalRun:
        nonlocal mutated
        captured = original_capture(*args, **kwargs)  # type: ignore[arg-type]
        if not mutated:
            temp_root = Path(phase0_harness.os.environ['TEMP'])
            if mutation == 'missing-sentinel':
                (temp_root / 'sentinel.txt').unlink()
            else:
                (temp_root / 'unexpected.tmp').write_bytes(b'unexpected')
            mutated = True
        return captured

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', mutate_canary)
    monkeypatch.setenv('TEMP', str(tmp_path / 'original-temp'))
    monkeypatch.setenv('TMP', str(tmp_path / 'original-tmp'))
    executable = tmp_path / 'reference.exe'
    executable.write_bytes(b'synthetic executable')
    smoke_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0h-smoke'

    with pytest.raises(HarnessFailure) as caught:
        run_phase0h_smoke(Phase0HSmokeRequest('gb', executable, executable, smoke_root))

    assert caught.value.verdict is HarnessVerdict.ENVIRONMENT_DRIFT
    assert phase0_harness.os.environ['TEMP'] == str(tmp_path / 'original-temp')
    assert phase0_harness.os.environ['TMP'] == str(tmp_path / 'original-tmp')
    assert not (smoke_root / 'temp-canary').exists()
    assert not list(smoke_root.rglob('*.xlsx'))


def test_phase0a_manifest_uses_external_output_size_for_base_reference(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)

    phase0_harness.capture_phase0a(request)

    payload = json.loads(request.output_path.read_text(encoding='utf-8'))
    assert payload['pipelines']['gb']['output_size_bytes'] == 321
    assert payload['pipelines']['sk']['output_size_bytes'] == 654


def test_phase0a_manifest_uses_conservative_decimal_median_for_one_byte_output_variation(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)

    def variable_group(*, pipeline: str, metric: str, **kwargs: object) -> CalibrationGroup:
        del kwargs
        base = 3_813_018 if pipeline == 'gb' else 654
        sizes = (base,) * 5 if metric == 'wall' else (base + 1,) * 5
        return _with_calibration_output_sizes(_calibration_group(pipeline, metric, base), sizes)

    monkeypatch.setattr(phase0_harness, '_capture_phase0a_group', variable_group)

    phase0_harness.capture_phase0a(request)

    payload = json.loads(request.output_path.read_text(encoding='utf-8'))
    assert payload['pipelines']['gb']['output_size_bytes'] == 3_813_019
    assert payload['pipelines']['sk']['output_size_bytes'] == 655


@pytest.mark.parametrize('invalid_size', (None, 0, -1, True))
def test_phase0a_manifest_rejects_missing_non_positive_or_boolean_output_size(
    invalid_size: int | None,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)

    def invalid_group(*, pipeline: str, metric: str, **kwargs: object) -> CalibrationGroup:
        del kwargs
        group = _calibration_group(pipeline, metric, 321 if pipeline == 'gb' else 654)
        if pipeline == 'gb' and metric == 'pws':
            return _with_calibration_output_sizes(group, (invalid_size,) * 5)
        return group

    monkeypatch.setattr(phase0_harness, '_capture_phase0a_group', invalid_group)

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.capture_phase0a(request)

    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert not request.output_path.exists()


def test_phase0a_manifest_contains_gb_sk_wall_pws_runtime_and_sheet_dimensions(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)

    phase0_harness.capture_phase0a(request)

    payload = json.loads(request.output_path.read_text(encoding='utf-8'))
    assert tuple(payload['pipelines']) == ('gb', 'sk')
    for pipeline in ('gb', 'sk'):
        assert set(payload['pipelines'][pipeline]['calibration']) == {'wall', 'pws'}
        assert payload['pipelines'][pipeline]['runtime']['sheet_count'] == 3
        assert payload['pipelines'][pipeline]['sheet_dimensions'] == ['A1:L8', 'A1:AZ3', 'A1:AZ1']


def test_phase0a_manifest_contains_no_paths_filenames_hostname_or_raw_logs(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)

    phase0_harness.capture_phase0a(request)

    text = request.output_path.read_text(encoding='utf-8').casefold()
    assert str(tmp_path).casefold() not in text
    assert request.gb_input_path.name.casefold() not in text
    assert request.sk_input_path.name.casefold() not in text
    assert socket.gethostname().casefold() not in text
    assert 'stdout' not in text and 'stderr' not in text


def test_phase0a_capture_refuses_existing_manifest(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)
    request.output_path.parent.mkdir(parents=True, exist_ok=True)
    request.output_path.write_text('{}', encoding='utf-8')

    with pytest.raises(FileExistsError):
        phase0_harness.capture_phase0a(request)


def test_phase0a_pws_uses_trusted_raw_root(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    input_path = tmp_path / 'synthetic.xlsx'
    executable = tmp_path / 'reference.exe'
    input_path.write_bytes(b'input')
    executable.write_bytes(b'executable')
    local_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0a'
    machine = MachineEvidence('build', 'x86_64', 'cpu', 1, 1, 'UNKNOWN', 1, '6' * 64)
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: '4' * 40)
    monkeypatch.setattr(phase0_harness, '_capture_machine_evidence', lambda: machine)

    def fake_pws(**kwargs: object) -> CapturedNormalRun:
        assert kwargs['local_root'] == phase0_harness._trusted_local_root()
        assert phase0_harness._is_sha256(kwargs['batch_id'])
        output = kwargs['output_path']
        assert isinstance(output, Path)
        _write_approved_test_workbook(output)
        return CapturedNormalRun(
            NormalRunEvidence(
                Decimal('1'),
                123,
                replace(_runtime('gb'), sheet_dimensions=_APPROVED_TEST_DIMENSIONS),
                '8' * 64,
            ),
            0,
            '7' * 64,
        )

    monkeypatch.setattr(phase0_harness, '_invoke_pws_single_sample', fake_pws)
    group = phase0_harness._capture_phase0a_group(
        pipeline='gb',
        metric='pws',
        input_path=input_path,
        executable=executable,
        local_root=local_root,
        machine=machine,
    )
    assert len(group.rounds) == 5


def test_phase0a_group_fills_empty_runtime_dimensions_from_actual_workbook(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    input_path = tmp_path / 'synthetic.xlsx'
    executable = tmp_path / 'reference.exe'
    input_path.write_bytes(b'input')
    executable.write_bytes(b'executable')
    local_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0a'
    machine = MachineEvidence('build', 'x86_64', 'cpu', 1, 1, 'UNKNOWN', 1, '6' * 64)
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: '4' * 40)
    monkeypatch.setattr(phase0_harness, '_capture_machine_evidence', lambda: machine)

    def fake_capture(
        executable: Path, pipeline: str, input_path: Path, output_path: Path, **kwargs: object
    ) -> CapturedNormalRun:
        del executable, input_path, kwargs
        _write_approved_test_workbook(output_path)
        return CapturedNormalRun(
            NormalRunEvidence(Decimal('1'), None, replace(_runtime(pipeline), sheet_dimensions=()), '8' * 64),
            0,
            '7' * 64,
        )

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', fake_capture)
    group = phase0_harness._capture_phase0a_group(
        pipeline='gb',
        metric='wall',
        input_path=input_path,
        executable=executable,
        local_root=local_root,
        machine=machine,
    )

    assert all(item.reference.normal_run.runtime.sheet_dimensions == _APPROVED_TEST_DIMENSIONS for item in group.rounds)
    sk_wall = _calibration_group('sk', 'wall', 8)
    sk_pws = _calibration_group('sk', 'pws', 8)
    payload = phase0_harness._phase0a_payload(
        phase0_harness.Phase0AManifest(
            '1' * 64,
            'a' * 40,
            '4' * 40,
            machine,
            group,
            replace(group, metric='pws'),
            sk_wall,
            sk_pws,
        ),
        _phase0a_request(tmp_path),
    )
    assert payload['pipelines']['gb']['sheet_dimensions'] == list(_APPROVED_TEST_DIMENSIONS)


def test_phase0a_group_rejects_runtime_dimensions_that_differ_from_actual_workbook(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    input_path = tmp_path / 'synthetic.xlsx'
    executable = tmp_path / 'reference.exe'
    input_path.write_bytes(b'input')
    executable.write_bytes(b'executable')
    local_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0a'
    machine = MachineEvidence('build', 'x86_64', 'cpu', 1, 1, 'UNKNOWN', 1, '6' * 64)
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: '4' * 40)
    monkeypatch.setattr(phase0_harness, '_capture_machine_evidence', lambda: machine)

    def fake_capture(
        executable: Path, pipeline: str, input_path: Path, output_path: Path, **kwargs: object
    ) -> CapturedNormalRun:
        del executable, input_path, kwargs
        _write_approved_test_workbook(output_path)
        mismatched = replace(_runtime(pipeline), sheet_dimensions=('A1:A1', 'A1:A1', 'A1:A1'))
        return CapturedNormalRun(NormalRunEvidence(Decimal('1'), None, mismatched, '8' * 64), 0, '7' * 64)

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', fake_capture)

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._capture_phase0a_group(
            pipeline='gb',
            metric='wall',
            input_path=input_path,
            executable=executable,
            local_root=local_root,
            machine=machine,
        )
    assert caught.value.verdict is HarnessVerdict.CORRECTNESS_FAILED


def _phase0a_request(tmp_path: Path) -> Phase0ARequest:
    gb = tmp_path / 'confidential-gb.xlsx'
    sk = tmp_path / 'confidential-sk.xlsx'
    executable = tmp_path / 'reference.exe'
    gb.write_bytes(b'gb')
    sk.write_bytes(b'sk')
    executable.write_bytes(b'executable')
    return Phase0ARequest(
        gb,
        sk,
        executable,
        'a' * 40,
        tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0a',
        tmp_path / 'docs' / 'performance' / 'phase0a.json',
    )


def _install_phase0a_capture(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setenv('COSTING_GB_SAMPLE', str(tmp_path / 'confidential-gb.xlsx'))
    monkeypatch.setenv('COSTING_SK_SAMPLE', str(tmp_path / 'confidential-sk.xlsx'))

    def fake_group(*, pipeline: str, metric: str, **kwargs: object) -> CalibrationGroup:
        del kwargs
        return _calibration_group(pipeline, metric, 321 if pipeline == 'gb' else 654)

    monkeypatch.setattr(phase0_harness, '_capture_phase0a_group', fake_group, raising=False)
    monkeypatch.setattr(
        phase0_harness,
        '_capture_machine_evidence',
        lambda: MachineEvidence('synthetic-build', 'x86_64', 'synthetic-cpu', 4, 8, 'SSD', 16, '6' * 64),
        raising=False,
    )
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: '4' * 40)


def test_phase0a_group_failure_cleans_residuals_and_allows_rerun(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)
    failed = False

    def flaky_group(*, pipeline: str, metric: str, local_root: Path, **kwargs: object) -> CalibrationGroup:
        nonlocal failed
        del kwargs
        raw_log = local_root / 'raw-logs' / 'partial.json'
        workbook = local_root / 'outputs' / pipeline / metric / 'partial.xlsx'
        raw_log.parent.mkdir(parents=True, exist_ok=True)
        workbook.parent.mkdir(parents=True, exist_ok=True)
        raw_log.write_text('{}', encoding='utf-8')
        workbook.write_bytes(b'partial')
        if not failed:
            failed = True
            raise RustNormalProcessError(9, 'f' * 64)
        return _calibration_group(pipeline, metric, 321 if pipeline == 'gb' else 654)

    monkeypatch.setattr(phase0_harness, '_capture_phase0a_group', flaky_group)

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.capture_phase0a(request)
    assert caught.value.verdict is HarnessVerdict.REFERENCE_FAILED
    assert caught.value.raw_log_sha256 == 'f' * 64
    assert not request.output_path.exists()
    assert not (request.local_root / 'raw-logs').exists()
    assert not (request.local_root / 'outputs').exists()

    phase0_harness.capture_phase0a(request)
    assert request.output_path.is_file()


@pytest.mark.parametrize('failure_point', ('scan', 'publish'))
def test_phase0a_scan_or_publish_failure_removes_staging_and_incomplete_manifest(
    failure_point: str, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)
    if failure_point == 'scan':
        monkeypatch.setattr(
            EvidenceSanitizer,
            'scan_tree',
            lambda self, *args, **kwargs: (_ for _ in ()).throw(ValueError('sensitive staging')),
        )
    else:
        monkeypatch.setattr(
            phase0_harness.os,
            'link',
            lambda *args, **kwargs: (_ for _ in ()).throw(OSError('publication failed')),
        )

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.capture_phase0a(request)

    assert caught.value.verdict is HarnessVerdict.SENSITIVE_EVIDENCE
    assert not request.output_path.exists()
    assert not list(request.local_root.glob('phase0a-sanitized-*'))
    assert not (request.local_root / 'raw-logs').exists()
    assert not (request.local_root / 'outputs').exists()


def test_phase0a_outer_cleanup_failure_preserves_nested_cleanup_primary_verdict(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)
    monkeypatch.setattr(
        phase0_harness,
        '_capture_phase0a_group',
        lambda **kwargs: (_ for _ in ()).throw(
            HarnessFailure(
                HarnessVerdict.CLEANUP_FAILED,
                'group cleanup failed',
                primary_verdict=HarnessVerdict.REFERENCE_FAILED,
                raw_log_sha256='f' * 64,
            )
        ),
    )
    monkeypatch.setattr(
        phase0_harness,
        '_remove_local_tree',
        lambda *args, **kwargs: (_ for _ in ()).throw(PermissionError('outer cleanup failed')),
    )

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.capture_phase0a(request)

    assert caught.value.verdict is HarnessVerdict.CLEANUP_FAILED
    assert caught.value.primary_verdict is HarnessVerdict.REFERENCE_FAILED
    assert caught.value.raw_log_sha256 == 'f' * 64
    assert not request.output_path.exists()


@pytest.mark.parametrize(
    ('error', 'expected_raw_log'),
    (
        (RustNormalProcessError(9, 'f' * 64), 'f' * 64),
        (RustNormalValidationError('invalid workbook', 'e' * 64), 'e' * 64),
        (RuntimeError('unexpected capture failure'), None),
    ),
)
def test_phase0a_maps_capture_boundary_to_reference_failure(
    error: Exception,
    expected_raw_log: str | None,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)
    monkeypatch.setattr(phase0_harness, '_capture_phase0a_group', lambda **kwargs: (_ for _ in ()).throw(error))

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.capture_phase0a(request)

    assert caught.value.verdict is HarnessVerdict.REFERENCE_FAILED
    assert caught.value.raw_log_sha256 == expected_raw_log


def test_phase0a_rechecks_machine_evidence_after_each_capture(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    initial = MachineEvidence('build', 'x86_64', 'cpu', 4, 8, 'SSD', 16, '6' * 64)
    changed = replace(initial, fingerprint_sha256='9' * 64)
    observed = iter((initial, changed))
    input_path = tmp_path / 'synthetic.xlsx'
    executable = tmp_path / 'reference.exe'
    input_path.write_bytes(b'input')
    executable.write_bytes(b'executable')
    local_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'phase0a'
    monkeypatch.setattr(phase0_harness, '_capture_machine_evidence', lambda: next(observed))
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: '4' * 40)

    def fake_capture(
        executable: Path, pipeline: str, input_path: Path, output_path: Path, **kwargs: object
    ) -> CapturedNormalRun:
        del executable, input_path, kwargs
        _write_approved_test_workbook(output_path)
        return CapturedNormalRun(
            NormalRunEvidence(
                Decimal('1'),
                None,
                replace(_runtime(pipeline), sheet_dimensions=_APPROVED_TEST_DIMENSIONS),
                '8' * 64,
            ),
            0,
            '7' * 64,
        )

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', fake_capture)

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._capture_phase0a_group(
            pipeline='gb',
            metric='wall',
            input_path=input_path,
            executable=executable,
            local_root=local_root,
            machine=initial,
        )
    assert caught.value.verdict is HarnessVerdict.ENVIRONMENT_DRIFT
    assert not list(local_root.rglob('*.xlsx'))


def test_main_maps_smoke_and_phase0a_capture_failures_without_traceback(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    executable = tmp_path / 'reference.exe'
    executable.write_bytes(b'executable')
    smoke_root = tmp_path / 'rust' / 'target' / 'perf-local' / 'smoke'
    monkeypatch.setattr(
        phase0_harness,
        'run_rust_normal_captured',
        lambda *args, **kwargs: (_ for _ in ()).throw(RustNormalProcessError(9, 'f' * 64)),
    )
    smoke_exit = phase0_harness.main(
        [
            'smoke',
            '--pipeline',
            'gb',
            '--reference-executable',
            str(executable),
            '--candidate-executable',
            str(executable),
            '--local-root',
            str(smoke_root),
        ]
    )
    assert smoke_exit == 3
    assert 'Traceback' not in capsys.readouterr().err

    _install_phase0a_capture(monkeypatch, tmp_path)
    request = _phase0a_request(tmp_path)
    monkeypatch.setattr(
        phase0_harness,
        '_capture_phase0a_group',
        lambda **kwargs: (_ for _ in ()).throw(RuntimeError('unexpected capture failure')),
    )
    phase0a_exit = phase0_harness.main(
        [
            'phase0a',
            '--gb-input',
            str(request.gb_input_path),
            '--sk-input',
            str(request.sk_input_path),
            '--reference-executable',
            str(request.reference_executable),
            '--fork-revision',
            request.fork_revision,
            '--local-root',
            str(request.local_root),
            '--output',
            str(request.output_path),
        ]
    )
    assert phase0a_exit == 3
    assert 'Traceback' not in capsys.readouterr().err


def test_paired_cli_exposes_no_batch_round_or_threshold_argument() -> None:
    parser = phase0_harness._argument_parser()
    paired = next(action for action in parser._actions if action.dest == 'command').choices['paired']
    option_strings = {option for action in paired._actions for option in action.option_strings}
    assert option_strings == {
        '-h',
        '--help',
        '--pipeline',
        '--input',
        '--reference-executable',
        '--candidate-executable',
        '--reference-label',
        '--candidate-label',
        '--comparison-profile',
        '--phase0a-manifest',
        '--local-root',
        '--evidence-path',
    }


def test_paired_cli_uses_closed_labels_profiles_and_exit_codes() -> None:
    parser = phase0_harness._argument_parser()
    paired = next(action for action in parser._actions if action.dest == 'command').choices['paired']
    actions = {action.dest: action for action in paired._actions}
    assert set(actions['reference_label'].choices) == {item.value for item in ClosedBinaryLabel}
    assert set(actions['candidate_label'].choices) == {item.value for item in ClosedBinaryLabel}
    assert set(actions['comparison_profile'].choices) == {item.value for item in ComparisonProfile}
    assert phase0_harness._exit_code(HarnessVerdict.VALIDATED) == 0
    assert phase0_harness._exit_code(HarnessVerdict.CANDIDATE_FAILED) == 2
    assert phase0_harness._exit_code(HarnessVerdict.REFERENCE_FAILED) == 3
    assert phase0_harness._exit_code(HarnessVerdict.CLEANUP_FAILED) == 4
    assert phase0_harness.main(['paired']) == 5


@pytest.mark.parametrize('pipeline', ('gb', 'sk'))
def test_sanitized_raw_fixture_runs_both_python_and_rust_to_three_sheets(tmp_path: Path, pipeline: str) -> None:
    from openpyxl import load_workbook

    from tests.rust_oracle.evidence import ApprovedSheet
    from tests.rust_oracle.sanitized_fixture import build_raw_fixture

    fixture = tmp_path / f'synthetic-{pipeline}.xlsx'
    python_output = tmp_path / f'python-{pipeline}.xlsx'
    rust_output = tmp_path / f'rust-{pipeline}.xlsx'
    build_raw_fixture(fixture, pipeline, 'small')  # type: ignore[arg-type]
    run_python_oracle(pipeline, fixture, python_output)
    run_rust_cli_release(build_rust_cli_release(), pipeline, fixture, rust_output)

    expected = [item.value for item in ApprovedSheet]
    for output in (python_output, rust_output):
        workbook = load_workbook(output, read_only=True, data_only=True)
        try:
            assert workbook.sheetnames == expected
        finally:
            workbook.close()


def test_normal_wall_group_omits_check_only_and_uses_unique_outputs(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    commands = _install_runner(monkeypatch, tmp_path)
    run_normal_wall_group(_group_request(tmp_path))
    paths = [command[3] for command in commands]
    assert len(paths) == len(set(paths)) == 10
    assert all('--check-only' not in command for command in commands)


def test_normal_wall_group_runs_global_ab_ba_order(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    commands = _install_runner(monkeypatch, tmp_path)
    run_normal_wall_group(_group_request(tmp_path))
    assert [item[0] for item in commands] == [
        'reference',
        'candidate',
        'candidate',
        'reference',
        'reference',
        'candidate',
        'candidate',
        'reference',
        'reference',
        'candidate',
    ]


def test_candidate_nonzero_rejects_candidate(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _install_runner(monkeypatch, tmp_path, fail_role='candidate')
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(_group_request(tmp_path))
    assert caught.value.verdict is HarnessVerdict.CANDIDATE_FAILED


def test_reference_nonzero_invalidates_whole_batch(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _install_runner(monkeypatch, tmp_path, fail_role='reference')
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(_group_request(tmp_path))
    assert caught.value.verdict is HarnessVerdict.REFERENCE_FAILED


@pytest.mark.parametrize('field', ('input_sha256', 'reference_sha256', 'candidate_sha256', 'git_head'))
def test_wall_group_rejects_input_reference_candidate_or_git_drift(
    field: str, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_runner(monkeypatch, tmp_path)
    values = [_identity(), replace(_identity(), **{field: 'drift'})]
    monkeypatch.setattr(
        phase0_harness,
        '_capture_identity',
        lambda request: values.pop(0) if values else replace(_identity(), **{field: 'drift'}),
    )
    with pytest.raises(HarnessFailure, match='drift'):
        run_normal_wall_group(_group_request(tmp_path))


def test_wall_group_deletes_workbooks_on_success(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _install_runner(monkeypatch, tmp_path)
    run_normal_wall_group(_group_request(tmp_path))
    assert not list((tmp_path / 'data').rglob('*.xlsx'))


def test_wall_group_deletes_workbooks_after_process_or_oracle_failure(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _install_runner(monkeypatch, tmp_path, fail_role='candidate')
    with pytest.raises(HarnessFailure):
        run_normal_wall_group(_group_request(tmp_path))
    assert not list((tmp_path / 'data').rglob('*.xlsx'))


def test_cleanup_failure_prevents_versionable_evidence(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _install_runner(monkeypatch, tmp_path)
    monkeypatch.setattr(
        Path,
        'unlink',
        lambda self, **kwargs: (_ for _ in ()).throw(PermissionError('locked')) if self.suffix == '.xlsx' else None,
    )
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(_group_request(tmp_path))
    assert caught.value.verdict is HarnessVerdict.CLEANUP_FAILED
    assert not _request(tmp_path).evidence_path.exists()


def test_first_formal_batch_requires_clean_worktree(tmp_path: Path) -> None:
    with pytest.raises(HarnessFailure, match='clean'):
        validate_formal_repository_state((' M tests/file.py',), evidence_root=tmp_path, prior_evidence_claims=())


def test_formal_batch_rejects_non_evidence_worktree_change(tmp_path: Path) -> None:
    approved = tmp_path / 'docs' / 'prior.json'
    approved.parent.mkdir()
    approved.write_bytes(b'{}')
    claim = _claim(tmp_path, approved, phase0_harness._sha256(approved))
    with pytest.raises(HarnessFailure, match='non-evidence'):
        validate_formal_repository_state(
            ('?? src/new.py',),
            evidence_root=tmp_path / 'docs',
            prior_evidence_claims=(claim,),
            root=tmp_path,
        )


def test_later_batch_accepts_only_create_new_sanitized_prior_evidence(tmp_path: Path) -> None:
    evidence = tmp_path / 'docs' / 'prior.json'
    evidence.parent.mkdir()
    evidence.write_bytes(b'{"sanitized":true}')
    digest = phase0_harness._sha256(evidence)
    validate_formal_repository_state(
        ('?? docs/prior.json',),
        evidence_root=tmp_path / 'docs',
        prior_evidence_claims=(_claim(tmp_path, evidence, digest),),
        root=tmp_path,
    )


def test_prior_evidence_content_change_invalidates_repository_state(tmp_path: Path) -> None:
    evidence = tmp_path / 'docs' / 'prior.json'
    evidence.parent.mkdir()
    evidence.write_bytes(b'changed')
    with pytest.raises(HarnessFailure, match='content'):
        validate_formal_repository_state(
            ('?? docs/prior.json',),
            evidence_root=tmp_path / 'docs',
            prior_evidence_claims=(_claim(tmp_path, evidence, '0' * 64),),
            root=tmp_path,
        )


def test_batch_id_is_derived_and_cannot_be_supplied(tmp_path: Path) -> None:
    request = _request(tmp_path)
    batch_id = derive_batch_id(request, _identity())
    assert batch_id == derive_batch_id(request, _identity())
    assert len(batch_id) == 64
    assert 'batch_id' not in PairedBenchmarkRequest.__dataclass_fields__


def test_batch_id_explicitly_contains_protocol_v2_identity(tmp_path: Path) -> None:
    request = _request(tmp_path)
    identity = BenchmarkIdentity('3' * 64, '1' * 64, '2' * 64, '4' * 40, '5' * 64, '6' * 64)
    current = derive_batch_id(request, identity)
    legacy_payload = {'profile': request.comparison_profile.value, 'pipeline': request.pipeline, **asdict(identity)}
    legacy = hashlib.sha256(phase0_harness._canonical_json(legacy_payload)).hexdigest()
    assert current != legacy


def test_v2_ledger_metadata_records_exact_integer_protocol_version(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    payload = json.loads((ledger.attempt_directory / 'metadata.json').read_text(encoding='utf-8'))
    assert payload['protocol_version'] == 2
    assert type(payload['protocol_version']) is int
    assert ledger.protocol_version == 2
    assert phase0_harness._batch_attempt(ledger, 'b' * 64, AttemptState.CREATED).protocol_version == 2


@pytest.mark.parametrize('invalid', (True, '2', 0, 3))
def test_ledger_rejects_unknown_or_non_integer_protocol_version(invalid: object, tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    _rewrite_metadata_and_matching_empty_journal(ledger.attempt_directory, protocol_version=invalid)
    with pytest.raises(HarnessFailure) as caught:
        AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity())
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_v1_metadata_without_protocol_version_loads_read_only(tmp_path: Path) -> None:
    attempt = _write_synthetic_v1_terminal(tmp_path, comparison_key='a' * 64)
    comparison = attempt.parent
    before = _file_snapshot(comparison)
    loaded = AppendOnlyAttemptLedger.load(attempt, _identity(), strict_identity=False)
    assert loaded.protocol_version == 1
    assert loaded.terminal_verdict is HarnessVerdict.INCONCLUSIVE
    assert _file_snapshot(comparison) == before


def test_v2_start_does_not_mutate_synthetic_v1_terminal(tmp_path: Path) -> None:
    v1_attempt = _write_synthetic_v1_terminal(tmp_path, comparison_key='a' * 64)
    terminal = v1_attempt / 'terminal.json'
    before = terminal.read_bytes()
    before_sha = hashlib.sha256(before).hexdigest()
    v2 = AppendOnlyAttemptLedger.create(
        tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
        _identity(),
        comparison_key='b' * 64,
    )
    assert v2.attempt_number == 1
    assert terminal.read_bytes() == before
    assert hashlib.sha256(terminal.read_bytes()).hexdigest() == before_sha


def test_v2_create_refuses_to_append_synthetic_v1_comparison_directory(tmp_path: Path) -> None:
    attempt = _write_synthetic_v1_terminal(tmp_path, comparison_key='a' * 64)
    comparison = attempt.parent
    before = _file_snapshot(comparison)
    with pytest.raises(HarnessFailure, match='protocol v1') as caught:
        AppendOnlyAttemptLedger.create(
            tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
            _identity(),
            comparison_key='a' * 64,
        )
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert _file_snapshot(comparison) == before


@pytest.mark.parametrize(
    ('missing_checkpoint', 'stale_journal'),
    ((True, False), (False, True)),
)
def test_v1_audit_recovery_edges_fail_without_writing(
    missing_checkpoint: bool,
    stale_journal: bool,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    attempt = _write_recoverable_v1_success(
        tmp_path,
        comparison_key='a' * 64,
        missing_checkpoint=missing_checkpoint,
        stale_journal=stale_journal,
    )
    comparison = attempt.parent
    before = _file_snapshot(comparison)
    monkeypatch.setattr(
        phase0_harness,
        '_validate_prepared_evidence',
        lambda record: {
            'artifact_basename': record['artifact_basename'],
            'artifact_sha256': record['artifact_sha256'],
            'artifact_content': record['artifact_content'],
        },
    )
    with pytest.raises(HarnessFailure, match='protocol v1') as caught:
        AppendOnlyAttemptLedger.load(attempt, _identity(), strict_identity=False)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert _file_snapshot(comparison) == before


def test_v2_create_does_not_repair_v1_before_rejecting(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    attempt = _write_recoverable_v1_success(
        tmp_path,
        comparison_key='a' * 64,
        missing_checkpoint=True,
        stale_journal=False,
    )
    comparison = attempt.parent
    before = _file_snapshot(comparison)
    monkeypatch.setattr(
        phase0_harness,
        '_validate_prepared_evidence',
        lambda record: {
            'artifact_basename': record['artifact_basename'],
            'artifact_sha256': record['artifact_sha256'],
            'artifact_content': record['artifact_content'],
        },
    )
    with pytest.raises(HarnessFailure, match='protocol v1'):
        AppendOnlyAttemptLedger.create(
            tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
            _identity(),
            comparison_key='a' * 64,
        )
    assert _file_snapshot(comparison) == before


@pytest.mark.parametrize('invalid', (True, '1', 1.0, 0, -1))
def test_attempt_metadata_rejects_non_exact_positive_integer_number(invalid: object, tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    _rewrite_metadata_and_matching_empty_journal(ledger.attempt_directory, attempt_number=invalid)
    with pytest.raises(HarnessFailure) as caught:
        AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity())
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_attempt_metadata_number_must_match_directory_basename(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    renamed = ledger.attempt_directory.parent / 'attempt-0002'
    ledger.attempt_directory.rename(renamed)
    with pytest.raises(HarnessFailure) as caught:
        AppendOnlyAttemptLedger.load(renamed, _identity())
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_attempt_metadata_comparison_key_must_match_parent_directory(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    _rewrite_metadata_and_matching_empty_journal(ledger.attempt_directory, comparison_key='b' * 64)
    with pytest.raises(HarnessFailure) as caught:
        AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity())
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_current_runner_refuses_v1_attempt_as_resume(tmp_path: Path) -> None:
    attempt = _write_synthetic_v1_terminal(tmp_path, comparison_key='a' * 64)
    load_current = getattr(phase0_harness, '_load_current_protocol_ledger', None)
    assert callable(load_current)
    with pytest.raises(HarnessFailure, match='protocol'):
        load_current(attempt, _identity())


def test_second_round_one_to_five_attempt_is_rejected(tmp_path: Path) -> None:
    with pytest.raises(ValueError):
        MetricGroupRequest(
            _request(tmp_path),
            'b' * 32,
            'wall',
            build_round_plan(global_round_start=1, round_count=5),
            tmp_path / 'attempt',
            first_group_sha256='a' * 64,
        )


def test_existing_round_record_cannot_be_overwritten(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.record_sample('wall', 1, 'reference', {'value': 1})
    with pytest.raises(HarnessFailure, match='overwrite'):
        ledger.record_sample('wall', 1, 'reference', {'value': 2})


def test_expanded_group_requires_original_first_group_sha(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    first = ledger.commit_first_group({'wall': 'one', 'pws': 'two'})
    with pytest.raises(HarnessFailure, match='first group'):
        ledger.commit_expanded_group({'wall': 'three'}, first_group_sha256='0' * 64)
    ledger.commit_expanded_group({'wall': 'three', 'pws': 'four'}, first_group_sha256=first)


def test_pws_only_resample_is_rejected(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    with pytest.raises(HarnessFailure, match='wall and pws'):
        ledger.commit_first_group({'pws': 'only'})


@pytest.mark.parametrize(('wall_near_limit', 'pws_near_limit'), ((True, False), (False, True), (True, True)))
def test_paired_batch_expands_wall_and_pws_together(wall_near_limit: bool, pws_near_limit: bool) -> None:
    wall_plans, pws_plans = phase0_harness._mandatory_paired_expansion_plans(
        wall_requires_expansion=wall_near_limit,
        pws_requires_expansion=pws_near_limit,
    )
    assert [plan.global_round for plan in wall_plans] == [6, 7, 8, 9, 10]
    assert wall_plans == pws_plans


def test_phase0b_v3_uses_wall_1_02(tmp_path: Path) -> None:
    wall = _metric_group('wall', reference='1', candidate='1.021')
    pws = _metric_group('pws', reference='1', candidate='1')

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._evaluate_closed_profile(_request(tmp_path), wall, pws, _phase0a_baseline())

    assert caught.value.verdict is HarnessVerdict.CANDIDATE_FAILED
    assert 'wall_ratio' in str(caught.value)


def test_phase0b_v3_has_no_direct_pws_gate(tmp_path: Path) -> None:
    wall = _metric_group('wall', reference='1', candidate='0.90')
    pws = _metric_group('pws', reference='1', candidate='9')

    phase0_harness._evaluate_closed_profile(_request(tmp_path), wall, pws, _phase0a_baseline())


@pytest.mark.parametrize('first_group_ratio', ('0.999', '1.021'))
def test_phase0b_v3_near_boundary_expands_wall_and_pws_from_round_six(
    first_group_ratio: str,
    tmp_path: Path,
) -> None:
    request = _request(tmp_path)
    wall = _metric_group('wall', reference='1', candidate=first_group_ratio)
    pws = _metric_group('pws', reference='1', candidate='1')

    assert phase0_harness._paired_groups_require_expansion(request, wall, pws)
    wall_plans, pws_plans = phase0_harness._mandatory_paired_expansion_plans(
        wall_requires_expansion=True,
        pws_requires_expansion=False,
    )
    assert [plan.global_round for plan in wall_plans] == [6, 7, 8, 9, 10]
    assert wall_plans == pws_plans


def test_phase0b_v3_combined_closed_failure_precedes_direction_conflict(tmp_path: Path) -> None:
    first_wall = _metric_group('wall', reference='1', candidate='0.99')
    second_wall = _metric_group('wall', start=6, reference='1', candidate='1.07')
    wall = phase0_harness.merge_metric_groups(first_wall, second_wall)
    pws = phase0_harness.merge_metric_groups(
        _metric_group('pws', reference='1', candidate='1'),
        _metric_group('pws', start=6, reference='1', candidate='1'),
    )
    diagnostic = phase0_harness.build_direction_diagnostic(
        first_wall,
        second_wall,
        limits=phase0_harness.COMPARISON_LIMITS[ComparisonProfile.PHASE0B_VS_PHASE0A]['gb'],
    )
    assert diagnostic.directions_conflict

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._evaluate_closed_profile(_request(tmp_path), wall, pws, _phase0a_baseline())

    assert caught.value.verdict is HarnessVerdict.CANDIDATE_FAILED


@pytest.mark.parametrize(
    ('profile', 'pipeline', 'stage', 'ratio'),
    (
        (ComparisonProfile.PHASE1_VS_PHASE0B, 'sk', 'writer_populate', Decimal('0.90')),
        (ComparisonProfile.PHASE2_B_VS_A, 'sk', 'xlsx_save', Decimal('0.85')),
        (ComparisonProfile.PHASE4_VS_PHASE3, 'gb', 'ingest', Decimal('1.05')),
    ),
)
def test_expansion_covers_each_stage_ratio_gate(
    profile: ComparisonProfile,
    pipeline: str,
    stage: str,
    ratio: Decimal,
    tmp_path: Path,
) -> None:
    request = replace(_request(tmp_path), comparison_profile=profile, pipeline=pipeline)
    wall = _with_candidate_stage_ratios(_metric_group('wall'), tuple({stage: ratio} for _ in range(5)))

    assert phase0_harness._paired_groups_require_expansion(request, wall, _metric_group('pws')) is True


def test_expansion_covers_ingest_or_pws_and_writer_export_minimum_wins(tmp_path: Path) -> None:
    ingest_request = replace(
        _request(tmp_path),
        comparison_profile=ComparisonProfile.PHASE4_VS_PHASE3,
        pipeline='sk',
    )
    ingest_wall = _with_candidate_stage_ratios(
        _metric_group('wall'),
        tuple({'ingest': Decimal('0.90')} for _ in range(5)),
    )
    assert phase0_harness._paired_groups_require_expansion(ingest_request, ingest_wall, _metric_group('pws')) is True

    wins_request = replace(
        _request(tmp_path),
        comparison_profile=ComparisonProfile.PHASE2_C_VS_A,
        pipeline='sk',
    )
    win_ratios = tuple(
        {'writer_populate': Decimal('0.50'), 'export': Decimal('1.50')}
        if index < 4
        else {'writer_populate': Decimal('1.50'), 'export': Decimal('1.50')}
        for index in range(5)
    )
    wins_wall = _with_candidate_stage_ratios(_metric_group('wall'), win_ratios)
    assert phase0_harness._paired_groups_require_expansion(wins_request, wins_wall, _metric_group('pws')) is True


@pytest.mark.parametrize(('wall_ratio', 'pws_ratio'), (('1.02', '1.20'), ('0.80', '0.98')))
def test_phase2_b_vs_c_expands_when_either_tie_break_ratio_is_within_three_percent_of_one(
    wall_ratio: str, pws_ratio: str, tmp_path: Path
) -> None:
    request = replace(
        _request(tmp_path),
        comparison_profile=ComparisonProfile.PHASE2_B_VS_C,
        pipeline='sk',
    )
    wall = _metric_group('wall', reference='1', candidate=wall_ratio)
    pws = _metric_group('pws', reference='1', candidate=pws_ratio)

    assert phase0_harness._paired_groups_require_expansion(request, wall, pws) is True


def test_writer_export_minimum_wins_requires_four_wins_in_each_five_round_window(tmp_path: Path) -> None:
    request = replace(
        _request(tmp_path),
        comparison_profile=ComparisonProfile.PHASE2_C_VS_A,
        pipeline='sk',
    )
    first_ratios = tuple(
        {'writer_populate': Decimal('0.50'), 'export': Decimal('1.50')}
        if index < 4
        else {'writer_populate': Decimal('1.50'), 'export': Decimal('1.50')}
        for index in range(5)
    )
    second_ratios = tuple({'writer_populate': Decimal('1.50'), 'export': Decimal('1.50')} for _ in range(5))
    first_wall = _with_candidate_stage_ratios(_metric_group('wall'), first_ratios)
    second_wall = _with_candidate_stage_ratios(_metric_group('wall', start=6), second_ratios)
    wall = phase0_harness.merge_metric_groups(first_wall, second_wall)
    pws = phase0_harness.merge_metric_groups(
        _metric_group('pws', reference='100', candidate='100'),
        _metric_group('pws', start=6, reference='100', candidate='100'),
    )
    baseline = phase0_harness._ApprovedPhase0ABaseline('6' * 64, 8, (Decimal('1'),) * 5, (Decimal('100'),) * 5)

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._evaluate_closed_profile(request, wall, pws, baseline)

    assert caught.value.verdict is HarnessVerdict.CANDIDATE_FAILED


@pytest.mark.parametrize(
    ('wall_sizes', 'pws_sizes', 'expected_verdict'),
    (
        ((100,) * 5, (101,) * 5, None),
        ((110,) * 5, (110,) * 5, None),
        ((111,) * 5, (111,) * 5, HarnessVerdict.ENVIRONMENT_DRIFT),
    ),
    ids=('volatile-within-limit', 'exactly-ten-percent', 'over-ten-percent'),
)
def test_paired_phase0a_output_drift_uses_rounded_median(
    wall_sizes: tuple[int, ...],
    pws_sizes: tuple[int, ...],
    expected_verdict: HarnessVerdict | None,
) -> None:
    wall = _with_output_sizes(
        _metric_group('wall'),
        reference_sizes=wall_sizes,
        candidate_sizes=wall_sizes,
    )
    pws = _with_output_sizes(
        _metric_group('pws', reference='100', candidate='100'),
        reference_sizes=pws_sizes,
        candidate_sizes=pws_sizes,
    )
    if expected_verdict is None:
        phase0_harness._validate_phase0a_drift(wall, pws, _phase0a_baseline(), _identity())
        return

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._validate_phase0a_drift(wall, pws, _phase0a_baseline(), _identity())

    assert caught.value.verdict is expected_verdict


@pytest.mark.parametrize(
    ('candidate_sizes', 'expected_verdict'),
    (
        ((100, 100, 100, 111, 111), None),
        ((100, 100, 111, 111, 111), HarnessVerdict.CANDIDATE_FAILED),
    ),
    ids=('raw-maximum-over-limit', 'median-over-limit'),
)
def test_closed_profile_candidate_output_gate_uses_median(
    candidate_sizes: tuple[int, ...],
    expected_verdict: HarnessVerdict | None,
    tmp_path: Path,
) -> None:
    wall = _with_output_sizes(
        _metric_group('wall'),
        reference_sizes=(100,) * 5,
        candidate_sizes=candidate_sizes,
    )
    pws = _with_output_sizes(
        _metric_group('pws', reference='100', candidate='100'),
        reference_sizes=(100,) * 5,
        candidate_sizes=candidate_sizes,
    )
    if expected_verdict is None:
        phase0_harness._evaluate_closed_profile(_request(tmp_path), wall, pws, _phase0a_baseline())
        return

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._evaluate_closed_profile(_request(tmp_path), wall, pws, _phase0a_baseline())

    assert caught.value.verdict is expected_verdict


def test_paired_evidence_serializes_median_rounded_output_bytes(tmp_path: Path) -> None:
    request = _request(tmp_path)
    identity = _full_identity()
    wall = _with_output_sizes(
        _metric_group('wall'),
        reference_sizes=(100,) * 5,
        candidate_sizes=(109,) * 5,
    )
    pws = _with_output_sizes(
        _metric_group('pws', reference='100', candidate='100'),
        reference_sizes=(101,) * 5,
        candidate_sizes=(110,) * 5,
    )
    attempt = BatchAttempt(
        protocol_version=2,
        comparison_key=_comparison_key(request, identity),
        batch_id=wall.batch_id,
        attempt_number=1,
        state=AttemptState.CLEANUP_COMPLETE,
        previous_attempt_head_sha256=None,
        first_group_sha256='a' * 64,
        expanded_group_sha256=None,
        ledger_head_sha256='b' * 64,
        attempt_directory=tmp_path,
    )

    evidence = phase0_harness._build_paired_evidence(request, wall, pws, attempt, identity, ())

    assert {item.role: item.value for item in evidence.output_bytes} == {'reference': 101, 'candidate': 110}


def test_expanded_paired_evidence_contains_exact_wall_then_pws_diagnostics(tmp_path: Path) -> None:
    request = _request(tmp_path)
    identity = _full_identity()
    wall_first = _metric_group('wall', reference='1', candidate='1.03')
    wall_second = _metric_group('wall', start=6, reference='1', candidate='0.95')
    pws_first = _metric_group('pws', reference='1', candidate='0.99')
    pws_second = _metric_group('pws', start=6, reference='1', candidate='1.01')
    wall = phase0_harness.merge_metric_groups(wall_first, wall_second)
    pws = phase0_harness.merge_metric_groups(pws_first, pws_second)
    attempt = BatchAttempt(
        protocol_version=2,
        comparison_key=_comparison_key(request, identity),
        batch_id=wall.batch_id,
        attempt_number=1,
        state=AttemptState.CLEANUP_COMPLETE,
        previous_attempt_head_sha256=None,
        first_group_sha256='a' * 64,
        expanded_group_sha256='b' * 64,
        ledger_head_sha256='c' * 64,
        attempt_directory=tmp_path,
    )

    limits = phase0_harness.COMPARISON_LIMITS[request.comparison_profile][request.pipeline]
    diagnostics = (
        phase0_harness.build_direction_diagnostic(wall_first, wall_second, limits=limits),
        phase0_harness.build_direction_diagnostic(pws_first, pws_second, limits=limits),
    )
    value = phase0_harness._build_paired_evidence(request, wall, pws, attempt, identity, diagnostics)
    artifact = EvidenceSanitizer.closed_policy().rebuild_audit_benchmark_manifest(value)

    assert tuple(item.metric for item in value.direction_diagnostics) == ('wall', 'pws')
    assert value.direction_diagnostics[0].first_group_ratio == Decimal('1.03')
    assert value.direction_diagnostics[0].second_group_ratio == Decimal('0.95')
    assert artifact.file_name.startswith('benchmark-v2-')


def test_v2_evidence_rejects_attempt_comparison_key_mismatch(tmp_path: Path) -> None:
    request = _request(tmp_path)
    identity = _full_identity()
    wall = _metric_group('wall')
    pws = _metric_group('pws')
    attempt = BatchAttempt(
        protocol_version=2,
        comparison_key='f' * 64,
        batch_id=wall.batch_id,
        attempt_number=1,
        state=AttemptState.CLEANUP_COMPLETE,
        previous_attempt_head_sha256=None,
        first_group_sha256='a' * 64,
        expanded_group_sha256=None,
        ledger_head_sha256='b' * 64,
        attempt_directory=tmp_path,
    )

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._build_paired_evidence(request, wall, pws, attempt, identity, ())

    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_loader_repairs_only_tail_committed_record_missing_checkpoint(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _request, ledger, artifact = _prepare_formal_evidence_recovery(monkeypatch, tmp_path)
    original_write = phase0_harness._write_create_new
    failed = False

    def fail_checkpoint(path: Path, content: bytes, *, allowed_root: Path) -> None:
        nonlocal failed
        if path.parent.name == 'checkpoints' and not failed:
            failed = True
            raise OSError('checkpoint interrupted')
        original_write(path, content, allowed_root=allowed_root)

    monkeypatch.setattr(phase0_harness, '_write_create_new', fail_checkpoint)
    with pytest.raises(OSError, match='checkpoint interrupted'):
        ledger.mark_evidence_committed(artifact_sha256=hashlib.sha256(artifact.content.encode()).hexdigest())

    recovered = AppendOnlyAttemptLedger.open_current_protocol_for_resume(ledger.attempt_directory, ledger.identity)
    assert recovered.state is AttemptState.EVIDENCE_COMMITTED
    assert len(tuple((ledger.attempt_directory / 'records').glob('*.json'))) == len(
        tuple((ledger.attempt_directory / 'checkpoints').glob('*.json'))
    )


def test_loader_repairs_only_missing_journal_anchor_after_committed_checkpoint(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _request, ledger, artifact = _prepare_formal_evidence_recovery(monkeypatch, tmp_path)
    failed = False
    original_anchor = ledger._append_journal_anchor

    def fail_anchor() -> str:
        nonlocal failed
        if not failed:
            failed = True
            raise OSError('journal interrupted')
        return original_anchor()

    monkeypatch.setattr(ledger, '_append_journal_anchor', fail_anchor)
    with pytest.raises(OSError, match='journal interrupted'):
        ledger.mark_evidence_committed(artifact_sha256=hashlib.sha256(artifact.content.encode()).hexdigest())

    recovered = AppendOnlyAttemptLedger.open_current_protocol_for_resume(ledger.attempt_directory, ledger.identity)
    assert recovered.state is AttemptState.EVIDENCE_COMMITTED


def _approved_phase0a_payload(*, wall_value: str = '1', pws_value: str = '100') -> dict[str, object]:
    calibration = {
        'wall': {'batch_id_sha256': 'a' * 64, 'values': [wall_value] * 5, 'local_log_sha256': ['7' * 64] * 5},
        'pws': {'batch_id_sha256': 'b' * 64, 'values': [pws_value] * 5, 'local_log_sha256': ['7' * 64] * 5},
    }
    pipeline = {
        'input_alias': '$GB_INPUT',
        'input_sha256': '3' * 64,
        'output_size_bytes': 8,
        'sheet_dimensions': ['1x1', '1x1', '1x1'],
        'runtime': {'sheet_count': 3, 'error_log_count': 0, 'run_counts': {}, 'stage_timings': {}},
        'calibration': calibration,
    }
    return {
        'schema_version': 1,
        'approval_state': 'APPROVED',
        'reference_exe_sha256': '1' * 64,
        'fork_revision': 'a' * 40,
        'git_head': '4' * 40,
        'machine': {
            'windows_build_sha256': '8' * 64,
            'architecture': 'x86_64',
            'cpu_model_sha256': '9' * 64,
            'logical_cpu_count': 4,
            'physical_memory_bytes': 8,
            'system_drive_media_type': 'SSD',
            'system_drive_size_bytes': 16,
            'fingerprint_sha256': '6' * 64,
        },
        'pipelines': {'gb': pipeline, 'sk': {**pipeline, 'input_alias': '$SK_INPUT'}},
    }


def _install_formal_paired(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    *,
    wall_reference: str = '1',
    wall_candidate: str = '1',
) -> tuple[PairedBenchmarkRequest, list[dict[str, object]]]:
    request = _request(tmp_path)
    identity = _full_identity()
    comparison_key = phase0_harness.derive_v2_comparison_key(
        pipeline=request.pipeline,
        comparison_profile=request.comparison_profile,
        reference_label=request.reference_label,
        candidate_label=request.candidate_label,
        input_sha256=identity.input_sha256,
        reference_sha256=identity.reference_sha256,
        candidate_sha256=identity.candidate_sha256,
    )
    evidence_name = expected_benchmark_artifact_name(protocol_version=2, comparison_key=comparison_key)
    request = replace(request, evidence_path=request.evidence_path.parent / evidence_name)
    request.phase0a_manifest.write_text(json.dumps(_approved_phase0a_payload()), encoding='utf-8')
    monkeypatch.setattr(phase0_harness, '_capture_identity', lambda benchmark: identity)
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: '')

    def wall(group_request: MetricGroupRequest) -> MetricGroup:
        group = _metric_group(
            'wall',
            start=group_request.plans[0].global_round,
            reference=wall_reference,
            candidate=wall_candidate,
        )
        return replace(group, batch_id=group_request.batch_id)

    def pws(group_request: MetricGroupRequest) -> MetricGroup:
        group = _metric_group('pws', start=group_request.plans[0].global_round, reference='100', candidate='100')
        return replace(group, batch_id=group_request.batch_id)

    publications: list[dict[str, object]] = []
    monkeypatch.setattr(phase0_harness, 'run_normal_wall_group', wall)
    monkeypatch.setattr(phase0_harness, 'run_pws_group', pws)
    monkeypatch.setattr(
        EvidenceSanitizer,
        'write_batch',
        lambda self, **kwargs: publications.append(kwargs),
    )
    return request, publications


def _direction_case_group(
    metric: str,
    *,
    start: int,
    candidate: str,
    pipeline: str,
    stage_ratios: dict[str, Decimal] | None = None,
) -> MetricGroup:
    group = _metric_group(metric, start=start, reference='1', candidate=candidate)
    if stage_ratios:
        group = _with_candidate_stage_ratios(group, tuple(dict(stage_ratios) for _ in group.rounds))
    rounds = tuple(
        replace(
            paired,
            reference=replace(
                paired.reference,
                normal_run=replace(
                    paired.reference.normal_run,
                    runtime=replace(paired.reference.normal_run.runtime, pipeline=pipeline),
                ),
            ),
            candidate=replace(
                paired.candidate,
                normal_run=replace(
                    paired.candidate.normal_run,
                    runtime=replace(paired.candidate.normal_run.runtime, pipeline=pipeline),
                ),
            ),
        )
        for paired in group.rounds
    )
    return replace(group, pipeline=pipeline, rounds=rounds)


def _install_v2_direction_case(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    *,
    profile: ComparisonProfile = ComparisonProfile.PHASE0B_VS_PHASE0A,
    pipeline: str = 'gb',
    wall: tuple[str, str],
    pws: tuple[str, str],
    stage_ratios: dict[str, Decimal] | None = None,
    force_expansion: bool = False,
) -> tuple[PairedBenchmarkRequest, BenchmarkIdentity, list[tuple[str, int]]]:
    labels = {
        ComparisonProfile.PHASE0B_VS_PHASE0A: (ClosedBinaryLabel.PHASE0A, ClosedBinaryLabel.PHASE0B),
        ComparisonProfile.PHASE1_VS_PHASE0A: (ClosedBinaryLabel.PHASE0A, ClosedBinaryLabel.PHASE1),
        ComparisonProfile.PHASE4_VS_PHASE3: (ClosedBinaryLabel.PHASE3, ClosedBinaryLabel.PHASE4),
    }
    reference_label, candidate_label = labels[profile]
    request = replace(
        _request(tmp_path),
        comparison_profile=profile,
        pipeline=pipeline,
        reference_label=reference_label,
        candidate_label=candidate_label,
    )
    input_path = tmp_path / 'confidential-source.xlsx'
    input_path.write_bytes(b'input')
    request = replace(request, input_path=input_path)
    identity = _full_identity()
    comparison_key = _comparison_key(request, identity)
    request = replace(
        request,
        evidence_path=request.evidence_path.parent
        / expected_benchmark_artifact_name(protocol_version=2, comparison_key=comparison_key),
    )
    request.phase0a_manifest.write_text(
        json.dumps(_approved_phase0a_payload(wall_value='1', pws_value='1')),
        encoding='utf-8',
    )
    monkeypatch.setattr(phase0_harness, '_capture_identity', lambda benchmark: identity)
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: '')
    monkeypatch.setattr(EvidenceSanitizer, 'scan_staged', lambda self, **kwargs: None)
    if force_expansion:
        monkeypatch.setattr(phase0_harness, '_paired_groups_require_expansion', lambda *_args: True)

    calls: list[tuple[str, int]] = []

    def group(group_request: MetricGroupRequest) -> MetricGroup:
        calls.append((group_request.metric, group_request.plans[0].global_round))
        values = wall if group_request.metric == 'wall' else pws
        index = 0 if group_request.plans[0].global_round == 1 else 1
        value = _direction_case_group(
            group_request.metric,
            start=group_request.plans[0].global_round,
            candidate=values[index],
            pipeline=pipeline,
            stage_ratios=stage_ratios if group_request.metric == 'wall' else None,
        )
        return replace(value, batch_id=group_request.batch_id)

    monkeypatch.setattr(phase0_harness, 'run_normal_wall_group', group)
    monkeypatch.setattr(phase0_harness, 'run_pws_group', group)
    return request, identity, calls


def _assert_v2_direction_verdict(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    *,
    expected: HarnessVerdict,
    **case: object,
) -> PairedBenchmarkResult | HarnessFailure:
    request, _identity, _calls = _install_v2_direction_case(monkeypatch, tmp_path, **case)  # type: ignore[arg-type]
    plans = build_round_plan(global_round_start=1, round_count=5)

    def group(metric: str, selected_plans: tuple[phase0_harness.RoundPlan, ...]) -> MetricGroup:
        runner = phase0_harness.run_normal_wall_group if metric == 'wall' else phase0_harness.run_pws_group
        return runner(
            MetricGroupRequest(
                request,
                '9' * 64,
                metric,  # type: ignore[arg-type]
                selected_plans,
                tmp_path / 'unused-attempt',
                first_group_sha256='a' * 64 if selected_plans[0].global_round == 6 else None,
            )
        )

    def evaluate() -> PairedBenchmarkResult:
        wall_first = group('wall', plans)
        pws_first = group('pws', plans)
        wall, pws = wall_first, pws_first
        diagnostics = ()
        if phase0_harness._paired_groups_require_expansion(request, wall_first, pws_first):
            expanded = build_round_plan(global_round_start=6, round_count=5)
            wall_second = group('wall', expanded)
            pws_second = group('pws', expanded)
            limits = phase0_harness.COMPARISON_LIMITS[request.comparison_profile][request.pipeline]
            diagnostics = (
                phase0_harness.build_direction_diagnostic(wall_first, wall_second, limits=limits),
                phase0_harness.build_direction_diagnostic(pws_first, pws_second, limits=limits),
            )
            wall = phase0_harness.merge_metric_groups(wall_first, wall_second)
            pws = phase0_harness.merge_metric_groups(pws_first, pws_second)
        phase0_harness._evaluate_closed_profile(request, wall, pws, _phase0a_baseline(8))
        if any(
            item.directions_conflict and item.direct_gate != 'none' and item.near_boundary is True
            for item in diagnostics
        ):
            raise HarnessFailure(HarnessVerdict.INCONCLUSIVE, 'active direct metric direction conflict')
        return PairedBenchmarkResult(
            wall,
            pws,
            BatchAttempt(
                3,
                'c' * 64,
                wall.batch_id,
                1,
                AttemptState.FIRST_GROUP_COMPLETE,
                None,
                'a' * 64,
                None,
                'b' * 64,
                tmp_path,
            ),
            HarnessVerdict.VALIDATED,
        )

    if expected is HarnessVerdict.VALIDATED:
        result = evaluate()
        assert result.verdict is HarnessVerdict.VALIDATED
        return result
    with pytest.raises(HarnessFailure) as caught:
        evaluate()
    assert caught.value.verdict is expected
    return caught.value


def test_active_wall_conflict_near_boundary_is_inconclusive_after_all_gates_pass(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _assert_v2_direction_verdict(
        monkeypatch,
        tmp_path,
        expected=HarnessVerdict.INCONCLUSIVE,
        wall=('1.03', '0.99'),
        pws=('1', '1'),
    )


def test_active_wall_conflict_decisive_pass_is_validated(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _assert_v2_direction_verdict(
        monkeypatch,
        tmp_path,
        expected=HarnessVerdict.VALIDATED,
        wall=('1.04', '0.90'),
        pws=('1', '1'),
    )


def test_active_wall_conflict_decisive_fail_is_candidate_failed(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _assert_v2_direction_verdict(
        monkeypatch,
        tmp_path,
        expected=HarnessVerdict.CANDIDATE_FAILED,
        wall=('0.99', '1.07'),
        pws=('1', '1'),
    )


def test_inactive_pws_conflict_is_diagnostic_only(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _assert_v2_direction_verdict(
        monkeypatch,
        tmp_path,
        expected=HarnessVerdict.VALIDATED,
        wall=('0.95', '0.95'),
        pws=('0.99', '1.01'),
        force_expansion=True,
    )


def test_near_wall_conflict_cannot_hide_direct_pws_failure(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _assert_v2_direction_verdict(
        monkeypatch,
        tmp_path,
        expected=HarnessVerdict.CANDIDATE_FAILED,
        profile=ComparisonProfile.PHASE1_VS_PHASE0A,
        wall=('1.12', '0.98'),
        pws=('1.06', '1.06'),
    )


@pytest.mark.parametrize(
    ('pipeline', 'wall', 'pws', 'stage_ratios'),
    (
        ('sk', ('1.02', '0.98'), ('0.95', '0.95'), {'ingest': Decimal('0.95')}),
        ('gb', ('1', '1'), ('1.12', '0.98'), {'ingest': Decimal('1.06')}),
    ),
    ids=('composite-failure', 'stage-only-failure'),
)
def test_near_wall_conflict_cannot_hide_composite_or_stage_failure(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    pipeline: str,
    wall: tuple[str, str],
    pws: tuple[str, str],
    stage_ratios: dict[str, Decimal],
) -> None:
    _assert_v2_direction_verdict(
        monkeypatch,
        tmp_path,
        expected=HarnessVerdict.CANDIDATE_FAILED,
        profile=ComparisonProfile.PHASE4_VS_PHASE3,
        pipeline=pipeline,
        wall=wall,
        pws=pws,
        stage_ratios=stage_ratios,
    )


def test_equal_to_one_group_direction_is_not_inconclusive(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _assert_v2_direction_verdict(
        monkeypatch,
        tmp_path,
        expected=HarnessVerdict.VALIDATED,
        wall=('1', '1.02'),
        pws=('1', '1'),
    )


def _prepare_formal_evidence_recovery(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> tuple[PairedBenchmarkRequest, AppendOnlyAttemptLedger, object]:
    request, _ = _install_formal_paired(monkeypatch, tmp_path)
    sanitized_input = tmp_path / 'confidential-gb-source.xlsx'
    sanitized_input.write_bytes(b'input')
    request = replace(request, input_path=sanitized_input)
    monkeypatch.setattr(EvidenceSanitizer, 'scan_staged', lambda self, **kwargs: None)
    identity = BenchmarkIdentity('3' * 64, '1' * 64, '2' * 64, '4' * 40, '5' * 64, '6' * 64)
    comparison_key = phase0_harness.derive_v2_comparison_key(
        pipeline=request.pipeline,
        comparison_profile=request.comparison_profile,
        reference_label=request.reference_label,
        candidate_label=request.candidate_label,
        input_sha256=identity.input_sha256,
        reference_sha256=identity.reference_sha256,
        candidate_sha256=identity.candidate_sha256,
    )
    ledger = AppendOnlyAttemptLedger.create(request.attempt_ledger_root, identity, comparison_key=comparison_key)
    ledger.commit_first_group({'wall': 'wall', 'pws': 'pws'})
    ledger.mark_cleanup_complete()
    batch_id = derive_batch_id(request, identity)
    wall = replace(_metric_group('wall'), batch_id=batch_id)
    pws = replace(_metric_group('pws', reference='100', candidate='100'), batch_id=batch_id)
    attempt = phase0_harness._batch_attempt(ledger, batch_id, AttemptState.CLEANUP_COMPLETE)
    evidence = phase0_harness._build_paired_evidence(request, wall, pws, attempt, identity, ())
    artifact = EvidenceSanitizer.closed_policy().rebuild_audit_benchmark_manifest(evidence)
    ledger.prepare_evidence(
        artifact_basename=artifact.file_name,
        artifact_sha256=hashlib.sha256(artifact.content.encode('utf-8')).hexdigest(),
        artifact_content=artifact.content,
    )
    return request, ledger, artifact


def test_prepared_evidence_recovery_rejects_v1_payload(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _request, _ledger, artifact = _prepare_formal_evidence_recovery(monkeypatch, tmp_path)
    assert isinstance(artifact.source, BenchmarkManifestEvidence)
    legacy_source = replace(
        artifact.source,
        schema_version=1,
        protocol_version=None,
        direction_diagnostics=(),
    )
    legacy_artifact = EvidenceSanitizer.closed_policy().rebuild_benchmark_manifest(legacy_source)
    payload = {
        'artifact_basename': legacy_artifact.file_name,
        'artifact_sha256': hashlib.sha256(legacy_artifact.content.encode('utf-8')).hexdigest(),
        'artifact_content': legacy_artifact.content,
    }

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness._rebuild_prepared_artifact(payload)

    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_paired_batch_rejects_evidence_basename_that_differs_from_typed_artifact(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    request, publications = _install_formal_paired(monkeypatch, tmp_path)
    request = replace(request, evidence_path=request.evidence_path.with_name('caller-chosen.json'))

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.run_paired_normal_batch(request)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert publications == []


def test_paired_batch_rejects_labels_that_do_not_match_closed_profile(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    request, publications = _install_formal_paired(monkeypatch, tmp_path)
    request = replace(request, candidate_label=ClosedBinaryLabel.PHASE1)

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.run_paired_normal_batch(request)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert publications == []


@pytest.mark.parametrize('mutation', ('extra', 'duplicate'))
def test_phase0a_loader_rejects_extra_or_duplicate_keys(
    mutation: str, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    request, _ = _install_formal_paired(monkeypatch, tmp_path)
    raw = json.dumps(_approved_phase0a_payload(), separators=(',', ':'))
    if mutation == 'extra':
        raw = raw[:-1] + ',"unexpected":1}'
    else:
        raw = raw[:-1] + ',"schema_version":1}'
    request.phase0a_manifest.write_text(raw, encoding='utf-8')

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.run_paired_normal_batch(request)
    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE


def test_wall_and_pws_groups_share_attempt_batch_and_n(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _install_runner(monkeypatch, tmp_path)
    wall = run_normal_wall_group(_group_request(tmp_path))
    pws = replace(wall, metric='pws')
    assert_same_benchmark_batch(wall, pws)


def test_interrupted_attempt_resumes_only_missing_samples(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    first_commands = _install_runner(monkeypatch, tmp_path, interrupt_role='candidate')
    with pytest.raises(KeyboardInterrupt):
        run_normal_wall_group(_group_request(tmp_path))
    assert [item[0] for item in first_commands] == ['reference', 'candidate']

    resumed_commands = _install_runner(monkeypatch, tmp_path)
    run_normal_wall_group(_group_request(tmp_path))
    assert resumed_commands[0][0] == 'candidate'
    assert len(resumed_commands) == 9


def test_failed_candidate_sha_cannot_be_retried(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.finish(HarnessVerdict.CANDIDATE_FAILED)
    with pytest.raises(HarnessFailure, match='candidate SHA'):
        AppendOnlyAttemptLedger.create(
            tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
            _identity(),
            comparison_key='a' * 64,
        )


def test_environment_recovery_attempt_links_previous_ledger_head(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.finish(HarnessVerdict.ENVIRONMENT_DRIFT)
    recovered = AppendOnlyAttemptLedger.create(
        tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
        _identity(),
        comparison_key='a' * 64,
    )
    assert recovered.previous_attempt_head_sha256 == ledger.head_sha256


def test_terminal_verdict_is_unique_and_cannot_change_retry_eligibility(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.finish(HarnessVerdict.CANDIDATE_FAILED, raw_log_sha256='d' * 64)
    assert ledger.state is AttemptState.FAILED
    assert ledger.terminal_verdict is HarnessVerdict.CANDIDATE_FAILED
    with pytest.raises(HarnessFailure, match='terminal'):
        ledger.finish(HarnessVerdict.ENVIRONMENT_DRIFT)
    with pytest.raises(HarnessFailure, match='terminal'):
        ledger.record_sample('wall', 2, 'reference', {'value': 2})


def test_candidate_failure_is_terminal_before_wall_runner_raises(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _group_request(tmp_path)
    _install_runner(monkeypatch, tmp_path, fail_role='candidate')
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(request)
    assert caught.value.verdict is HarnessVerdict.CANDIDATE_FAILED
    loaded = AppendOnlyAttemptLedger.load(request.attempt_directory, _identity())
    assert loaded.terminal_verdict is HarnessVerdict.CANDIDATE_FAILED
    assert loaded.terminal_raw_log_sha256 == 'f' * 64


def test_correctness_failure_is_terminal_with_raw_log_sha(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _group_request(tmp_path)
    _install_runner(monkeypatch, tmp_path, validation_fail_role='candidate')
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(request)
    assert caught.value.verdict is HarnessVerdict.CORRECTNESS_FAILED
    loaded = AppendOnlyAttemptLedger.load(request.attempt_directory, _identity())
    assert loaded.terminal_verdict is HarnessVerdict.CORRECTNESS_FAILED
    assert loaded.terminal_raw_log_sha256 == 'e' * 64


def test_wall_dimension_mismatch_is_terminal_and_cleans_captured_artifacts(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _group_request(tmp_path)
    _install_runner(monkeypatch, tmp_path)
    raw_log_sha256 = 'd' * 64
    raw_log = request.benchmark.local_root / 'raw-logs' / f'{raw_log_sha256}.json'

    def mismatched_capture(*args: object, **_kwargs: object) -> CapturedNormalRun:
        output = Path(args[3])
        _write_approved_test_workbook(output)
        raw_log.parent.mkdir(parents=True, exist_ok=True)
        raw_log.write_text('{}', encoding='utf-8')
        runtime = replace(_runtime(), sheet_dimensions=('A1:A1', 'A1:A1', 'A1:A1'))
        return CapturedNormalRun(
            NormalRunEvidence(Decimal('1'), None, runtime, '8' * 64),
            0,
            raw_log_sha256,
        )

    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', mismatched_capture)

    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(request)

    assert caught.value.verdict is HarnessVerdict.CORRECTNESS_FAILED
    assert caught.value.raw_log_sha256 == raw_log_sha256
    loaded = AppendOnlyAttemptLedger.load(request.attempt_directory, _identity())
    assert loaded.terminal_verdict is HarnessVerdict.CORRECTNESS_FAILED
    assert loaded.terminal_raw_log_sha256 == raw_log_sha256
    assert not raw_log.exists()
    assert not list((tmp_path / 'data').rglob('*.xlsx'))
    assert not request.benchmark.evidence_path.exists()


def test_capture_boundary_exception_maps_to_closed_terminal_verdict(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _group_request(tmp_path)
    _install_runner(monkeypatch, tmp_path, exception_role='candidate')
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(request)
    assert caught.value.verdict is HarnessVerdict.CORRECTNESS_FAILED
    assert isinstance(caught.value.__cause__, RuntimeError)
    assert (
        AppendOnlyAttemptLedger.load(request.attempt_directory, _identity()).terminal_verdict
        is HarnessVerdict.CORRECTNESS_FAILED
    )


def test_sealed_attempt_rejects_deleted_tail_sample(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.record_sample('wall', 1, 'candidate', {'value': 1})
    ledger.finish(HarnessVerdict.CANDIDATE_FAILED)
    sample_record = next((ledger.attempt_directory / 'records').glob('*-sample.json'))
    sample_record.unlink()
    with pytest.raises(HarnessFailure, match='checkpoint|sealed|record count'):
        AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity())


def test_comparison_journal_rejects_deleted_terminal(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.finish(HarnessVerdict.CANDIDATE_FAILED)
    (ledger.attempt_directory / 'terminal.json').unlink()
    with pytest.raises(HarnessFailure, match='journal'):
        AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity())


def test_comparison_journal_rejects_tail_record_and_checkpoint_rollback(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.record_sample('wall', 1, 'candidate', {'value': 1})
    next((ledger.attempt_directory / 'records').glob('*-sample.json')).unlink()
    next((ledger.attempt_directory / 'checkpoints').glob('*.json')).unlink()
    with pytest.raises(HarnessFailure, match='journal'):
        AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity())


def test_unsealed_attempt_does_not_turn_deleted_candidate_sample_into_missing(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.record_sample('wall', 1, 'candidate', {'value': 1})
    sample_record = next((ledger.attempt_directory / 'records').glob('*-sample.json'))
    sample_record.unlink()
    with pytest.raises(HarnessFailure, match='checkpoint'):
        AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity())


def test_transient_cleanup_failure_is_closed_and_outer_retry_removes_workbook(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _install_runner(monkeypatch, tmp_path)
    real_remove = phase0_harness._remove_workbook
    failures_left = 1

    def transient(path: Path) -> None:
        nonlocal failures_left
        if failures_left:
            failures_left -= 1
            raise PermissionError('transient lock')
        real_remove(path)

    monkeypatch.setattr(phase0_harness, '_remove_workbook', transient)
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(_group_request(tmp_path))
    assert caught.value.verdict is HarnessVerdict.ENVIRONMENT_DRIFT
    assert not list((tmp_path / 'data').rglob('*.xlsx'))


def test_persistent_cleanup_failure_preserves_primary_verdict(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    _install_runner(monkeypatch, tmp_path, fail_role='candidate')
    real_remove = phase0_harness._remove_workbook
    calls = 0

    def fail_after_reference(path: Path) -> None:
        nonlocal calls
        calls += 1
        if calls == 1:
            real_remove(path)
            return
        raise PermissionError('persistent lock')

    monkeypatch.setattr(phase0_harness, '_remove_workbook', fail_after_reference)
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(_group_request(tmp_path))
    assert caught.value.verdict is HarnessVerdict.CLEANUP_FAILED
    assert caught.value.primary_verdict is HarnessVerdict.CANDIDATE_FAILED


def test_cleanup_recovery_attempt_cleans_historical_planned_outputs_before_resume(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    ledger = _ledger(tmp_path)
    ledger.record_planned_output(
        'wall',
        1,
        'reference',
        {
            'pipeline': 'gb',
            'batch_id': 'b' * 32,
            'metric': 'wall',
            'binary_sha256': '1' * 64,
            'global_round': 1,
            'role': 'reference',
            'relative_path': 'gb/.perf-runs/' + '/'.join(('b' * 32, 'wall', '1' * 64, '1', 'reference.xlsx')),
        },
    )
    monkeypatch.setattr(phase0_harness, 'repo_root', lambda: tmp_path)
    historical = phase0_harness._planned_paths(ledger.all_planned_output_payloads())[0]
    historical.write_bytes(b'partial')
    ledger.finish(HarnessVerdict.CLEANUP_FAILED, primary_verdict=HarnessVerdict.CANDIDATE_FAILED)
    recovered = AppendOnlyAttemptLedger.create(
        tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
        _identity(),
        comparison_key='a' * 64,
    )
    assert recovered.all_planned_output_payloads()
    assert recovered.cleanup_only is True
    commands = _install_runner(monkeypatch, tmp_path)
    request = MetricGroupRequest(
        _request(tmp_path),
        'b' * 8,
        'wall',
        build_round_plan(global_round_start=1, round_count=5),
        recovered.attempt_directory,
    )
    with pytest.raises(HarnessFailure) as caught:
        run_normal_wall_group(request)
    assert caught.value.verdict is HarnessVerdict.CANDIDATE_FAILED
    assert commands == []
    assert not historical.exists()
    with pytest.raises(HarnessFailure, match='candidate SHA'):
        AppendOnlyAttemptLedger.create(
            tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
            _identity(),
            comparison_key='a' * 64,
        )


def test_cleanup_only_ledger_rejects_benchmark_records_but_can_finish(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.cleanup_only = True

    with pytest.raises(HarnessFailure, match='cleanup-only'):
        ledger.record_sample('wall', 1, 'reference', {'value': 1})

    ledger.finish(HarnessVerdict.ENVIRONMENT_DRIFT)
    assert ledger.terminal_verdict is HarnessVerdict.ENVIRONMENT_DRIFT


def test_success_ledger_rejects_benchmark_records_after_cleanup(tmp_path: Path) -> None:
    ledger = _ledger(tmp_path)
    ledger.commit_first_group({'wall': {}, 'pws': {}})
    ledger.mark_cleanup_complete()

    with pytest.raises(HarnessFailure, match='cleanup|success'):
        ledger.record_sample('wall', 1, 'reference', {'value': 1})

    with pytest.raises(HarnessFailure, match='evidence success'):
        ledger.mark_evidence_committed(artifact_sha256='8' * 64)
    assert AppendOnlyAttemptLedger.load(ledger.attempt_directory, _identity()).state is AttemptState.CLEANUP_COMPLETE


@pytest.mark.parametrize('field', ('local_root', 'attempt_ledger_root'))
def test_wall_runner_requires_repository_trusted_local_roots(
    field: str,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    ledger = _ledger(tmp_path)
    benchmark = replace(_request(tmp_path), **{field: tmp_path / 'docs' / 'performance'})
    request = MetricGroupRequest(
        benchmark,
        'b' * 8,
        'wall',
        build_round_plan(global_round_start=1, round_count=5),
        ledger.attempt_directory,
    )
    commands = _install_runner(monkeypatch, tmp_path)
    with pytest.raises(HarnessFailure, match='trusted'):
        run_normal_wall_group(request)
    assert commands == []


@pytest.mark.parametrize('comparison_key', ('../escape', 'a/b', 'C:/absolute', 'A' * 64, 'a' * 63))
def test_attempt_ledger_rejects_non_closed_comparison_key(comparison_key: str, tmp_path: Path) -> None:
    local_root = tmp_path / 'rust' / 'target' / 'perf-local'
    with pytest.raises(HarnessFailure, match='comparison_key'):
        AppendOnlyAttemptLedger.create(
            local_root / 'batches',
            _identity(),
            comparison_key=comparison_key,
        )


def test_attempt_ledger_rejects_root_outside_local_root(tmp_path: Path) -> None:
    with pytest.raises(HarnessFailure, match='local root'):
        AppendOnlyAttemptLedger.create(
            tmp_path / 'outside',
            _identity(),
            comparison_key='a' * 64,
        )


def test_attempt_ledger_rejects_raw_symlink_component(tmp_path: Path) -> None:
    local_root = tmp_path / 'rust' / 'target' / 'perf-local'
    local_root.mkdir(parents=True)
    outside = tmp_path / 'outside'
    outside.mkdir()
    linked = local_root / 'batches'
    try:
        linked.symlink_to(outside, target_is_directory=True)
    except OSError as exc:
        pytest.skip(f'symlink creation is unavailable: {exc}')
    with pytest.raises(HarnessFailure, match='reparse|symlink'):
        AppendOnlyAttemptLedger.create(
            linked,
            _identity(),
            comparison_key='a' * 64,
        )


def test_prior_evidence_claim_content_participates_in_identity(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _request(tmp_path)
    evidence = tmp_path / 'docs' / 'performance' / 'prior.json'
    evidence.parent.mkdir(parents=True, exist_ok=True)
    evidence.write_bytes(b'first')
    claim = _claim(tmp_path, evidence, phase0_harness._sha256(evidence))
    request = replace(request, prior_evidence_claims=(claim,))
    monkeypatch.setattr(phase0_harness, 'repo_root', lambda: tmp_path)
    monkeypatch.setattr(phase0_harness, '_run_git', lambda root, *args: 'head' if args[0] == 'rev-parse' else '')

    first = phase0_harness._capture_identity(request)
    evidence.write_bytes(b'second')
    with pytest.raises(HarnessFailure, match='prior evidence'):
        phase0_harness._capture_identity(request)
    assert first.repository_state_sha256


def _v3_batch_id_kwargs(
    *,
    comparison_key: str = 'c' * 64,
    pipeline: str = 'gb',
    recovery_provenance: RecoveryProvenance | None = None,
    upstream_gate_provenance: UpstreamGateProvenance | None = None,
) -> dict[str, object]:
    if recovery_provenance is None and upstream_gate_provenance is None:
        recovery_provenance = _test_recovery_provenance() if pipeline == 'gb' else None
        upstream_gate_provenance = _test_upstream_gate_provenance() if pipeline == 'sk' else None
    return {
        'comparison_key': comparison_key,
        'profile': ComparisonProfile.PHASE0B_VS_PHASE0A,
        'pipeline': pipeline,
        'phase0a_manifest_sha256': '9' * 64,
        'identity': _full_identity(),
        'recovery_provenance': recovery_provenance,
        'upstream_gate_provenance': upstream_gate_provenance,
    }


def _v3_ledger(
    tmp_path: Path,
    *,
    batch_id: str | None = None,
    comparison_key: str = 'c' * 64,
) -> AppendOnlyAttemptLedger:
    ledger = AppendOnlyAttemptLedger.create_v3_once(
        tmp_path / 'rust' / 'target' / 'perf-local' / 'batches',
        _full_identity(),
        comparison_key=comparison_key,
        phase0a_manifest_sha256='9' * 64,
        recovery_provenance=_test_recovery_provenance(),
        upstream_gate_provenance=None,
    )
    if batch_id is not None:
        assert ledger.batch_id == batch_id
    return ledger


def _v3_group_request(tmp_path: Path, *, metric: str) -> MetricGroupRequest:
    ledger = _v3_ledger(tmp_path)
    return MetricGroupRequest(
        _request(tmp_path),
        ledger.batch_id,
        metric,  # type: ignore[arg-type]
        build_round_plan(global_round_start=1, round_count=5),
        ledger.attempt_directory,
    )


def test_derive_expected_formal_v3_identity_is_read_only_for_gb(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    derive = getattr(phase0_harness, 'derive_expected_formal_v3_identity', None)
    assert callable(derive)
    parent = _write_synthetic_v2_recovery_parent(tmp_path)
    request = _request(tmp_path)
    monkeypatch.setattr(phase0_harness, 'APPROVED_RECOVERY_PARENTS', (parent.approved,))
    monkeypatch.setattr(phase0_harness, '_capture_static_comparison_inputs', lambda _request: parent.static)
    before = _file_snapshot(parent.comparison)
    expected_v3_root = request.attempt_ledger_root

    derived = derive(request)

    assert derived.static == parent.static
    assert derived.recovery_provenance == RecoveryProvenance(
        parent_protocol_version=parent.approved.parent_protocol_version,
        parent_comparison_key=parent.approved.parent_comparison_key,
        parent_attempt=parent.approved.parent_attempt,
        parent_terminal_sha256=parent.approved.parent_terminal_sha256,
        parent_comparison_tree_sha256=parent.approved.parent_comparison_tree_sha256,
        parent_journal_head_sha256=parent.approved.parent_journal_head_sha256,
        parent_inventory_entry_count=parent.approved.parent_inventory_entry_count,
        reason=parent.approved.reason,
    )
    assert derived.comparison_key == phase0_harness.derive_v3_comparison_key(
        pipeline=parent.static.pipeline,
        comparison_profile=parent.static.comparison_profile,
        reference_label=parent.static.reference_label,
        candidate_label=parent.static.candidate_label,
        phase0a_manifest_sha256=parent.static.phase0a_manifest_sha256,
        input_sha256=parent.static.input_sha256,
        reference_sha256=parent.static.reference_sha256,
        candidate_sha256=parent.static.candidate_sha256,
        recovery_provenance=derived.recovery_provenance,
        upstream_gate_provenance=None,
    )
    assert derived.artifact_basename == expected_benchmark_artifact_name(
        protocol_version=3,
        comparison_key=derived.comparison_key,
    )
    assert _file_snapshot(parent.comparison) == before
    assert not (expected_v3_root / derived.comparison_key).exists()
    assert not request.evidence_path.exists()


def test_task5_keeps_sk_v3_fail_closed_before_parent_or_child(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = replace(_request(tmp_path), pipeline='sk')
    static = replace(
        _write_synthetic_v2_recovery_parent(tmp_path).static,
        pipeline='sk',
    )
    monkeypatch.setattr(phase0_harness, '_capture_static_comparison_inputs', lambda _request: static)
    monkeypatch.setattr(
        phase0_harness,
        'authorize_v3_recovery',
        lambda _static: (_ for _ in ()).throw(AssertionError('SK must not authorize a GB recovery parent')),
    )
    monkeypatch.setattr(
        phase0_harness,
        '_capture_identity',
        lambda _request: (_ for _ in ()).throw(AssertionError('SK must fail before full identity or child capture')),
    )

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.derive_expected_formal_v3_identity(request)

    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert 'Task 7' in str(caught.value)


def test_v3_verdict_never_reads_v2_metric_values(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    ledger.commit_first_group(_v3_groups())
    wall = replace(_metric_group('wall', reference='1', candidate='0.90'), batch_id=ledger.batch_id)
    pws = replace(_metric_group('pws', reference='1', candidate='9'), batch_id=ledger.batch_id)
    attempt = phase0_harness._batch_attempt(ledger, ledger.batch_id, AttemptState.FIRST_GROUP_COMPLETE)

    evidence = phase0_harness._build_paired_evidence(
        _request(tmp_path),
        wall,
        pws,
        attempt,
        _full_identity(),
        (),
        ledger=ledger,
    )

    assert evidence.schema_version == evidence.protocol_version == 3
    assert evidence.comparison_key == ledger.comparison_key
    assert evidence.batch_id == ledger.batch_id
    assert evidence.recovery_provenance == ledger.recovery_provenance
    assert {item.metric.value: item.value for item in evidence.metrics}['wall_ratio'] == Decimal('0.90')
    assert {item.metric.value: item.value for item in evidence.metrics}['pws_ratio'] == Decimal('9')
    assert not any('metric' in field.name or 'median' in field.name for field in fields(RecoveryProvenance))


@pytest.mark.parametrize('metric', ('wall', 'pws'))
@pytest.mark.parametrize('role', ('reference', 'candidate'))
def test_inner_interruption_after_sample_started_is_mapped_and_never_retried(
    metric: str,
    role: str,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _v3_group_request(tmp_path, metric=metric)
    target_calls = 0

    def capture(*args: object, **kwargs: object) -> CapturedNormalRun:
        nonlocal target_calls
        executable = kwargs.get('executable', args[0] if args else None)
        captured_role = 'reference' if Path(executable).name.startswith('reference') else 'candidate'
        if captured_role == role:
            target_calls += 1
            snapshot = AppendOnlyAttemptLedger.load_read_only(request.attempt_directory, _full_identity())
            started = snapshot.sample_started(metric, 1, role)  # type: ignore[arg-type]
            assert started['planned_output_record_sha256'] == snapshot._plan_record_sha256s[(metric, 1, role)]
            assert snapshot.journal_head_sha256
            raise KeyboardInterrupt('simulated after durable sample start')
        return CapturedNormalRun(
            NormalRunEvidence(
                Decimal('1'),
                100 if metric == 'pws' else None,
                _runtime(),
                hashlib.sha256(f'{metric}:1'.encode()).hexdigest(),
            ),
            0,
            hashlib.sha256(f'{metric}:{captured_role}:log'.encode()).hexdigest(),
        )

    monkeypatch.setattr(phase0_harness, '_capture_identity', lambda _request: _full_identity())
    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', capture)
    monkeypatch.setattr(phase0_harness, '_invoke_pws_single_sample', capture)
    monkeypatch.setattr(phase0_harness, '_with_actual_workbook_dimensions', lambda captured, _output: captured)

    runner = run_normal_wall_group if metric == 'wall' else phase0_harness.run_pws_group
    with pytest.raises(HarnessFailure) as first:
        runner(request)
    assert first.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    snapshot = AppendOnlyAttemptLedger.load_read_only(request.attempt_directory, _full_identity())
    assert snapshot.sample_started_sha(metric, 1, role) is not None  # type: ignore[arg-type]
    assert snapshot.sample_payload(metric, 1, role) is None  # type: ignore[arg-type]
    assert snapshot.terminal_verdict is None

    with pytest.raises(HarnessFailure):
        runner(request)
    assert target_calls == 1


def test_outer_paired_runner_cleans_and_seals_interrupted_started_sample_once(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _request(tmp_path)
    comparison_key = 'c' * 64
    evidence_root = tmp_path / 'docs' / 'performance' / 'runs' / 'phase0b-v3'
    evidence_root.mkdir(parents=True)
    request = replace(
        request,
        evidence_path=evidence_root
        / expected_benchmark_artifact_name(protocol_version=3, comparison_key=comparison_key),
    )
    static = StaticComparisonInputs(
        pipeline='gb',
        comparison_profile=request.comparison_profile,
        reference_label=request.reference_label,
        candidate_label=request.candidate_label,
        phase0a_manifest_sha256='9' * 64,
        input_sha256=_full_identity().input_sha256,
        reference_sha256=_full_identity().reference_sha256,
        candidate_sha256=_full_identity().candidate_sha256,
    )
    monkeypatch.setattr(
        phase0_harness,
        'derive_expected_formal_v3_identity',
        lambda _request: phase0_harness.ExpectedFormalV3Identity(
            static,
            _test_recovery_provenance(),
            comparison_key,
            request.evidence_path.name,
            'NEW',
            None,
        ),
    )
    monkeypatch.setattr(phase0_harness, '_run_git', lambda *_args: '')
    monkeypatch.setattr(phase0_harness, '_capture_identity', lambda _request: _full_identity())
    monkeypatch.setattr(phase0_harness, '_load_approved_phase0a_baseline', lambda *_args: _phase0a_baseline(8))
    cleanup_calls = 0
    original_cleanup = phase0_harness._cleanup_all

    def cleanup(paths: list[Path]) -> tuple[str, ...]:
        nonlocal cleanup_calls
        cleanup_calls += 1
        return original_cleanup(paths)

    def interrupt(*args: object, **_kwargs: object) -> CapturedNormalRun:
        output = Path(args[3])
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_bytes(b'partial')
        raise KeyboardInterrupt('simulated child interruption')

    monkeypatch.setattr(phase0_harness, '_cleanup_all', cleanup)
    monkeypatch.setattr(phase0_harness, 'run_rust_normal_captured', interrupt)
    monkeypatch.setattr(
        phase0_harness,
        'run_pws_group',
        lambda _request: (_ for _ in ()).throw(AssertionError('PWS must not start after wall interruption')),
    )

    with pytest.raises(HarnessFailure) as caught:
        phase0_harness.run_paired_normal_batch(request)

    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert cleanup_calls == 1
    attempt = request.attempt_ledger_root / comparison_key / 'attempt-0001'
    sealed = AppendOnlyAttemptLedger.load_read_only(attempt, _full_identity())
    assert sealed.terminal_verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert sealed.sample_started_sha('wall', 1, 'reference') is not None
    assert sealed.sample_payload('wall', 1, 'reference') is None
    assert not list((tmp_path / 'data' / 'processed').rglob('*.xlsx'))


def test_v3_outer_cleanup_failure_preserves_primary_verdict(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    ledger = _v3_ledger(tmp_path)
    _record_v3_sample_start(ledger)
    primary = HarnessFailure(HarnessVerdict.CANDIDATE_FAILED, 'candidate failed')
    monkeypatch.setattr(phase0_harness, '_cleanup_all', lambda _paths: ('OSError:28',))

    final = phase0_harness._seal_v3_outer_failure(_request(tmp_path), ledger, primary)

    assert final.verdict is HarnessVerdict.CLEANUP_FAILED
    assert final.primary_verdict is HarnessVerdict.CANDIDATE_FAILED
    sealed = AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, _full_identity())
    assert sealed.terminal_verdict is HarnessVerdict.CLEANUP_FAILED
    assert sealed.terminal_primary_verdict is HarnessVerdict.CANDIDATE_FAILED


def test_formal_v3_runner_publishes_current_typed_evidence_after_cleanup(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    request = _request(tmp_path)
    comparison_key = 'c' * 64
    evidence_root = tmp_path / 'docs' / 'performance' / 'runs' / 'phase0b-v3'
    evidence_root.mkdir(parents=True)
    request = replace(
        request,
        evidence_path=evidence_root
        / expected_benchmark_artifact_name(protocol_version=3, comparison_key=comparison_key),
    )
    static = StaticComparisonInputs(
        pipeline='gb',
        comparison_profile=request.comparison_profile,
        reference_label=request.reference_label,
        candidate_label=request.candidate_label,
        phase0a_manifest_sha256='9' * 64,
        input_sha256=_full_identity().input_sha256,
        reference_sha256=_full_identity().reference_sha256,
        candidate_sha256=_full_identity().candidate_sha256,
    )
    monkeypatch.setattr(
        phase0_harness,
        'derive_expected_formal_v3_identity',
        lambda _request: phase0_harness.ExpectedFormalV3Identity(
            static,
            _test_recovery_provenance(),
            comparison_key,
            request.evidence_path.name,
            'NEW',
            None,
        ),
    )
    monkeypatch.setattr(phase0_harness, '_run_git', lambda *_args: '')
    monkeypatch.setattr(phase0_harness, '_capture_identity', lambda _request: _full_identity())
    monkeypatch.setattr(
        phase0_harness,
        '_load_approved_phase0a_baseline',
        lambda *_args: phase0_harness._ApprovedPhase0ABaseline(
            _full_identity().machine_fingerprint_sha256,
            8,
            (Decimal('1'),) * 5,
            (Decimal('100'),) * 5,
        ),
    )

    def group(group_request: MetricGroupRequest) -> MetricGroup:
        if group_request.metric == 'wall':
            value = _metric_group('wall', reference='1', candidate='0.90')
        else:
            value = _metric_group('pws', reference='100', candidate='900')
        return replace(value, batch_id=group_request.batch_id)

    publications: list[dict[str, object]] = []
    monkeypatch.setattr(phase0_harness, 'run_normal_wall_group', group)
    monkeypatch.setattr(phase0_harness, 'run_pws_group', group)
    monkeypatch.setattr(
        EvidenceSanitizer,
        'write_batch',
        lambda self, **kwargs: publications.append(kwargs),
    )

    result = phase0_harness.run_paired_normal_batch(request)

    assert result.verdict is HarnessVerdict.VALIDATED
    assert result.attempt.protocol_version == 3
    assert result.attempt.comparison_key == comparison_key
    assert result.wall is not None and result.pws is not None
    assert len(publications) == 1
    artifacts = publications[0]['artifacts']
    assert isinstance(artifacts, tuple)
    assert artifacts[0].source.schema_version == artifacts[0].source.protocol_version == 3
    committed = AppendOnlyAttemptLedger.load_read_only(result.attempt.attempt_directory, _full_identity())
    assert phase0_harness.classify_v3_attempt_state(committed) == 'EVIDENCE_COMMITTED'


def _v3_planned_payload(
    *,
    batch_id: str,
    role: str = 'reference',
    metric: str = 'wall',
    global_round: int = 1,
) -> dict[str, object]:
    identity = _full_identity()
    binary_sha256 = identity.reference_sha256 if role == 'reference' else identity.candidate_sha256
    return {
        'pipeline': 'gb',
        'batch_id': batch_id,
        'metric': metric,
        'binary_sha256': binary_sha256,
        'global_round': global_round,
        'role': role,
        'relative_path': (f'gb/.perf-runs/{batch_id}/{metric}/{binary_sha256}/{global_round}/{role}.xlsx'),
    }


def _v3_sample_payload(*, batch_id: str, role: str = 'reference') -> dict[str, object]:
    return {
        **_synthetic_v2_sample_payload(_full_identity(), metric='wall', global_round=1, role=role),
        'batch_id': batch_id,
    }


def _record_v3_sample_start(ledger: AppendOnlyAttemptLedger, *, role: str = 'reference') -> tuple[str, str]:
    plan_sha = ledger.record_planned_output(
        'wall',
        1,
        role,  # type: ignore[arg-type]
        _v3_planned_payload(batch_id=ledger.batch_id, role=role),
    )
    started_sha = ledger.record_sample_started(
        batch_id=ledger.batch_id,
        metric='wall',
        global_round=1,
        role=role,  # type: ignore[arg-type]
        order=('reference', 'candidate'),
        input_sha256=_full_identity().input_sha256,
        binary_sha256=(_full_identity().reference_sha256 if role == 'reference' else _full_identity().candidate_sha256),
        planned_output_record_sha256=plan_sha,
    )
    return plan_sha, started_sha


def test_v3_sample_started_is_durable_before_capture(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    _plan_sha, started_sha = _record_v3_sample_start(ledger)

    assert tuple((ledger.attempt_directory / 'records').glob('*sample-started.json'))
    loaded = AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, _full_identity())
    assert loaded.sample_started_sha('wall', 1, 'reference') == started_sha


def test_reloaded_plan_returns_original_record_sha_for_sample_start(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    original_plan_sha = ledger.record_planned_output(
        'wall', 1, 'reference', _v3_planned_payload(batch_id=ledger.batch_id)
    )
    ledger.record_planned_output(
        'wall', 1, 'candidate', _v3_planned_payload(batch_id=ledger.batch_id, role='candidate')
    )

    reloaded = AppendOnlyAttemptLedger.open_v3_for_resume(ledger.attempt_directory, _full_identity())

    assert (
        reloaded.record_planned_output('wall', 1, 'reference', _v3_planned_payload(batch_id=ledger.batch_id))
        == original_plan_sha
    )


def test_v3_batch_id_matches_exact_13_key_payload_and_is_bound_everywhere(tmp_path: Path) -> None:
    recovery = _test_recovery_provenance()
    identity = _full_identity()
    comparison_key = 'c' * 64
    expected_payload = {
        'protocol_version': 3,
        'comparison_key': comparison_key,
        'profile': 'phase0b-vs-phase0a',
        'pipeline': 'gb',
        'phase0a_manifest_sha256': '9' * 64,
        'input_sha256': identity.input_sha256,
        'reference_sha256': identity.reference_sha256,
        'candidate_sha256': identity.candidate_sha256,
        'git_head': identity.git_head,
        'repository_state_sha256': identity.repository_state_sha256,
        'machine_fingerprint_sha256': identity.machine_fingerprint_sha256,
        'recovery_provenance': _recovery_payload(recovery),
        'upstream_gate_provenance': None,
    }
    expected = hashlib.sha256(
        json.dumps(expected_payload, ensure_ascii=True, sort_keys=True, separators=(',', ':')).encode('utf-8')
    ).hexdigest()
    actual = phase0_harness.derive_v3_batch_id(**_v3_batch_id_kwargs(comparison_key=comparison_key))
    assert actual == expected
    ledger = _v3_ledger(tmp_path, batch_id=actual, comparison_key=comparison_key)
    assert ledger.metadata['batch_id'] == actual
    plan_sha, started_sha = _record_v3_sample_start(ledger)
    assert plan_sha
    ledger.record_sample(
        'wall',
        1,
        'reference',
        _v3_sample_payload(batch_id=actual),
        sample_started_record_sha256=started_sha,
    )
    reloaded = AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, identity)
    assert reloaded.planned_output('wall', 1, 'reference')['batch_id'] == actual
    assert reloaded.sample_started('wall', 1, 'reference')['batch_id'] == actual
    assert reloaded.sample('wall', 1, 'reference')['batch_id'] == actual
    assert reloaded.sample('wall', 1, 'reference')['sample_started_record_sha256'] == started_sha


@pytest.mark.parametrize('record_kind', ('planned-output', 'sample-started', 'sample'))
def test_v3_ledger_rejects_record_from_wrong_batch(record_kind: str, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    wrong_batch = 'd' * 64
    with pytest.raises(HarnessFailure, match='batch'):
        if record_kind == 'planned-output':
            ledger.record_planned_output('wall', 1, 'reference', _v3_planned_payload(batch_id=wrong_batch))
        else:
            _plan_sha, started_sha = _record_v3_sample_start(ledger)
            if record_kind == 'sample-started':
                ledger.record_sample_started(
                    batch_id=wrong_batch,
                    metric='wall',
                    global_round=2,
                    role='reference',
                    order=('candidate', 'reference'),
                    input_sha256=_full_identity().input_sha256,
                    binary_sha256=_full_identity().reference_sha256,
                    planned_output_record_sha256=started_sha,
                )
            else:
                ledger.record_sample(
                    'wall',
                    1,
                    'reference',
                    _v3_sample_payload(batch_id=wrong_batch),
                    sample_started_record_sha256=started_sha,
                )


@pytest.mark.parametrize('mutation', ('parent_tree', 'parent_journal', 'gb_artifact', 'gb_marker', 'gb_commit'))
def test_v3_batch_id_changes_for_every_recovery_or_upstream_anchor(mutation: str) -> None:
    if mutation.startswith('parent_'):
        recovery = _test_recovery_provenance()
        field = {
            'parent_tree': 'parent_comparison_tree_sha256',
            'parent_journal': 'parent_journal_head_sha256',
        }[mutation]
        before = _v3_batch_id_kwargs(recovery_provenance=recovery)
        after = _v3_batch_id_kwargs(recovery_provenance=replace(recovery, **{field: 'f' * 64}))
    else:
        upstream = _test_upstream_gate_provenance()
        field = {
            'gb_artifact': 'artifact_sha256',
            'gb_marker': 'marker_sha256',
            'gb_commit': 'validated_commit_sha',
        }[mutation]
        replacement = 'f' * (40 if field == 'validated_commit_sha' else 64)
        before = _v3_batch_id_kwargs(pipeline='sk', recovery_provenance=None, upstream_gate_provenance=upstream)
        after = _v3_batch_id_kwargs(
            pipeline='sk',
            recovery_provenance=None,
            upstream_gate_provenance=replace(upstream, **{field: replacement}),
        )
    assert phase0_harness.derive_v3_batch_id(**before) != phase0_harness.derive_v3_batch_id(**after)


def test_v3_batch_payload_rejects_unknown_key_string_number_bool_as_int_and_illegal_inventory() -> None:
    recovery = _recovery_payload(_test_recovery_provenance())
    payload = {
        'protocol_version': 3,
        'comparison_key': 'c' * 64,
        'profile': 'phase0b-vs-phase0a',
        'pipeline': 'gb',
        'phase0a_manifest_sha256': '9' * 64,
        **asdict(_full_identity()),
        'recovery_provenance': recovery,
        'upstream_gate_provenance': None,
    }
    invalid_payloads = (
        payload | {'unknown': None},
        payload | {'protocol_version': '3'},
        payload | {'recovery_provenance': recovery | {'parent_attempt': True}},
        payload | {'recovery_provenance': recovery | {'parent_inventory_entry_count': 135}},
    )
    for invalid in invalid_payloads:
        with pytest.raises(ValueError):
            phase0_harness._parse_exact_v3_batch_payload(invalid)


def test_started_without_sample_is_terminal_and_never_reinvokes_capture(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    ledger = _v3_ledger(tmp_path)
    _record_v3_sample_start(ledger)
    monkeypatch.setattr(
        phase0_harness,
        'run_rust_normal_captured',
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError('capture must not rerun')),
    )

    with pytest.raises(HarnessFailure) as caught:
        AppendOnlyAttemptLedger.create_v3_once(
            ledger.attempt_directory.parents[1],
            _full_identity(),
            comparison_key=ledger.comparison_key,
            phase0a_manifest_sha256='9' * 64,
            recovery_provenance=_test_recovery_provenance(),
            upstream_gate_provenance=None,
        )

    assert caught.value.verdict is HarnessVerdict.INCOMPLETE_EVIDENCE
    assert (
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, _full_identity()).terminal_verdict
        is HarnessVerdict.INCOMPLETE_EVIDENCE
    )


@pytest.mark.parametrize(
    'terminal',
    tuple(
        verdict
        for verdict in HarnessVerdict
        if verdict not in (HarnessVerdict.VALIDATED, HarnessVerdict.CLEANUP_FAILED)
    ),
)
def test_v3_failure_terminal_never_creates_sampling_successor(terminal: HarnessVerdict, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    ledger.finish(terminal)
    with pytest.raises(HarnessFailure):
        AppendOnlyAttemptLedger.create_v3_once(
            ledger.attempt_directory.parents[1],
            _full_identity(),
            comparison_key=ledger.comparison_key,
            phase0a_manifest_sha256='9' * 64,
            recovery_provenance=_test_recovery_provenance(),
            upstream_gate_provenance=None,
        )


def test_cleanup_only_successor_prohibits_all_benchmark_records(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    ledger.finish(HarnessVerdict.CLEANUP_FAILED, primary_verdict=HarnessVerdict.CANDIDATE_FAILED)
    successor = AppendOnlyAttemptLedger.create_v3_once(
        ledger.attempt_directory.parents[1],
        _full_identity(),
        comparison_key=ledger.comparison_key,
        phase0a_manifest_sha256='9' * 64,
        recovery_provenance=_test_recovery_provenance(),
        upstream_gate_provenance=None,
    )
    for operation in (
        lambda: successor.record_planned_output('wall', 1, 'reference', {}),
        lambda: successor.record_sample_started(
            batch_id=successor.batch_id,
            metric='wall',
            global_round=1,
            role='reference',
            order=('reference', 'candidate'),
            input_sha256='3' * 64,
            binary_sha256='1' * 64,
            planned_output_record_sha256='4' * 64,
        ),
        lambda: successor.record_sample('wall', 1, 'reference', {}, sample_started_record_sha256='5' * 64),
    ):
        with pytest.raises(HarnessFailure, match='cleanup-only'):
            operation()


def test_committed_state_binds_artifact_and_marker_and_is_sealed(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    ledger.commit_first_group(_v3_groups())
    ledger.mark_cleanup_complete()
    artifact_content = '{}'
    artifact_sha256 = hashlib.sha256(artifact_content.encode('utf-8')).hexdigest()
    artifact_basename = f'benchmark-v3-{ledger.comparison_key[:16]}.json'
    marker_basename, marker_sha256 = _expected_v3_marker_identity(artifact_basename, artifact_content)
    ledger.prepare_evidence(
        artifact_basename=artifact_basename,
        artifact_sha256=artifact_sha256,
        artifact_content=artifact_content,
        marker_basename=marker_basename,
        marker_sha256=marker_sha256,
    )
    ledger.mark_evidence_committed(artifact_sha256=artifact_sha256, marker_sha256=marker_sha256)
    loaded = AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, _full_identity())
    assert loaded.state is AttemptState.EVIDENCE_COMMITTED
    with pytest.raises(HarnessFailure):
        loaded.record_sample('wall', 1, 'reference', {}, sample_started_record_sha256='5' * 64)


@pytest.mark.parametrize('link_kind', ('file', 'directory'))
def test_v3_read_only_loader_rejects_reparse_inventory(link_kind: str, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    ledger.record_planned_output('wall', 1, 'reference', _v3_planned_payload(batch_id=ledger.batch_id))
    link = (
        next((ledger.attempt_directory / 'records').glob('*.json'))
        if link_kind == 'file'
        else ledger.attempt_directory / 'records'
    )
    target = tmp_path / f'outside-{link_kind}'
    link.rename(target)
    try:
        link.symlink_to(target, target_is_directory=link_kind == 'directory')
    except OSError as exc:
        pytest.skip(f'symlink creation is unavailable: {exc}')

    with pytest.raises(HarnessFailure, match='reparse'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, _full_identity())


@pytest.mark.parametrize('field', ('input_sha256', 'binary_sha256'))
def test_v3_sample_identity_must_match_durable_start(field: str, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    _plan_sha, started_sha = _record_v3_sample_start(ledger)
    payload = _v3_sample_payload(batch_id=ledger.batch_id)
    payload[field] = 'f' * 64

    with pytest.raises(HarnessFailure, match='identity'):
        ledger.record_sample(
            'wall',
            1,
            'reference',
            payload,
            sample_started_record_sha256=started_sha,
        )


def _expected_v3_marker_identity(artifact_basename: str, artifact_content: str) -> tuple[str, str]:
    artifact_sha256 = hashlib.sha256(artifact_content.encode('utf-8')).hexdigest()
    basis = {
        'schema_version': 1,
        'artifacts': [{'kind': 'benchmark', 'file_name': artifact_basename, 'sha256': artifact_sha256}],
    }
    batch_sha256 = hashlib.sha256(
        json.dumps(basis, ensure_ascii=False, separators=(',', ':'), allow_nan=False).encode('utf-8')
    ).hexdigest()
    marker_content = (
        json.dumps({**basis, 'batch_sha256': batch_sha256}, ensure_ascii=False, indent=2, allow_nan=False) + '\n'
    )
    return f'batch-{batch_sha256[:16]}.commit.json', hashlib.sha256(marker_content.encode('utf-8')).hexdigest()


@pytest.mark.parametrize('mutation', ('artifact_basename', 'marker_basename'))
def test_v3_prepared_evidence_binds_exact_artifact_and_marker_names(mutation: str, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    ledger.commit_first_group(_v3_groups())
    ledger.mark_cleanup_complete()
    artifact_content = '{}'
    artifact_basename = f'benchmark-v3-{ledger.comparison_key[:16]}.json'
    if mutation == 'artifact_basename':
        artifact_basename = 'benchmark-v3-ffffffffffffffff.json'
    marker_basename, marker_sha256 = _expected_v3_marker_identity(artifact_basename, artifact_content)
    if mutation == 'marker_basename':
        marker_basename = 'batch-ffffffffffffffff.commit.json'

    with pytest.raises(HarnessFailure, match='basename|marker'):
        ledger.prepare_evidence(
            artifact_basename=artifact_basename,
            artifact_sha256=hashlib.sha256(artifact_content.encode('utf-8')).hexdigest(),
            artifact_content=artifact_content,
            marker_basename=marker_basename,
            marker_sha256=marker_sha256,
        )


def _append_v3_test_journal_anchor(
    ledger: AppendOnlyAttemptLedger,
    *,
    record_count: int,
    record_head_sha256: str,
    checkpoint_head_sha256: str,
    terminal_head_sha256: str | None = None,
    verdict: HarnessVerdict | None = None,
) -> str:
    journal = sorted((ledger.attempt_directory.parent / 'journal').glob('*.json'))
    previous = hashlib.sha256(journal[-1].read_bytes()).hexdigest()
    return _write_synthetic_journal_entry(
        ledger.attempt_directory.parent,
        previous_journal_sha256=previous,
        record_count=record_count,
        record_head_sha256=record_head_sha256,
        checkpoint_head_sha256=checkpoint_head_sha256,
        terminal_head_sha256=terminal_head_sha256,
        verdict=verdict,
    )


def _write_resealed_v3_terminal(
    ledger: AppendOnlyAttemptLedger,
    verdict: HarnessVerdict,
    *,
    record_count: object | None = None,
    primary_verdict: object = None,
) -> None:
    terminal = {
        'verdict': verdict.value,
        'primary_verdict': primary_verdict,
        'raw_log_sha256': None,
        'record_count': ledger._record_count if record_count is None else record_count,
        'record_head_sha256': ledger._record_head_sha256,
        'checkpoint_head_sha256': ledger._checkpoint_head_sha256,
    }
    raw = phase0_harness._canonical_json(terminal)
    (ledger.attempt_directory / 'terminal.json').write_bytes(raw)
    _append_v3_test_journal_anchor(
        ledger,
        record_count=ledger._record_count,
        record_head_sha256=ledger._record_head_sha256,
        checkpoint_head_sha256=ledger._checkpoint_head_sha256,
        terminal_head_sha256=hashlib.sha256(raw).hexdigest(),
        verdict=verdict,
    )


def _commit_v3_test_evidence(ledger: AppendOnlyAttemptLedger) -> None:
    ledger.commit_first_group(_v3_groups())
    ledger.mark_cleanup_complete()
    artifact_content = '{}'
    artifact_basename = f'benchmark-v3-{ledger.comparison_key[:16]}.json'
    marker_basename, marker_sha256 = _expected_v3_marker_identity(artifact_basename, artifact_content)
    artifact_sha256 = hashlib.sha256(artifact_content.encode('utf-8')).hexdigest()
    ledger.prepare_evidence(
        artifact_basename=artifact_basename,
        artifact_sha256=artifact_sha256,
        artifact_content=artifact_content,
        marker_basename=marker_basename,
        marker_sha256=marker_sha256,
    )
    ledger.mark_evidence_committed(artifact_sha256=artifact_sha256, marker_sha256=marker_sha256)


@pytest.mark.parametrize(
    ('mutation', 'value'),
    (('pipeline', 'sk'), ('binary_sha256', 'f' * 64)),
)
def test_v3_planned_output_binds_pipeline_and_role_binary(mutation: str, value: str, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    payload = _v3_planned_payload(batch_id=ledger.batch_id)
    payload[mutation] = value
    if mutation == 'binary_sha256':
        payload['relative_path'] = f'gb/.perf-runs/{ledger.batch_id}/wall/{value}/1/reference.xlsx'
    elif mutation == 'pipeline':
        payload['relative_path'] = (
            f'sk/.perf-runs/{ledger.batch_id}/wall/{_full_identity().reference_sha256}/1/reference.xlsx'
        )

    with pytest.raises(HarnessFailure, match='pipeline|binary|identity'):
        ledger.record_planned_output('wall', 1, 'reference', payload)


def test_v3_loader_rejects_resealed_plan_start_sample_with_one_wrong_binary(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    wrong_binary = 'f' * 64
    plan = _v3_planned_payload(batch_id=ledger.batch_id)
    plan['binary_sha256'] = wrong_binary
    plan['relative_path'] = f'gb/.perf-runs/{ledger.batch_id}/wall/{wrong_binary}/1/reference.xlsx'
    record_head, checkpoint_head = _write_synthetic_record(
        ledger.attempt_directory,
        sequence=1,
        kind='planned-output',
        previous_record_sha256=ledger._record_head_sha256,
        previous_checkpoint_sha256=ledger._checkpoint_head_sha256,
        metric='wall',
        global_round=1,
        role='reference',
        payload=plan,
    )
    _append_v3_test_journal_anchor(
        ledger, record_count=1, record_head_sha256=record_head, checkpoint_head_sha256=checkpoint_head
    )
    started = {
        'batch_id': ledger.batch_id,
        'metric': 'wall',
        'global_round': 1,
        'role': 'reference',
        'order': ['reference', 'candidate'],
        'input_sha256': ledger.identity.input_sha256,
        'binary_sha256': wrong_binary,
        'planned_output_record_sha256': record_head,
    }
    record_head, checkpoint_head = _write_synthetic_record(
        ledger.attempt_directory,
        sequence=2,
        kind='sample-started',
        previous_record_sha256=record_head,
        previous_checkpoint_sha256=checkpoint_head,
        **started,
    )
    _append_v3_test_journal_anchor(
        ledger, record_count=2, record_head_sha256=record_head, checkpoint_head_sha256=checkpoint_head
    )
    sample = _v3_sample_payload(batch_id=ledger.batch_id)
    sample['binary_sha256'] = wrong_binary
    sample['sample_started_record_sha256'] = record_head
    record_head, checkpoint_head = _write_synthetic_record(
        ledger.attempt_directory,
        sequence=3,
        kind='sample',
        previous_record_sha256=record_head,
        previous_checkpoint_sha256=checkpoint_head,
        metric='wall',
        global_round=1,
        role='reference',
        payload=sample,
    )
    _append_v3_test_journal_anchor(
        ledger, record_count=3, record_head_sha256=record_head, checkpoint_head_sha256=checkpoint_head
    )

    with pytest.raises(HarnessFailure, match='binary|identity|linkage'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


@pytest.mark.parametrize('verdict', (HarnessVerdict.INCOMPLETE_EVIDENCE, HarnessVerdict.CLEANUP_FAILED))
def test_v3_committed_success_rejects_fully_resealed_terminal(verdict: HarnessVerdict, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    _commit_v3_test_evidence(ledger)
    _write_resealed_v3_terminal(ledger, verdict)

    with pytest.raises(HarnessFailure, match='committed|terminal|sealed'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


def test_v3_terminal_rejects_validated_verdict(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    _write_resealed_v3_terminal(ledger, HarnessVerdict.VALIDATED)

    with pytest.raises(HarnessFailure, match='verdict|terminal'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


def test_v3_committed_success_never_creates_cleanup_only_successor(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    _commit_v3_test_evidence(ledger)
    _write_resealed_v3_terminal(ledger, HarnessVerdict.CLEANUP_FAILED)

    with pytest.raises(HarnessFailure):
        AppendOnlyAttemptLedger.create_v3_once(
            ledger.attempt_directory.parents[1],
            ledger.identity,
            comparison_key=ledger.comparison_key,
            phase0a_manifest_sha256=ledger.phase0a_manifest_sha256,
            recovery_provenance=ledger.recovery_provenance,
            upstream_gate_provenance=ledger.upstream_gate_provenance,
        )


def test_v3_read_only_rejects_extra_comparison_sibling(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    (ledger.attempt_directory.parent / 'unexpected.json').write_text('{}', encoding='utf-8')

    with pytest.raises(HarnessFailure, match='inventory'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


def test_v3_read_only_rejects_journal_directory_reparse_before_journal_read(tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    journal = ledger.attempt_directory.parent / 'journal'
    target = tmp_path / 'outside-journal'
    journal.rename(target)
    try:
        journal.symlink_to(target, target_is_directory=True)
    except OSError as exc:
        pytest.skip(f'symlink creation is unavailable: {exc}')

    with pytest.raises(HarnessFailure, match='inventory|reparse'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


@pytest.mark.parametrize(
    ('case', 'expected'),
    (
        ('new', 'NEW'),
        ('sampling', 'SAMPLING_RESUMABLE'),
        ('started', 'STARTED_WITHOUT_SAMPLE'),
        ('cleanup', 'CLEANUP_COMPLETE'),
        ('prepared', 'EVIDENCE_PREPARED'),
        ('committed', 'EVIDENCE_COMMITTED'),
        ('cleanup_only', 'CLEANUP_ONLY'),
        ('failed', 'FAILED_TERMINAL'),
        ('committed_terminal', 'INVALID'),
        ('cleanup_only_benchmark', 'INVALID'),
    ),
)
def test_v3_classifier_has_exact_closed_state_contract(case: str, expected: str, tmp_path: Path) -> None:
    if case == 'new':
        ledger = None
    else:
        ledger = _v3_ledger(tmp_path)
        if case == 'started':
            _record_v3_sample_start(ledger)
        elif case in ('cleanup', 'prepared', 'committed', 'committed_terminal'):
            ledger.commit_first_group(_v3_groups())
            ledger.mark_cleanup_complete()
            if case in ('prepared', 'committed', 'committed_terminal'):
                artifact_content = '{}'
                artifact_basename = f'benchmark-v3-{ledger.comparison_key[:16]}.json'
                marker_basename, marker_sha256 = _expected_v3_marker_identity(artifact_basename, artifact_content)
                artifact_sha256 = hashlib.sha256(artifact_content.encode('utf-8')).hexdigest()
                ledger.prepare_evidence(
                    artifact_basename=artifact_basename,
                    artifact_sha256=artifact_sha256,
                    artifact_content=artifact_content,
                    marker_basename=marker_basename,
                    marker_sha256=marker_sha256,
                )
                if case in ('committed', 'committed_terminal'):
                    ledger.mark_evidence_committed(artifact_sha256=artifact_sha256, marker_sha256=marker_sha256)
                if case == 'committed_terminal':
                    ledger.terminal_verdict = HarnessVerdict.CLEANUP_FAILED
        elif case in ('cleanup_only', 'cleanup_only_benchmark'):
            ledger.finish(HarnessVerdict.CLEANUP_FAILED)
            ledger = AppendOnlyAttemptLedger.create_v3_once(
                ledger.attempt_directory.parents[1],
                ledger.identity,
                comparison_key=ledger.comparison_key,
                phase0a_manifest_sha256=ledger.phase0a_manifest_sha256,
                recovery_provenance=ledger.recovery_provenance,
                upstream_gate_provenance=ledger.upstream_gate_provenance,
            )
            if case == 'cleanup_only_benchmark':
                ledger._sample_payloads[('wall', 1, 'reference')] = {}
        elif case == 'failed':
            ledger.finish(HarnessVerdict.INCOMPLETE_EVIDENCE)

    assert phase0_harness.classify_v3_attempt_state(ledger) == expected


def _rewrite_latest_v3_journal(ledger: AppendOnlyAttemptLedger, **changes: object) -> None:
    path = sorted((ledger.attempt_directory.parent / 'journal').glob('*.json'))[-1]
    payload = json.loads(path.read_text(encoding='utf-8'))
    payload.update(changes)
    path.write_bytes(phase0_harness._canonical_json(payload))


@pytest.mark.parametrize(
    ('target', 'field', 'value'),
    (
        ('checkpoint', 'record_count', True),
        ('terminal', 'record_count', False),
        ('terminal', 'primary_verdict', False),
        ('metadata', 'recovery_primary_verdict', False),
        ('journal', 'attempt_number', 1.0),
        ('journal', 'record_count', 0.0),
        ('journal', 'terminal_present', 0),
    ),
)
def test_v3_scalar_grammar_rejects_bool_number_and_null_confusion(
    target: str, field: str, value: object, tmp_path: Path
) -> None:
    ledger = _v3_ledger(tmp_path)
    if target == 'checkpoint':
        ledger.record_planned_output('wall', 1, 'reference', _v3_planned_payload(batch_id=ledger.batch_id))
        checkpoint = ledger.attempt_directory / 'checkpoints' / '0001.json'
        payload = json.loads(checkpoint.read_text(encoding='utf-8'))
        payload[field] = value
        raw = phase0_harness._canonical_json(payload)
        checkpoint.write_bytes(raw)
        _rewrite_latest_v3_journal(ledger, checkpoint_head_sha256=hashlib.sha256(raw).hexdigest())
    elif target == 'terminal':
        _write_resealed_v3_terminal(
            ledger,
            HarnessVerdict.INCOMPLETE_EVIDENCE,
            record_count=value if field == 'record_count' else None,
            primary_verdict=value if field == 'primary_verdict' else None,
        )
    elif target == 'metadata':
        _rewrite_metadata_and_matching_empty_journal(ledger.attempt_directory, **{field: value})
    else:
        _rewrite_latest_v3_journal(ledger, **{field: value})

    with pytest.raises(HarnessFailure):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


def test_v2_committed_success_still_rejects_appended_records(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _request, ledger, artifact = _prepare_formal_evidence_recovery(monkeypatch, tmp_path)
    ledger.mark_evidence_committed(artifact_sha256=hashlib.sha256(artifact.content.encode()).hexdigest())

    with pytest.raises(HarnessFailure, match='committed'):
        ledger.record_sample('wall', 10, 'reference', {'value': 1})


def test_v2_committed_success_preserves_legacy_finish_behavior(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _request, ledger, artifact = _prepare_formal_evidence_recovery(monkeypatch, tmp_path)
    ledger.mark_evidence_committed(artifact_sha256=hashlib.sha256(artifact.content.encode()).hexdigest())

    ledger.finish(HarnessVerdict.INCONCLUSIVE)

    assert ledger.terminal_verdict is HarnessVerdict.INCONCLUSIVE


def _v3_groups() -> dict[str, str]:
    return {'wall': 'a' * 64, 'pws': 'b' * 64}


def _append_unchecked_v3_record(
    ledger: AppendOnlyAttemptLedger,
    kind: str,
    **payload: object,
) -> tuple[str, str]:
    sequence = ledger._record_count + 1
    record_head, checkpoint_head = _write_synthetic_record(
        ledger.attempt_directory,
        sequence=sequence,
        kind=kind,
        previous_record_sha256=ledger._record_head_sha256,
        previous_checkpoint_sha256=ledger._checkpoint_head_sha256,
        **payload,
    )
    _append_v3_test_journal_anchor(
        ledger,
        record_count=sequence,
        record_head_sha256=record_head,
        checkpoint_head_sha256=checkpoint_head,
    )
    return record_head, checkpoint_head


@pytest.mark.parametrize(
    ('kind', 'field', 'value'),
    (
        ('planned-output', 'global_round', True),
        ('planned-output', 'metric', 1),
        ('planned-output', 'role', False),
        ('sample', 'global_round', True),
        ('sample', 'metric', 1),
        ('sample', 'role', False),
    ),
)
def test_v3_loader_rejects_non_exact_outer_sample_and_plan_keys(
    kind: str, field: str, value: object, tmp_path: Path
) -> None:
    ledger = _v3_ledger(tmp_path)
    outer: dict[str, object] = {'metric': 'wall', 'global_round': 1, 'role': 'reference'}
    if kind == 'planned-output':
        payload = _v3_planned_payload(batch_id=ledger.batch_id)
    else:
        _plan_sha, started_sha = _record_v3_sample_start(ledger)
        payload = _v3_sample_payload(batch_id=ledger.batch_id)
        payload['sample_started_record_sha256'] = started_sha
    outer[field] = value
    _append_unchecked_v3_record(ledger, kind, **outer, payload=payload)

    with pytest.raises(HarnessFailure):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


@pytest.mark.parametrize(
    ('kind', 'groups'),
    (
        ('first', ['wall', 'pws']),
        ('first', {'wall': 'a' * 64, 'pws': False}),
        ('expanded', ['wall', 'pws']),
        ('expanded', {'wall': 'a' * 64, 'pws': False}),
    ),
)
def test_v3_group_writer_rejects_non_exact_groups(kind: str, groups: object, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    with pytest.raises(HarnessFailure, match='group'):
        if kind == 'first':
            ledger.commit_first_group(groups)  # type: ignore[arg-type]
        else:
            first_sha = ledger.commit_first_group(_v3_groups())
            ledger.commit_expanded_group(groups, first_group_sha256=first_sha)  # type: ignore[arg-type]


@pytest.mark.parametrize(
    ('kind', 'mutation'),
    (
        ('first', 'list'),
        ('first', 'bad_hash'),
        ('expanded', 'list'),
        ('expanded', 'bad_hash'),
        ('expanded', 'wrong_first_sha'),
        ('expanded', 'non_sha_first'),
    ),
)
def test_v3_loader_rejects_non_exact_group_records(kind: str, mutation: str, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    groups: object = _v3_groups()
    if mutation == 'list':
        groups = ['wall', 'pws']
    elif mutation == 'bad_hash':
        groups = {'wall': 'a' * 64, 'pws': False}
    if kind == 'first':
        _append_unchecked_v3_record(ledger, 'first-group', groups=groups)
    else:
        first_sha = ledger.commit_first_group(_v3_groups())
        linked_sha = first_sha
        if mutation == 'wrong_first_sha':
            linked_sha = 'f' * 64
        elif mutation == 'non_sha_first':
            linked_sha = False  # type: ignore[assignment]
        _append_unchecked_v3_record(
            ledger,
            'expanded-group',
            first_group_sha256=linked_sha,
            groups=groups,
        )

    with pytest.raises(HarnessFailure, match='group|SHA|link'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


@pytest.mark.parametrize('planned_output_count', (True, 0.0, 1))
def test_v3_loader_rejects_non_exact_cleanup_count(planned_output_count: object, tmp_path: Path) -> None:
    ledger = _v3_ledger(tmp_path)
    ledger.commit_first_group(_v3_groups())
    _append_unchecked_v3_record(
        ledger,
        'cleanup-complete',
        planned_output_count=planned_output_count,
    )

    with pytest.raises(HarnessFailure, match='cleanup|count'):
        AppendOnlyAttemptLedger.load_read_only(ledger.attempt_directory, ledger.identity)


@pytest.mark.parametrize('loader_name', ('read', 'resume'))
@pytest.mark.parametrize('mutation', ('unknown_key', 'bool_hash', 'reference_mismatch'))
def test_v3_identity_is_exact_even_when_strict_identity_is_false(
    loader_name: str, mutation: str, tmp_path: Path
) -> None:
    ledger = _v3_ledger(tmp_path)
    stored = asdict(ledger.identity)
    if mutation == 'unknown_key':
        stored['unknown'] = None
    elif mutation == 'bool_hash':
        stored['input_sha256'] = False
    else:
        stored['reference_sha256'] = 'f' * 64
    _rewrite_metadata_and_matching_empty_journal(ledger.attempt_directory, identity=stored)

    with pytest.raises(HarnessFailure, match='identity|metadata'):
        if loader_name == 'read':
            AppendOnlyAttemptLedger.load_read_only(
                ledger.attempt_directory,
                ledger.identity,
                strict_identity=False,
            )
        else:
            AppendOnlyAttemptLedger.open_v3_for_resume(
                ledger.attempt_directory,
                ledger.identity,
                strict_identity=False,
            )


def _write_test_journal_entry(
    comparison: Path,
    *,
    sequence: int,
    previous_journal_sha256: str | None,
    attempt_number: int,
    record_count: int,
    record_head_sha256: str,
    checkpoint_head_sha256: str,
    terminal_head_sha256: str | None = None,
    verdict: HarnessVerdict | None = None,
) -> str:
    state = phase0_harness._journal_state_payload(
        attempt_number=attempt_number,
        record_count=record_count,
        record_head_sha256=record_head_sha256,
        checkpoint_head_sha256=checkpoint_head_sha256,
        terminal_present=terminal_head_sha256 is not None,
        terminal_head_sha256=terminal_head_sha256,
        verdict=verdict,
    )
    raw = phase0_harness._canonical_json({'previous_journal_sha256': previous_journal_sha256, **state})
    (comparison / 'journal' / f'{sequence:06d}.json').write_bytes(raw)
    return hashlib.sha256(raw).hexdigest()


def test_v3_recursive_cleanup_successor_never_relaxes_parent_identity(tmp_path: Path) -> None:
    parent = _v3_ledger(tmp_path)
    parent.finish(HarnessVerdict.CLEANUP_FAILED)
    successor = AppendOnlyAttemptLedger.create_v3_once(
        parent.attempt_directory.parents[1],
        parent.identity,
        comparison_key=parent.comparison_key,
        phase0a_manifest_sha256=parent.phase0a_manifest_sha256,
        recovery_provenance=parent.recovery_provenance,
        upstream_gate_provenance=parent.upstream_gate_provenance,
    )
    parent_metadata_path = parent.attempt_directory / 'metadata.json'
    parent_metadata = json.loads(parent_metadata_path.read_text(encoding='utf-8'))
    parent_metadata['identity']['reference_sha256'] = 'f' * 64
    parent_metadata_raw = phase0_harness._canonical_json(parent_metadata)
    parent_metadata_path.write_bytes(parent_metadata_raw)
    parent_metadata_sha = hashlib.sha256(parent_metadata_raw).hexdigest()
    parent_terminal_path = parent.attempt_directory / 'terminal.json'
    parent_terminal = json.loads(parent_terminal_path.read_text(encoding='utf-8'))
    parent_terminal['record_head_sha256'] = parent_metadata_sha
    parent_terminal['checkpoint_head_sha256'] = parent_metadata_sha
    parent_terminal_raw = phase0_harness._canonical_json(parent_terminal)
    parent_terminal_path.write_bytes(parent_terminal_raw)
    parent_terminal_sha = hashlib.sha256(parent_terminal_raw).hexdigest()
    successor_metadata_path = successor.attempt_directory / 'metadata.json'
    successor_metadata = json.loads(successor_metadata_path.read_text(encoding='utf-8'))
    successor_metadata['previous_attempt_head_sha256'] = parent_terminal_sha
    successor_metadata_raw = phase0_harness._canonical_json(successor_metadata)
    successor_metadata_path.write_bytes(successor_metadata_raw)
    successor_metadata_sha = hashlib.sha256(successor_metadata_raw).hexdigest()
    journal_directory = parent.attempt_directory.parent / 'journal'
    for path in journal_directory.glob('*.json'):
        path.unlink()
    first_journal = _write_test_journal_entry(
        parent.attempt_directory.parent,
        sequence=1,
        previous_journal_sha256=None,
        attempt_number=1,
        record_count=0,
        record_head_sha256=parent_metadata_sha,
        checkpoint_head_sha256=parent_metadata_sha,
    )
    second_journal = _write_test_journal_entry(
        parent.attempt_directory.parent,
        sequence=2,
        previous_journal_sha256=first_journal,
        attempt_number=1,
        record_count=0,
        record_head_sha256=parent_metadata_sha,
        checkpoint_head_sha256=parent_metadata_sha,
        terminal_head_sha256=parent_terminal_sha,
        verdict=HarnessVerdict.CLEANUP_FAILED,
    )
    _write_test_journal_entry(
        parent.attempt_directory.parent,
        sequence=3,
        previous_journal_sha256=second_journal,
        attempt_number=2,
        record_count=0,
        record_head_sha256=successor_metadata_sha,
        checkpoint_head_sha256=successor_metadata_sha,
    )

    with pytest.raises(HarnessFailure, match='identity|metadata'):
        AppendOnlyAttemptLedger.load_read_only(
            successor.attempt_directory,
            successor.identity,
            strict_identity=False,
        )
