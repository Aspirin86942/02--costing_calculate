from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tests.rust_oracle.workbook_compare import compare_workbooks, values_equal


def test_numeric_strings_are_not_equal_to_numbers() -> None:
    assert not values_equal('00123', 123)
    assert not values_equal('2025', 2025)


def test_numbers_use_decimal_tolerance() -> None:
    assert values_equal(1, 1.0000001)


def test_compare_workbooks_detects_value_mismatch(tmp_path: Path) -> None:
    expected = tmp_path / 'expected.xlsx'
    actual = tmp_path / 'actual.xlsx'
    _write_workbook(expected, '00123')
    _write_workbook(actual, 123)

    report = compare_workbooks(expected, actual)

    assert not report['passed']
    assert report['errors'] == ["value mismatch Sheet!1,1: expected='00123', actual=123.0"]


def test_compare_workbooks_rejects_forbidden_sheet(tmp_path: Path) -> None:
    expected = tmp_path / 'expected.xlsx'
    actual = tmp_path / 'actual.xlsx'
    _write_workbook(expected, 'ok')
    _write_workbook(actual, 'ok', extra_sheet='成本分析产品维度')

    report = compare_workbooks(expected, actual)

    assert not report['passed']
    assert 'actual workbook contains forbidden product dimension sheet' in report['errors']


def test_compare_workbooks_detects_shape_mismatch(tmp_path: Path) -> None:
    expected = tmp_path / 'expected.xlsx'
    actual = tmp_path / 'actual.xlsx'
    _write_workbook(expected, 'ok')
    _write_workbook(actual, 'ok', second_value='extra')

    report = compare_workbooks(expected, actual)

    assert not report['passed']
    assert report['errors'] == ['shape mismatch Sheet: expected=1x1, actual=1x2']


def test_compare_workbooks_detects_sheet_metadata_mismatch(tmp_path: Path) -> None:
    expected = tmp_path / 'expected.xlsx'
    actual = tmp_path / 'actual.xlsx'
    _write_workbook(expected, 'ok', freeze_panes='A2', auto_filter='A1:A1')
    _write_workbook(actual, 'ok')

    report = compare_workbooks(expected, actual)

    assert not report['passed']
    assert 'freeze panes mismatch Sheet: expected=A2, actual=None' in report['errors']
    assert 'auto filter mismatch Sheet: expected=A1:A1, actual=None' in report['errors']


def test_compare_workbooks_detects_number_format_width_and_header_style_mismatch(tmp_path: Path) -> None:
    expected = tmp_path / 'expected.xlsx'
    actual = tmp_path / 'actual.xlsx'
    _write_workbook(expected, 'Header', second_row_value=1, styled=True)
    _write_workbook(actual, 'Header', second_row_value=1)

    report = compare_workbooks(expected, actual)

    assert not report['passed']
    assert any(error.startswith('column widths mismatch Sheet') for error in report['errors'])
    assert any(error.startswith('number formats mismatch Sheet') for error in report['errors'])
    assert any(error.startswith('header styles mismatch Sheet') for error in report['errors'])


def test_compare_workbooks_detects_data_cell_style_mismatch(tmp_path: Path) -> None:
    expected = tmp_path / 'expected.xlsx'
    actual = tmp_path / 'actual.xlsx'
    _write_workbook(expected, 'Header', second_row_value=1, data_cell_styled=True)
    _write_workbook(actual, 'Header', second_row_value=1)

    report = compare_workbooks(expected, actual)

    assert not report['passed']
    assert any(error.startswith('data styles mismatch Sheet') for error in report['errors'])


def test_compare_workbooks_normalizes_explicit_and_inherited_blank_cell_styles(tmp_path: Path) -> None:
    expected = tmp_path / 'expected.xlsx'
    actual = tmp_path / 'actual.xlsx'
    _write_blank_numeric_style_workbook(expected, explicit_blank_style=True)
    _write_blank_numeric_style_workbook(actual, explicit_blank_style=False)

    report = compare_workbooks(expected, actual)

    assert report == {'passed': True, 'errors': []}


def _write_workbook(
    path: Path,
    value: object,
    *,
    second_value: object | None = None,
    second_row_value: object | None = None,
    extra_sheet: str | None = None,
    freeze_panes: str | None = None,
    auto_filter: str | None = None,
    styled: bool = False,
    data_cell_styled: bool = False,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet'
    sheet['A1'] = value
    if second_row_value is not None:
        sheet['A2'] = second_row_value
    if second_value is not None:
        sheet['B1'] = second_value
    if freeze_panes is not None:
        sheet.freeze_panes = freeze_panes
    if auto_filter is not None:
        sheet.auto_filter.ref = auto_filter
    if styled:
        sheet.column_dimensions['A'].width = 15
        sheet['A1'].font = Font(bold=True)
        sheet['A1'].fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin')
        sheet['A1'].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet['A2'].number_format = '#,##0.00'
    if data_cell_styled:
        thin = Side(style='thin')
        sheet['A2'].alignment = Alignment(horizontal='right', vertical='center')
        sheet['A2'].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if extra_sheet is not None:
        workbook.create_sheet(extra_sheet)
    workbook.save(path)


def _write_blank_numeric_style_workbook(path: Path, *, explicit_blank_style: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet'
    sheet['A1'] = '数值列'
    sheet['B1'] = '标记列'
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['B2'] = '保留数据行'
    sheet.column_dimensions['A'].width = 15
    if explicit_blank_style:
        sheet['A2'].number_format = '#,##0.00'
    else:
        sheet.column_dimensions['A'].number_format = '#,##0.00'
    workbook.save(path)
