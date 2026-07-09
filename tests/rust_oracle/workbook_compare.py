from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

FORBIDDEN_SHEETS = {'成本分析产品维度'}
DECIMAL_TOLERANCE = Decimal('0.000001')
MAX_ERRORS = 20


def compare_workbooks(expected_path: Path, actual_path: Path) -> dict[str, Any]:
    expected = load_workbook(expected_path, data_only=False)
    actual = load_workbook(actual_path, data_only=False)
    errors: list[str] = []
    try:
        if FORBIDDEN_SHEETS.intersection(actual.sheetnames):
            errors.append('actual workbook contains forbidden product dimension sheet')
        if expected.sheetnames != actual.sheetnames:
            errors.append(f'sheet names differ: expected={expected.sheetnames}, actual={actual.sheetnames}')

        for sheet_name in expected.sheetnames:
            if sheet_name not in actual.sheetnames:
                continue
            expected_ws = expected[sheet_name]
            actual_ws = actual[sheet_name]
            if expected_ws.max_row != actual_ws.max_row or expected_ws.max_column != actual_ws.max_column:
                errors.append(
                    f'shape mismatch {sheet_name}: '
                    f'expected={expected_ws.max_row}x{expected_ws.max_column}, '
                    f'actual={actual_ws.max_row}x{actual_ws.max_column}'
                )
                continue
            if expected_ws.freeze_panes != actual_ws.freeze_panes:
                errors.append(f'freeze panes mismatch {sheet_name}')
            if expected_ws.auto_filter.ref != actual_ws.auto_filter.ref:
                errors.append(f'auto filter mismatch {sheet_name}')
            for row in range(1, expected_ws.max_row + 1):
                for col in range(1, expected_ws.max_column + 1):
                    expected_value = expected_ws.cell(row, col).value
                    actual_value = actual_ws.cell(row, col).value
                    if not values_equal(expected_value, actual_value):
                        errors.append(
                            f'value mismatch {sheet_name}!{row},{col}: '
                            f'expected={expected_value!r}, actual={actual_value!r}'
                        )
                        if len(errors) >= MAX_ERRORS:
                            return {'passed': False, 'errors': errors}
        return {'passed': not errors, 'errors': errors}
    finally:
        expected.close()
        actual.close()


def values_equal(expected: object, actual: object) -> bool:
    if is_blank(expected) and is_blank(actual):
        return True
    if is_number_like(expected) and is_number_like(actual):
        expected_decimal = as_decimal(expected)
        actual_decimal = as_decimal(actual)
        if expected_decimal is not None and actual_decimal is not None:
            return abs(expected_decimal - actual_decimal) <= DECIMAL_TOLERANCE
    return expected == actual


def is_blank(value: object) -> bool:
    return value is None or value == ''


def is_number_like(value: object) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def as_decimal(value: object) -> Decimal | None:
    if is_blank(value):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
