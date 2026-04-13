"""测试主 ETL 输出与基础行为。"""

import logging
from decimal import Decimal
from pathlib import Path
from unittest.mock import Mock, patch

import pandas as pd
import polars as pl
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.analytics.contracts import (
    AnalysisArtifacts,
    ConditionalFormatRule,
    FactBundle,
    FlatSheet,
    ProductAnomalySection,
    QualityMetric,
    SheetModel,
    WorkbookPayload,
)
from src.analytics.presentation_builder import build_sheet_models, dataframe_to_sheet_model
from src.analytics.qty_enricher import build_report_artifacts
from src.etl.costing_etl import CostingWorkbookETL
from src.excel.fast_writer import FastSheetWriter
from src.excel.workbook_writer import CostingWorkbookWriter
from tests.contracts._workbook_contract_helper import extract_highlight_semantics


def _build_header_map(worksheet, header_row: int = 1) -> dict[str, int]:
    return {
        worksheet.cell(header_row, col_idx).value: col_idx
        for col_idx in range(1, worksheet.max_column + 1)
        if worksheet.cell(header_row, col_idx).value is not None
    }


def _find_title_row(worksheet, title: str) -> int:
    for row_idx in range(1, worksheet.max_row + 1):
        if worksheet.cell(row_idx, 1).value == title:
            return row_idx
    raise AssertionError(f'未找到标题行: {title}')


def _assert_header_columns_fixed_width(
    worksheet,
    header_map: dict[str, int],
    *,
    expected_width: float = 15.0,
) -> None:
    dimensions = list(worksheet.column_dimensions.values())
    for col_idx in header_map.values():
        width = None
        for dimension in dimensions:
            if dimension.width is None:
                continue
            if dimension.min <= col_idx <= dimension.max:
                width = float(dimension.width)
                break
        if width is None:
            column_letter = get_column_letter(col_idx)
            fallback_width = worksheet.column_dimensions[column_letter].width
            width = None if fallback_width is None else float(fallback_width)
        assert width == expected_width


def _build_workbook_payload_from_artifacts(
    *,
    detail_df: pd.DataFrame | pl.DataFrame,
    artifacts: AnalysisArtifacts,
    stage_timings: dict[str, float] | None = None,
) -> WorkbookPayload:
    error_log = artifacts.error_log.copy()
    for column_name in error_log.columns:
        error_log[column_name] = pd.Series(
            [_normalize_error_log_value(value) for value in error_log[column_name].tolist()],
            dtype='object',
        )
    sheet_models = build_sheet_models(
        detail_df=detail_df,
        qty_sheet_df=artifacts.qty_sheet_df,
        fact_bundle=artifacts.fact_bundle,
        work_order_sheet=artifacts.work_order_sheet,
        product_anomaly_sections=artifacts.product_anomaly_sections,
    )
    return WorkbookPayload(
        sheet_models=sheet_models,
        quality_metrics=artifacts.quality_metrics,
        error_log_count=len(artifacts.error_log),
        stage_timings=stage_timings
        or {
            'ingest': 1.0,
            'normalize': 2.0,
            'fact': 3.0,
            'analysis': 4.0,
            'presentation': 5.0,
        },
        error_log_export=error_log,
    )


def _normalize_error_log_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value, 'f')
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    return str(value)


class TestCostingWorkbookETL:
    """测试 CostingWorkbookETL 类。"""

    def test_etl_initialization(self) -> None:
        """测试 ETL 初始化。"""
        etl = CostingWorkbookETL(skip_rows=2)
        assert etl.skip_rows == 2
        assert hasattr(etl, 'process_file')
        assert hasattr(etl, 'FILL_COLS')
        assert hasattr(etl, 'DETAIL_COLS')
        assert hasattr(etl, 'QTY_COLS')

    def test_standalone_cost_items_are_normalized(self) -> None:
        """Ensure standalone cost items are stripped and empty entries removed."""
        etl = CostingWorkbookETL(
            skip_rows=2,
            standalone_cost_items=(' 委外加工费 ', '', '  软件费用  '),
        )
        assert etl.standalone_cost_items == ('委外加工费', '软件费用')

    def test_process_file_not_found(self) -> None:
        """测试文件不存在时返回 False。"""
        etl = CostingWorkbookETL(skip_rows=2)
        assert etl.process_file(Path('missing.xlsx'), Path('output.xlsx')) is False

    def test_auto_rename_columns(self) -> None:
        """测试列名自动识别。"""
        etl = CostingWorkbookETL(skip_rows=2)
        df = pd.DataFrame(columns=['物料编码', '成本项目', '其他列'])

        col_map = etl._auto_rename_columns(df)

        assert col_map['物料编码'] == '子项物料编码'
        assert col_map['成本项目'] == '成本项目名称'

    def test_remove_total_rows(self) -> None:
        """测试剔除合计行。"""
        etl = CostingWorkbookETL(skip_rows=2)
        df = pd.DataFrame({'年期': ['2024年01期', '合计', '2024年03期'], '数据': [1, 2, 3]})

        result = etl._remove_total_rows(df)

        assert len(result) == 2
        assert '合计' not in result['年期'].tolist()

    def test_forward_fill_with_rules_skip_vendor_for_integrated_workshop(self) -> None:
        """测试集成车间下供应商字段不向下填充。"""
        etl = CostingWorkbookETL(skip_rows=2)
        df_raw = pd.DataFrame(
            {
                '成本中心名称': ['集成车间', None],
                '产品编码': ['P001', None],
                '供应商编码': ['V001', None],
                '供应商名称': ['供应商A', None],
            }
        )

        result = etl._forward_fill_with_rules(df_raw)

        assert result.loc[1, '成本中心名称'] == '集成车间'
        assert result.loc[1, '产品编码'] == 'P001'
        assert pd.isna(result.loc[1, '供应商编码'])
        assert pd.isna(result.loc[1, '供应商名称'])

    def test_filter_fact_df_for_analysis_uses_whitelist_order(self) -> None:
        """测试分析白名单过滤和顺序。"""
        etl = CostingWorkbookETL(skip_rows=2)
        fact_df = pd.DataFrame(
            [
                {
                    'period': '2025-01',
                    'product_code': 'GB_C.D.B0041AA',
                    'product_name': 'BMS-1100W驱动器',
                    'cost_bucket': 'direct_material',
                    'amount': 220,
                    'qty': 20,
                    'price': 11,
                },
                {
                    'period': '2025-01',
                    'product_code': 'GB_C.D.B0040AA',
                    'product_name': 'BMS-750W驱动器',
                    'cost_bucket': 'direct_material',
                    'amount': 100,
                    'qty': 10,
                    'price': 10,
                },
                {
                    'period': '2025-01',
                    'product_code': 'P001',
                    'product_name': '产品A',
                    'cost_bucket': 'direct_material',
                    'amount': 100,
                    'qty': 10,
                    'price': 10,
                },
            ]
        )

        result = etl._filter_fact_df_for_analysis(fact_df)

        assert result['product_code'].tolist() == ['GB_C.D.B0040AA', 'GB_C.D.B0041AA']

    def test_filter_fact_df_for_analysis_allows_empty_product_order(self) -> None:
        """测试显式空白白名单会跳过过滤，不使用默认顺序。"""
        etl = CostingWorkbookETL(skip_rows=2, product_order=())
        fact_df = pd.DataFrame(
            [
                {
                    'period': '2025-01',
                    'product_code': 'CUSTOM-A',
                    'product_name': '产品A',
                    'cost_bucket': 'direct_material',
                    'amount': 100,
                    'qty': 5,
                    'price': 20,
                },
                {
                    'period': '2025-01',
                    'product_code': 'CUSTOM-B',
                    'product_name': '产品B',
                    'cost_bucket': 'direct_material',
                    'amount': 200,
                    'qty': 10,
                    'price': 20,
                },
            ]
        )

        result = etl._filter_fact_df_for_analysis(fact_df)

        pd.testing.assert_frame_equal(result, fact_df)

    def test_filter_fact_df_for_analysis_honors_injected_product_order(self) -> None:
        """测试注入的产品顺序会被用于分析过滤与排序。"""
        custom_order = (
            ('CUSTOM-002', '产品乙'),
            ('CUSTOM-001', '产品甲'),
        )
        etl = CostingWorkbookETL(skip_rows=2, product_order=custom_order)
        fact_df = pd.DataFrame(
            [
                {
                    'period': '2025-01',
                    'product_code': 'CUSTOM-001',
                    'product_name': '产品甲',
                    'cost_bucket': 'direct_material',
                    'amount': 100,
                    'qty': 10,
                    'price': 10,
                },
                {
                    'period': '2025-01',
                    'product_code': 'CUSTOM-002',
                    'product_name': '产品乙',
                    'cost_bucket': 'direct_material',
                    'amount': 200,
                    'qty': 20,
                    'price': 10,
                },
                {
                    'period': '2025-01',
                    'product_code': 'CUSTOM-003',
                    'product_name': '产品丙',
                    'cost_bucket': 'direct_material',
                    'amount': 300,
                    'qty': 30,
                    'price': 10,
                },
            ]
        )

        result = etl._filter_fact_df_for_analysis(fact_df)

        assert result['product_code'].tolist() == ['CUSTOM-002', 'CUSTOM-001']


def test_workbook_writer_routes_hot_sheets_to_fast_writer(tmp_path) -> None:
    """热点 sheet 应路由到 write_dataframe_fast。"""
    workbook_writer = CostingWorkbookWriter()
    output_path = tmp_path / 'routed.xlsx'

    detail_df = pd.DataFrame([{'本期完工单位成本': 10.0, '本期完工金额': 100.0}])
    qty_sheet_df = pd.DataFrame(
        [
            {
                '本期完工金额': 165.0,
                '本期完工直接材料合计完工金额': 100.0,
            }
        ]
    )
    work_order_sheet = FlatSheet(data=pd.DataFrame([{'月份': '2025年01期'}]), column_types={'月份': 'text'})

    with (
        patch.object(workbook_writer.sheet_writer, 'write_dataframe_fast') as fast_writer_mock,
        patch.object(workbook_writer.sheet_writer, 'write_dataframe_sheet') as dataframe_writer_mock,
    ):
        workbook_writer.write_workbook(
            output_path,
            detail_df=detail_df,
            qty_sheet_df=qty_sheet_df,
            analysis_tables={},
            work_order_sheet=work_order_sheet,
            product_anomaly_sections=[],
        )

    assert [call.args[1] for call in fast_writer_mock.call_args_list] == ['成本明细', '产品数量统计']
    dataframe_writer_mock.assert_not_called()


def test_process_file_uses_workbook_payload_and_logs_all_new_stage_timings(caplog, tmp_path: Path) -> None:
    caplog.set_level(logging.INFO)
    etl = CostingWorkbookETL(skip_rows=2, product_order=())
    payload = WorkbookPayload(
        sheet_models=(
            SheetModel(
                sheet_name='成本明细',
                columns=('产品编码',),
                rows_factory=lambda: iter([('P001',)]),
                column_types={'产品编码': 'text'},
                number_formats={},
            ),
        ),
        quality_metrics=(),
        error_log_count=0,
        stage_timings={
            'ingest': 1.0,
            'normalize': 2.0,
            'fact': 3.0,
            'analysis': 4.0,
            'presentation': 5.0,
        },
    )

    with (
        patch.object(etl.pipeline, 'build_workbook_payload', return_value=payload) as payload_mock,
        patch.object(etl.workbook_writer, 'write_workbook_from_models') as writer_mock,
    ):
        assert etl.process_file(tmp_path / 'input.xlsx', tmp_path / 'output.xlsx') is True

    payload_mock.assert_called_once()
    writer_mock.assert_called_once()
    messages = [record.message for record in caplog.records]
    assert any('Timing | stage=ingest' in message for message in messages)
    assert any('Timing | stage=normalize' in message for message in messages)
    assert any('Timing | stage=fact' in message for message in messages)
    assert any('Timing | stage=analysis' in message for message in messages)
    assert any('Timing | stage=presentation' in message for message in messages)
    assert any('Timing | stage=export' in message for message in messages)


def test_workbook_writer_can_export_sheet_models_with_conditional_formats(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_models.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = (
        SheetModel(
            sheet_name='按工单按产品异常值分析',
            columns=('直接材料单位完工成本', '直接材料异常标记'),
            rows_factory=lambda: iter([(18.0, '关注')]),
            column_types={'直接材料单位完工成本': 'price', '直接材料异常标记': 'text'},
            number_formats={'直接材料单位完工成本': '#,##0.00'},
            freeze_panes='A2',
            auto_filter=True,
            fixed_width=15.0,
            conditional_formats=(
                ConditionalFormatRule(
                    target_range='A2:A1048576',
                    formula='=$B2="关注"',
                    format_key='attention',
                ),
            ),
        ),
    )

    writer.write_workbook_from_models(output_path, sheet_models=sheet_models)

    workbook = load_workbook(output_path)
    worksheet = workbook['按工单按产品异常值分析']
    assert worksheet.freeze_panes == 'A2'
    assert worksheet['A2'].number_format == '#,##0.00'
    assert worksheet.conditional_formatting


def test_build_sheet_models_avoids_pyarrow_dependency_for_pandas_inputs() -> None:
    detail_df = pd.DataFrame([{'月份': '2025年01期', '产品编码': 'P001'}])
    qty_sheet_df = pd.DataFrame(
        [{'月份': '2025年01期', '产品编码': 'P001', '本期完工数量': 10.0, '本期完工金额': 100.0}]
    )
    work_order_sheet = FlatSheet(
        data=pd.DataFrame(
            [{'月份': '2025年01期', '产品编码': 'P001', '直接材料单位完工成本': 10.0, '直接材料异常标记': '关注'}]
        ),
        column_types={'月份': 'text', '产品编码': 'text', '直接材料单位完工成本': 'price', '直接材料异常标记': 'text'},
    )
    product_sections = [
        ProductAnomalySection(
            product_code='P001',
            product_name='产品A',
            data=pd.DataFrame([{'月份': '2025年01期', '总成本': 100.0, '完工数量': 10.0, '单位成本': 10.0}]),
            column_types={'月份': 'text', '总成本': 'amount', '完工数量': 'qty', '单位成本': 'price'},
            amount_columns=['总成本'],
            outlier_cells=set(),
        )
    ]

    with patch(
        'src.analytics.presentation_builder.pl.from_pandas',
        side_effect=AssertionError('should not use pyarrow'),
    ):
        models = build_sheet_models(
            detail_df=detail_df,
            qty_sheet_df=qty_sheet_df,
            fact_bundle=None,
            work_order_sheet=work_order_sheet,
            product_anomaly_sections=product_sections,
        )

    assert len(models) == 7
    product_model = next(model for model in models if model.sheet_name == '按产品异常值分析')
    assert product_model.freeze_panes == 'A6'
    assert list(product_model.rows_factory())[0][0:2] == ('P001', '产品A')


def test_build_sheet_models_handles_leading_nan_before_text_in_pandas_object_column() -> None:
    detail_df = pd.DataFrame([{'月份': '2025年01期', '产品编码': 'P001'}])
    qty_sheet_df = pd.DataFrame(
        [
            {'月份': '2025年01期', '成本中心名称': float('nan'), '本期完工数量': 1.0, '本期完工金额': 100.0},
            {'月份': '2025年02期', '成本中心名称': '集成检测部', '本期完工数量': 2.0, '本期完工金额': 200.0},
        ]
    )

    models = build_sheet_models(
        detail_df=detail_df,
        qty_sheet_df=qty_sheet_df,
        fact_bundle=None,
        work_order_sheet=FlatSheet(data=pd.DataFrame([{'月份': '2025年01期'}]), column_types={'月份': 'text'}),
        product_anomaly_sections=[],
    )

    qty_model = next(model for model in models if model.sheet_name == '产品数量统计')
    rows = list(qty_model.rows_factory())
    assert rows[0][1] is None
    assert rows[1][1] == '集成检测部'


def test_build_sheet_models_marks_detail_and_qty_as_fast_flat_sheets() -> None:
    detail_df = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '产品编码': 'P001',
                '本期完工单位成本': 10.0,
                '本期完工金额': 100.0,
            }
        ]
    )
    qty_sheet_df = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '产品编码': 'P001',
                '本期完工数量': 10.0,
                '本期完工金额': 100.0,
            }
        ]
    )
    work_order_sheet = FlatSheet(
        data=pd.DataFrame([{'月份': '2025年01期', '产品编码': 'P001'}]),
        column_types={'月份': 'text', '产品编码': 'text'},
    )

    models = build_sheet_models(
        detail_df=detail_df,
        qty_sheet_df=qty_sheet_df,
        fact_bundle=None,
        work_order_sheet=work_order_sheet,
        product_anomaly_sections=[],
    )

    detail_model = next(model for model in models if model.sheet_name == '成本明细')
    qty_model = next(model for model in models if model.sheet_name == '产品数量统计')
    work_order_model = next(model for model in models if model.sheet_name == '按工单按产品异常值分析')
    product_anomaly_model = next(model for model in models if model.sheet_name == '按产品异常值分析')
    analysis_models = [
        model for model in models if model.sheet_name in ('直接材料_价量比', '直接人工_价量比', '制造费用_价量比')
    ]

    assert detail_model.write_mode == 'dataframe_fast'
    assert detail_model.style_profile == 'lightweight_flat'
    assert isinstance(detail_model.source_frame, pl.DataFrame)
    assert detail_model.source_frame.to_dicts() == pl.DataFrame(
        detail_df.to_dict(orient='list'),
        strict=False,
    ).to_dicts()

    assert qty_model.write_mode == 'dataframe_fast'
    assert qty_model.style_profile == 'lightweight_flat'
    assert isinstance(qty_model.source_frame, pl.DataFrame)
    assert qty_model.source_frame.to_dicts() == pl.DataFrame(
        qty_sheet_df.to_dict(orient='list'),
        strict=False,
    ).to_dicts()

    assert work_order_model.write_mode is None
    assert work_order_model.style_profile is None
    assert work_order_model.source_frame is None
    assert product_anomaly_model.write_mode is None
    assert product_anomaly_model.style_profile is None
    assert product_anomaly_model.source_frame is None
    assert analysis_models
    for model in analysis_models:
        assert model.write_mode is None
        assert model.style_profile is None
        assert model.source_frame is None


def test_dataframe_to_sheet_model_rejects_invalid_fast_metadata() -> None:
    frame = pl.DataFrame({'a': [1]})

    with pytest.raises(ValueError, match='fast export metadata'):
        dataframe_to_sheet_model(
            sheet_name='测试',
            frame=frame,
            column_types={'a': 'text'},
            number_formats={},
            write_mode='dataframe_fast',
        )

    with pytest.raises(ValueError, match='source_frame'):
        dataframe_to_sheet_model(
            sheet_name='测试',
            frame=frame,
            column_types={'a': 'text'},
            number_formats={},
            write_mode='dataframe_fast',
            style_profile='lightweight_flat',
            source_frame=pl.DataFrame({'b': [1]}),
        )


def test_dataframe_to_sheet_model_keeps_rows_factory_in_sync_with_source_frame() -> None:
    frame = pl.DataFrame({'a': [1]})
    source_frame = pl.DataFrame({'a': [2]})

    model = dataframe_to_sheet_model(
        sheet_name='测试',
        frame=frame,
        column_types={'a': 'text'},
        number_formats={},
        write_mode='dataframe_fast',
        style_profile='lightweight_flat',
        source_frame=source_frame,
    )

    assert list(model.rows_factory()) == list(source_frame.iter_rows())


def test_sheet_model_rejects_partial_fast_metadata() -> None:
    with pytest.raises(ValueError, match='fast export metadata'):
        SheetModel(
            sheet_name='测试',
            columns=('a',),
            rows_factory=lambda: iter([(1,)]),
            column_types={'a': 'text'},
            number_formats={},
            write_mode='dataframe_fast',
        )


def test_sheet_model_fast_metadata_keeps_rows_factory_in_sync() -> None:
    source_frame = pl.DataFrame({'a': [2]})

    model = SheetModel(
        sheet_name='测试',
        columns=('a',),
        rows_factory=lambda: iter([(1,)]),
        column_types={'a': 'text'},
        number_formats={},
        write_mode='dataframe_fast',
        style_profile='lightweight_flat',
        source_frame=source_frame,
    )

    assert list(model.rows_factory()) == list(source_frame.iter_rows())


def test_workbook_writer_sheet_model_preserves_product_anomaly_legacy_layout(tmp_path: Path) -> None:
    output_path = tmp_path / 'product_anomaly_model.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = build_sheet_models(
        detail_df=pd.DataFrame([{'月份': '2025年01期', '产品编码': 'P001'}]),
        qty_sheet_df=pd.DataFrame(
            [{'月份': '2025年01期', '产品编码': 'P001', '本期完工数量': 10.0, '本期完工金额': 100.0}]
        ),
        fact_bundle=None,
        work_order_sheet=FlatSheet(data=pd.DataFrame([{'月份': '2025年01期'}]), column_types={'月份': 'text'}),
        product_anomaly_sections=[
            ProductAnomalySection(
                product_code='P001',
                product_name='产品A',
                data=pd.DataFrame([{'月份': '2025年01期', '总成本': 100.0, '完工数量': 10.0, '单位成本': 10.0}]),
                column_types={'月份': 'text', '总成本': 'amount', '完工数量': 'qty', '单位成本': 'price'},
                amount_columns=['总成本'],
                outlier_cells=set(),
            )
        ],
    )

    writer.write_workbook_from_models(output_path, sheet_models=sheet_models)

    workbook = load_workbook(output_path)
    worksheet = workbook['按产品异常值分析']
    assert worksheet['A1'].value == '四、按单个产品异常值分析'
    assert worksheet['A3'].value == '产品编码'
    assert worksheet['A4'].value == 'P001'
    assert worksheet['B3'].value == '产品名称'
    assert worksheet['B4'].value == '产品A'
    assert worksheet.freeze_panes == 'A6'


def test_sheet_model_writer_preserves_detail_and_qty_number_formats(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_models_formats.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = build_sheet_models(
        detail_df=pd.DataFrame(
            [
                {
                    '月份': '2025年01期',
                    '本期完工单位成本': 10.0,
                    '本期完工金额': 100.0,
                }
            ]
        ),
        qty_sheet_df=pd.DataFrame(
            [
                {
                    '月份': '2025年01期',
                    '本期完工金额': 165.0,
                    '本期完工直接材料合计完工金额': 100.0,
                }
            ]
        ),
        fact_bundle=None,
        work_order_sheet=FlatSheet(data=pd.DataFrame([{'月份': '2025年01期'}]), column_types={'月份': 'text'}),
        product_anomaly_sections=[],
    )

    writer.write_workbook_from_models(output_path, sheet_models=sheet_models)

    workbook = load_workbook(output_path)
    detail_sheet = workbook['成本明细']
    qty_sheet = workbook['产品数量统计']
    detail_headers = _build_header_map(detail_sheet)
    qty_headers = _build_header_map(qty_sheet)

    assert detail_sheet.cell(2, detail_headers['本期完工单位成本']).number_format == '#,##0.00'
    assert detail_sheet.cell(2, detail_headers['本期完工金额']).number_format == '#,##0.00'
    assert qty_sheet.cell(2, qty_headers['本期完工金额']).number_format == '#,##0.00'
    assert qty_sheet.cell(2, qty_headers['本期完工直接材料合计完工金额']).number_format == '#,##0.00'


def test_write_workbook_from_models_routes_hot_sheet_models_to_fast_tabular_writer(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_models_route_fast.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = (
        SheetModel(
            sheet_name='成本明细',
            columns=('文本列', '金额列'),
            rows_factory=lambda: iter([('A', 10.0)]),
            column_types={'文本列': 'text', '金额列': 'amount'},
            number_formats={'金额列': '#,##0.00'},
            write_mode='dataframe_fast',
            style_profile='lightweight_flat',
            source_frame=pl.DataFrame({'文本列': ['A'], '金额列': [10.0]}),
        ),
        SheetModel(
            sheet_name='按工单按产品异常值分析',
            columns=('直接材料单位完工成本', '直接材料异常标记'),
            rows_factory=lambda: iter([(18.0, '关注')]),
            column_types={'直接材料单位完工成本': 'price', '直接材料异常标记': 'text'},
            number_formats={'直接材料单位完工成本': '#,##0.00'},
            freeze_panes='A2',
            auto_filter=True,
            fixed_width=15.0,
        ),
    )

    with (
        patch.object(
            writer.sheet_writer,
            'write_sheet_model_as_lightweight_table',
            wraps=writer.sheet_writer.write_sheet_model_as_lightweight_table,
        ) as fast_tabular_mock,
        patch.object(
            writer.sheet_writer,
            'write_sheet_model',
            wraps=writer.sheet_writer.write_sheet_model,
        ) as generic_mock,
    ):
        writer.write_workbook_from_models(output_path, sheet_models=sheet_models)

    assert [call.args[1].sheet_name for call in fast_tabular_mock.call_args_list] == ['成本明细']
    assert [call.args[1].sheet_name for call in generic_mock.call_args_list] == ['按工单按产品异常值分析']


def test_sheet_model_fast_tabular_writer_lightweight_data_cells(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_model_lightweight_data_cells.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = (
        SheetModel(
            sheet_name='成本明细',
            columns=('文本列', '金额列'),
            rows_factory=lambda: iter([('A', 10.0)]),
            column_types={'文本列': 'text', '金额列': 'amount'},
            number_formats={'金额列': '#,##0.00'},
            freeze_panes='A2',
            auto_filter=True,
            fixed_width=15.0,
            write_mode='dataframe_fast',
            style_profile='lightweight_flat',
            source_frame=pl.DataFrame({'文本列': ['A'], '金额列': [10.0]}),
        ),
    )

    writer.write_workbook_from_models(output_path, sheet_models=sheet_models)

    workbook = load_workbook(output_path)
    worksheet = workbook['成本明细']
    assert worksheet.freeze_panes == 'A2'
    assert worksheet.auto_filter.ref == 'A1:B2'
    assert worksheet['B2'].number_format == '#,##0.00'
    _assert_header_columns_fixed_width(worksheet, _build_header_map(worksheet))
    assert worksheet['A2'].border.left.style is None
    assert worksheet['A2'].border.right.style is None
    assert worksheet['A2'].border.top.style is None
    assert worksheet['A2'].border.bottom.style is None


def test_sheet_model_fast_tabular_writer_keeps_blank_numeric_cell_format_lightweight(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_model_lightweight_blank_numeric.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = (
        SheetModel(
            sheet_name='成本明细',
            columns=('文本列', '金额列'),
            rows_factory=lambda: iter([('A', None)]),
            column_types={'文本列': 'text', '金额列': 'amount'},
            number_formats={'金额列': '#,##0.00'},
            freeze_panes='A2',
            auto_filter=True,
            fixed_width=15.0,
            write_mode='dataframe_fast',
            style_profile='lightweight_flat',
            source_frame=pl.DataFrame({'文本列': ['A'], '金额列': [None]}, strict=False),
        ),
    )

    writer.write_workbook_from_models(output_path, sheet_models=sheet_models)

    workbook = load_workbook(output_path)
    worksheet = workbook['成本明细']
    value_cell = worksheet['B2']

    assert value_cell.value is None
    assert value_cell.number_format == '#,##0.00'
    assert value_cell.alignment.horizontal == 'right'
    assert value_cell.border.left.style is None
    assert value_cell.border.right.style is None
    assert value_cell.border.top.style is None
    assert value_cell.border.bottom.style is None


def test_sheet_model_fast_tabular_writer_skips_non_blank_numeric_rewrite(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_model_lightweight_single_write.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = (
        SheetModel(
            sheet_name='成本明细',
            columns=('文本列', '金额列'),
            rows_factory=lambda: iter([('A', 10.0)]),
            column_types={'文本列': 'text', '金额列': 'amount'},
            number_formats={'金额列': '#,##0.00'},
            freeze_panes='A2',
            auto_filter=True,
            fixed_width=15.0,
            write_mode='dataframe_fast',
            style_profile='lightweight_flat',
            source_frame=pl.DataFrame({'文本列': ['A'], '金额列': [10.0]}, strict=False),
        ),
    )

    with patch('src.excel.fast_writer._write_cell') as write_cell_mock:
        writer.write_workbook_from_models(output_path, sheet_models=sheet_models)

    write_cell_mock.assert_not_called()


def test_sheet_model_fast_tabular_writer_rejects_conditional_formats(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_model_fast_reject_conditional.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = (
        SheetModel(
            sheet_name='成本明细',
            columns=('文本列', '金额列'),
            rows_factory=lambda: iter([('A', 10.0)]),
            column_types={'文本列': 'text', '金额列': 'amount'},
            number_formats={'金额列': '#,##0.00'},
            write_mode='dataframe_fast',
            style_profile='lightweight_flat',
            source_frame=pl.DataFrame({'文本列': ['A'], '金额列': [10.0]}),
            conditional_formats=(
                ConditionalFormatRule(
                    target_range='A2:A1048576',
                    formula='=$B2>0',
                    format_key='attention',
                ),
            ),
        ),
    )

    with pytest.raises(ValueError, match='conditional_formats.*成本明细'):
        writer.write_workbook_from_models(output_path, sheet_models=sheet_models)


def test_sheet_model_fast_tabular_writer_rejects_product_anomaly_special_layout(tmp_path: Path) -> None:
    output_path = tmp_path / 'sheet_model_fast_reject_product_anomaly.xlsx'
    writer = CostingWorkbookWriter()
    sheet_models = (
        SheetModel(
            sheet_name='按产品异常值分析',
            columns=('产品编码', '产品名称', '月份', '总成本'),
            rows_factory=lambda: iter([('P001', '产品A', '2025年01期', 100.0)]),
            column_types={'产品编码': 'text', '产品名称': 'text', '月份': 'text', '总成本': 'amount'},
            number_formats={'总成本': '#,##0.00'},
            write_mode='dataframe_fast',
            style_profile='lightweight_flat',
            source_frame=pl.DataFrame(
                {'产品编码': ['P001'], '产品名称': ['产品A'], '月份': ['2025年01期'], '总成本': [100.0]}
            ),
        ),
    )

    with pytest.raises(ValueError, match='按产品异常值分析'):
        writer.write_workbook_from_models(output_path, sheet_models=sheet_models)


def test_build_sheet_models_handles_fact_bundle_summary_without_pyarrow() -> None:
    fact_bundle = FactBundle(
        detail_fact=pl.DataFrame(),
        qty_fact=pl.DataFrame(),
        work_order_fact=pl.DataFrame(),
        product_summary_fact=pl.DataFrame(
            [
                {
                    'product_code': 'P001',
                    'product_name': '产品A',
                    'period_display': '2025年01期',
                    'total_cost': 100.0,
                    'completed_qty': 10.0,
                    'unit_cost': 10.0,
                    'dm_cost': 70.0,
                    'dm_unit_cost': 7.0,
                    'dm_contrib': 0.7,
                    'dl_cost': 20.0,
                    'dl_unit_cost': 2.0,
                    'dl_contrib': 0.2,
                    'moh_cost': 10.0,
                    'moh_unit_cost': 1.0,
                    'moh_contrib': 0.1,
                }
            ]
        ),
        error_fact=pl.DataFrame(),
    )

    with patch(
        'src.analytics.presentation_builder.pl.from_pandas',
        side_effect=AssertionError('should not use pyarrow'),
    ):
        models = build_sheet_models(
            detail_df=pd.DataFrame([{'月份': '2025年01期', '本期完工金额': 100.0}]),
            qty_sheet_df=pd.DataFrame([{'月份': '2025年01期', '本期完工金额': 100.0}]),
            fact_bundle=fact_bundle,
            work_order_sheet=FlatSheet(data=pd.DataFrame([{'月份': '2025年01期'}]), column_types={'月份': 'text'}),
            product_anomaly_sections=[],
        )

    direct_material_model = next(model for model in models if model.sheet_name == '直接材料_价量比')
    rows = list(direct_material_model.rows_factory())
    assert rows
    assert rows[0][0] == 'P001'
    assert rows[0][3] == 70.0


def test_write_dataframe_fast_keeps_blank_numeric_cell_format(tmp_path) -> None:
    """快路径下空数值单元格仍应保留数字格式和边框。"""
    output_path = tmp_path / 'fast_blank_numeric.xlsx'
    sheet_writer = FastSheetWriter()
    sheet_df = pd.DataFrame([{'文本列': 'A', '数值列': None}])

    with pd.ExcelWriter(
        output_path,
        engine='xlsxwriter',
        engine_kwargs={'options': {'constant_memory': True, 'strings_to_urls': False}},
    ) as writer:
        sheet_writer.write_dataframe_fast(
            writer,
            '热点sheet',
            sheet_df,
            numeric_columns={'数值列'},
            freeze_panes='A2',
            fixed_width=15,
        )

    workbook = load_workbook(output_path)
    worksheet = workbook['热点sheet']
    value_cell = worksheet['B2']

    assert value_cell.value is None
    assert value_cell.number_format == '#,##0.00'
    assert value_cell.alignment.horizontal == 'right'
    assert value_cell.border.left.style == 'thin'
    assert value_cell.border.right.style == 'thin'
    assert value_cell.border.top.style == 'thin'
    assert value_cell.border.bottom.style == 'thin'


def test_write_dataframe_fast_error_log_style_without_column_widths(tmp_path) -> None:
    """error_log 在 apply_column_widths=False 下数据区样式不退化。"""
    output_path = tmp_path / 'fast_error_log.xlsx'
    sheet_writer = FastSheetWriter()
    error_log_df = pd.DataFrame(
        [
            {
                'row_id': '1',
                'error_type': 'MISSING_AMOUNT',
                'message': 'missing amount',
            }
        ]
    )

    with pd.ExcelWriter(
        output_path,
        engine='xlsxwriter',
        engine_kwargs={'options': {'constant_memory': True, 'strings_to_urls': False}},
    ) as writer:
        sheet_writer.write_dataframe_fast(
            writer,
            'error_log',
            error_log_df,
            numeric_columns=set(),
            freeze_panes=None,
            auto_filter=False,
            apply_column_widths=False,
        )

    workbook = load_workbook(output_path)
    worksheet = workbook['error_log']
    cell = worksheet['A2']

    assert worksheet.freeze_panes is None
    assert worksheet.auto_filter.ref is None
    assert cell.alignment.horizontal == 'left'
    assert cell.border.left.style == 'thin'
    assert cell.border.right.style == 'thin'
    assert cell.border.top.style == 'thin'
    assert cell.border.bottom.style == 'thin'


def test_process_file_writes_v3_analysis_sheets(tmp_path) -> None:
    """测试 process_file 会输出 v3 相关 sheet 与基础样式。"""
    etl = CostingWorkbookETL(skip_rows=2)
    input_path = tmp_path / 'input.xlsx'
    output_path = tmp_path / 'output.xlsx'

    df_detail = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接材料',
                '本期完工单位成本': 10,
                '本期完工金额': 100,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接人工',
                '本期完工单位成本': 2,
                '本期完工金额': 20,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '制造费用-人工',
                '本期完工单位成本': 3,
                '本期完工金额': 30,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '委外加工费',
                '本期完工单位成本': 1.5,
                '本期完工金额': 15,
            },
        ]
    )
    df_qty = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '本期完工数量': 10,
                '本期完工金额': 165,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-02',
                '工单编号': 'WO-002',
                '工单行号': 1,
                '基本单位': 'PCS',
                '本期完工数量': 0,
                '本期完工金额': 100,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-03',
                '工单编号': 'WO-003',
                '工单行号': 1,
                '基本单位': 'PCS',
                '本期完工数量': 5,
                '本期完工金额': None,
            },
        ]
    )

    artifacts = build_report_artifacts(df_detail, df_qty, standalone_cost_items=etl.standalone_cost_items)
    payload = _build_workbook_payload_from_artifacts(detail_df=df_detail, artifacts=artifacts)

    with patch.object(etl.pipeline, 'build_workbook_payload', return_value=payload):
        assert etl.process_file(input_path, output_path) is True

    xls = pd.ExcelFile(output_path, engine='openpyxl')
    expected_sheets = {
        '成本明细',
        '产品数量统计',
        '直接材料_价量比',
        '直接人工_价量比',
        '制造费用_价量比',
        '按工单按产品异常值分析',
        '按产品异常值分析',
    }
    assert set(xls.sheet_names) == expected_sheets
    assert len(xls.sheet_names) == 7

    wb = load_workbook(output_path)
    ws_detail = wb['成本明细']
    detail_headers = _build_header_map(ws_detail)
    assert ws_detail.freeze_panes == 'A2'
    assert ws_detail.cell(2, detail_headers['本期完工单位成本']).number_format == '#,##0.00'
    assert ws_detail.cell(2, detail_headers['本期完工金额']).number_format == '#,##0.00'
    assert isinstance(ws_detail.cell(2, detail_headers['本期完工单位成本']).value, (int, float))

    ws_qty = wb['产品数量统计']
    qty_headers = _build_header_map(ws_qty)
    qty_decimal_columns = [
        '本期完工金额',
        '本期完工直接材料合计完工金额',
        '本期完工直接人工合计完工金额',
        '本期完工制造费用合计完工金额',
        '本期完工制造费用_其他合计完工金额',
        '本期完工制造费用_人工合计完工金额',
        '本期完工制造费用_机物料及低耗合计完工金额',
        '本期完工制造费用_折旧合计完工金额',
        '本期完工制造费用_水电费合计完工金额',
        '本期完工委外加工费合计完工金额',
        '直接材料单位完工金额',
        '直接人工单位完工金额',
        '制造费用单位完工金额',
        '制造费用_其他单位完工成本',
        '制造费用_人工单位完工成本',
        '制造费用_机物料及低耗单位完工成本',
        '制造费用_折旧单位完工成本',
        '制造费用_水电费单位完工成本',
        '委外加工费单位完工成本',
    ]
    assert ws_qty.freeze_panes == 'A2'
    assert ws_qty.max_row == 2
    assert '完工数量是否有效' not in qty_headers
    assert '完工数量是否小于等于0' not in qty_headers
    assert '是否存在空值' not in qty_headers
    for column_name in qty_decimal_columns:
        cell = ws_qty.cell(2, qty_headers[column_name])
        assert cell.number_format == '#,##0.00'
        assert isinstance(cell.value, (int, float))

    ws_price = wb['直接材料_价量比']
    price_headers = _build_header_map(ws_price)
    assert ws_price['A1'].value == '产品编码'
    assert ws_price.freeze_panes == 'C3'
    assert ws_price.auto_filter.ref is not None
    assert ws_price.cell(2, price_headers['直接材料成本']).number_format == '#,##0.00'
    assert ws_price.cell(2, price_headers['完工数量']).number_format == '#,##0.00'
    assert ws_price.cell(2, price_headers['单位直接材料成本']).number_format == '#,##0.00'
    assert isinstance(ws_price.cell(2, price_headers['直接材料成本']).value, (int, float))

    ws_work_order = wb['按工单按产品异常值分析']
    assert ws_work_order['A1'].value == '月份'
    assert ws_work_order['I2'].value == 10
    work_order_headers = _build_header_map(ws_work_order)
    assert '委外加工费合计完工金额' in work_order_headers
    assert '委外加工费单位完工成本' in work_order_headers
    assert '委外加工费异常标记' not in work_order_headers
    assert ws_work_order.max_row == 2
    assert ws_work_order.freeze_panes == 'A2'
    assert ws_work_order.auto_filter.ref is not None

    ws_product = wb['按产品异常值分析']
    assert ws_product['A1'].value == '四、按单个产品异常值分析'
    assert ws_product['A3'].value == '产品编码'
    assert ws_product['A4'].value == 'GB_C.D.B0040AA'
    assert ws_product.freeze_panes == 'A6'
    assert 'error_log' not in wb.sheetnames

    quality_metrics = {metric.metric: metric.value for metric in etl.last_quality_metrics}
    assert '本期完工数量缺失率' not in quality_metrics
    assert '本期完工金额缺失率' not in quality_metrics
    assert '完工数量小于等于0行数' not in quality_metrics
    assert str(quality_metrics['产品数量统计输出行数']) == '1'
    assert str(quality_metrics['工单异常分析输出行数']) == '1'
    assert str(quality_metrics['因完工数量无效被过滤行数']) == '1'
    assert str(quality_metrics['因总完工成本为空被过滤行数']) == '1'
    assert etl.last_error_log_count == len(etl.last_error_log_frame)
    assert not etl.last_error_log_frame.empty
    assert 'issue_type' in etl.last_error_log_frame.columns


def test_process_file_logs_new_payload_stage_timings(caplog, tmp_path) -> None:
    """process_file 应输出 payload 路径各阶段与导出耗时日志。"""
    caplog.set_level(logging.INFO, logger='src.etl.costing_etl')
    etl = CostingWorkbookETL(skip_rows=2)
    payload = WorkbookPayload(
        sheet_models=(
            SheetModel(
                sheet_name='成本明细',
                columns=('产品编码',),
                rows_factory=lambda: iter([('P001',)]),
                column_types={'产品编码': 'text'},
                number_formats={},
            ),
        ),
        quality_metrics=(),
        error_log_count=0,
        stage_timings={
            'ingest': 1.0,
            'normalize': 2.0,
            'fact': 3.0,
            'analysis': 4.0,
            'presentation': 5.0,
        },
    )

    with patch.object(etl.pipeline, 'build_workbook_payload', return_value=payload):
        assert etl.process_file(tmp_path / 'input.xlsx', tmp_path / 'output.xlsx') is True

    messages = [record.message for record in caplog.records]
    assert any('Timing | stage=ingest' in message for message in messages)
    assert any('Timing | stage=normalize' in message for message in messages)
    assert any('Timing | stage=fact' in message for message in messages)
    assert any('Timing | stage=analysis' in message for message in messages)
    assert any('Timing | stage=presentation' in message for message in messages)
    assert any('Timing | stage=export' in message for message in messages)


def test_lightweight_export_writes_workbook_skeleton(tmp_path) -> None:
    """轻量导出骨架：仍写出7张sheet，关键明细页保留A2冻结和数值格式。"""
    etl = CostingWorkbookETL(skip_rows=2)
    input_path = tmp_path / 'input.xlsx'
    output_path = tmp_path / 'output.xlsx'

    df_detail = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接材料',
                '本期完工单位成本': 10.0,
                '本期完工金额': 100.0,
            }
        ]
    )
    df_qty = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '本期完工数量': 10.0,
                '本期完工金额': 165.0,
            }
        ]
    )

    artifacts = build_report_artifacts(df_detail, df_qty, standalone_cost_items=etl.standalone_cost_items)
    payload = _build_workbook_payload_from_artifacts(detail_df=df_detail, artifacts=artifacts)

    with patch.object(etl.pipeline, 'build_workbook_payload', return_value=payload):
        assert etl.process_file(input_path, output_path) is True

    xls = pd.ExcelFile(output_path, engine='openpyxl')
    assert len(xls.sheet_names) == 7
    assert set(xls.sheet_names) == {
        '成本明细',
        '产品数量统计',
        '直接材料_价量比',
        '直接人工_价量比',
        '制造费用_价量比',
        '按工单按产品异常值分析',
        '按产品异常值分析',
    }

    wb = load_workbook(output_path)
    ws_detail = wb['成本明细']
    detail_headers = _build_header_map(ws_detail)
    assert ws_detail.freeze_panes == 'A2'
    assert ws_detail.cell(2, detail_headers['本期完工单位成本']).number_format == '#,##0.00'
    assert ws_detail.cell(2, detail_headers['本期完工金额']).number_format == '#,##0.00'
    _assert_header_columns_fixed_width(ws_detail, detail_headers)

    ws_qty = wb['产品数量统计']
    qty_headers = _build_header_map(ws_qty)
    assert ws_qty.freeze_panes == 'A2'
    assert ws_qty.cell(2, qty_headers['本期完工金额']).number_format == '#,##0.00'
    assert ws_qty.cell(2, qty_headers['本期完工直接材料合计完工金额']).number_format == '#,##0.00'
    _assert_header_columns_fixed_width(ws_qty, qty_headers)


def test_process_file_writes_work_order_conditional_format_rules(tmp_path) -> None:
    """测试工单异常页会写出条件格式规则，而不是直接回填单元格颜色。"""
    etl = CostingWorkbookETL(skip_rows=2)
    input_path = tmp_path / 'input.xlsx'
    output_path = tmp_path / 'output.xlsx'

    df_detail = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接材料',
                '本期完工金额': 100,
            }
        ]
    )
    df_qty = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '本期完工数量': 10,
                '本期完工金额': 100,
            }
        ]
    )
    work_order_df = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心': '中心A',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行': '1',
                '基本单位': 'PCS',
                '本期完工数量': 10,
                '直接材料单位完工成本': 18.0,
                '直接人工单位完工成本': 2.0,
                '制造费用单位完工成本': 3.0,
                '制造费用_其他单位完工成本': 0.0,
                '制造费用_人工单位完工成本': 30.0,
                '制造费用_机物料及低耗单位完工成本': 0.0,
                '制造费用_折旧单位完工成本': 0.0,
                '制造费用_水电费单位完工成本': 0.0,
                '直接材料异常标记': '关注',
                '直接人工异常标记': '正常',
                '制造费用异常标记': '正常',
                '制造费用_其他异常标记': '正常',
                '制造费用_人工异常标记': '高度可疑',
                '制造费用_机物料及低耗异常标记': '正常',
                '制造费用_折旧异常标记': '正常',
                '制造费用_水电费异常标记': '正常',
            }
        ]
    )
    work_order_column_types = {
        '月份': 'text',
        '成本中心': 'text',
        '产品编码': 'text',
        '产品名称': 'text',
        '规格型号': 'text',
        '工单编号': 'text',
        '工单行': 'text',
        '基本单位': 'text',
        '本期完工数量': 'qty',
        '直接材料单位完工成本': 'price',
        '直接人工单位完工成本': 'price',
        '制造费用单位完工成本': 'price',
        '制造费用_其他单位完工成本': 'price',
        '制造费用_人工单位完工成本': 'price',
        '制造费用_机物料及低耗单位完工成本': 'price',
        '制造费用_折旧单位完工成本': 'price',
        '制造费用_水电费单位完工成本': 'price',
        '直接材料异常标记': 'text',
        '直接人工异常标记': 'text',
        '制造费用异常标记': 'text',
        '制造费用_其他异常标记': 'text',
        '制造费用_人工异常标记': 'text',
        '制造费用_机物料及低耗异常标记': 'text',
        '制造费用_折旧异常标记': 'text',
        '制造费用_水电费异常标记': 'text',
    }
    artifacts = AnalysisArtifacts(
        fact_df=pd.DataFrame(
            [
                {
                    'period': '2025-01',
                    'product_code': 'GB_C.D.B0040AA',
                    'product_name': 'BMS-750W驱动器',
                    'cost_bucket': 'direct_material',
                    'amount': 100,
                    'qty': 10,
                    'price': 10,
                }
            ]
        ),
        qty_sheet_df=df_qty.copy(),
        work_order_sheet=FlatSheet(data=work_order_df, column_types=work_order_column_types),
        product_anomaly_sections=[
            ProductAnomalySection(
                product_code='GB_C.D.B0040AA',
                product_name='BMS-750W驱动器',
                data=pd.DataFrame([{'月份': '2025年01期', '总成本': 100.0, '完工数量': 10, '单位成本': 10.0}]),
                column_types={'月份': 'text', '总成本': 'amount', '完工数量': 'qty', '单位成本': 'price'},
                amount_columns=['总成本'],
                outlier_cells=set(),
            )
        ],
        quality_metrics=(
            QualityMetric(
                category='行数勾稽',
                metric='样例',
                value='1',
                description='测试',
            ),
        ),
        error_log=pd.DataFrame(),
    )
    payload = _build_workbook_payload_from_artifacts(detail_df=df_detail, artifacts=artifacts)

    with patch.object(etl.pipeline, 'build_workbook_payload', return_value=payload):
        assert etl.process_file(input_path, output_path) is True

    highlight_semantics = extract_highlight_semantics(output_path)

    assert {
        'sqref': 'J2:J1048576',
        'formula': ['=EXACT($R2,UNICHAR(20851)&UNICHAR(27880))'],
        'fill': 'DDEBF7',
        'font': None,
    } in highlight_semantics['rules']
    assert {
        'sqref': 'R2:R1048576',
        'formula': ['=EXACT($R2,UNICHAR(20851)&UNICHAR(27880))'],
        'fill': 'DDEBF7',
        'font': None,
    } in highlight_semantics['rules']
    assert {
        'sqref': 'N2:N1048576',
        'formula': ['=EXACT($V2,UNICHAR(39640)&UNICHAR(24230)&UNICHAR(21487)&UNICHAR(30097))'],
        'fill': '4472C4',
        'font': 'FFFFFF',
    } in highlight_semantics['rules']
    assert {
        'sqref': 'V2:V1048576',
        'formula': ['=EXACT($V2,UNICHAR(39640)&UNICHAR(24230)&UNICHAR(21487)&UNICHAR(30097))'],
        'fill': '4472C4',
        'font': 'FFFFFF',
    } in highlight_semantics['rules']


def test_process_file_passes_standalone_cost_items_to_pipeline_payload_builder(tmp_path) -> None:
    """process_file 调用 payload 编排时应透传 standalone_cost_items。"""
    etl = CostingWorkbookETL(skip_rows=2, product_order=(), standalone_cost_items=('委外加工费', '软件费用'))
    payload = WorkbookPayload(
        sheet_models=(
            SheetModel(
                sheet_name='成本明细',
                columns=('产品编码',),
                rows_factory=lambda: iter([('DP.C.P0197AA',)]),
                column_types={'产品编码': 'text'},
                number_formats={},
            ),
        ),
        quality_metrics=(),
        error_log_count=0,
        stage_timings={},
    )

    with patch.object(etl.pipeline, 'build_workbook_payload', return_value=payload) as payload_mock:
        assert etl.process_file(tmp_path / 'input.xlsx', tmp_path / 'output.xlsx') is True

    assert payload_mock.call_count == 1
    assert payload_mock.call_args.kwargs['standalone_cost_items'] == ('委外加工费', '软件费用')


def test_process_file_filters_whitelist_before_presentation_and_preserves_numeric_order_line_sort(tmp_path) -> None:
    etl = CostingWorkbookETL(skip_rows=2)
    input_path = tmp_path / 'input.xlsx'
    output_path = tmp_path / 'output.xlsx'

    detail_df = pd.DataFrame([{'月份': '2025年01期', '产品编码': 'GB_C.D.B0040AA', '本期完工金额': 100.0}])
    qty_df = pd.DataFrame([{'月份': '2025年01期', '产品编码': 'GB_C.D.B0040AA', '本期完工金额': 100.0}])
    work_order_df = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '工单编号': 'WO-001',
                '工单行': 10,
            },
            {
                '月份': '2025年01期',
                '产品编码': 'GB_C.D.B0040AA',
                '产品名称': 'BMS-750W驱动器',
                '工单编号': 'WO-001',
                '工单行': 2,
            },
            {
                '月份': '2025年01期',
                '产品编码': 'CUSTOM-001',
                '产品名称': '非白名单产品',
                '工单编号': 'WO-002',
                '工单行': 1,
            },
        ]
    )
    artifacts = AnalysisArtifacts(
        fact_df=pd.DataFrame(
            [
                {
                    'period': '2025-01',
                    'product_code': 'GB_C.D.B0040AA',
                    'product_name': 'BMS-750W驱动器',
                    'cost_bucket': 'direct_material',
                    'amount': 100.0,
                    'qty': 10.0,
                    'price': 10.0,
                },
                {
                    'period': '2025-01',
                    'product_code': 'CUSTOM-001',
                    'product_name': '非白名单产品',
                    'cost_bucket': 'direct_material',
                    'amount': 50.0,
                    'qty': 5.0,
                    'price': 10.0,
                },
            ]
        ),
        qty_sheet_df=qty_df,
        work_order_sheet=FlatSheet(
            data=work_order_df,
            column_types={'月份': 'text', '产品编码': 'text', '产品名称': 'text', '工单编号': 'text', '工单行': 'text'},
        ),
        product_anomaly_sections=[
            ProductAnomalySection(
                product_code='GB_C.D.B0040AA',
                product_name='BMS-750W驱动器',
                data=pd.DataFrame([{'月份': '2025年01期', '总成本': 100.0}]),
                column_types={'月份': 'text', '总成本': 'amount'},
                amount_columns=['总成本'],
                outlier_cells=set(),
            ),
            ProductAnomalySection(
                product_code='CUSTOM-001',
                product_name='非白名单产品',
                data=pd.DataFrame([{'月份': '2025年01期', '总成本': 50.0}]),
                column_types={'月份': 'text', '总成本': 'amount'},
                amount_columns=['总成本'],
                outlier_cells=set(),
            ),
        ],
        quality_metrics=(),
        error_log=pd.DataFrame(),
        fact_bundle=FactBundle(
            detail_fact=pl.DataFrame(),
            qty_fact=pl.DataFrame(),
            work_order_fact=pl.DataFrame(),
            product_summary_fact=pl.DataFrame(
                [
                    {
                        'product_code': 'GB_C.D.B0040AA',
                        'product_name': 'BMS-750W驱动器',
                        'period': '2025-01',
                        'period_display': '2025年01期',
                        'total_cost': 100.0,
                        'completed_qty': 10.0,
                        'unit_cost': 10.0,
                        'dm_cost': 100.0,
                        'dm_unit_cost': 10.0,
                        'dm_contrib': 1.0,
                        'dl_cost': 0.0,
                        'dl_unit_cost': 0.0,
                        'dl_contrib': 0.0,
                        'moh_cost': 0.0,
                        'moh_unit_cost': 0.0,
                        'moh_contrib': 0.0,
                    },
                    {
                        'product_code': 'CUSTOM-001',
                        'product_name': '非白名单产品',
                        'period': '2025-01',
                        'period_display': '2025年01期',
                        'total_cost': 50.0,
                        'completed_qty': 5.0,
                        'unit_cost': 10.0,
                        'dm_cost': 50.0,
                        'dm_unit_cost': 10.0,
                        'dm_contrib': 1.0,
                        'dl_cost': 0.0,
                        'dl_unit_cost': 0.0,
                        'dl_contrib': 0.0,
                        'moh_cost': 0.0,
                        'moh_unit_cost': 0.0,
                        'moh_contrib': 0.0,
                    },
                ]
            ),
            error_fact=pl.DataFrame(),
        ),
    )

    with (
        patch.object(etl.pipeline, 'load_raw_workbook_frame', return_value=Mock()),
        patch.object(etl.pipeline, 'build_normalized_cost_frame', return_value=Mock()),
        patch.object(etl.pipeline, 'split_normalized_frames', return_value=Mock(detail_df=detail_df, qty_df=qty_df)),
        patch('src.etl.pipeline.build_report_artifacts', return_value=artifacts),
        patch(
            'src.etl.pipeline.build_sheet_models',
            return_value=(
                SheetModel(
                    sheet_name='成本明细',
                    columns=('产品编码',),
                    rows_factory=lambda: iter([('GB_C.D.B0040AA',)]),
                    column_types={'产品编码': 'text'},
                    number_formats={},
                ),
            ),
        ) as build_models_mock,
        patch.object(etl.workbook_writer, 'write_workbook_from_models'),
    ):
        assert etl.process_file(input_path, output_path) is True

    work_order_arg = build_models_mock.call_args.kwargs['work_order_sheet'].data
    assert work_order_arg['产品编码'].tolist() == ['GB_C.D.B0040AA', 'GB_C.D.B0040AA']
    assert work_order_arg['工单行'].tolist() == [2, 10]
    product_summary_arg = build_models_mock.call_args.kwargs['fact_bundle'].product_summary_fact
    assert product_summary_arg['product_code'].to_list() == ['GB_C.D.B0040AA']
    product_sections_arg = build_models_mock.call_args.kwargs['product_anomaly_sections']
    assert [(section.product_code, section.product_name) for section in product_sections_arg] == [
        ('GB_C.D.B0040AA', 'BMS-750W驱动器')
    ]


def test_process_file_sk_workbook_renders_software_fee_columns_without_polluting_gb(tmp_path) -> None:
    """真实链路校验：SK 应输出软件费用列，GB 不应输出软件费用列。"""
    input_path = tmp_path / 'input.xlsx'
    sk_output_path = tmp_path / 'sk_output.xlsx'
    gb_output_path = tmp_path / 'gb_output.xlsx'

    df_detail = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'DP.C.P0197AA',
                '产品名称': '动力线',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接材料',
                '本期完工金额': 80.0,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'DP.C.P0197AA',
                '产品名称': '动力线',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '直接人工',
                '本期完工金额': 20.0,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'DP.C.P0197AA',
                '产品名称': '动力线',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '制造费用-人工',
                '本期完工金额': 10.0,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'DP.C.P0197AA',
                '产品名称': '动力线',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '委外加工费',
                '本期完工金额': 30.0,
            },
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'DP.C.P0197AA',
                '产品名称': '动力线',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '成本项目名称': '软件费用',
                '本期完工金额': 40.0,
            },
        ]
    )
    df_qty = pd.DataFrame(
        [
            {
                '月份': '2025年01期',
                '成本中心名称': '中心A',
                '产品编码': 'DP.C.P0197AA',
                '产品名称': '动力线',
                '规格型号': 'S-01',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '基本单位': 'PCS',
                '本期完工数量': 10.0,
                '本期完工金额': 180.0,
            }
        ]
    )

    etl_sk = CostingWorkbookETL(skip_rows=2, product_order=(), standalone_cost_items=('委外加工费', '软件费用'))
    sk_artifacts = build_report_artifacts(df_detail, df_qty, standalone_cost_items=etl_sk.standalone_cost_items)
    sk_payload = _build_workbook_payload_from_artifacts(detail_df=df_detail, artifacts=sk_artifacts)
    with patch.object(etl_sk.pipeline, 'build_workbook_payload', return_value=sk_payload):
        assert etl_sk.process_file(input_path, sk_output_path) is True

    sk_wb = load_workbook(sk_output_path)
    sk_qty_ws = sk_wb['产品数量统计']
    sk_qty_headers = _build_header_map(sk_qty_ws)
    assert '本期完工软件费用合计完工金额' in sk_qty_headers
    assert '软件费用单位完工成本' in sk_qty_headers
    assert sk_qty_ws.cell(2, sk_qty_headers['本期完工软件费用合计完工金额']).number_format == '#,##0.00'
    assert sk_qty_ws.cell(2, sk_qty_headers['软件费用单位完工成本']).number_format == '#,##0.00'

    sk_work_order_ws = sk_wb['按工单按产品异常值分析']
    sk_work_order_headers = _build_header_map(sk_work_order_ws)
    assert '软件费用合计完工金额' in sk_work_order_headers
    assert '软件费用单位完工成本' in sk_work_order_headers
    assert '软件费用异常标记' not in sk_work_order_headers
    assert 'log_软件费用单位完工成本' not in sk_work_order_headers
    assert 'Modified Z-score_软件费用' not in sk_work_order_headers

    etl_gb = CostingWorkbookETL(skip_rows=2, product_order=(), standalone_cost_items=('委外加工费',))
    gb_artifacts = build_report_artifacts(df_detail, df_qty, standalone_cost_items=etl_gb.standalone_cost_items)
    gb_payload = _build_workbook_payload_from_artifacts(detail_df=df_detail, artifacts=gb_artifacts)
    with patch.object(etl_gb.pipeline, 'build_workbook_payload', return_value=gb_payload):
        assert etl_gb.process_file(input_path, gb_output_path) is True

    gb_wb = load_workbook(gb_output_path)
    gb_qty_ws = gb_wb['产品数量统计']
    gb_qty_headers = _build_header_map(gb_qty_ws)
    assert '本期完工软件费用合计完工金额' not in gb_qty_headers
    assert '软件费用单位完工成本' not in gb_qty_headers

    gb_work_order_ws = gb_wb['按工单按产品异常值分析']
    gb_work_order_headers = _build_header_map(gb_work_order_ws)
    assert '软件费用合计完工金额' not in gb_work_order_headers
    assert '软件费用单位完工成本' not in gb_work_order_headers
