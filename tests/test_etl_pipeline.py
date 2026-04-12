"""测试 ETL 阶段 pipeline 契约。"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import get_args, get_origin, get_type_hints
from unittest.mock import patch

import pandas as pd
import polars as pl
from openpyxl import Workbook

from src.analytics.contracts import (
    AnalysisArtifacts,
    ConditionalFormatRule,
    FactBundle,
    FlatSheet,
    NormalizedCostFrame,
    RawWorkbookFrame,
    ResolvedColumns,
    SheetModel,
    SplitResult,
    WorkbookPayload,
)
from src.etl.costing_etl import CostingWorkbookETL
from src.etl.stages.normalizer import build_normalized_cost_frame
from src.etl.stages.reader import load_raw_workbook
from src.etl.stages.splitter import split_normalized_frames
from src.etl.stages.workbook_ingestor import WorkbookIngestor


def test_polars_pipeline_contract_objects_are_constructible() -> None:
    raw = RawWorkbookFrame(
        sheet_name='成本计算单',
        header_rows=(('年期', '产品编码'), ('', '')),
        frame=pl.DataFrame({'column_0': ['2025年1期'], 'column_1': ['P001']}),
    )
    normalized = NormalizedCostFrame(
        frame=pl.DataFrame({'年期': ['2025-01'], '产品编码': ['P001']}),
        key_columns=('年期', '产品编码'),
    )
    model = SheetModel(
        sheet_name='成本明细',
        columns=('年期', '产品编码'),
        rows_factory=lambda: iter([('2025-01', 'P001')]),
        column_types={'年期': 'text', '产品编码': 'text'},
        number_formats={},
        freeze_panes='A2',
        auto_filter=True,
        fixed_width=15.0,
        conditional_formats=(
            ConditionalFormatRule(
                target_range='A2:A1048576',
                formula='=$B2="高度可疑"',
                format_key='suspicious',
            ),
        ),
    )
    payload = WorkbookPayload(sheet_models=(model,), quality_metrics=(), error_log_count=0, stage_timings={})

    assert raw.sheet_name == '成本计算单'
    assert normalized.key_columns == ('年期', '产品编码')
    assert list(model.rows_factory()) == [('2025-01', 'P001')]
    assert payload.sheet_models[0].sheet_name == '成本明细'


def test_fact_bundle_is_constructible_and_exposes_expected_frames() -> None:
    bundle = FactBundle(
        detail_fact=pl.DataFrame({'detail_id': ['D-001']}),
        qty_fact=pl.DataFrame({'qty_id': ['Q-001']}),
        work_order_fact=pl.DataFrame({'work_order_id': ['WO-001']}),
        product_summary_fact=pl.DataFrame({'product_id': ['P-001']}),
        error_fact=pl.DataFrame({'error_code': ['MISSING_AMOUNT']}),
    )

    assert bundle.detail_fact.columns == ['detail_id']
    assert bundle.qty_fact.item(0, 'qty_id') == 'Q-001'
    assert bundle.work_order_fact.item(0, 'work_order_id') == 'WO-001'
    assert bundle.product_summary_fact.item(0, 'product_id') == 'P-001'
    assert bundle.error_fact.item(0, 'error_code') == 'MISSING_AMOUNT'


def test_analysis_artifacts_remains_constructible_without_fact_bundle() -> None:
    artifacts = AnalysisArtifacts(
        fact_df=pd.DataFrame({'产品编码': ['P001']}),
        qty_sheet_df=pd.DataFrame({'产品编码': ['P001']}),
        work_order_sheet=FlatSheet(
            data=pd.DataFrame({'工单编号': ['WO-001']}),
            column_types={'工单编号': 'text'},
        ),
        product_anomaly_sections=[],
        quality_metrics=(),
        error_log=pd.DataFrame({'错误码': ['MISSING_AMOUNT']}),
    )

    assert artifacts.fact_bundle is None
    assert artifacts.work_order_sheet.column_types == {'工单编号': 'text'}
    assert artifacts.error_log.iloc[0]['错误码'] == 'MISSING_AMOUNT'


def test_analysis_artifacts_accepts_fact_bundle_without_breaking_existing_fields() -> None:
    bundle = FactBundle(
        detail_fact=pl.DataFrame({'detail_id': ['D-001']}),
        qty_fact=pl.DataFrame({'qty_id': ['Q-001']}),
        work_order_fact=pl.DataFrame({'work_order_id': ['WO-001']}),
        product_summary_fact=pl.DataFrame({'product_id': ['P-001']}),
        error_fact=pl.DataFrame({'error_code': ['MISSING_AMOUNT']}),
    )
    artifacts = AnalysisArtifacts(
        fact_df=pd.DataFrame({'产品编码': ['P001']}),
        qty_sheet_df=pd.DataFrame({'产品编码': ['P001']}),
        work_order_sheet=FlatSheet(
            data=pd.DataFrame({'工单编号': ['WO-001']}),
            column_types={'工单编号': 'text'},
        ),
        product_anomaly_sections=[],
        quality_metrics=(),
        error_log=pd.DataFrame({'错误码': ['MISSING_AMOUNT']}),
        fact_bundle=bundle,
    )

    assert artifacts.fact_bundle is bundle
    assert artifacts.fact_df.iloc[0]['产品编码'] == 'P001'
    assert artifacts.fact_bundle.error_fact.item(0, 'error_code') == 'MISSING_AMOUNT'


def test_pipeline_resolve_columns_returns_resolved_columns_contract() -> None:
    etl = CostingWorkbookETL(skip_rows=2)
    df = pd.DataFrame(columns=['物料编码', '成本项目', '其他列'])

    resolved = etl.pipeline.resolve_columns(df)

    assert isinstance(resolved, ResolvedColumns)
    assert resolved.child_material_column == '子项物料编码'
    assert resolved.cost_item_column == '成本项目名称'
    assert resolved.rename_map == {'物料编码': '子项物料编码', '成本项目': '成本项目名称'}


def test_pipeline_split_sheets_returns_split_result_and_keeps_month_column() -> None:
    etl = CostingWorkbookETL(skip_rows=2)
    df_raw = pd.DataFrame(
        [
            {
                '年期': '2025年1期',
                '成本中心名称': '中心A',
                '产品编码': 'P001',
                '产品名称': '产品A',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '子项物料编码': '',
                '成本项目名称': '',
                '本期完工数量': 10,
                '本期完工金额': 100,
            },
            {
                '年期': '2025年1期',
                '成本中心名称': '中心A',
                '产品编码': 'P001',
                '产品名称': '产品A',
                '工单编号': 'WO-001',
                '工单行号': 1,
                '子项物料编码': 'MAT-001',
                '成本项目名称': '直接材料',
                '本期完工数量': 10,
                '本期完工金额': 100,
            },
        ]
    )
    df_filled = df_raw.copy()
    df_filled['Filled_成本项目'] = df_filled['成本项目名称'].replace('', pd.NA).ffill()

    split_result = etl.pipeline.split_sheets(df_raw, df_filled)

    assert isinstance(split_result, SplitResult)
    assert split_result.qty_df.columns.tolist()[0:2] == ['年期', '月份']
    assert split_result.detail_df.columns.tolist()[0:2] == ['年期', '月份']
    assert split_result.qty_df.iloc[0]['工单编号'] == 'WO-001'
    assert split_result.detail_df.iloc[0]['成本项目名称'] == '直接材料'


def test_workbook_ingestor_falls_back_once_when_fast_reader_fails(tmp_path: Path) -> None:
    ingestor = WorkbookIngestor()
    fallback = RawWorkbookFrame(
        sheet_name='Sheet1',
        header_rows=(('年期', '产品编码'), ('', '')),
        frame=pl.DataFrame({'column_0': ['2025年1期'], 'column_1': ['P001']}),
    )

    with (
        patch.object(ingestor, '_load_with_calamine', side_effect=RuntimeError('boom')),
        patch.object(ingestor, '_load_with_openpyxl', return_value=fallback) as fallback_mock,
    ):
        result = ingestor.load(tmp_path / 'input.xlsx', skip_rows=2)

    assert result.sheet_name == 'Sheet1'
    fallback_mock.assert_called_once()


def test_load_raw_workbook_delegates_to_workbook_ingestor(tmp_path: Path) -> None:
    raw = RawWorkbookFrame(
        sheet_name='Sheet1',
        header_rows=(('年期', '产品编码'), ('', '')),
        frame=pl.DataFrame({'column_0': ['2025年1期'], 'column_1': ['P001']}),
    )

    with patch.object(WorkbookIngestor, 'load', return_value=raw) as load_mock:
        result = load_raw_workbook(tmp_path / 'input.xlsx', skip_rows=2)

    assert result is raw
    load_mock.assert_called_once_with(tmp_path / 'input.xlsx', skip_rows=2)


def test_workbook_ingestor_calamine_fast_path_preserves_contract(tmp_path: Path) -> None:
    workbook_path = tmp_path / 'fast.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'FastSheet'
    worksheet.append(['metadata'])
    worksheet.append(['metadata2'])
    worksheet.append(['年期', '产品编码'])
    worksheet.append(['', '产品名称'])
    worksheet.append(['2026年3期', 'P002'])
    workbook.save(workbook_path)

    ingestor = WorkbookIngestor()
    result = ingestor.load(workbook_path, skip_rows=2)

    assert result.sheet_name == 'FastSheet'
    assert result.header_rows == (('年期', '产品编码'), ('', '产品名称'))
    assert result.frame.columns == ['column_0', 'column_1']
    assert result.frame.row(0) == ('2026年3期', 'P002')


def test_workbook_ingestor_calamine_fast_path_handles_late_blank_cells_without_fallback(tmp_path: Path) -> None:
    class _FakeSheet:
        name = 'FastSheet'

        def to_python(self, skip_empty_area: bool = False) -> list[list[object]]:
            return [
                ['metadata'],
                ['metadata2'],
                ['本期完工数量'],
                ['数量'],
                *[[index] for index in range(1, 151)],
                [''],
            ]

    class _FakeWorkbook:
        def get_sheet_by_index(self, index: int) -> _FakeSheet:
            assert index == 0
            return _FakeSheet()

    ingestor = WorkbookIngestor()
    with (
        patch('src.etl.stages.workbook_ingestor.CalamineWorkbook.from_path', return_value=_FakeWorkbook()),
        patch.object(ingestor, '_load_with_openpyxl', side_effect=AssertionError('should not fallback')),
    ):
        result = ingestor.load(tmp_path / 'input.xlsx', skip_rows=2)

    assert result.sheet_name == 'FastSheet'
    assert result.header_rows == (('本期完工数量',), ('数量',))
    assert result.frame.columns == ['column_0']
    assert result.frame.row(0) == (1,)
    assert result.frame.row(149) == (150,)
    assert result.frame.row(150) == (None,)


def test_workbook_ingestor_openpyxl_fallback_preserves_sheet_name_and_headers(tmp_path: Path, caplog) -> None:
    workbook_path = tmp_path / 'fallback.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'FallbackSheet'
    worksheet.append(['metadata'])
    worksheet.append(['unused'])
    worksheet.append(['主表头', '产品编码'])
    worksheet.append(['次表头', '产品名称'])
    worksheet.append(['2025年1期', 'P001'])
    workbook.save(workbook_path)

    ingestor = WorkbookIngestor()
    caplog.set_level(logging.WARNING, logger=WorkbookIngestor.__module__)
    with patch.object(ingestor, '_load_with_calamine', side_effect=RuntimeError('boom')):
        result = ingestor.load(workbook_path, skip_rows=2)

    assert result.sheet_name == 'FallbackSheet'
    assert result.header_rows == (('主表头', '产品编码'), ('次表头', '产品名称'))
    assert isinstance(result.frame, pl.DataFrame)
    assert result.frame.columns == ['column_0', 'column_1']
    assert result.frame.row(0) == ('2025年1期', 'P001')
    assert any(
        'falling back to openpyxl' in record.message and record.levelno >= logging.WARNING
        for record in caplog.records
    )


def test_workbook_ingestor_openpyxl_fallback_normalizes_nan_in_text_columns(tmp_path: Path) -> None:
    workbook_path = tmp_path / 'fallback-with-null.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'FallbackSheet'
    worksheet.append(['metadata'])
    worksheet.append(['unused'])
    worksheet.append(['年期', '产品编码'])
    worksheet.append(['月份', '产品名称'])
    worksheet.append(['2025年1期', 'P001'])
    worksheet.append(['2025年2期', None])
    workbook.save(workbook_path)

    ingestor = WorkbookIngestor()
    with patch.object(ingestor, '_load_with_calamine', side_effect=RuntimeError('boom')):
        result = ingestor.load(workbook_path, skip_rows=2)

    assert result.sheet_name == 'FallbackSheet'
    assert result.frame.columns == ['column_0', 'column_1']
    assert result.frame.row(0) == ('2025年1期', 'P001')
    assert result.frame.row(1) == ('2025年2期', None)


def test_workbook_ingestor_openpyxl_fallback_allows_mixed_int_and_float_columns(tmp_path: Path) -> None:
    workbook_path = tmp_path / 'fallback-with-mixed-numeric.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'FallbackSheet'
    worksheet.append(['metadata'])
    worksheet.append(['unused'])
    worksheet.append(['年期', '本期完工数量'])
    worksheet.append(['月份', '数量'])
    worksheet.append(['2025年1期', 1])
    worksheet.append(['2025年2期', 0.8459])
    workbook.save(workbook_path)

    ingestor = WorkbookIngestor()
    with patch.object(ingestor, '_load_with_calamine', side_effect=RuntimeError('boom')):
        result = ingestor.load(workbook_path, skip_rows=2)

    assert result.frame.columns == ['column_0', 'column_1']
    assert result.frame.row(0) == ('2025年1期', 1.0)
    assert result.frame.row(1) == ('2025年2期', 0.8459)


def test_build_normalized_cost_frame_removes_totals_and_skips_integrated_vendor_fill() -> None:
    raw = RawWorkbookFrame(
        sheet_name='成本计算单',
        header_rows=(
            (
                '年期',
                '成本中心名称',
                '产品编码',
                '供应商编码',
                '成本项目名称',
                '工单编号',
                '子项物料编码',
                '本期完工金额',
            ),
            ('', '', '', '', '', '', '', ''),
        ),
        frame=pl.DataFrame(
            {
                'column_0': ['2025年1期', '2025年1期', '合计'],
                'column_1': ['集成车间', '中心A', '中心A'],
                'column_2': ['P001', 'P001', 'P001'],
                'column_3': ['V001', None, 'V999'],
                'column_4': [None, '直接材料', '直接材料'],
                'column_5': ['WO-001', 'WO-001', 'WO-TOTAL'],
                'column_6': [None, 'MAT-001', None],
                'column_7': [None, 100, 100],
            }
        ),
    )

    normalized = build_normalized_cost_frame(
        raw,
        child_material_column='子项物料编码',
        cost_item_column='成本项目名称',
        period_column='年期',
        fill_columns=['年期', '成本中心名称', '产品编码', '供应商编码'],
        vendor_columns=['供应商编码'],
        cost_center_column='成本中心名称',
        integrated_workshop_name='集成车间',
    )

    rows = normalized.frame.select(['月份', '成本中心名称', '供应商编码']).to_dicts()
    assert rows == [
        {'月份': '2025年01期', '成本中心名称': '集成车间', '供应商编码': 'V001'},
        {'月份': '2025年01期', '成本中心名称': '中心A', '供应商编码': None},
    ]


def test_split_normalized_frames_keeps_qty_and_detail_contracts() -> None:
    normalized = build_normalized_cost_frame(
        RawWorkbookFrame(
            sheet_name='成本计算单',
            header_rows=(
                (
                    '年期',
                    '产品编码',
                    '产品名称',
                    '工单编号',
                    '工单行号',
                    '子项物料编码',
                    '成本项目名称',
                    '本期完工数量',
                    '本期完工金额',
                ),
                ('', '', '', '', '', '', '', '', ''),
            ),
            frame=pl.DataFrame(
                {
                    'column_0': ['2025年1期', '2025年1期'],
                    'column_1': ['P001', 'P001'],
                    'column_2': ['产品A', '产品A'],
                    'column_3': ['WO-001', 'WO-001'],
                    'column_4': [1, 1],
                    'column_5': [None, 'MAT-001'],
                    'column_6': [None, '直接材料'],
                    'column_7': [10, 10],
                    'column_8': [100, 100],
                }
            ),
        ),
        child_material_column='子项物料编码',
        cost_item_column='成本项目名称',
        period_column='年期',
        fill_columns=['年期', '产品编码', '产品名称', '工单编号', '工单行号'],
        vendor_columns=[],
        cost_center_column='成本中心名称',
        integrated_workshop_name='集成车间',
    )

    split = split_normalized_frames(
        normalized,
        child_material_column='子项物料编码',
        cost_item_column='成本项目名称',
        order_number_column='工单编号',
        filled_cost_item_column='Filled_成本项目',
        qty_columns=['年期', '月份', '产品编码', '工单编号', '本期完工数量', '本期完工金额'],
        detail_columns=['年期', '月份', '产品编码', '工单编号', '成本项目名称', '本期完工金额'],
    )

    assert split.qty_df.columns == ['年期', '月份', '产品编码', '工单编号', '本期完工数量', '本期完工金额']
    assert split.detail_df.columns == ['年期', '月份', '产品编码', '工单编号', '成本项目名称', '本期完工金额']
    assert split.qty_df.height == 1
    assert split.detail_df.height == 1
    assert isinstance(split.qty_df, pl.DataFrame)
    assert isinstance(split.detail_df, pl.DataFrame)


def test_split_result_contract_accepts_pandas_and_polars_frames() -> None:
    type_hints = get_type_hints(SplitResult)
    detail_type = type_hints['detail_df']
    qty_type = type_hints['qty_df']

    assert get_origin(detail_type) is not None
    assert get_origin(qty_type) is not None
    assert pd.DataFrame in get_args(detail_type)
    assert pl.DataFrame in get_args(detail_type)
    assert pd.DataFrame in get_args(qty_type)
    assert pl.DataFrame in get_args(qty_type)


def test_pipeline_load_raw_dataframe_keeps_legacy_pandas_contract(tmp_path: Path) -> None:
    legacy_df = pd.DataFrame(
        [['2025年1期', 'P001']],
        columns=pd.Index([('年期', ''), ('产品编码', '产品名称')]),
    )
    etl = CostingWorkbookETL(skip_rows=2)

    with patch('src.etl.pipeline.pd.read_excel', return_value=legacy_df):
        loaded = etl.pipeline.load_raw_dataframe(tmp_path / 'input.xlsx')

    assert isinstance(loaded, pd.DataFrame)
    assert list(loaded.columns) == [('年期', ''), ('产品编码', '产品名称')]
    assert loaded.iloc[0, 0] == '2025年1期'


def test_pipeline_remove_total_rows_pandas_path_does_not_call_pl_from_pandas() -> None:
    etl = CostingWorkbookETL(skip_rows=2)
    frame = pd.DataFrame({'年期': ['2025年01期', '合计'], '成本中心名称': ['中心A', '中心A']})

    with patch('src.etl.pipeline.pl.from_pandas', side_effect=AssertionError('should not be called')):
        result = etl.pipeline.remove_total_rows(frame)

    assert isinstance(result, pd.DataFrame)
    assert result['年期'].tolist() == ['2025年01期']


def test_pipeline_forward_fill_with_rules_pandas_path_does_not_call_pl_from_pandas() -> None:
    etl = CostingWorkbookETL(skip_rows=2)
    frame = pd.DataFrame(
        {
            '成本中心名称': ['集成车间', None],
            '产品编码': ['P001', None],
            '供应商编码': ['V001', None],
            '供应商名称': ['供应商A', None],
        }
    )

    with patch('src.etl.pipeline.pl.from_pandas', side_effect=AssertionError('should not be called')):
        result = etl.pipeline.forward_fill_with_rules(frame)

    assert isinstance(result, pd.DataFrame)
    assert result.loc[1, '成本中心名称'] == '集成车间'
    assert result.loc[1, '产品编码'] == 'P001'
    assert pd.isna(result.loc[1, '供应商编码'])
    assert pd.isna(result.loc[1, '供应商名称'])
