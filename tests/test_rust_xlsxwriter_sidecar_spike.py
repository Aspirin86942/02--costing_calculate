from __future__ import annotations

import importlib.util
import json
import shutil
import subprocess
import sys
from pathlib import Path

import polars as pl
import pytest
from openpyxl import Workbook

from src.analytics.contracts import SheetModel

SPIKE_DIR = Path(__file__).resolve().parents[1] / 'spikes' / '001-rust-xlsxwriter-sidecar'
PYTHON_DIR = SPIKE_DIR / 'python'


def _load_spike_module(module_name: str):
    module_path = PYTHON_DIR / f'{module_name}.py'
    python_dir_text = str(PYTHON_DIR)
    if python_dir_text not in sys.path:
        sys.path.insert(0, python_dir_text)
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _cargo_available() -> bool:
    return shutil.which('cargo') is not None


def _sheet_model(
    *,
    sheet_name: str = '成本计算单总表',
    source_frame: pl.DataFrame | None = None,
) -> SheetModel:
    frame = source_frame or pl.DataFrame({'月份': ['2025年01期'], '金额': [12.5]})
    return SheetModel(
        sheet_name=sheet_name,
        columns=tuple(frame.columns),
        rows_factory=lambda: frame.iter_rows(),
        column_types={'月份': 'text', '金额': 'amount'},
        number_formats={'金额': '#,##0.00'},
        write_mode='dataframe_fast',
        style_profile='lightweight_flat',
        source_frame=frame,
    )


def test_export_sheet_models_to_payload_uses_polars_csv_and_manifest(tmp_path: Path) -> None:
    sidecar_payload = _load_spike_module('sidecar_payload')
    model = _sheet_model()

    result = sidecar_payload.export_sheet_models_to_payload(
        (model,),
        tmp_path,
        output_path=tmp_path / 'rust_output.xlsx',
    )

    assert result.manifest_path == tmp_path / 'manifest.json'
    assert result.manifest['workbook']['output_path'] == str(tmp_path / 'rust_output.xlsx')
    assert result.manifest['sheets'][0]['sheet_name'] == '成本计算单总表'
    assert result.manifest['sheets'][0]['columns'] == ['月份', '金额']
    assert result.manifest['sheets'][0]['column_types']['金额'] == 'amount'
    assert result.manifest['sheets'][0]['number_formats']['金额'] == '#,##0.00'
    assert result.manifest['sheets'][0]['write_types'] == {'月份': 'text', '金额': 'number'}
    assert result.manifest['sheets'][0]['csv_export_method'] == 'polars'
    assert result.manifest['sheets'][0]['intermediate_export_seconds'] >= 0
    csv_path = Path(result.manifest['sheets'][0]['csv_path'])
    assert csv_path.exists()
    assert csv_path.read_text(encoding='utf-8').splitlines() == ['月份,金额', '2025年01期,12.5']

    persisted_manifest = json.loads(result.manifest_path.read_text(encoding='utf-8'))
    assert persisted_manifest == result.manifest


def test_export_sheet_models_to_payload_rejects_product_dimension_sheet(tmp_path: Path) -> None:
    sidecar_payload = _load_spike_module('sidecar_payload')
    product_model = _sheet_model(sheet_name='成本分析产品维度')

    with pytest.raises(ValueError, match='成本分析产品维度'):
        sidecar_payload.export_sheet_models_to_payload(
            (product_model,),
            tmp_path,
            output_path=tmp_path / 'rust_output.xlsx',
        )


def test_export_sheet_models_to_payload_records_polars_source_dtypes(tmp_path: Path) -> None:
    sidecar_payload = _load_spike_module('sidecar_payload')
    source_frame = pl.DataFrame({'产品编码': ['A001'], '计划产量': [70.0]})
    model = SheetModel(
        sheet_name='成本计算单总表',
        columns=tuple(source_frame.columns),
        rows_factory=lambda: source_frame.iter_rows(),
        column_types={'产品编码': 'text', '计划产量': 'text'},
        number_formats={},
        write_mode='dataframe_fast',
        style_profile='lightweight_flat',
        source_frame=source_frame,
    )

    result = sidecar_payload.export_sheet_models_to_payload(
        (model,),
        tmp_path,
        output_path=tmp_path / 'rust_output.xlsx',
    )

    assert result.manifest['sheets'][0]['source_dtypes'] == {'产品编码': 'String', '计划产量': 'Float64'}
    assert result.manifest['sheets'][0]['write_types'] == {'产品编码': 'text', '计划产量': 'number'}


def test_validate_workbooks_normalizes_empty_and_numeric_cells(tmp_path: Path) -> None:
    sidecar_validation = _load_spike_module('sidecar_validation')
    manifest = {
        'sheets': [
            {
                'sheet_name': '成本计算单总表',
                'columns': ['文本', '金额'],
                'column_types': {'文本': 'text', '金额': 'amount'},
                'number_formats': {'金额': '#,##0.00'},
                'write_types': {'文本': 'text', '金额': 'number'},
                'freeze_panes': 'A2',
                'auto_filter': True,
                'row_count': 1,
                'column_count': 2,
            }
        ]
    }
    python_path = tmp_path / 'python.xlsx'
    rust_path = tmp_path / 'rust.xlsx'

    for path, text_value, amount_value in (
        (python_path, None, 1.0),
        (rust_path, '', 1.0000000000001),
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '成本计算单总表'
        worksheet.append(['文本', '金额'])
        worksheet.append([text_value, amount_value])
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = 'A1:B2'
        worksheet['B2'].number_format = '#,##0.00'
        workbook.save(path)

    report = sidecar_validation.validate_workbooks(python_path, rust_path, manifest)

    assert report['passed'] is True
    assert report['sheet_count'] == 1
    assert report['checked_cells'] == 4
    assert report['number_format_matched'] is True
    assert report['shape_matched'] is True
    assert report['auto_filter_matched'] is True


def test_validate_workbooks_rejects_shape_and_filter_mismatch(tmp_path: Path) -> None:
    sidecar_validation = _load_spike_module('sidecar_validation')
    manifest = {
        'sheets': [
            {
                'sheet_name': '成本计算单总表',
                'columns': ['文本', '金额'],
                'column_types': {'文本': 'text', '金额': 'amount'},
                'number_formats': {'金额': '#,##0.00'},
                'write_types': {'文本': 'text', '金额': 'number'},
                'freeze_panes': 'A2',
                'auto_filter': True,
                'row_count': 1,
                'column_count': 2,
            }
        ]
    }
    python_path = tmp_path / 'python.xlsx'
    rust_path = tmp_path / 'rust.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '成本计算单总表'
    worksheet.append(['文本', '金额'])
    worksheet.append(['A', 1.0])
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = 'A1:B2'
    worksheet['B2'].number_format = '#,##0.00'
    workbook.save(python_path)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '成本计算单总表'
    worksheet.append(['文本', '金额', '多余列'])
    worksheet.append(['A', 1.0, 'extra'])
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = 'A1:A2'
    worksheet['B2'].number_format = '#,##0.00'
    workbook.save(rust_path)

    report = sidecar_validation.validate_workbooks(python_path, rust_path, manifest)

    assert report['passed'] is False
    assert report['shape_matched'] is False
    assert report['auto_filter_matched'] is False
    assert report['error_count'] >= 2


def test_validate_workbooks_does_not_let_source_dtype_override_text_semantics(tmp_path: Path) -> None:
    sidecar_validation = _load_spike_module('sidecar_validation')
    manifest = {
        'sheets': [
            {
                'sheet_name': '成本计算单总表',
                'columns': ['工单行号'],
                'column_types': {'工单行号': 'text'},
                'source_dtypes': {'工单行号': 'Float64'},
                'number_formats': {},
                'freeze_panes': 'A2',
                'auto_filter': True,
                'row_count': 1,
                'column_count': 1,
            }
        ]
    }
    python_path = tmp_path / 'python.xlsx'
    rust_path = tmp_path / 'rust.xlsx'

    for path, value in ((python_path, '001'), (rust_path, 1)):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '成本计算单总表'
        worksheet.append(['工单行号'])
        worksheet.append([value])
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = 'A1:A2'
        workbook.save(path)

    report = sidecar_validation.validate_workbooks(python_path, rust_path, manifest)

    assert report['passed'] is False
    assert report['error_count'] == 1
    assert 'value mismatch' in report['errors'][0]


def test_classify_verdict_uses_sidecar_export_seconds_and_protocol_bottleneck() -> None:
    sidecar_benchmark = _load_spike_module('sidecar_benchmark')

    assert (
        sidecar_benchmark.classify_verdict(
            median_python_3sheet_export_seconds=10.0,
            median_intermediate_export_seconds=1.0,
            median_rust_export_seconds=3.0,
        )
        == 'VALIDATED'
    )
    assert (
        sidecar_benchmark.classify_verdict(
            median_python_3sheet_export_seconds=10.0,
            median_intermediate_export_seconds=4.5,
            median_rust_export_seconds=1.0,
        )
        == 'PARTIAL_PROTOCOL_BOTTLENECK'
    )
    assert (
        sidecar_benchmark.classify_verdict(
            median_python_3sheet_export_seconds=4.0,
            median_intermediate_export_seconds=1.0,
            median_rust_export_seconds=3.0,
        )
        == 'PARTIAL'
    )
    assert (
        sidecar_benchmark.classify_verdict(
            median_python_3sheet_export_seconds=20.0,
            median_intermediate_export_seconds=2.5,
            median_rust_export_seconds=4.0,
        )
        == 'PARTIAL'
    )
    assert (
        sidecar_benchmark.classify_verdict(
            median_python_3sheet_export_seconds=10.0,
            median_intermediate_export_seconds=4.5,
            median_rust_export_seconds=4.0,
        )
        == 'INVALIDATED'
    )


def test_benchmark_cli_returns_zero_only_for_validated(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    benchmark_cli = _load_spike_module('benchmark_rust_writer')

    monkeypatch.setattr(
        benchmark_cli,
        'run_benchmark',
        lambda **_kwargs: {'validation_passed': True, 'verdict': 'PARTIAL'},
    )
    assert benchmark_cli.main(['gb', '--tmp-dir', str(tmp_path / 'partial')]) == 1

    monkeypatch.setattr(
        benchmark_cli,
        'run_benchmark',
        lambda **_kwargs: {'validation_passed': True, 'verdict': 'VALIDATED'},
    )
    assert benchmark_cli.main(['gb', '--tmp-dir', str(tmp_path / 'validated')]) == 0


@pytest.mark.skipif(not _cargo_available(), reason='cargo is required for Rust writer integration tests')
def test_rust_writer_generates_valid_workbook_from_manifest(tmp_path: Path) -> None:
    sidecar_validation = _load_spike_module('sidecar_validation')
    csv_path = tmp_path / 'detail.csv'
    output_path = tmp_path / 'rust.xlsx'
    manifest_path = tmp_path / 'manifest.json'
    csv_path.write_text('月份,金额\n2025年01期,12.5\n', encoding='utf-8')
    manifest = {
        'version': 1,
        'workbook': {'output_path': str(output_path)},
        'sheets': [
            {
                'sheet_name': '成本计算单总表',
                'csv_path': str(csv_path),
                'columns': ['月份', '金额'],
                'column_types': {'月份': 'text', '金额': 'amount'},
                'number_formats': {'金额': '#,##0.00'},
                'write_types': {'月份': 'text', '金额': 'number'},
                'freeze_panes': 'A2',
                'auto_filter': True,
                'fixed_width': 15.0,
                'row_count': 1,
                'csv_export_method': 'polars',
            }
        ],
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False), encoding='utf-8')

    python_path = tmp_path / 'python.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '成本计算单总表'
    worksheet.append(['月份', '金额'])
    worksheet.append(['2025年01期', 12.5])
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = 'A1:B2'
    worksheet['B2'].number_format = '#,##0.00'
    workbook.save(python_path)

    rust_writer_dir = SPIKE_DIR / 'rust-writer'
    completed = subprocess.run(  # noqa: S603
        [  # noqa: S607
            'cargo',
            'run',
            '--quiet',
            '--manifest-path',
            str(rust_writer_dir / 'Cargo.toml'),
            '--',
            '--manifest',
            str(manifest_path),
            '--output',
            str(output_path),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    assert output_path.exists()
    report = sidecar_validation.validate_workbooks(python_path, output_path, manifest)
    assert report['passed'] is True, report['errors']


@pytest.mark.skipif(not _cargo_available(), reason='cargo is required for Rust writer integration tests')
def test_rust_writer_fails_when_number_column_cannot_parse(tmp_path: Path) -> None:
    csv_path = tmp_path / 'detail.csv'
    output_path = tmp_path / 'rust.xlsx'
    manifest_path = tmp_path / 'manifest.json'
    csv_path.write_text('月份,金额\n2025年01期,not-a-number\n', encoding='utf-8')
    manifest = {
        'version': 1,
        'workbook': {'output_path': str(output_path)},
        'sheets': [
            {
                'sheet_name': '成本计算单总表',
                'csv_path': str(csv_path),
                'columns': ['月份', '金额'],
                'column_types': {'月份': 'text', '金额': 'amount'},
                'number_formats': {'金额': '#,##0.00'},
                'write_types': {'月份': 'text', '金额': 'number'},
                'freeze_panes': 'A2',
                'auto_filter': True,
                'fixed_width': 15.0,
                'row_count': 1,
                'column_count': 2,
                'csv_export_method': 'polars',
            }
        ],
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False), encoding='utf-8')
    rust_writer_dir = SPIKE_DIR / 'rust-writer'

    completed = subprocess.run(  # noqa: S603
        [  # noqa: S607
            'cargo',
            'run',
            '--quiet',
            '--manifest-path',
            str(rust_writer_dir / 'Cargo.toml'),
            '--',
            '--manifest',
            str(manifest_path),
            '--output',
            str(output_path),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode != 0
    assert 'failed to parse number' in completed.stderr
    assert '金额' in completed.stderr
    assert not output_path.exists()
