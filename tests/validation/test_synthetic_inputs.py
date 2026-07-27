from pathlib import Path

from openpyxl import load_workbook

from tools.validation.synthetic_inputs import RAW_COLUMNS, build_raw_fixture


def test_synthetic_inputs_cover_both_pipeline_specific_cost_items(tmp_path: Path) -> None:
    for pipeline, expected_item in (('gb', '委外加工费'), ('sk', '软件费用')):
        path = tmp_path / f'{pipeline}.xlsx'

        build_raw_fixture(path, pipeline, 'small')

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            assert tuple(cell.value for cell in sheet[3]) == RAW_COLUMNS
            values = {row[8].value for row in sheet.iter_rows(min_row=5)}
            assert expected_item in values
        finally:
            workbook.close()
