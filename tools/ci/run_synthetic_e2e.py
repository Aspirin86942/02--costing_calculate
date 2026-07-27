"""Run public GB/SK smoke tests against generated, synthetic-only workbooks."""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

from tools.validation.synthetic_inputs import build_raw_fixture  # noqa: E402
from tools.validation.types import PipelineName  # noqa: E402

WORKBOOK_BASELINE = PROJECT_ROOT / 'tests' / 'contracts' / 'baselines' / 'workbook_semantics.json'
PIPELINES: tuple[PipelineName, ...] = ('gb', 'sk')


def load_expected_sheet_order() -> tuple[str, ...]:
    """Read the single frozen workbook-contract source of truth."""
    baseline = json.loads(WORKBOOK_BASELINE.read_text(encoding='utf-8'))
    sheet_order = baseline.get('default_workbook', {}).get('sheet_order')
    if not isinstance(sheet_order, list) or not sheet_order or not all(isinstance(name, str) for name in sheet_order):
        raise RuntimeError(f'invalid default_workbook.sheet_order in {WORKBOOK_BASELINE}')
    return tuple(sheet_order)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--binary', type=Path, required=True)
    return parser.parse_args()


def _run(binary: Path, pipeline: PipelineName, cwd: Path, *extra: str) -> subprocess.CompletedProcess[str]:
    # `binary` is resolved and verified as a local file by main before this helper is called.
    return subprocess.run(  # noqa: S603
        [str(binary), pipeline, *extra],
        cwd=cwd,
        check=False,
        capture_output=True,
        text=True,
        encoding='utf-8',
    )


def _assert_success(
    result: subprocess.CompletedProcess[str],
    pipeline: PipelineName,
    expected_sheet_count: int,
) -> dict[str, object]:
    if result.returncode != 0:
        raise RuntimeError(f'{pipeline} synthetic run failed with {result.returncode}: {result.stderr.strip()}')
    payload = json.loads(result.stdout)
    if payload.get('status') != 'succeeded' or payload.get('pipeline') != pipeline:
        raise RuntimeError(f'{pipeline} returned an unexpected RunSummary: {payload}')
    if payload.get('sheet_count') != expected_sheet_count:
        raise RuntimeError(f'{pipeline} returned an unexpected sheet count: {payload}')
    return payload


def _assert_config_valid(result: subprocess.CompletedProcess[str], pipeline: PipelineName) -> None:
    if result.returncode != 0:
        raise RuntimeError(f'{pipeline} embedded config validation failed: {result.stderr.strip()}')
    payload = json.loads(result.stdout)
    if (
        payload.get('status') != 'valid'
        or payload.get('pipeline') != pipeline
        or payload.get('schema_version') != 1
        or payload.get('source') != 'embedded-default'
    ):
        raise RuntimeError(f'{pipeline} returned an unexpected config validation record: {payload}')


def _exercise_pipeline(
    binary: Path,
    root: Path,
    pipeline: PipelineName,
    expected_sheets: tuple[str, ...],
) -> None:
    _assert_config_valid(_run(binary, pipeline, root, '--validate-config'), pipeline)
    input_path = root / 'data' / 'raw' / pipeline / f'{pipeline}-synthetic.xlsx'
    build_raw_fixture(input_path, pipeline, 'small')

    check_summary = _assert_success(
        _run(binary, pipeline, root, '--check-only'),
        pipeline,
        len(expected_sheets),
    )
    if check_summary.get('output_written') is not False:
        raise RuntimeError(f'{pipeline} check-only unexpectedly wrote an output')

    normal_summary = _assert_success(_run(binary, pipeline, root), pipeline, len(expected_sheets))
    output_path_text = normal_summary.get('workbook_path')
    if not isinstance(output_path_text, str):
        raise RuntimeError(f'{pipeline} did not report a workbook path')
    output_path = Path(output_path_text)
    if not output_path.is_absolute():
        output_path = root / output_path
    if not output_path.is_file() or output_path.stat().st_size == 0:
        raise RuntimeError(f'{pipeline} did not create a non-empty workbook')

    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        if tuple(workbook.sheetnames) != expected_sheets:
            raise RuntimeError(f'{pipeline} workbook sheets differ: {workbook.sheetnames}')
    finally:
        workbook.close()

    second_run = _run(binary, pipeline, root)
    if second_run.returncode == 0:
        raise RuntimeError(f'{pipeline} unexpectedly overwrote an existing workbook')
    error_payload = json.loads(second_run.stderr)
    if error_payload.get('code') != 'OUTPUT_EXISTS':
        raise RuntimeError(f'{pipeline} returned an unexpected overwrite error: {error_payload}')


def main() -> int:
    args = _parse_args()
    binary = args.binary.resolve()
    if not binary.is_file():
        raise FileNotFoundError(binary)

    pytest_tmp = PROJECT_ROOT / '.pytest-tmp'
    pytest_tmp.mkdir(exist_ok=True)
    with tempfile.TemporaryDirectory(prefix='synthetic-e2e-', dir=pytest_tmp) as temporary:
        root = Path(temporary)
        expected_sheets = load_expected_sheet_order()
        for pipeline in PIPELINES:
            _exercise_pipeline(binary, root, pipeline, expected_sheets)
    print('synthetic GB/SK config, check-only, workbook, and no-overwrite smokes passed')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
